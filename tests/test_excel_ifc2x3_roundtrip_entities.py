import pandas as pd
import ifcopenshell
from ifcopenshell.guid import new as new_guid
from openpyxl import load_workbook

from app import extract_to_excel, update_ifc_from_excel


def _build_ifc2x3_model(tmp_path):
    model = ifcopenshell.file(schema="IFC2X3")
    project = model.create_entity("IfcProject", GlobalId=new_guid(), Name="P", LongName="PN-001")
    site = model.create_entity("IfcSite", GlobalId=new_guid(), Name="S")
    building = model.create_entity("IfcBuilding", GlobalId=new_guid(), Name="B")
    model.create_entity("IfcRelAggregates", GlobalId=new_guid(), RelatingObject=project, RelatedObjects=[site])
    model.create_entity("IfcRelAggregates", GlobalId=new_guid(), RelatingObject=site, RelatedObjects=[building])

    proxy = model.create_entity("IfcBuildingElementProxy", GlobalId=new_guid(), Name="ProxyOcc")
    proxy_type = model.create_entity("IfcBuildingElementProxyType", GlobalId=new_guid(), Name="ProxyType", PredefinedType="NOTDEFINED")
    wall_type = model.create_entity("IfcWallType", GlobalId=new_guid(), Name="WallType", PredefinedType="NOTDEFINED")
    model.create_entity("IfcRelDefinesByType", GlobalId=new_guid(), RelatedObjects=[proxy], RelatingType=proxy_type)
    return model, proxy, proxy_type, wall_type


def test_export_includes_types_raw_entities_and_dropdowns(tmp_path):
    model, proxy, proxy_type, _ = _build_ifc2x3_model(tmp_path)
    src = tmp_path / "src.ifc"
    xlsx = tmp_path / "out.xlsx"
    model.write(str(src))

    extract_to_excel(str(src), str(xlsx))

    xls = pd.ExcelFile(xlsx)
    assert "Types" in xls.sheet_names
    assert "RawEntities" in xls.sheet_names

    elements = pd.read_excel(xls, "Elements")
    types = pd.read_excel(xls, "Types")
    raw = pd.read_excel(xls, "RawEntities")

    assert "IfcBuildingElementProxy" in set(elements["CurrentEntity"])
    assert proxy.GlobalId in set(elements["GlobalId"])
    assert "IfcBuildingElementProxyType" in set(types["CurrentEntity"])
    assert proxy_type.GlobalId in set(types["GlobalId"])
    assert "RawStepLine" in raw.columns
    for required_col in ("ExpressLine", "GlobalId", "IfcEntity", "PredefinedType", "Name", "ObjectType", "TypeName", "SourceFile"):
        assert required_col in elements.columns

    wb = load_workbook(str(xlsx))
    assert wb["_Lookups_IFC2X3_Entities"].sheet_state == "hidden"
    assert wb["_Lookups_IFC2X3_Predefs"].sheet_state == "hidden"
    assert wb["_Lookups_IFC2X3_Map"].sheet_state == "hidden"


def test_workbook_lookup_validations_roundtrip_safe(tmp_path):
    model, _, _, _ = _build_ifc2x3_model(tmp_path)
    src = tmp_path / "src.ifc"
    xlsx = tmp_path / "out.xlsx"
    rewritten = tmp_path / "rewritten.xlsx"
    model.write(str(src))

    extract_to_excel(str(src), str(xlsx))

    wb = load_workbook(str(xlsx))
    assert {"Elements", "Properties", "_Lookups_IFC2X3_Entities", "_Lookups_IFC2X3_Predefs", "_Lookups_IFC2X3_Map"}.issubset(set(wb.sheetnames))
    assert wb.defined_names.get("IfcEntityList") is not None
    assert wb.defined_names.get("PredefinedTypeList") is not None

    elements_ws = wb["Elements"]
    validations = list(elements_ws.data_validations.dataValidation)
    formulas = {dv.formula1 for dv in validations}
    assert "=IfcEntityList" in formulas
    assert "=PredefinedTypeList" in formulas
    assert all(dv.sqref for dv in validations)

    wb.save(str(rewritten))


def test_reupload_predefined_only_and_valid_type_reclassification(tmp_path):
    model, _, proxy_type, _ = _build_ifc2x3_model(tmp_path)
    src = tmp_path / "src.ifc"
    xlsx = tmp_path / "out.xlsx"
    updated = tmp_path / "updated.ifc"
    model.write(str(src))
    extract_to_excel(str(src), str(xlsx))

    xls = pd.ExcelFile(xlsx)
    elements = pd.read_excel(xls, "Elements")
    types = pd.read_excel(xls, "Types")
    props = pd.read_excel(xls, "Properties")
    cobie = pd.read_excel(xls, "COBieMapping")
    project = pd.read_excel(xls, "ProjectData")
    raw_entities = pd.read_excel(xls, "RawEntities")
    uniclass_pr = pd.read_excel(xls, "Uniclass_Pr")
    uniclass_ss = pd.read_excel(xls, "Uniclass_Ss")
    uniclass_ef = pd.read_excel(xls, "Uniclass_EF")
    xls.close()

    wall_type_row = types[types["CurrentEntity"] == "IfcWallType"].iloc[0]
    types.loc[types["RowKey"] == wall_type_row["RowKey"], "TargetPredefinedType"] = "USERDEFINED"
    types.loc[types["RowKey"] == wall_type_row["RowKey"], "ApplyChange"] = "Yes"

    type_row = types[types["CurrentEntity"] == "IfcBuildingElementProxyType"].iloc[0]
    types.loc[types["RowKey"] == type_row["RowKey"], "TargetEntity"] = "IfcWallType"
    types.loc[types["RowKey"] == type_row["RowKey"], "ApplyChange"] = "Yes"

    with pd.ExcelWriter(xlsx, engine="openpyxl") as writer:
        project.to_excel(writer, sheet_name="ProjectData", index=False)
        elements.to_excel(writer, sheet_name="Elements", index=False)
        types.to_excel(writer, sheet_name="Types", index=False)
        raw_entities.to_excel(writer, sheet_name="RawEntities", index=False)
        props.to_excel(writer, sheet_name="Properties", index=False)
        cobie.to_excel(writer, sheet_name="COBieMapping", index=False)
        uniclass_pr.to_excel(writer, sheet_name="Uniclass_Pr", index=False)
        uniclass_ss.to_excel(writer, sheet_name="Uniclass_Ss", index=False)
        uniclass_ef.to_excel(writer, sheet_name="Uniclass_EF", index=False)

    update_ifc_from_excel(str(src), str(xlsx), str(updated), add_new="yes")
    revised = ifcopenshell.open(str(updated))

    updated_proxy_type = revised.by_guid(proxy_type.GlobalId)
    assert updated_proxy_type.is_a() == "IfcWallType"

    updated_wall_type = revised.by_guid(wall_type_row["GlobalId"])
    assert updated_wall_type.PredefinedType == "USERDEFINED"


def test_reject_invalid_entity_predefined_combo(tmp_path):
    model, _, proxy_type, _ = _build_ifc2x3_model(tmp_path)
    src = tmp_path / "src.ifc"
    xlsx = tmp_path / "out.xlsx"
    updated = tmp_path / "updated.ifc"
    model.write(str(src))
    extract_to_excel(str(src), str(xlsx))

    xls = pd.ExcelFile(xlsx)
    elements = pd.read_excel(xls, "Elements")
    types = pd.read_excel(xls, "Types")
    project = pd.read_excel(xls, "ProjectData")
    raw_entities = pd.read_excel(xls, "RawEntities")
    props = pd.read_excel(xls, "Properties")
    cobie = pd.read_excel(xls, "COBieMapping")
    uniclass_pr = pd.read_excel(xls, "Uniclass_Pr")
    uniclass_ss = pd.read_excel(xls, "Uniclass_Ss")
    uniclass_ef = pd.read_excel(xls, "Uniclass_EF")
    xls.close()

    type_row = types[types["CurrentEntity"] == "IfcBuildingElementProxyType"].iloc[0]
    types.loc[types["RowKey"] == type_row["RowKey"], "TargetEntity"] = "IfcWall"  # type -> occurrence
    types.loc[types["RowKey"] == type_row["RowKey"], "TargetPredefinedType"] = "NOT_A_VALID_TYPE"
    types.loc[types["RowKey"] == type_row["RowKey"], "ApplyChange"] = "Yes"

    with pd.ExcelWriter(xlsx, engine="openpyxl") as writer:
        project.to_excel(writer, sheet_name="ProjectData", index=False)
        elements.to_excel(writer, sheet_name="Elements", index=False)
        types.to_excel(writer, sheet_name="Types", index=False)
        raw_entities.to_excel(writer, sheet_name="RawEntities", index=False)
        props.to_excel(writer, sheet_name="Properties", index=False)
        cobie.to_excel(writer, sheet_name="COBieMapping", index=False)
        uniclass_pr.to_excel(writer, sheet_name="Uniclass_Pr", index=False)
        uniclass_ss.to_excel(writer, sheet_name="Uniclass_Ss", index=False)
        uniclass_ef.to_excel(writer, sheet_name="Uniclass_EF", index=False)

    update_ifc_from_excel(str(src), str(xlsx), str(updated), add_new="yes")
    log_df = pd.read_excel(xlsx, sheet_name="ChangeLog")
    assert any("type to occurrence" in str(msg) for msg in log_df["Message"].tolist())

    revised = ifcopenshell.open(str(updated))
    assert revised.by_guid(proxy_type.GlobalId).is_a() == "IfcBuildingElementProxyType"
