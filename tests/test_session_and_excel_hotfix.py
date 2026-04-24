import asyncio
import io
from pathlib import Path

import ifcopenshell
from ifcopenshell.guid import new as new_guid
from fastapi import UploadFile
from openpyxl import load_workbook

import app


def _upload(name: str, content: bytes) -> UploadFile:
    return UploadFile(filename=name, file=io.BytesIO(content))


def _write_min_ifc(path: Path, schema: str = "IFC4") -> None:
    model = ifcopenshell.file(schema=schema)
    model.create_entity("IfcProject", GlobalId=new_guid(), Name="Proj")
    model.write(str(path))


def test_session_route_contract_roundtrip():
    payload = app.create_session({})
    session_id = payload["session_id"]

    info = app.get_session(session_id)
    assert info["session_id"] == session_id

    uploaded = asyncio.run(app.upload_files(session_id, [_upload("sample.ifc", b"ISO-10303-21;\n")]))
    assert uploaded["files"][0]["id"] == "sample.ifc"

    listing = app.list_files(session_id)
    assert listing["files"]
    first = listing["files"][0]
    for key in ("id", "filename", "size", "mime_type", "extension", "created_at", "path"):
        assert key in first

    deleted = app.delete_session_file(session_id, "sample.ifc")
    assert deleted["deleted"] is True


def test_ifc2x3_schema_detected_from_file_schema_header(tmp_path):
    ifc_path = tmp_path / "model.ifc"
    _write_min_ifc(ifc_path, schema="IFC2X3")

    detected, warning = app.detect_ifc_schema_from_header(str(ifc_path))
    assert detected == "IFC2X3"
    assert warning == ""


def test_excel_export_reopen_and_lookup_sheet_for_ifc4(tmp_path):
    src = tmp_path / "source.ifc"
    out = tmp_path / "export.xlsx"
    _write_min_ifc(src, schema="IFC4")

    result = app.extract_to_excel(str(src), str(out))
    assert result["schema_detected"].startswith("IFC4")

    wb = load_workbook(str(out))
    assert "_Lookups_IFC4_Entities" in wb.sheetnames
    assert wb.defined_names.get("IfcEntityList") is not None
    wb.close()
