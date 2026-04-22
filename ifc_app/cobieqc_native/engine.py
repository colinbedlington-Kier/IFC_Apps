import logging
import os
import shlex
import shutil
import subprocess
import hashlib
from datetime import date, datetime
from collections import Counter
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel
from xml.etree import ElementTree as ET

LOGGER = logging.getLogger("ifc_app.cobieqc.native")
APP_ROOT = Path(__file__).resolve().parents[2]
COBIE_ENTITY_NAMES = ["Contact", "Facility", "Floor", "Space", "Zone", "Type", "Component"]
KEY_FIELD_DIAGNOSTIC_FIELDS: Dict[str, List[str]] = {
    "Contact": ["Email", "CreatedBy", "CreatedOn"],
    "Facility": ["Name", "CreatedOn", "Currency"],
    "Floor": ["Name", "CreatedOn"],
    "Space": ["Name", "FloorName", "CreatedOn"],
    "Type": ["Name", "CreatedBy", "Manufacturer", "WarrantyGuarantorParts", "WarrantyGuarantorLabor"],
    "Component": ["Name", "TypeName", "Space", "CreatedBy"],
}
KEY_FIELD_SOURCE_ALIASES: Dict[str, List[str]] = {
    "Email": ["Email", "EMail", "ContactEmail"],
    "Name": ["Name", "Component", "Space", "Type", "Floor", "Facility"],
    "CreatedBy": ["CreatedBy", "Created By"],
    "CreatedOn": ["CreatedOn", "Created On"],
    "FloorName": ["FloorName", "Floor"],
    "TypeName": ["TypeName", "Type"],
    "Space": ["Space", "SpaceName"],
}
DESKTOP_IDENTITY_COLUMN_ALIASES: Dict[str, List[str]] = {
    "Contact": ["Email", "Name"],
    "Facility": ["Name", "Facility"],
    "Floor": ["Name", "FloorName", "Floor"],
    "Space": ["Name", "SpaceName", "Space"],
    "Type": ["Name", "TypeName", "Type"],
    "Component": ["Name", "ComponentName", "Component"],
}


def _clean_tag(value: str) -> str:
    cleaned = "".join(ch if ch.isalnum() else "_" for ch in str(value).strip())
    return cleaned or "Column"


def _xml_escape(value: Any) -> str:
    text = "" if value is None else str(value)
    return (
        text.replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
        .replace('"', "&quot;")
        .replace("'", "&apos;")
    )


def _local_name(tag: str) -> str:
    if "}" in tag:
        return tag.rsplit("}", 1)[-1]
    return tag


def _namespace_uri(tag: str) -> str:
    if tag.startswith("{") and "}" in tag:
        return tag[1:].split("}", 1)[0]
    return ""


def _safe_iterparse_namespaces(xml_path: Path) -> Dict[str, str]:
    namespaces: Dict[str, str] = {}
    try:
        for _, payload in ET.iterparse(str(xml_path), events=("start-ns",)):
            prefix, uri = payload
            namespaces[prefix or "(default)"] = uri
    except Exception:
        return {}
    return namespaces


def _build_element_parent_map(root: ET.Element) -> Dict[ET.Element, Optional[ET.Element]]:
    parent_map: Dict[ET.Element, Optional[ET.Element]] = {root: None}
    for parent in root.iter():
        for child in list(parent):
            parent_map[child] = parent
    return parent_map


def _element_path(element: ET.Element, parent_map: Dict[ET.Element, Optional[ET.Element]]) -> str:
    parts = []
    cursor: Optional[ET.Element] = element
    while cursor is not None:
        parts.append(_local_name(cursor.tag))
        cursor = parent_map.get(cursor)
    return "/" + "/".join(reversed(parts))


def _find_reference_xml_path(out_dir: Path) -> Optional[Path]:
    configured = os.getenv("COBIEQC_REFERENCE_XML_PATH", "").strip()
    if configured:
        candidate = Path(configured).expanduser().resolve()
        if candidate.exists() and candidate.is_file():
            return candidate
    for filename in ("legacy_generated_cobie.xml", "legacy_cobie.xml", "known_good_cobie.xml", "bsn.xml"):
        candidate = out_dir / filename
        if candidate.exists() and candidate.is_file():
            return candidate
    return None


def _find_reference_svrl_path(out_dir: Path) -> Optional[Path]:
    configured = os.getenv("COBIEQC_REFERENCE_SVRL_PATH", "").strip()
    if configured:
        candidate = Path(configured).expanduser().resolve()
        if candidate.exists() and candidate.is_file():
            return candidate
    for filename in ("legacy_validation_result.svrl.xml", "reference_validation_result.svrl.xml", "legacy.svrl.xml"):
        candidate = out_dir / filename
        if candidate.exists() and candidate.is_file():
            return candidate
    return None


def _inspect_cobie_xml(cobie_xml_path: Path) -> Dict[str, Any]:
    doc = ET.parse(str(cobie_xml_path))
    root = doc.getroot()
    parent_map = _build_element_parent_map(root)
    first_level = [_local_name(child.tag) for child in list(root)]
    lower_names = {name.lower() for name in COBIE_ENTITY_NAMES}
    counts: Dict[str, int] = {}
    sample_paths: Dict[str, List[str]] = {}
    sheet_to_rows: Dict[str, int] = {}
    created_on_samples: List[str] = []
    for entity_name in COBIE_ENTITY_NAMES:
        needle = entity_name.lower()
        count = 0
        samples: List[str] = []
        for node in root.iter():
            node_name = _local_name(node.tag).lower()
            if node_name == needle or node_name == f"{needle}s":
                count += 1
                if len(samples) < 3:
                    samples.append(_element_path(node, parent_map))
        counts[entity_name] = count
        sample_paths[entity_name] = samples
    for node in root.iter():
        source_sheet = str(node.attrib.get("sourceSheet", "")).strip()
        if source_sheet:
            sheet_to_rows[source_sheet] = sheet_to_rows.get(source_sheet, 0) + 1
        if _local_name(node.tag).lower() == "createdon":
            text = (node.text or "").strip()
            if text and len(created_on_samples) < 10:
                created_on_samples.append(text)
    model_path_count = sum(1 for node in root.iter() if _local_name(node.tag).lower() in lower_names)
    cross_ref = _cross_reference_created_by(root)
    return {
        "root_element_name": _local_name(root.tag),
        "root_namespace": _namespace_uri(root.tag) or "(none)",
        "first_level_children": first_level,
        "entity_counts": counts,
        "sample_paths": sample_paths,
        "sheet_row_counts": sheet_to_rows,
        "model_path_count": model_path_count,
        "created_on_samples": created_on_samples,
        "cross_reference": cross_ref,
    }


def _find_child_text(parent: ET.Element, child_name: str) -> str:
    needle = child_name.strip().lower()
    for child in list(parent):
        if _local_name(child.tag).strip().lower() == needle:
            return (child.text or "").strip()
    return ""


def _sheet_records_by_row(workbook_path: Path) -> Dict[str, Dict[int, Dict[str, Any]]]:
    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    records: Dict[str, Dict[int, Dict[str, Any]]] = {}
    for sheet in wb.worksheets:
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            continue
        headers: List[str] = []
        for value in rows[0]:
            text = "" if value is None else str(value).strip()
            headers.append(text)
        row_map: Dict[int, Dict[str, Any]] = {}
        for row_idx, row in enumerate(rows[1:], start=2):
            if all(cell in (None, "") for cell in row):
                continue
            row_map[row_idx] = {headers[col_idx]: row[col_idx] for col_idx in range(min(len(headers), len(row)))}
        records[sheet.title] = row_map
    return records


def _resolve_source_column_name(source_row: Dict[str, Any], field_name: str) -> str:
    candidates = KEY_FIELD_SOURCE_ALIASES.get(field_name, [field_name])
    for candidate in candidates:
        for column_name in source_row.keys():
            if column_name.strip().lower() == candidate.strip().lower():
                return column_name
    return ""


def _resolve_source_value_by_aliases(source_row: Dict[str, Any], aliases: List[str]) -> Tuple[str, str]:
    for alias in aliases:
        col = _resolve_source_column_name(source_row, alias)
        if not col:
            continue
        value = source_row.get(col)
        if value is None:
            return col, ""
        return col, str(value).strip()
    return "", ""


def _entity_nodes(root: ET.Element, entity_name: str) -> List[ET.Element]:
    plural = f"{entity_name[:-1]}ies" if entity_name.endswith("y") else f"{entity_name}s"
    return root.findall(f".//{{*}}{plural}/{{*}}{entity_name}")


def _name_profile(entity_nodes: List[ET.Element]) -> Dict[str, Any]:
    names: List[Tuple[str, str, str]] = []
    blank_samples: List[str] = []
    for node in entity_nodes:
        name_value = _find_child_text(node, "Name")
        row_ref = f"{node.attrib.get('sourceSheet', '')}#{node.attrib.get('rowNumber', '')}"
        if not name_value and len(blank_samples) < 10:
            blank_samples.append(row_ref)
        names.append((name_value, row_ref, node.attrib.get("sourceSheet", "")))
    duplicates = [name for name, count in Counter([name for name, _, _ in names if name]).items() if count > 1]
    duplicate_samples: List[str] = []
    if duplicates:
        dup_set = set(duplicates)
        for name, row_ref, _ in names:
            if name in dup_set and len(duplicate_samples) < 10:
                duplicate_samples.append(f"{name}@{row_ref}")
    return {
        "blank_name_count": sum(1 for name, _, _ in names if not name),
        "duplicate_name_count": len(duplicates),
        "blank_name_samples": blank_samples,
        "duplicate_name_samples": duplicate_samples,
    }


def _cross_reference_diagnostics(root: ET.Element) -> Dict[str, Any]:
    floor_names = {_find_child_text(node, "Name") for node in _entity_nodes(root, "Floor")}
    type_names = {_find_child_text(node, "Name") for node in _entity_nodes(root, "Type")}
    space_names = {_find_child_text(node, "Name") for node in _entity_nodes(root, "Space")}
    contact_names = {
        _find_child_text(node, "Email") or _find_child_text(node, "Name") for node in _entity_nodes(root, "Contact")
    }
    floor_names.discard("")
    type_names.discard("")
    space_names.discard("")
    contact_names.discard("")

    def _collect_unresolved(nodes: List[ET.Element], field_name: str, valid_values: set[str]) -> Dict[str, Any]:
        unresolved: List[str] = []
        for node in nodes:
            value = _find_child_text(node, field_name)
            if value and value in valid_values:
                continue
            if not value:
                unresolved.append(f"(blank)@{node.attrib.get('sourceSheet', '')}#{node.attrib.get('rowNumber', '')}")
            else:
                unresolved.append(f"{value}@{node.attrib.get('sourceSheet', '')}#{node.attrib.get('rowNumber', '')}")
        return {"count": len(unresolved), "samples": unresolved[:20]}

    created_by_unresolved: List[str] = []
    for entity in ("Facility", "Floor", "Space", "Type", "Component", "Zone"):
        for node in _entity_nodes(root, entity):
            created_by = _find_child_text(node, "CreatedBy")
            if created_by and created_by in contact_names:
                continue
            created_by_unresolved.append(
                f"{created_by or '(blank)'}@{node.attrib.get('sourceSheet', '')}#{node.attrib.get('rowNumber', '')}"
            )

    return {
        "space_floor_name": _collect_unresolved(_entity_nodes(root, "Space"), "FloorName", floor_names),
        "component_type_name": _collect_unresolved(_entity_nodes(root, "Component"), "TypeName", type_names),
        "component_space": _collect_unresolved(_entity_nodes(root, "Component"), "Space", space_names),
        "created_by": {"count": len(created_by_unresolved), "samples": created_by_unresolved[:20]},
    }


def _build_key_field_diagnostics(workbook_path: Path, cobie_xml_path: Path) -> Dict[str, Any]:
    root = ET.parse(str(cobie_xml_path)).getroot()
    source_rows = _sheet_records_by_row(workbook_path)
    entity_diagnostics: Dict[str, Any] = {}
    for entity, fields in KEY_FIELD_DIAGNOSTIC_FIELDS.items():
        rows: List[Dict[str, Any]] = []
        for node in _entity_nodes(root, entity):
            source_sheet = str(node.attrib.get("sourceSheet", ""))
            row_number = int(str(node.attrib.get("rowNumber", "0") or "0"))
            source_row = source_rows.get(source_sheet, {}).get(row_number, {})
            field_values: Dict[str, Dict[str, Any]] = {}
            for field in fields:
                xml_value = _find_child_text(node, field)
                source_column = _resolve_source_column_name(source_row, field)
                source_value = source_row.get(source_column) if source_column else None
                field_values[field] = {
                    "xml_value": xml_value,
                    "source_column": source_column or "(missing)",
                    "source_value": "" if source_value is None else str(source_value).strip(),
                }
            identity_aliases = DESKTOP_IDENTITY_COLUMN_ALIASES.get(entity, ["Name"])
            source_identity_column, source_identity_value = _resolve_source_value_by_aliases(source_row, identity_aliases)
            xml_identity_value = ""
            for alias in identity_aliases:
                xml_identity_value = _find_child_text(node, alias)
                if xml_identity_value:
                    break
            rows.append(
                {
                    "source_sheet": source_sheet or "(unknown)",
                    "source_row_number": row_number,
                    "fields": field_values,
                    "identity": {
                        "aliases": identity_aliases,
                        "source_column": source_identity_column or "(missing)",
                        "source_value": source_identity_value,
                        "xml_value": xml_identity_value,
                        "matches": source_identity_value == xml_identity_value,
                    },
                }
            )
        entity_diagnostics[entity] = {
            "row_count": len(rows),
            "rows": rows,
            "samples": rows[:20],
            **_name_profile(_entity_nodes(root, entity)),
        }
    cross_references = _cross_reference_diagnostics(root)
    mismatch_report = {
        "floor_name": cross_references["space_floor_name"]["samples"][:20],
        "type_name": cross_references["component_type_name"]["samples"][:20],
        "space": cross_references["component_space"]["samples"][:20],
        "created_by": cross_references["created_by"]["samples"][:20],
        "blank_name_rows": [],
    }
    for entity, profile in entity_diagnostics.items():
        for sample in profile.get("samples", []):
            name_data = sample.get("fields", {}).get("Name")
            if name_data and not name_data.get("xml_value"):
                mismatch_report["blank_name_rows"].append(
                    f"{entity}:{sample.get('source_sheet', '')}#{sample.get('source_row_number', 0)}"
                )
                if len(mismatch_report["blank_name_rows"]) >= 20:
                    break
        if len(mismatch_report["blank_name_rows"]) >= 20:
            break
    return {"entities": entity_diagnostics, "cross_references": cross_references, "mismatch_report": mismatch_report}


def _cross_reference_created_by(root: ET.Element) -> Dict[str, Any]:
    contact_names = {
        (name_node.text or "").strip()
        for name_node in root.findall(".//{*}Contacts/{*}Contact/{*}Email")
        if (name_node.text or "").strip()
    }
    if not contact_names:
        contact_names = {
            (name_node.text or "").strip()
            for name_node in root.findall(".//{*}Contacts/{*}Contact/{*}Name")
            if (name_node.text or "").strip()
        }
    unmatched: List[str] = []
    total = 0
    matched = 0
    for parent in root.iter():
        for child in list(parent):
            if _local_name(child.tag).lower() != "createdby":
                continue
            total += 1
            created_by = (child.text or "").strip()
            if created_by and created_by in contact_names:
                matched += 1
                continue
            parent_sheet = parent.attrib.get("sourceSheet", "")
            parent_row = parent.attrib.get("rowNumber", "")
            unmatched.append(f"{created_by or '(blank)'}@{parent_sheet}#{parent_row}")
    return {
        "contact_name_count": len(contact_names),
        "created_by_total": total,
        "created_by_matched": matched,
        "created_by_unmatched": max(total - matched, 0),
        "unmatched_samples": unmatched[:20],
    }


def _collect_xml_structure_snapshot(xml_path: Path) -> Dict[str, Any]:
    doc = ET.parse(str(xml_path))
    root = doc.getroot()
    parent_map = _build_element_parent_map(root)
    path_counter: Counter[str] = Counter()
    frequency: Counter[str] = Counter()
    for node in root.iter():
        name = _local_name(node.tag)
        frequency[name] += 1
        path_counter[_element_path(node, parent_map)] += 1
    return {
        "root_name": _local_name(root.tag),
        "root_namespace": _namespace_uri(root.tag) or "(none)",
        "namespaces": _safe_iterparse_namespaces(xml_path),
        "element_frequency": dict(frequency),
        "xpath_inventory": dict(path_counter),
    }


def _compare_xml_structure(generated_xml_path: Path, reference_xml_path: Path) -> Dict[str, Any]:
    generated = _collect_xml_structure_snapshot(generated_xml_path)
    reference = _collect_xml_structure_snapshot(reference_xml_path)

    generated_paths = set(generated["xpath_inventory"].keys())
    reference_paths = set(reference["xpath_inventory"].keys())
    generated_elements = set(generated["element_frequency"].keys())
    reference_elements = set(reference["element_frequency"].keys())

    missing_paths = sorted(reference_paths - generated_paths)
    extra_paths = sorted(generated_paths - reference_paths)
    missing_elements = sorted(reference_elements - generated_elements)
    extra_elements = sorted(generated_elements - reference_elements)
    frequency_deltas: List[str] = []
    for key in sorted(generated_elements | reference_elements):
        gen_count = int(generated["element_frequency"].get(key, 0))
        ref_count = int(reference["element_frequency"].get(key, 0))
        if gen_count != ref_count:
            frequency_deltas.append(f"{key}:generated={gen_count},reference={ref_count}")

    return {
        "reference_path": str(reference_xml_path),
        "generated_root": generated["root_name"],
        "reference_root": reference["root_name"],
        "generated_namespace": generated["root_namespace"],
        "reference_namespace": reference["root_namespace"],
        "generated_namespaces": generated["namespaces"],
        "reference_namespaces": reference["namespaces"],
        "missing_paths": missing_paths[:100],
        "extra_paths": extra_paths[:100],
        "missing_elements": missing_elements[:100],
        "extra_elements": extra_elements[:100],
        "frequency_deltas": frequency_deltas[:100],
    }


def _inspect_svrl(svrl_output_path: Path) -> Dict[str, Any]:
    root = ET.parse(str(svrl_output_path)).getroot()
    active_patterns = root.findall(".//{*}active-pattern")
    active_pattern_ids = [str(node.attrib.get("id", "")).strip() for node in active_patterns if node.attrib.get("id")]
    fired_rules = root.findall(".//{*}fired-rule")
    failed_asserts = root.findall(".//{*}failed-assert")
    successful_reports = root.findall(".//{*}successful-report")
    diagnostics = root.findall(".//{*}diagnostic-reference")
    failed_assert_samples: List[Dict[str, Any]] = []
    failed_assert_counts: Counter[str] = Counter()
    failed_assert_rows: Counter[str] = Counter()
    for node in failed_asserts:
        rule_id = str(node.attrib.get("id", "")).strip() or str(node.attrib.get("test", "")).strip() or "(unknown)"
        failed_assert_counts[rule_id] += 1
        location = str(node.attrib.get("location", "")).strip() or "(unknown)"
        failed_assert_rows[location] += 1
        if len(failed_assert_samples) < 5:
            text_value = ""
            diagnostics_values: List[str] = []
            for child in list(node):
                child_name = _local_name(child.tag)
                if child_name == "text":
                    text_value = (child.text or "").strip()
                if child_name == "diagnostic-reference":
                    diag_text = "".join(child.itertext()).strip()
                    diag_id = str(child.attrib.get("diagnostic", "")).strip()
                    diagnostics_values.append(f"{diag_id}:{diag_text}" if diag_id else diag_text)
            failed_assert_samples.append(
                {
                    "id": str(node.attrib.get("id", "")).strip() or "(none)",
                    "location": str(node.attrib.get("location", "")).strip() or "(none)",
                    "role": str(node.attrib.get("role", "")).strip() or "(none)",
                    "text": text_value or "(none)",
                    "diagnostics": diagnostics_values,
                }
            )
    top_failing_rules = [{"rule": key, "count": count} for key, count in failed_assert_counts.most_common(10)]
    return {
        "root_element_name": _local_name(root.tag),
        "root_namespace": _namespace_uri(root.tag) or "(none)",
        "namespaces": _safe_iterparse_namespaces(svrl_output_path),
        "fired_rules": len(fired_rules),
        "failed_asserts": len(failed_asserts),
        "successful_reports": len(successful_reports),
        "rules_matched": len(fired_rules),
        "rules_evaluated": len(failed_asserts) + len(successful_reports),
        "active_patterns": len(active_patterns),
        "active_pattern_ids": active_pattern_ids[:5],
        "diagnostics": len(diagnostics),
        "failed_assert_count_by_rule": dict(failed_assert_counts),
        "top_failing_rules": top_failing_rules,
        "affected_rows_sample": [key for key, _ in failed_assert_rows.most_common(10)],
        "failed_assert_samples": failed_assert_samples,
    }


def _normalize_svrl_for_html_report(svrl_output_path: Path) -> Dict[str, int]:
    root = ET.parse(str(svrl_output_path)).getroot()
    changed = {"fired_rule_role_defaults": 0}
    root_children = list(root)
    for idx, node in enumerate(root_children):
        if _local_name(node.tag) != "fired-rule":
            continue
        if str(node.attrib.get("role", "")).strip():
            continue
        has_failed_assert = False
        has_successful_report = False
        for trailing in root_children[idx + 1 :]:
            trailing_name = _local_name(trailing.tag)
            if trailing_name in {"fired-rule", "active-pattern"}:
                break
            if trailing_name == "failed-assert":
                has_failed_assert = True
            elif trailing_name == "successful-report":
                has_successful_report = True
        if has_failed_assert:
            node.set("role", "WorksheetErrors")
            changed["fired_rule_role_defaults"] += 1
        elif has_successful_report:
            node.set("role", "WorksheetCount")
            changed["fired_rule_role_defaults"] += 1

    if changed["fired_rule_role_defaults"] > 0:
        svrl_output_path.write_bytes(ET.tostring(root, encoding="utf-8", xml_declaration=True))
    return changed


def _first_lines(path: Path, line_count: int = 50) -> List[str]:
    if not path.exists():
        return [f"(missing) {path}"]
    return path.read_text(encoding="utf-8", errors="replace").splitlines()[:line_count]


def _compare_svrl_files(primary_path: Path, secondary_path: Path) -> Dict[str, Any]:
    primary_bytes = primary_path.read_bytes() if primary_path.exists() else b""
    secondary_bytes = secondary_path.read_bytes() if secondary_path.exists() else b""
    primary_root = ET.parse(str(primary_path)).getroot() if primary_path.exists() else None
    secondary_root = ET.parse(str(secondary_path)).getroot() if secondary_path.exists() else None
    primary_children = len(list(primary_root)) if primary_root is not None else 0
    secondary_children = len(list(secondary_root)) if secondary_root is not None else 0
    return {
        "primary_path": str(primary_path),
        "secondary_path": str(secondary_path),
        "exact_copy": primary_bytes == secondary_bytes,
        "primary_size": len(primary_bytes),
        "secondary_size": len(secondary_bytes),
        "primary_sha256": hashlib.sha256(primary_bytes).hexdigest() if primary_bytes else "",
        "secondary_sha256": hashlib.sha256(secondary_bytes).hexdigest() if secondary_bytes else "",
        "primary_root": _local_name(primary_root.tag) if primary_root is not None else "(missing)",
        "secondary_root": _local_name(secondary_root.tag) if secondary_root is not None else "(missing)",
        "primary_children": primary_children,
        "secondary_children": secondary_children,
    }


def _extract_svrl_failure_index(root: ET.Element) -> Dict[str, Any]:
    per_rule: Counter[str] = Counter()
    per_rule_rows: Dict[str, set] = {}
    for node in root.findall(".//{*}failed-assert"):
        rule_id = str(node.attrib.get("id", "")).strip() or str(node.attrib.get("test", "")).strip() or "(unknown)"
        per_rule[rule_id] += 1
        row = str(node.attrib.get("location", "")).strip() or "(unknown)"
        if rule_id not in per_rule_rows:
            per_rule_rows[rule_id] = set()
        per_rule_rows[rule_id].add(row)
    return {"counts": per_rule, "rows": per_rule_rows}


def _compare_svrl_outputs(generated_svrl_path: Path, reference_svrl_path: Path) -> Dict[str, Any]:
    generated_root = ET.parse(str(generated_svrl_path)).getroot()
    reference_root = ET.parse(str(reference_svrl_path)).getroot()
    generated = _extract_svrl_failure_index(generated_root)
    reference = _extract_svrl_failure_index(reference_root)
    generated_rules = set(generated["counts"].keys())
    reference_rules = set(reference["counts"].keys())
    all_rules = sorted(generated_rules | reference_rules)
    count_deltas: List[str] = []
    row_deltas: List[str] = []
    for rule_id in all_rules:
        gen_count = int(generated["counts"].get(rule_id, 0))
        ref_count = int(reference["counts"].get(rule_id, 0))
        if gen_count != ref_count:
            count_deltas.append(f"{rule_id}:generated={gen_count},reference={ref_count}")
        gen_rows = generated["rows"].get(rule_id, set())
        ref_rows = reference["rows"].get(rule_id, set())
        if gen_rows != ref_rows:
            missing = sorted(ref_rows - gen_rows)[:3]
            extra = sorted(gen_rows - ref_rows)[:3]
            row_deltas.append(f"{rule_id}:missing_rows={missing},extra_rows={extra}")
    return {
        "reference_svrl_path": str(reference_svrl_path),
        "generated_failed_asserts": sum(generated["counts"].values()),
        "reference_failed_asserts": sum(reference["counts"].values()),
        "missing_rules": sorted(reference_rules - generated_rules),
        "extra_rules": sorted(generated_rules - reference_rules),
        "count_deltas": count_deltas[:100],
        "row_deltas": row_deltas[:100],
    }


def _inspect_svrl_html_xslt(xslt_path: Path, svrl_path: Path) -> Dict[str, Any]:
    details: Dict[str, Any] = {
        "xslt_path": str(xslt_path),
        "exists": xslt_path.exists(),
        "error_group_xpaths": [],
        "affected_row_xpaths": [],
        "checks_performed_xpaths": [],
        "xpath_evaluations": [],
        "assumptions": [],
        "svrl_root": "",
        "svrl_namespaces": {},
    }
    if not xslt_path.exists():
        return details
    try:
        from lxml import etree
    except Exception as exc:
        details["error"] = f"lxml unavailable: {exc}"
        return details

    try:
        xslt_doc = etree.parse(str(xslt_path))
        svrl_doc = etree.parse(str(svrl_path))
    except Exception as exc:
        details["error"] = f"parse failure: {exc}"
        return details

    svrl_root = svrl_doc.getroot()
    details["svrl_root"] = etree.QName(svrl_root.tag).localname if svrl_root is not None else "(missing)"
    details["svrl_namespaces"] = {str(k or "(default)"): str(v) for k, v in (svrl_root.nsmap or {}).items()}

    expressions: List[str] = []
    for node in xslt_doc.xpath("//*"):
        for attr_name in ("select", "match"):
            value = str(node.attrib.get(attr_name, "")).strip()
            if value:
                expressions.append(value)
    unique_expr: List[str] = []
    for expr in expressions:
        if expr not in unique_expr:
            unique_expr.append(expr)
    details["error_group_xpaths"] = [expr for expr in unique_expr if "failed-assert" in expr][:20]
    details["affected_row_xpaths"] = [expr for expr in unique_expr if "location" in expr][:20]
    details["checks_performed_xpaths"] = [
        expr for expr in unique_expr if any(token in expr for token in ("fired-rule", "successful-report", "active-pattern"))
    ][:20]
    if any("svrl:" in expr for expr in unique_expr):
        details["assumptions"].append("expects_svrl_prefix")
    if any("WorksheetErrors" in expr for expr in unique_expr):
        details["assumptions"].append("expects_WorksheetErrors_role")
    if any("WorksheetCount" in expr for expr in unique_expr):
        details["assumptions"].append("expects_WorksheetCount_role")
    if any("*:text/*:location" in expr or "svrl:text/" in expr for expr in unique_expr):
        details["assumptions"].append("expects_location_below_text")
    if any("sps:" in expr for expr in unique_expr):
        details["assumptions"].append("expects_sps_functions")

    xpath_ns = {k: v for k, v in (svrl_root.nsmap or {}).items() if k}
    if "svrl" not in xpath_ns:
        xpath_ns["svrl"] = "http://purl.oclc.org/dsdl/svrl"
    xpath_ns["sps"] = "http://www.schematron-quickfix.com/validator/process"
    for expr in [*details["error_group_xpaths"], *details["affected_row_xpaths"], *details["checks_performed_xpaths"]]:
        try:
            result = svrl_doc.xpath(expr, namespaces=xpath_ns)
            details["xpath_evaluations"].append(
                {"expression": expr, "result_count": len(result) if isinstance(result, list) else 1, "ok": True}
            )
        except Exception as exc:
            details["xpath_evaluations"].append({"expression": expr, "ok": False, "error": str(exc)})
    return details


@dataclass
class CobieQcNativeResult:
    ok: bool
    output_filename: str
    output_html: str
    cobie_xml: str
    svrl_xml: str
    summary: Dict[str, Any]
    stdout: str
    stderr: str
    error: str


class CobieWorkbookXmlBuilder:
    def __init__(self, workbook_path: Path, stage: str, template_path: Path | None = None) -> None:
        self.workbook_path = workbook_path
        self.stage = stage
        self.template_path = template_path

    def build(self) -> Tuple[bytes, List[str]]:
        warnings: List[str] = []
        wb = load_workbook(self.workbook_path, read_only=True, data_only=True)
        root = ET.Element("COBie")
        root.set("stage", self.stage)
        root.set("source", self.workbook_path.name)

        if self.template_path and self.template_path.exists():
            root.set("template", self.template_path.name)
        else:
            warnings.append("COBieExcelTemplate.xml missing; using direct entity sheet mapping")

        canonical_entity_map: Dict[str, Tuple[str, str]] = {
            "contact": ("Contacts", "Contact"),
            "facility": ("Facilities", "Facility"),
            "floor": ("Floors", "Floor"),
            "space": ("Spaces", "Space"),
            "zone": ("Zones", "Zone"),
            "type": ("Types", "Type"),
            "component": ("Components", "Component"),
        }
        container_elements: Dict[str, ET.Element] = {}

        def _resolve_container(sheet_title: str) -> Tuple[str, str]:
            normalized = _clean_tag(sheet_title).lower()
            if normalized.endswith("s"):
                normalized = normalized[:-1]
            if normalized in canonical_entity_map:
                return canonical_entity_map[normalized]
            entity = _clean_tag(sheet_title)
            if entity.endswith("s"):
                entity = entity[:-1] or entity
            return f"{entity}s", entity

        for sheet in wb.worksheets:
            rows = list(sheet.iter_rows(values_only=True))
            if not rows:
                warnings.append(f"Sheet '{sheet.title}' is empty")
                continue

            container_name, entity_name = _resolve_container(sheet.title)
            container_el = container_elements.get(container_name)
            if container_el is None:
                container_el = ET.SubElement(root, container_name)
                container_elements[container_name] = container_el

            header = [_clean_tag(v if v is not None else "Column") for v in rows[0]]
            entity_ordinal = 0
            for row_idx, row in enumerate(rows[1:], start=2):
                if all(cell in (None, "") for cell in row):
                    continue
                entity_ordinal += 1
                row_el = ET.SubElement(container_el, entity_name)
                row_el.set("sourceSheet", sheet.title)
                row_el.set("rowNumber", str(row_idx))
                row_el.set("sheetOrdinal", str(entity_ordinal))
                for col_idx, cell in enumerate(row):
                    col_name = header[col_idx] if col_idx < len(header) else f"Column_{col_idx + 1}"
                    cell_el = ET.SubElement(row_el, col_name)
                    normalized = self._normalize_cell_value(
                        cell=cell,
                        column_name=col_name,
                        workbook_epoch=wb.epoch,
                        warnings=warnings,
                    )
                    if normalized is not None:
                        cell_el.text = normalized

        return ET.tostring(root, encoding="utf-8", xml_declaration=True), warnings

    def _normalize_cell_value(
        self,
        cell: Any,
        column_name: str,
        workbook_epoch: datetime,
        warnings: List[str],
    ) -> Optional[str]:
        if cell is None:
            return None
        name = column_name.strip().lower()
        if name == "createdon":
            normalized_date = self._normalize_created_on(cell, workbook_epoch, warnings)
            return normalized_date
        if isinstance(cell, str):
            trimmed = cell.strip()
            return trimmed or None
        return str(cell).strip() or None

    def _normalize_created_on(self, value: Any, workbook_epoch: datetime, warnings: List[str]) -> Optional[str]:
        if isinstance(value, datetime):
            return value.strftime("%Y-%m-%dT%H:%M:%S")
        if isinstance(value, date):
            return datetime(value.year, value.month, value.day).strftime("%Y-%m-%dT%H:%M:%S")
        if isinstance(value, (float, int)):
            try:
                dt = from_excel(value, epoch=workbook_epoch)
                if isinstance(dt, datetime):
                    return dt.strftime("%Y-%m-%dT%H:%M:%S")
                if isinstance(dt, date):
                    return datetime(dt.year, dt.month, dt.day).strftime("%Y-%m-%dT%H:%M:%S")
            except Exception as exc:
                warnings.append(f"CreatedOn numeric value conversion failed value={value} error={exc}")
        text = str(value).strip()
        if not text:
            return None
        try:
            parsed = datetime.fromisoformat(text.replace("Z", "+00:00"))
            return parsed.strftime("%Y-%m-%dT%H:%M:%S")
        except Exception:
            warnings.append(f"CreatedOn invalid ISO8601 value dropped value='{text}'")
            return None


class SchematronPipeline:
    def __init__(self, resources_dir: Path, stage: str) -> None:
        self.resources_dir = resources_dir
        self.stage = stage

    def validate(
        self,
        cobie_xml_path: Path,
        svrl_output_path: Path,
        compiled_xslt_output_path: Path,
    ) -> Tuple[Dict[str, int], List[str], str, List[str]]:
        engine = _get_xslt_engine()
        logs: List[str] = [f"xslt_engine={engine}"]
        if engine == "saxon":
            return self._validate_with_saxon(
                cobie_xml_path=cobie_xml_path,
                svrl_output_path=svrl_output_path,
                compiled_xslt_output_path=compiled_xslt_output_path,
                logs=logs,
            )
        if engine == "lxml":
            return self._validate_with_lxml(
                cobie_xml_path=cobie_xml_path,
                svrl_output_path=svrl_output_path,
                compiled_xslt_output_path=compiled_xslt_output_path,
                logs=logs,
            )
        raise RuntimeError(
            f"Unsupported COBIEQC_XSLT_ENGINE='{engine}'. Expected one of: saxon, lxml."
        )

    def _phase_catalog(self, sch_root: Any, use_xpath: bool = True) -> Dict[str, Any]:
        if use_xpath:
            ns = {"sch": "http://purl.oclc.org/dsdl/schematron"}
            pattern_ids = [str(v) for v in sch_root.xpath("/sch:schema/sch:pattern/@id", namespaces=ns)]
            phase_nodes = sch_root.xpath("/sch:schema/sch:phase", namespaces=ns)
            phase_map: Dict[str, List[str]] = {}
            for phase in phase_nodes:
                phase_id = str(phase.attrib.get("id", "")).strip()
                if not phase_id:
                    continue
                refs = [str(v) for v in phase.xpath("./sch:active/@pattern", namespaces=ns)]
                phase_map[phase_id] = refs
            return {"phase_ids": list(phase_map.keys()), "phase_patterns": phase_map, "pattern_ids": pattern_ids}
        ns = {"sch": "http://purl.oclc.org/dsdl/schematron"}
        pattern_ids = [str(el.attrib.get("id", "")).strip() for el in sch_root.findall("sch:pattern", ns)]
        phase_map = {}
        for phase in sch_root.findall("sch:phase", ns):
            phase_id = str(phase.attrib.get("id", "")).strip()
            if not phase_id:
                continue
            refs = [str(active.attrib.get("pattern", "")).strip() for active in phase.findall("sch:active", ns)]
            phase_map[phase_id] = [ref for ref in refs if ref]
        return {"phase_ids": list(phase_map.keys()), "phase_patterns": phase_map, "pattern_ids": pattern_ids}

    def _resolve_phase_ids(self, phase_ids: List[str]) -> List[str]:
        configured = os.getenv("COBIEQC_SCHEMATRON_PHASES", "").strip()
        if configured:
            desired = [part.strip() for part in configured.split(",") if part.strip()]
            selected = [phase for phase in desired if phase in phase_ids]
            if selected:
                return selected
        if not phase_ids:
            return []
        stage_value = (self.stage or "").strip().upper()
        keyword_map = {"D": ("design", "information"), "C": ("construction",)}
        keywords = keyword_map.get(stage_value, (stage_value.lower(),))
        exact = [phase for phase in phase_ids if phase.upper() == stage_value]
        matching = [phase for phase in phase_ids if any(keyword in phase.lower() for keyword in keywords if keyword)]
        selected = []
        for item in [*exact, *matching]:
            if item not in selected:
                selected.append(item)
        return selected or [phase_ids[0]]

    def _validate_with_lxml(
        self,
        cobie_xml_path: Path,
        svrl_output_path: Path,
        compiled_xslt_output_path: Path,
        logs: List[str],
    ) -> Tuple[Dict[str, int], List[str], str, List[str]]:
        warnings: List[str] = []
        try:
            from lxml import etree
        except Exception as exc:
            warnings.append(f"lxml unavailable for Schematron validation: {exc}")
            self._write_fallback_svrl(svrl_output_path, warnings, [])
            return {"failed_asserts": 0, "successful_reports": 0, "diagnostics": 1}, warnings, "", logs

        sch_path = self.resources_dir / "COBieRules.sch"
        compile_xslt_path = self.resources_dir / "iso_svrl_for_xslt2.xsl"
        skeleton_path = self.resources_dir / "iso_schematron_skeleton_for_saxon.xsl"
        functions_path = self.resources_dir / "COBieRules_Functions.xsl"
        if not sch_path.exists():
            warnings.append("COBieRules.sch missing; generated fallback SVRL")
            self._write_fallback_svrl(svrl_output_path, warnings, [])
            return {"failed_asserts": 0, "successful_reports": 0, "diagnostics": 1}, warnings, "", logs
        if not compile_xslt_path.exists():
            warnings.append("iso_svrl_for_xslt2.xsl missing; generated fallback SVRL")
            self._write_fallback_svrl(svrl_output_path, warnings, [])
            return {"failed_asserts": 0, "successful_reports": 0, "diagnostics": 1}, warnings, "", logs

        try:
            sch_doc = etree.parse(str(sch_path))
        except Exception as exc:
            warnings.append(f"Schematron execution fallback used: compilation failed during schematron compile: {exc}")
            warnings.extend(self._safe_artifact_preview(sch_path))
            self._write_fallback_svrl(svrl_output_path, warnings, [str(exc)])
            return {"failed_asserts": 0, "successful_reports": 0, "diagnostics": len(warnings)}, warnings, str(exc), logs
        logs.append(self._phase_log("schematron_source_loaded", sch_path))
        if skeleton_path.exists():
            logs.append(self._phase_log("schematron_skeleton_loaded", skeleton_path))
        if functions_path.exists():
            logs.append(self._phase_log("schematron_functions_loaded", functions_path))

        catalog = self._phase_catalog(sch_doc, use_xpath=True)
        logs.append(f"schematron_available_phases={','.join(catalog['phase_ids']) or '(none)'}")
        for phase_id, patterns in catalog["phase_patterns"].items():
            logs.append(f"schematron_phase_patterns phase={phase_id} active_patterns={','.join(patterns) or '(none)'}")
        phase_ids = self._resolve_phase_ids(catalog["phase_ids"])
        logs.append(f"schematron_phase_selected stage={self.stage} phases={','.join(phase_ids) or '(default)'}")
        if not phase_ids:
            warnings.append(f"No phases discovered for stage '{self.stage}'. Running default Schematron phase.")

        if not phase_ids:
            phase_ids = [""]
        phase_svrl_roots: List[Any] = []
        phase_compiled_paths: List[Path] = []
        compile_step = "schematron compile"
        validation_step = "XML validation"
        for idx, phase_id in enumerate(phase_ids):
            phase_suffix = f".phase_{idx + 1}"
            phase_compiled_path = (
                compiled_xslt_output_path.parent / f"{compiled_xslt_output_path.stem}{phase_suffix}{compiled_xslt_output_path.suffix}"
            )
            phase_svrl_path = svrl_output_path.parent / f"{svrl_output_path.stem}{phase_suffix}{svrl_output_path.suffix}"
            try:
                compiler_doc = etree.parse(str(compile_xslt_path))
                compiler = etree.XSLT(compiler_doc)
                compile_args = {"phase": etree.XSLT.strparam(phase_id)} if phase_id else {}
                compiled_validation_doc = compiler(sch_doc, **compile_args)
                phase_compiled_path.write_bytes(
                    etree.tostring(compiled_validation_doc, encoding="utf-8", pretty_print=True, xml_declaration=True)
                )
                phase_compiled_paths.append(phase_compiled_path)
                logs.append(self._phase_log("schematron_compiled_to_xslt", phase_compiled_path))
                logs.append(f"compiled_validation_xslt phase={phase_id or '(default)'} path={phase_compiled_path}")
                detected_hrefs = self._collect_xslt_dependency_hrefs(phase_compiled_path)
                logs.append(f"compiled_xslt_dependencies_detected phase={phase_id or '(default)'} hrefs={','.join(detected_hrefs) or '(none)'}")
                rewritten = self._rewrite_xslt_dependency_hrefs(phase_compiled_path)
                for source_href, target_href in rewritten:
                    logs.append(f"compiled_xslt_dependency_rewritten phase={phase_id or '(default)'} from={source_href} to={target_href}")
                unresolved = self._find_unresolved_relative_hrefs(phase_compiled_path)
                if unresolved:
                    warnings.append(
                        f"compiled_xslt_unresolved_relative_dependencies={','.join(unresolved)} "
                        f"base_dir={phase_compiled_path.parent}"
                    )
                    logs.append(f"compiled_xslt_unresolved_relative_dependencies phase={phase_id or '(default)'} hrefs={','.join(unresolved)}")
                logs.append(f"validation_resolver_base path={phase_compiled_path.parent}")
                xml_doc = etree.parse(str(cobie_xml_path))
                validation_doc = etree.parse(str(phase_compiled_path))
                validation_transform = etree.XSLT(validation_doc)
                svrl_doc = validation_transform(xml_doc)
                phase_svrl_path.write_bytes(
                    etree.tostring(svrl_doc, encoding="utf-8", pretty_print=True, xml_declaration=True)
                )
                phase_root = etree.parse(str(phase_svrl_path)).getroot()
                phase_root.attrib["data-phase"] = phase_id or "(default)"
                phase_svrl_roots.append(phase_root)
                logs.append(self._phase_log("validation_xslt_applied", phase_compiled_path))
                logs.append(self._phase_log("svrl_generated", phase_svrl_path))
            except Exception as exc:
                diagnostics = self._collect_schematron_diagnostics(sch_doc)
                warnings.append(f"Schematron execution fallback used: failed during {compile_step}/{validation_step}: {exc}")
                warnings.extend(diagnostics)
                self._write_fallback_svrl(svrl_output_path, warnings, [str(exc)])
                return ({"failed_asserts": 0, "successful_reports": 0, "diagnostics": len(warnings)}, warnings, str(exc), logs)

        if phase_svrl_roots:
            merged_root = phase_svrl_roots[0]
            for extra_root in phase_svrl_roots[1:]:
                for child in list(extra_root):
                    merged_root.append(child)
            svrl_output_path.write_bytes(
                etree.tostring(merged_root, encoding="utf-8", pretty_print=True, xml_declaration=True)
            )
            if phase_compiled_paths:
                shutil.copy2(phase_compiled_paths[-1], compiled_xslt_output_path)
        logs.append(self._phase_log("svrl_generated", svrl_output_path))

        root = etree.parse(str(svrl_output_path)).getroot()
        failed_asserts = len(root.xpath("//*[local-name()='failed-assert']"))
        successful_reports = len(root.xpath("//*[local-name()='successful-report']"))
        diagnostics = len(root.xpath("//*[local-name()='diagnostic-reference']"))
        return (
            {
                "failed_asserts": failed_asserts,
                "successful_reports": successful_reports,
                "diagnostics": diagnostics,
            },
            warnings,
            "",
            logs,
        )

    def _validate_with_saxon(
        self,
        cobie_xml_path: Path,
        svrl_output_path: Path,
        compiled_xslt_output_path: Path,
        logs: List[str],
    ) -> Tuple[Dict[str, int], List[str], str, List[str]]:
        warnings: List[str] = []
        sch_path = self.resources_dir / "COBieRules.sch"
        compile_xslt_path = self.resources_dir / "iso_svrl_for_xslt2.xsl"
        skeleton_path = self.resources_dir / "iso_schematron_skeleton_for_saxon.xsl"
        functions_path = self.resources_dir / "COBieRules_Functions.xsl"
        if not sch_path.exists():
            raise RuntimeError(f"COBieRules.sch missing at {sch_path}")
        if not compile_xslt_path.exists():
            raise RuntimeError(f"iso_svrl_for_xslt2.xsl missing at {compile_xslt_path}")

        sch_root = ET.parse(str(sch_path)).getroot()
        catalog = self._phase_catalog(sch_root, use_xpath=False)
        logs.append(f"schematron_available_phases={','.join(catalog['phase_ids']) or '(none)'}")
        for phase_id, patterns in catalog["phase_patterns"].items():
            logs.append(f"schematron_phase_patterns phase={phase_id} active_patterns={','.join(patterns) or '(none)'}")
        phase_ids = self._resolve_phase_ids(catalog["phase_ids"])
        logs.append(f"schematron_phase_selected stage={self.stage} phases={','.join(phase_ids) or '(default)'}")
        if not phase_ids:
            phase_ids = [""]
        logs.append(self._phase_log("schematron_source_loaded", sch_path))
        if skeleton_path.exists():
            logs.append(self._phase_log("schematron_skeleton_loaded", skeleton_path))
        if functions_path.exists():
            logs.append(self._phase_log("schematron_functions_loaded", functions_path))

        phase_svrl_paths: List[Path] = []
        phase_compiled_paths: List[Path] = []
        for idx, phase_id in enumerate(phase_ids):
            phase_suffix = f".phase_{idx + 1}"
            phase_compiled_path = (
                compiled_xslt_output_path.parent / f"{compiled_xslt_output_path.stem}{phase_suffix}{compiled_xslt_output_path.suffix}"
            )
            phase_svrl_path = svrl_output_path.parent / f"{svrl_output_path.stem}{phase_suffix}{svrl_output_path.suffix}"
            compile_stdout, compile_stderr = _run_saxon_xslt(
                xml_input_path=sch_path,
                stylesheet_path=compile_xslt_path,
                output_path=phase_compiled_path,
                params={"phase": phase_id} if phase_id else {},
                logs=logs,
            )
            phase_compiled_paths.append(phase_compiled_path)
            logs.append(self._phase_log("schematron_compiled_to_xslt", phase_compiled_path))
            logs.append(f"compiled_validation_xslt phase={phase_id or '(default)'} path={phase_compiled_path}")
            if compile_stdout:
                logs.append(f"saxon_compile_stdout phase={phase_id or '(default)'} stdout={compile_stdout.strip()}")
            if compile_stderr:
                logs.append(f"saxon_compile_stderr phase={phase_id or '(default)'} stderr={compile_stderr.strip()}")
            detected_hrefs = self._collect_xslt_dependency_hrefs(phase_compiled_path)
            logs.append(f"compiled_xslt_dependencies_detected phase={phase_id or '(default)'} hrefs={','.join(detected_hrefs) or '(none)'}")
            rewritten = self._rewrite_xslt_dependency_hrefs(phase_compiled_path)
            for source_href, target_href in rewritten:
                logs.append(f"compiled_xslt_dependency_rewritten phase={phase_id or '(default)'} from={source_href} to={target_href}")
            unresolved = self._find_unresolved_relative_hrefs(phase_compiled_path)
            if unresolved:
                warnings.append(
                    f"compiled_xslt_unresolved_relative_dependencies={','.join(unresolved)} "
                    f"base_dir={phase_compiled_path.parent}"
                )
                logs.append(f"compiled_xslt_unresolved_relative_dependencies phase={phase_id or '(default)'} hrefs={','.join(unresolved)}")
            logs.append(f"validation_resolver_base path={phase_compiled_path.parent}")
            validation_stdout, validation_stderr = _run_saxon_xslt(
                xml_input_path=cobie_xml_path,
                stylesheet_path=phase_compiled_path,
                output_path=phase_svrl_path,
                params={},
                logs=logs,
            )
            phase_svrl_paths.append(phase_svrl_path)
            logs.append(self._phase_log("validation_xslt_applied", phase_compiled_path))
            logs.append(self._phase_log("svrl_generated", phase_svrl_path))
            if validation_stdout:
                logs.append(f"saxon_validation_stdout phase={phase_id or '(default)'} stdout={validation_stdout.strip()}")
            if validation_stderr:
                logs.append(f"saxon_validation_stderr phase={phase_id or '(default)'} stderr={validation_stderr.strip()}")

        if phase_svrl_paths:
            base_root = ET.parse(str(phase_svrl_paths[0])).getroot()
            for extra in phase_svrl_paths[1:]:
                extra_root = ET.parse(str(extra)).getroot()
                for child in list(extra_root):
                    base_root.append(child)
            svrl_output_path.write_bytes(ET.tostring(base_root, encoding="utf-8", xml_declaration=True))
            if phase_compiled_paths:
                shutil.copy2(phase_compiled_paths[-1], compiled_xslt_output_path)
        logs.append(self._phase_log("svrl_generated", svrl_output_path))

        root = ET.parse(str(svrl_output_path)).getroot()
        failed_asserts = len(root.findall(".//{*}failed-assert"))
        successful_reports = len(root.findall(".//{*}successful-report"))
        diagnostics = len(root.findall(".//{*}diagnostic-reference"))
        return (
            {
                "failed_asserts": failed_asserts,
                "successful_reports": successful_reports,
                "diagnostics": diagnostics,
            },
            warnings,
            "",
            logs,
        )

    def _collect_schematron_diagnostics(self, sch_doc: Any) -> List[str]:
        ns = {"sch": "http://purl.oclc.org/dsdl/schematron"}
        pattern_ids = [str(v) for v in sch_doc.xpath("/sch:schema/sch:pattern/@id", namespaces=ns)]
        rule_ids = [str(v) for v in sch_doc.xpath("//sch:rule/@id", namespaces=ns)]
        phase_refs = [str(v) for v in sch_doc.xpath("/sch:schema/sch:phase/sch:active/@pattern", namespaces=ns)]
        unresolved_refs = sorted(set(ref for ref in phase_refs if ref not in pattern_ids))
        details = [
            f"Schematron diagnostics: pattern_ids={','.join(pattern_ids) or '(none)'}",
            f"Schematron diagnostics: rule_ids={','.join(rule_ids) or '(none)'}",
            f"Schematron diagnostics: phase_pattern_refs={','.join(phase_refs) or '(none)'}",
        ]
        if unresolved_refs:
            details.append(f"Schematron diagnostics: unresolved_pattern_refs={','.join(unresolved_refs)}")
        return details

    def _phase_log(self, phase: str, artifact: Path) -> str:
        size = artifact.stat().st_size if artifact.exists() else 0
        return f"{phase} path={artifact} size_bytes={size}"

    def _collect_xslt_dependency_hrefs(self, xslt_path: Path) -> List[str]:
        try:
            from lxml import etree
        except Exception:
            return []
        doc = etree.parse(str(xslt_path))
        nodes = doc.xpath("//*[local-name()='import' or local-name()='include'][@href]")
        return [str(node.get("href", "")).strip() for node in nodes]

    def _rewrite_xslt_dependency_hrefs(self, xslt_path: Path) -> List[Tuple[str, str]]:
        try:
            from lxml import etree
        except Exception:
            return []
        parser = etree.XMLParser(remove_blank_text=False)
        doc = etree.parse(str(xslt_path), parser)
        rewritten: List[Tuple[str, str]] = []
        for node in doc.xpath("//*[local-name()='import' or local-name()='include'][@href]"):
            href = str(node.get("href", "")).strip()
            if not href or "://" in href or href.startswith("/"):
                continue
            candidate = (self.resources_dir / href).resolve()
            if candidate.exists():
                absolute_href = candidate.as_uri()
                node.set("href", absolute_href)
                rewritten.append((href, absolute_href))
        if rewritten:
            doc.write(str(xslt_path), encoding="utf-8", pretty_print=True, xml_declaration=True)
        return rewritten

    def _find_unresolved_relative_hrefs(self, xslt_path: Path) -> List[str]:
        unresolved: List[str] = []
        for href in self._collect_xslt_dependency_hrefs(xslt_path):
            if not href or "://" in href or href.startswith("/"):
                continue
            unresolved.append(href)
        return unresolved

    def _safe_artifact_preview(self, artifact: Path) -> List[str]:
        if not artifact.exists():
            return [f"artifact_preview unavailable path={artifact} reason=missing"]
        lines = artifact.read_text(encoding="utf-8", errors="replace").splitlines()[:20]
        preview = " | ".join(lines)
        return [f"artifact_preview path={artifact} lines_1_20={preview}"]

    def _write_fallback_svrl(self, output_path: Path, warnings: List[str], errors: List[str]) -> None:
        root = ET.Element("svrl:schematron-output", {"xmlns:svrl": "http://purl.oclc.org/dsdl/svrl"})
        for warning in warnings:
            warn_el = ET.SubElement(root, "svrl:successful-report")
            warn_el.set("role", "warning")
            text_el = ET.SubElement(warn_el, "svrl:text")
            text_el.text = warning
        for error in errors:
            err_el = ET.SubElement(root, "svrl:failed-assert")
            text_el = ET.SubElement(err_el, "svrl:text")
            text_el.text = error
        output_path.write_bytes(ET.tostring(root, encoding="utf-8", xml_declaration=True))


class SvrlHtmlRenderer:
    def __init__(self, resources_dir: Path) -> None:
        self.resources_dir = resources_dir

    def render(
        self,
        svrl_path: Path,
        html_output_path: Path,
        summary: Dict[str, Any],
        warnings: List[str],
    ) -> Tuple[List[str], str]:
        engine = _get_xslt_engine()
        logs: List[str] = [f"html_transform_input path={svrl_path} exists={svrl_path.exists()}"]
        error = ""
        css_path = self.resources_dir / "SpaceReport.css"
        target_css = html_output_path.parent / "SpaceReport.css"
        if css_path.exists():
            shutil.copy2(css_path, target_css)
            logs.append(f"html_css_copied path={target_css} size_bytes={target_css.stat().st_size}")

        xslt_candidates = [
            self.resources_dir / "SVRL_HTML_altLocation.xslt",
            self.resources_dir / "_SVRL_HTML_altLocation.xslt",
        ]

        for xslt_path in xslt_candidates:
            if not xslt_path.exists():
                continue
            try:
                if engine == "saxon":
                    _run_saxon_xslt(
                        xml_input_path=svrl_path,
                        stylesheet_path=xslt_path,
                        output_path=html_output_path,
                        params={},
                        logs=logs,
                    )
                else:
                    from lxml import etree
                    svrl_doc = etree.parse(str(svrl_path))
                    xslt_doc = etree.parse(str(xslt_path))
                    transform = etree.XSLT(xslt_doc)
                    html_doc = transform(svrl_doc)
                    html_output_path.write_bytes(
                        etree.tostring(html_doc, encoding="utf-8", pretty_print=True, method="html")
                    )
                logs.append(f"html_generated path={html_output_path} size_bytes={html_output_path.stat().st_size}")
                return logs, error
            except Exception as exc:
                error = f"SVRL-to-HTML transform failed: {exc}"
                continue

        self._write_fallback_html(html_output_path, summary, warnings)
        logs.append(f"html_generated path={html_output_path} size_bytes={html_output_path.stat().st_size}")
        if error and svrl_path.exists():
            preview = svrl_path.read_text(encoding="utf-8", errors="replace").splitlines()[:20]
            warnings.append(f"svrl_preview_first_20={' | '.join(preview)}")
        return logs, error

    def _write_fallback_html(self, html_output_path: Path, summary: Dict[str, Any], warnings: List[str]) -> None:
        warning_items = "".join(f"<li>{_xml_escape(w)}</li>" for w in warnings)
        html = f"""<!doctype html>
<html>
<head>
  <meta charset=\"utf-8\" />
  <title>COBieQC Report</title>
  <link rel=\"stylesheet\" href=\"SpaceReport.css\" />
</head>
<body>
  <h1>COBieQC Validation Report</h1>
  <ul>
    <li>Failed asserts: {summary.get('failed_asserts', 0)}</li>
    <li>Successful reports: {summary.get('successful_reports', 0)}</li>
    <li>Diagnostics: {summary.get('diagnostics', 0)}</li>
  </ul>
  <h2>Pipeline Notes</h2>
  <ul>{warning_items or '<li>None</li>'}</ul>
</body>
</html>
"""
        html_output_path.write_text(html, encoding="utf-8")


def _get_xslt_engine() -> str:
    return os.getenv("COBIEQC_XSLT_ENGINE", "saxon").strip().lower() or "saxon"


def _resolve_required_jar_path(env_name: str) -> Path:
    configured = os.getenv(env_name, "").strip()
    if not configured:
        raise RuntimeError(f"Saxon configuration error: environment variable '{env_name}' is required but was not set.")
    jar_path = Path(configured).expanduser().resolve()
    if not jar_path.exists() or not jar_path.is_file():
        raise RuntimeError(
            f"Saxon configuration error: {env_name} points to a missing or non-file path: {jar_path}"
        )
    return jar_path


def _resolve_saxon_runtime_jars() -> Tuple[Path, Path, Path]:
    saxon_jar = _resolve_required_jar_path("COBIEQC_SAXON_JAR_PATH")
    xmlresolver_jar = _resolve_required_jar_path("COBIEQC_SAXON_XMLRESOLVER_JAR_PATH")
    xmlresolver_data_jar = _resolve_required_jar_path("COBIEQC_SAXON_XMLRESOLVER_DATA_JAR_PATH")
    return saxon_jar, xmlresolver_jar, xmlresolver_data_jar


def _resolve_saxon_command() -> List[str]:
    configured = os.getenv("COBIEQC_SAXON_CMD", "").strip()
    if configured:
        return shlex.split(configured)
    java_bin = os.getenv("JAVA_BIN", "java").strip() or "java"
    saxon_jar, xmlresolver_jar, xmlresolver_data_jar = _resolve_saxon_runtime_jars()
    classpath = os.pathsep.join([str(saxon_jar), str(xmlresolver_jar), str(xmlresolver_data_jar)])
    return [java_bin, "-cp", classpath, "net.sf.saxon.Transform"]


def _run_saxon_xslt(
    xml_input_path: Path,
    stylesheet_path: Path,
    output_path: Path,
    params: Dict[str, str],
    logs: List[str],
) -> Tuple[str, str]:
    logs.append("xslt_engine=saxon")
    logs.append("saxon_main_class=net.sf.saxon.Transform")
    configured_override = os.getenv("COBIEQC_SAXON_CMD", "").strip()
    if configured_override:
        command = [
            *shlex.split(configured_override),
            f"-s:{xml_input_path}",
            f"-xsl:{stylesheet_path}",
            f"-o:{output_path}",
        ]
        logs.append("saxon_command_source=override:COBIEQC_SAXON_CMD")
    else:
        saxon_jar, xmlresolver_jar, xmlresolver_data_jar = _resolve_saxon_runtime_jars()
        logs.append(f"saxon_jar_path={saxon_jar}")
        logs.append(f"saxon_jar_size_bytes={saxon_jar.stat().st_size}")
        logs.append(f"xmlresolver_jar_path={xmlresolver_jar}")
        logs.append(f"xmlresolver_jar_size_bytes={xmlresolver_jar.stat().st_size}")
        logs.append(f"xmlresolver_data_jar_path={xmlresolver_data_jar}")
        logs.append(f"xmlresolver_data_jar_size_bytes={xmlresolver_data_jar.stat().st_size}")
        classpath = os.pathsep.join([str(saxon_jar), str(xmlresolver_jar), str(xmlresolver_data_jar)])
        java_bin = os.getenv("JAVA_BIN", "java").strip() or "java"
        command = [
            java_bin,
            "-cp",
            classpath,
            "net.sf.saxon.Transform",
            f"-s:{xml_input_path}",
            f"-xsl:{stylesheet_path}",
            f"-o:{output_path}",
        ]
        logs.append(f"computed_classpath={classpath}")
        logs.append("saxon_command_source=classpath_env_vars")
    for key, value in params.items():
        command.append(f"{key}={value}")
    logs.append(f"saxon_argv={command}")
    logs.append(f"xslt_engine=saxon input_xml_path={xml_input_path}")
    logs.append(f"xslt_engine=saxon stylesheet_path={stylesheet_path}")
    logs.append(f"xslt_engine=saxon output_path={output_path}")
    completed = subprocess.run(command, capture_output=True, text=True)
    if completed.stdout.strip():
        logs.append(f"xslt_engine=saxon stdout={completed.stdout.strip()}")
    if completed.stderr.strip():
        logs.append(f"xslt_engine=saxon stderr={completed.stderr.strip()}")
    if completed.returncode != 0:
        raise RuntimeError(
            f"Saxon XSLT failed (exit={completed.returncode}) for stylesheet={stylesheet_path} input={xml_input_path}. "
            f"stderr={completed.stderr.strip()}"
        )
    if not output_path.exists():
        raise RuntimeError(
            f"Saxon XSLT did not produce expected output file: {output_path} "
            f"(stylesheet={stylesheet_path}, input={xml_input_path})"
        )
    return completed.stdout, completed.stderr


def run_cobieqc_native(input_xlsx_path: str, stage: str, job_dir: str, resources_dir: Path) -> CobieQcNativeResult:
    input_path = Path(input_xlsx_path).resolve()
    out_dir = Path(job_dir).resolve()
    out_dir.mkdir(parents=True, exist_ok=True)

    cobie_xml_path = out_dir / "generated_cobie.xml"
    compiled_validation_xslt_path = out_dir / "compiled_validation.xsl"
    svrl_xml_path = out_dir / "validation_result.svrl.xml"
    html_path = out_dir / "final_report.html"
    logs: List[str] = []

    try:
        logs.append("workbook parsed")
        builder = CobieWorkbookXmlBuilder(
            workbook_path=input_path,
            stage=stage,
            template_path=resources_dir / "COBieExcelTemplate.xml",
        )
        cobie_xml_bytes, parse_warnings = builder.build()
        cobie_xml_path.write_bytes(cobie_xml_bytes)
        logs.append(f"generated_cobie_xml path={cobie_xml_path} size_bytes={cobie_xml_path.stat().st_size}")
        cobie_diagnostics = _inspect_cobie_xml(cobie_xml_path)
        logs.append(
            "generated_cobie_xml_diagnostics "
            f"root_element={cobie_diagnostics['root_element_name']} "
            f"root_namespace={cobie_diagnostics['root_namespace']} "
            f"first_level_children={','.join(cobie_diagnostics['first_level_children']) or '(none)'}"
        )
        entity_counts = cobie_diagnostics["entity_counts"]
        logs.append(
            "generated_cobie_xml_entity_counts "
            + " ".join(f"{key}={entity_counts.get(key, 0)}" for key in COBIE_ENTITY_NAMES)
        )
        for entity_name in COBIE_ENTITY_NAMES:
            sample_paths = cobie_diagnostics["sample_paths"].get(entity_name, [])
            logs.append(
                f"generated_cobie_xml_sample_paths entity={entity_name} "
                f"paths={','.join(sample_paths) if sample_paths else '(none)'}"
            )
        sheet_row_counts = cobie_diagnostics["sheet_row_counts"]
        logs.append(
            "generated_cobie_xml_sheet_row_counts "
            + " ".join(f"{sheet}:{count}" for sheet, count in sorted(sheet_row_counts.items()))
        )
        logs.append(
            "generated_created_on_samples "
            f"values={','.join(cobie_diagnostics['created_on_samples']) or '(none)'}"
        )
        cross_ref = cobie_diagnostics["cross_reference"]
        logs.append(
            "generated_cross_reference_summary "
            f"contact_names={cross_ref['contact_name_count']} "
            f"created_by_total={cross_ref['created_by_total']} "
            f"created_by_unmatched={cross_ref['created_by_unmatched']}"
        )
        logs.append(
            "generated_cross_reference_unmatched_samples "
            f"samples={','.join(cross_ref['unmatched_samples']) or '(none)'}"
        )
        key_field_diag = _build_key_field_diagnostics(input_path, cobie_xml_path)
        for entity_name in KEY_FIELD_DIAGNOSTIC_FIELDS:
            profile = key_field_diag["entities"].get(entity_name, {})
            logs.append(
                "key_field_name_diagnostics "
                f"entity={entity_name} "
                f"blank_name_count={profile.get('blank_name_count', 0)} "
                f"duplicate_name_count={profile.get('duplicate_name_count', 0)} "
                f"blank_name_samples={','.join(profile.get('blank_name_samples', [])) or '(none)'} "
                f"duplicate_name_samples={','.join(profile.get('duplicate_name_samples', [])) or '(none)'}"
            )
            logs.append(
                "key_field_audit_summary "
                f"entity={entity_name} "
                f"rows={profile.get('row_count', 0)} "
                f"audit_samples_logged={len(profile.get('samples', []))}"
            )
            for sample_idx, sample in enumerate(profile.get("samples", []), start=1):
                field_chunks = []
                for field_name, field_data in sample.get("fields", {}).items():
                    field_chunks.append(
                        f"{field_name}[xml={field_data.get('xml_value', '')}|source_col={field_data.get('source_column', '')}|source={field_data.get('source_value', '')}]"
                    )
                logs.append(
                    "key_field_sample "
                    f"entity={entity_name} "
                    f"sample={sample_idx} "
                    f"source_sheet={sample.get('source_sheet', '(unknown)')} "
                    f"source_row={sample.get('source_row_number', 0)} "
                    f"identity_source_col={sample.get('identity', {}).get('source_column', '(missing)')} "
                    f"identity_source={sample.get('identity', {}).get('source_value', '')} "
                    f"identity_xml={sample.get('identity', {}).get('xml_value', '')} "
                    f"identity_matches={sample.get('identity', {}).get('matches', False)} "
                    f"values={';'.join(field_chunks) or '(none)'}"
                )
        cross_diag = key_field_diag["cross_references"]
        name_diag_parts = []
        for entity_name in KEY_FIELD_DIAGNOSTIC_FIELDS:
            profile = key_field_diag["entities"].get(entity_name, {})
            name_diag_parts.append(f"{entity_name}:{profile.get('blank_name_count', 0)}")
        logs.append("blank_name_diagnostics_by_entity " + " ".join(name_diag_parts))
        logs.append(
            "cross_reference_diagnostics "
            f"space_floor_unresolved={cross_diag['space_floor_name']['count']} "
            f"component_type_unresolved={cross_diag['component_type_name']['count']} "
            f"component_space_unresolved={cross_diag['component_space']['count']} "
            f"created_by_unresolved={cross_diag['created_by']['count']}"
        )
        logs.append(
            "cross_reference_samples "
            f"space_floor={','.join(cross_diag['space_floor_name']['samples']) or '(none)'} "
            f"component_type={','.join(cross_diag['component_type_name']['samples']) or '(none)'} "
            f"component_space={','.join(cross_diag['component_space']['samples']) or '(none)'} "
            f"created_by={','.join(cross_diag['created_by']['samples']) or '(none)'}"
        )
        mismatch_report = key_field_diag["mismatch_report"]
        logs.append(
            "cross_reference_mismatch_report "
            f"floor_name_first20={','.join(mismatch_report['floor_name']) or '(none)'} "
            f"type_name_first20={','.join(mismatch_report['type_name']) or '(none)'} "
            f"space_first20={','.join(mismatch_report['space']) or '(none)'} "
            f"created_by_first20={','.join(mismatch_report['created_by']) or '(none)'} "
            f"blank_name_first20={','.join(mismatch_report['blank_name_rows']) or '(none)'}"
        )
        workbook_style = cobie_diagnostics["root_element_name"] == "COBieWorkbook" or any(
            child == "Sheet" for child in cobie_diagnostics["first_level_children"]
        )
        if workbook_style:
            raise RuntimeError(
                "Workbook XML shape is not compatible with COBie Schematron validation; "
                "expected COBie entity containers instead of COBieWorkbook/Sheet wrappers."
            )
        if cobie_diagnostics["root_element_name"] != "COBie":
            parse_warnings.append(
                "Generated COBie XML root element is not 'COBie'; rules may expect a COBie root container."
            )
        if cobie_diagnostics["root_namespace"] == "(none)":
            parse_warnings.append(
                "Generated COBie XML has no root namespace; rules may rely on COBie namespaces."
            )
        if cobie_diagnostics["model_path_count"] == 0:
            parse_warnings.append(
                "Generated COBie XML did not expose Contact/Facility/Floor/Space/Zone/Type/Component paths."
            )
        reference_xml_path = _find_reference_xml_path(out_dir)
        if reference_xml_path:
            comparison = _compare_xml_structure(cobie_xml_path, reference_xml_path)
            logs.append(
                "cobie_xml_comparison_summary "
                f"reference_path={comparison['reference_path']} "
                f"generated_root={comparison['generated_root']} reference_root={comparison['reference_root']} "
                f"generated_namespace={comparison['generated_namespace']} "
                f"reference_namespace={comparison['reference_namespace']}"
            )
            logs.append(
                "cobie_xml_comparison_namespace_map "
                f"generated={comparison['generated_namespaces']} "
                f"reference={comparison['reference_namespaces']}"
            )
            logs.append(
                "cobie_xml_comparison_paths "
                f"missing_count={len(comparison['missing_paths'])} extra_count={len(comparison['extra_paths'])}"
            )
            logs.append(
                "cobie_xml_comparison_elements "
                f"missing={','.join(comparison['missing_elements']) or '(none)'} "
                f"extra={','.join(comparison['extra_elements']) or '(none)'}"
            )
            logs.append(
                "cobie_xml_comparison_frequency_deltas "
                f"deltas={';'.join(comparison['frequency_deltas']) or '(none)'}"
            )
        else:
            logs.append(
                "cobie_xml_comparison_skipped reason=no_reference_xml "
                "hint=Set_COBIEQC_REFERENCE_XML_PATH_or_place_legacy_cobie.xml_in_job_dir"
            )

        logs.append("XML generated")
        validator = SchematronPipeline(resources_dir=resources_dir, stage=stage)
        summary, svrl_warnings, svrl_error, schematron_logs = validator.validate(
            cobie_xml_path,
            svrl_xml_path,
            compiled_validation_xslt_path,
        )
        logs.extend(schematron_logs)

        warnings = [*parse_warnings, *svrl_warnings]
        normalization_changes = _normalize_svrl_for_html_report(svrl_xml_path)
        if normalization_changes["fired_rule_role_defaults"] > 0:
            logs.append(
                "svrl_html_normalization "
                f"fired_rule_role_defaults={normalization_changes['fired_rule_role_defaults']}"
            )
        svrl_diagnostics = _inspect_svrl(svrl_xml_path)
        svrl_first_20 = _first_lines(svrl_xml_path, line_count=20)
        logs.append(f"svrl_first_20 path={svrl_xml_path} lines={' | '.join(svrl_first_20)}")
        logs.append(f"svrl_namespace_map {svrl_diagnostics['namespaces']}")
        logs.append(
            "svrl_diagnostics "
            f"root_element={svrl_diagnostics['root_element_name']} "
            f"root_namespace={svrl_diagnostics['root_namespace']} "
            f"fired_rules={svrl_diagnostics['fired_rules']} "
            f"failed_asserts={svrl_diagnostics['failed_asserts']} "
            f"successful_reports={svrl_diagnostics['successful_reports']} "
            f"rules_matched={svrl_diagnostics['rules_matched']} "
            f"rules_evaluated={svrl_diagnostics['rules_evaluated']} "
            f"active_patterns={svrl_diagnostics['active_patterns']} "
            f"active_pattern_ids={','.join(svrl_diagnostics['active_pattern_ids']) or '(none)'}"
        )
        logs.append(
            "svrl_rule_failure_top10 "
            + ",".join(
                f"{item['rule']}:{item['count']}" for item in svrl_diagnostics["top_failing_rules"]
            )
        )
        logs.append(
            "svrl_failed_assert_count_by_rule "
            + ",".join(
                f"{rule}:{count}" for rule, count in sorted(svrl_diagnostics["failed_assert_count_by_rule"].items())
            )
        )
        for idx, sample in enumerate(svrl_diagnostics["failed_assert_samples"], start=1):
            logs.append(
                "svrl_failed_assert_sample "
                f"index={idx} id={sample['id']} location={sample['location']} role={sample['role']} "
                f"text={sample['text']} diagnostics={';'.join(sample['diagnostics']) or '(none)'}"
            )

        diagnostic_mode = os.getenv("COBIEQC_SVRL_DIAGNOSTIC_MODE", "").strip().lower() in {"1", "true", "yes", "on"}
        if diagnostic_mode:
            phase_1_path = svrl_xml_path.parent / f"{svrl_xml_path.stem}.phase_1{svrl_xml_path.suffix}"
            phase_1_lines = _first_lines(phase_1_path, line_count=50)
            final_lines = _first_lines(svrl_xml_path, line_count=50)
            logs.append(f"svrl_phase_1_first_50 path={phase_1_path} lines={' | '.join(phase_1_lines)}")
            logs.append(f"svrl_final_first_50 path={svrl_xml_path} lines={' | '.join(final_lines)}")
            phase_compare = _compare_svrl_files(svrl_xml_path, phase_1_path)
            logs.append(
                "svrl_phase_compare "
                f"final_path={phase_compare['primary_path']} phase_1_path={phase_compare['secondary_path']} "
                f"exact_copy={phase_compare['exact_copy']} "
                f"final_size={phase_compare['primary_size']} phase_1_size={phase_compare['secondary_size']} "
                f"final_root={phase_compare['primary_root']} phase_1_root={phase_compare['secondary_root']} "
                f"final_children={phase_compare['primary_children']} phase_1_children={phase_compare['secondary_children']} "
                f"final_sha256={phase_compare['primary_sha256']} phase_1_sha256={phase_compare['secondary_sha256']}"
            )

        reference_svrl_path = _find_reference_svrl_path(out_dir)
        if reference_svrl_path:
            svrl_comparison = _compare_svrl_outputs(svrl_xml_path, reference_svrl_path)
            logs.append(
                "svrl_comparison_summary "
                f"reference_svrl_path={svrl_comparison['reference_svrl_path']} "
                f"generated_failed_asserts={svrl_comparison['generated_failed_asserts']} "
                f"reference_failed_asserts={svrl_comparison['reference_failed_asserts']}"
            )
            logs.append(
                "svrl_comparison_rules "
                f"missing_rules={','.join(svrl_comparison['missing_rules']) or '(none)'} "
                f"extra_rules={','.join(svrl_comparison['extra_rules']) or '(none)'}"
            )
            logs.append(
                "svrl_comparison_count_deltas "
                f"deltas={';'.join(svrl_comparison['count_deltas']) or '(none)'}"
            )
            logs.append(
                "svrl_comparison_row_deltas "
                f"deltas={';'.join(svrl_comparison['row_deltas']) or '(none)'}"
            )
            summary["svrl_comparison"] = svrl_comparison
        else:
            logs.append(
                "svrl_comparison_skipped reason=no_reference_svrl "
                "hint=Set_COBIEQC_REFERENCE_SVRL_PATH_or_place_legacy_validation_result.svrl.xml_in_job_dir"
            )
        svrl_size_bytes = svrl_xml_path.stat().st_size if svrl_xml_path.exists() else 0
        non_trivial_workbook = sum(cobie_diagnostics["sheet_row_counts"].values()) >= 10
        if svrl_diagnostics["fired_rules"] == 0 and (non_trivial_workbook or svrl_size_bytes < 2048):
            warnings.append(
                "SVRL diagnostics indicate zero fired rules for a non-trivial workbook or tiny output; "
                "rule contexts likely did not match the generated COBie XML structure."
            )
        summary["warnings"] = warnings
        summary["stage"] = stage
        summary["svrl_diagnostics"] = svrl_diagnostics
        summary["cobie_xml_diagnostics"] = {
            "root_element_name": cobie_diagnostics["root_element_name"],
            "root_namespace": cobie_diagnostics["root_namespace"],
            "entity_counts": cobie_diagnostics["entity_counts"],
            "key_field_diagnostics": key_field_diag,
        }

        renderer = SvrlHtmlRenderer(resources_dir=resources_dir)
        html_logs, html_error = renderer.render(svrl_xml_path, html_path, summary, warnings)
        logs.extend(html_logs)
        html_text = html_path.read_text(encoding="utf-8", errors="replace") if html_path.exists() else ""
        if svrl_diagnostics["failed_asserts"] > 0 and "no errors" in html_text.lower():
            sample_ids = [item["id"] for item in svrl_diagnostics["failed_assert_samples"][:5]]
            warning = (
                "HARD WARNING: SVRL contains failed-asserts but HTML rendered 'No Errors'. "
                f"failed_assert_count={svrl_diagnostics['failed_asserts']} sample_ids={sample_ids}"
            )
            warnings.append(warning)
            logs.append(f"svrl_html_mismatch {warning}")
            root_for_dump = ET.parse(str(svrl_xml_path)).getroot()
            failed_nodes = root_for_dump.findall(".//{*}failed-assert")[:5]
            fired_nodes = root_for_dump.findall(".//{*}fired-rule")[:5]
            for idx, node in enumerate(failed_nodes, start=1):
                logs.append(
                    "svrl_integrity_failed_assert "
                    f"index={idx} attrs={dict(node.attrib)} text={' '.join(''.join(node.itertext()).split())[:300]}"
                )
            for idx, node in enumerate(fired_nodes, start=1):
                logs.append(f"svrl_integrity_fired_rule index={idx} attrs={dict(node.attrib)}")
            logs.append(f"svrl_integrity_namespace_map {_safe_iterparse_namespaces(svrl_xml_path)}")
            warnings.append("SVRL/HTML integrity check failed; replacing misleading HTML report with fallback diagnostic report.")
            renderer._write_fallback_html(html_path, summary, warnings)
            html_text = html_path.read_text(encoding='utf-8', errors='replace') if html_path.exists() else html_text

        html_xslt_path = None
        for candidate in (
            resources_dir / "SVRL_HTML_altLocation.xslt",
            resources_dir / "_SVRL_HTML_altLocation.xslt",
        ):
            if candidate.exists():
                html_xslt_path = candidate
                break
        if html_xslt_path:
            xslt_diag = _inspect_svrl_html_xslt(html_xslt_path, svrl_xml_path)
            logs.append(
                "svrl_html_xslt_expectations "
                f"path={xslt_diag['xslt_path']} "
                f"svrl_root={xslt_diag.get('svrl_root', '(unknown)')} "
                f"svrl_namespaces={xslt_diag.get('svrl_namespaces', {})} "
                f"assumptions={';'.join(xslt_diag.get('assumptions', [])) or '(none)'} "
                f"error_group_xpaths={';'.join(xslt_diag['error_group_xpaths']) or '(none)'} "
                f"affected_row_xpaths={';'.join(xslt_diag['affected_row_xpaths']) or '(none)'} "
                f"checks_performed_xpaths={';'.join(xslt_diag['checks_performed_xpaths']) or '(none)'}"
            )
            for eval_item in xslt_diag["xpath_evaluations"][:20]:
                if eval_item.get("ok"):
                    logs.append(
                        f"svrl_html_xpath_eval expression={eval_item['expression']} "
                        f"result_count={eval_item['result_count']}"
                    )
                else:
                    logs.append(
                        f"svrl_html_xpath_eval expression={eval_item['expression']} "
                        f"error={eval_item.get('error', '(unknown)')}"
                    )
            if xslt_diag.get("error"):
                warnings.append(f"SVRL/HTML XPath diagnostic skipped: {xslt_diag['error']}")
            worksheet_fired = 0
            worksheet_grouped_failed = 0
            try:
                from lxml import etree

                svrl_doc = etree.parse(str(svrl_xml_path))
                ns = {"svrl": "http://purl.oclc.org/dsdl/svrl"}
                worksheet_fired = len(svrl_doc.xpath("//svrl:fired-rule[@role='WorksheetErrors']", namespaces=ns))
                worksheet_grouped_failed = len(
                    svrl_doc.xpath(
                        "//svrl:failed-assert[preceding-sibling::svrl:fired-rule[1]/@role='WorksheetErrors']",
                        namespaces=ns,
                    )
                )
            except Exception as exc:
                warnings.append(f"WorksheetErrors SVRL probe failed: {exc}")
            logs.append(
                "svrl_worksheet_errors_probe "
                f"fired_rules={worksheet_fired} grouped_failed_asserts={worksheet_grouped_failed}"
            )

        if html_error:
            warnings.append(f"Schematron execution fallback used: html generation failed during SVRL-to-HTML transform: {html_error}")

        return CobieQcNativeResult(
            ok=True,
            output_filename=html_path.name,
            output_html=str(html_path),
            cobie_xml=str(cobie_xml_path),
            svrl_xml=str(svrl_xml_path),
            summary=summary,
            stdout="\n".join(logs),
            stderr=svrl_error,
            error="",
        )
    except Exception as exc:
        stage_hint = logs[-1] if logs else "input handling"
        return CobieQcNativeResult(
            ok=False,
            output_filename=html_path.name,
            output_html=str(html_path),
            cobie_xml=str(cobie_xml_path),
            svrl_xml=str(svrl_xml_path),
            summary={"failed_asserts": 0, "successful_reports": 0, "diagnostics": 0, "warnings": []},
            stdout="\n".join(logs),
            stderr="",
            error=f"Pipeline failed during stage '{stage_hint}': {exc}",
        )
