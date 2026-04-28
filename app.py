import argparse
import configparser
import csv
import datetime
import difflib
import hashlib
import json
import logging
import mimetypes
import os
import re
import shutil
import subprocess
import sys
import tempfile
import threading
import traceback
import time
import uuid
import zipfile
from contextlib import asynccontextmanager, contextmanager
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Callable, Dict, List, Optional, Set, Tuple

import ifcopenshell
import ifcopenshell.api
import ifcopenshell.util.element
import ifcopenshell.util.placement
import pandas as pd
from openpyxl.workbook.defined_name import DefinedName
from openpyxl import load_workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from fastapi import BackgroundTasks, Body, FastAPI, File, Form, HTTPException, Request, UploadFile
from fastapi.responses import FileResponse, HTMLResponse, JSONResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from ifcopenshell.guid import new as new_guid

from check_definitions_loader import load_check_definitions, summarize_sections
from expression_engine import ExpressionEngine
from field_access import FieldDescriptor, FieldKind, get_value, set_value
from mapping_store import (
    EXPRESSIONS_PATH,
    MAPPINGS_PATH,
    load_expression_config,
    load_mapping_config,
    save_expression_for_check,
    save_mapping_for_check,
)
from validation import validate_value
from cobieqc_service.jobs import (
    STATUS_DONE,
    STATUS_ERROR,
    STATUS_RUNNING,
    CobieQcJobStore,
)
from cobieqc_service.bootstrap import bootstrap_cobieqc_assets, get_cobieqc_bootstrap_status
from cobieqc_service import runner as cobieqc_runner_module
from cobieqc_service.runner import get_cobieqc_runtime_diagnostics, get_cobieqc_engine, resolve_java_executable, run_cobieqc
from cobieqc_service.security import sanitize_filename as sanitize_upload_filename
from cobieqc_service.security import validate_upload
from ifc_qa_service import REGISTRY as IFC_QA_V2_REGISTRY
from ifc_qa_service import read_session_summary as read_ifc_qa_session_summary
from ifc_app.ifc_qa.jobs import start_ifc_qa_session_job
from backend.ifc_qa.config_loader import (
    REQUIRED_TOP_LEVEL_KEYS,
    build_config_indexes,
    load_default_config as load_ifc_qa_default_config,
    merge_config_override,
    validate_config_structure,
)
from backend.ifc_jobs import create_job as create_ifc_job, get_job as get_ifc_job, update_job as update_ifc_job
from backend.ifc_file_size_reducer import (
    IfcFileSizeReducerError,
    analyze_ifc_file,
    run_reduction as run_ifc_size_reduction,
)
from backend.ifc_area_spaces import (
    AreaSpaceError,
    AREA_SPACE_JOB_SEMAPHORE,
    AREA_SPACE_MAX_FILE_MB,
    AREA_SPACE_MAX_PURGE_FILE_MB,
    AREA_SPACE_MEMORY_ABORT_PERCENT,
    get_memory_status,
    is_memory_high,
    package_outputs as package_area_space_outputs,
    result_to_log_payload as area_space_log_payload,
    scan_ifc_for_area_spaces,
)
from backend.ifc_area_spaces_router import build_area_spaces_router
from backend.ifc_move_rotate import TransformRequest, transform_ifc_file
from backend.project_tables import get_tables_for_project_slug

try:
    import gc
    import psutil  # type: ignore
except Exception:  # pragma: no cover - optional runtime dependency
    gc = None  # type: ignore
    psutil = None  # type: ignore

STEP2IFC_ROOT = Path(__file__).resolve().parent / "step2ifc"
if STEP2IFC_ROOT.exists():
    sys.path.append(str(STEP2IFC_ROOT))

STEP2IFC_AVAILABLE = False
STEP2IFC_IMPORT_ERROR = None
STEP2IFC_RUN_CONVERT = None
try:
    from step2ifc.auto import auto_convert
    from step2ifc.cli import run_convert as step2ifc_run_convert
    STEP2IFC_AVAILABLE = True
    STEP2IFC_RUN_CONVERT = step2ifc_run_convert
except Exception as exc:  # pragma: no cover - runtime dependency checks
    STEP2IFC_IMPORT_ERROR = str(exc)

APP_LOGGER = logging.getLogger("ifc_app")
RESOURCE_DIR = Path(__file__).resolve().parent / "resources"
QA_RESOURCE_DIR = Path(__file__).resolve().parent / "app" / "resources" / "ifc_qa"
DATA_DIR = Path(__file__).resolve().parent / "data"
BACKEND_CONFIG_DIR = Path(__file__).resolve().parent / "backend" / "config"
UTC = datetime.timezone.utc
HEAVY_JOB_SEMAPHORE = threading.Semaphore(1)
MAX_UPLOAD_BYTES = 1_200_000_000
MAX_UPLOAD_GB = 1.2
MAX_UPLOAD_DISPLAY = "1.2 GB"
REQUEST_BODY_LIMIT_HEADROOM_BYTES = 25_000_000
MAX_REQUEST_BODY_BYTES = MAX_UPLOAD_BYTES + REQUEST_BODY_LIMIT_HEADROOM_BYTES
MAX_IFC_BYTES = int(os.getenv("MAX_IFC_BYTES", str(80 * 1024 * 1024)))
MAX_EXCEL_BYTES = int(os.getenv("MAX_EXCEL_BYTES", str(25 * 1024 * 1024)))
HEAVY_JOB_TIMEOUT_SECONDS = int(os.getenv("HEAVY_JOB_TIMEOUT_SECONDS", "900"))
STATIC_DIR = Path(__file__).resolve().parent / "static"
HASHED_STATIC_DIR = STATIC_DIR / "_hashed"
HASHED_NAME_RE = re.compile(r"\.[0-9a-f]{8,}\.")
ASSET_VERSIONED_FILES = ("app.js", "ifc_qa_app.js", "session_shared.js", "style.css")
SESSION_ID_RE = re.compile(r"^[0-9a-f]{32}$")


IFC_QA_JOB_STARTER = start_ifc_qa_session_job
if not callable(IFC_QA_JOB_STARTER):
    raise RuntimeError("IFC QA job starter is not callable")

def utc_now() -> datetime.datetime:
    return datetime.datetime.now(UTC)


def resolve_server_host_port() -> tuple[str, int]:
    host = os.getenv("HOST", "0.0.0.0").strip() or "0.0.0.0"
    raw_port = (os.getenv("PORT", "").strip() or os.getenv("APP_PORT", "").strip() or "8000")
    try:
        port = int(raw_port)
    except ValueError:
        APP_LOGGER.warning("Invalid PORT value '%s'; falling back to 8000", raw_port)
        port = 8000
    return host, port


def _resolve_git_commit_sha() -> str:
    env_sha = os.getenv("GIT_COMMIT_SHA", "").strip()
    if env_sha:
        return env_sha
    try:
        proc = subprocess.run(
            ["git", "rev-parse", "HEAD"],
            capture_output=True,
            text=True,
            check=False,
            timeout=5,
            cwd=str(Path(__file__).resolve().parent),
        )
        if proc.returncode == 0:
            return (proc.stdout or "").strip() or "unknown"
    except Exception:
        pass
    return "unknown"


def _build_versioned_asset_map() -> Dict[str, str]:
    HASHED_STATIC_DIR.mkdir(parents=True, exist_ok=True)
    assets: Dict[str, str] = {}
    for relative_name in ASSET_VERSIONED_FILES:
        source_path = STATIC_DIR / relative_name
        if not source_path.exists():
            APP_LOGGER.warning("Static asset missing for versioning: %s", source_path)
            continue
        digest = hashlib.sha256(source_path.read_bytes()).hexdigest()[:12]
        stem = source_path.stem
        hashed_name = f"{stem}.{digest}{source_path.suffix}"
        hashed_rel = f"_hashed/{hashed_name}"
        target_path = STATIC_DIR / hashed_rel
        if not target_path.exists():
            target_path.write_bytes(source_path.read_bytes())
        assets[relative_name] = hashed_rel
    return assets


VERSIONED_ASSET_MAP = _build_versioned_asset_map()
BUILD_TIMESTAMP_UTC = utc_now().strftime("%Y-%m-%dT%H:%M:%SZ")
GIT_SHA = _resolve_git_commit_sha()
FRONTEND_BUNDLE_VERSION = VERSIONED_ASSET_MAP.get("ifc_qa_app.js", "unknown")
FRONTEND_BUILD_ID = f"{BUILD_TIMESTAMP_UTC}|{GIT_SHA[:8]}|{FRONTEND_BUNDLE_VERSION}"


def resolve_asset_url(asset_name: str) -> str:
    versioned = VERSIONED_ASSET_MAP.get(asset_name, asset_name)
    return f"/static/{versioned}"


class CacheControlledStaticFiles(StaticFiles):
    def file_response(self, full_path, stat_result, scope, status_code=200):
        response = super().file_response(full_path, stat_result, scope, status_code=status_code)
        request_path = scope.get("path", "")
        is_hashed = bool(HASHED_NAME_RE.search(request_path))
        if is_hashed:
            response.headers["Cache-Control"] = "public, max-age=31536000, immutable"
        else:
            response.headers["Cache-Control"] = "no-cache, must-revalidate"
        return response


# ----------------------------------------------------------------------------
# Session handling
# ----------------------------------------------------------------------------

def sanitize_filename(base: str) -> str:
    for c in '<>:"/\\|?*':
        base = base.replace(c, "_")
    return base


def human_size(n: int) -> str:
    for unit in ["B", "KB", "MB", "GB"]:
        if n < 1024:
            return f"{n:.0f} {unit}"
        n /= 1024
    return f"{n:.1f} TB"


def _rss_mb() -> Optional[float]:
    if psutil is None:
        return None
    try:
        process = psutil.Process(os.getpid())
        return round(process.memory_info().rss / (1024 * 1024), 2)
    except Exception:
        return None


def log_memory_stage(
    *,
    stage: str,
    session_id: Optional[str],
    file_name: Optional[str],
    file_size: Optional[int],
    endpoint: str,
    started_at: float,
) -> None:
    APP_LOGGER.info(
        "memory_stage stage=%s rss_mb=%s session_id=%s file_name=%s file_size_bytes=%s endpoint=%s elapsed_s=%.3f",
        stage,
        _rss_mb(),
        session_id or "-",
        file_name or "-",
        file_size if file_size is not None else -1,
        endpoint,
        time.monotonic() - started_at,
    )


def enforce_upload_limits(file_path: str, *, endpoint: str) -> None:
    size = os.path.getsize(file_path)
    if size > MAX_UPLOAD_BYTES:
        raise_upload_too_large(
            endpoint=endpoint,
            filename=os.path.basename(file_path),
            actual_size=size,
            rejection_reason="file_size_exceeds_limit",
        )
    APP_LOGGER.info("upload_limit_check endpoint=%s file=%s bytes=%s", endpoint, file_path, size)


def upload_too_large_payload() -> Dict[str, Any]:
    return {
        "code": "UPLOAD_TOO_LARGE",
        "message": "File exceeds the maximum upload size of 1.2 GB.",
        "max_bytes": MAX_UPLOAD_BYTES,
        "max_gb": MAX_UPLOAD_GB,
    }


def log_upload_rejection(
    *,
    endpoint: str,
    filename: str,
    content_length: Optional[int],
    actual_size: Optional[int],
    rejection_reason: str,
) -> None:
    APP_LOGGER.warning(
        "upload_rejected_413 endpoint=%s filename=%s content_length=%s actual_size=%s configured_max_upload_bytes=%s rejection_reason=%s",
        endpoint,
        filename,
        content_length if content_length is not None else -1,
        actual_size if actual_size is not None else -1,
        MAX_UPLOAD_BYTES,
        rejection_reason,
    )


def raise_upload_too_large(
    *,
    endpoint: str,
    filename: str,
    actual_size: Optional[int] = None,
    content_length: Optional[int] = None,
    rejection_reason: str = "file_too_large",
) -> None:
    log_upload_rejection(
        endpoint=endpoint,
        filename=filename,
        content_length=content_length,
        actual_size=actual_size,
        rejection_reason=rejection_reason,
    )
    raise HTTPException(status_code=413, detail=upload_too_large_payload())


@contextmanager
def single_flight_heavy_job(endpoint: str):
    acquired = HEAVY_JOB_SEMAPHORE.acquire(blocking=False)
    if not acquired:
        raise HTTPException(status_code=429, detail="Another heavy IFC job is already running. Please retry shortly.")
    try:
        yield
    finally:
        HEAVY_JOB_SEMAPHORE.release()


def assert_heavy_capacity(endpoint: str) -> None:
    acquired = HEAVY_JOB_SEMAPHORE.acquire(blocking=False)
    if not acquired:
        raise HTTPException(status_code=429, detail=f"{endpoint} is busy; another heavy IFC job is already running.")
    HEAVY_JOB_SEMAPHORE.release()


def has_active_ifc_qa_job() -> bool:
    jobs = getattr(IFC_QA_V2_REGISTRY, "jobs", {})
    for job in jobs.values():
        if job.get("status") in {"queued", "running"}:
            return True
    return False


def _require_valid_session_id(session_id: str) -> str:
    normalized = str(session_id or "").strip().lower()
    if not SESSION_ID_RE.fullmatch(normalized):
        raise HTTPException(status_code=400, detail="Invalid session id format")
    return normalized


def _ensure_session_dir_for_upload(session_id: str) -> str:
    normalized = _require_valid_session_id(session_id)
    root = SESSION_STORE.session_path(normalized)
    os.makedirs(root, exist_ok=True)
    SESSION_STORE.sessions[normalized] = utc_now()
    return root


class SessionStore:
    def __init__(self, base_dir: str, ttl_hours: int = 6) -> None:
        self.base_dir = base_dir
        self.ttl_hours = ttl_hours
        os.makedirs(self.base_dir, exist_ok=True)
        self.sessions: Dict[str, datetime.datetime] = {}

    def create(self) -> str:
        session_id = uuid.uuid4().hex
        os.makedirs(self.session_path(session_id), exist_ok=True)
        now = utc_now()
        self.sessions[session_id] = now
        return session_id

    def session_path(self, session_id: str) -> str:
        return os.path.join(self.base_dir, session_id)

    def touch(self, session_id: str) -> None:
        if not self.exists(session_id):
            raise HTTPException(status_code=404, detail="Session not found")
        self.sessions[session_id] = utc_now()

    def exists(self, session_id: str) -> bool:
        if session_id in self.sessions and os.path.isdir(self.session_path(session_id)):
            return True
        # Recover sessions persisted on disk even when in-memory bookkeeping was lost.
        path = self.session_path(session_id)
        if os.path.isdir(path):
            self.sessions[session_id] = utc_now()
            return True
        return False

    def cleanup_stale(self) -> None:
        cutoff = utc_now() - datetime.timedelta(hours=self.ttl_hours)
        stale = [sid for sid, ts in self.sessions.items() if ts < cutoff]
        for sid in stale:
            self.drop(sid)
        # Remove stray dirs without bookkeeping
        for entry in os.listdir(self.base_dir):
            path = os.path.join(self.base_dir, entry)
            if os.path.isdir(path) and entry not in self.sessions:
                shutil.rmtree(path, ignore_errors=True)

    def drop(self, session_id: str) -> None:
        path = self.session_path(session_id)
        shutil.rmtree(path, ignore_errors=True)
        self.sessions.pop(session_id, None)

    def ensure(self, session_id: str) -> str:
        if not self.exists(session_id):
            raise HTTPException(status_code=404, detail="Session not found")
        self.touch(session_id)
        return self.session_path(session_id)


SESSION_STORE = SessionStore(os.path.join(tempfile.gettempdir(), "ifc_app_sessions"))
STEP2IFC_JOBS: Dict[str, Dict[str, Any]] = {}
DATA_EXTRACT_JOBS: Dict[str, Dict[str, Any]] = {}
IFC_QA_JOBS: Dict[str, Dict[str, Any]] = {}
COBIEQC_JOB_STORE = CobieQcJobStore()
EXCEL_SCAN_CACHE: Dict[str, Dict[str, Any]] = {}


# ----------------------------------------------------------------------------
# Data extractor helpers
# ----------------------------------------------------------------------------

def load_default_config() -> Dict[str, str]:
    config_path = RESOURCE_DIR / "configfile.ini"
    parser = configparser.RawConfigParser()
    if not config_path.exists():
        return {
            "regex_ifc_name": "",
            "regex_ifc_type": "",
            "regex_ifc_system": "",
            "regex_ifc_layer": "",
            "regex_ifc_name_code": "",
            "regex_ifc_type_code": "",
            "regex_ifc_system_code": "",
        }
    parser.read(config_path)
    section = "RegularExpressions"
    return {
        "regex_ifc_name": parser.get(section, "regex_ifc_name", fallback=""),
        "regex_ifc_type": parser.get(section, "regex_ifc_type", fallback=""),
        "regex_ifc_system": parser.get(section, "regex_ifc_system", fallback=""),
        "regex_ifc_layer": parser.get(section, "regex_ifc_layer", fallback=""),
        "regex_ifc_name_code": parser.get(section, "regex_ifc_name_code", fallback=""),
        "regex_ifc_type_code": parser.get(section, "regex_ifc_type_code", fallback=""),
        "regex_ifc_system_code": parser.get(section, "regex_ifc_system_code", fallback=""),
    }


def _clean_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (list, tuple)):
        return ", ".join(_clean_value(v) for v in value if v is not None)
    return str(value)


def _normalize_ifc_value(
    value: Any,
    *,
    prop_name: str = "",
    entity_type: str = "",
    _depth: int = 0,
) -> Any:
    if _depth > 8:
        APP_LOGGER.warning(
            "IFC value normalization max depth reached property=%s entity_type=%s python_type=%s fallback_conversion=%s",
            prop_name,
            entity_type,
            type(value).__name__,
            True,
        )
        return str(value) if value is not None else None
    if value is None:
        return None
    if hasattr(value, "wrappedValue"):
        try:
            return getattr(value, "wrappedValue")
        except Exception:
            APP_LOGGER.warning(
                "IFC wrappedValue read failed property=%s entity_type=%s python_type=%s fallback_conversion=%s",
                prop_name,
                entity_type,
                type(value).__name__,
                True,
            )
    if isinstance(value, tuple):
        return " | ".join("" if v is None else str(_normalize_ifc_value(v, prop_name=prop_name, entity_type=entity_type, _depth=_depth + 1)) for v in value)
    if isinstance(value, list):
        return " | ".join("" if v is None else str(_normalize_ifc_value(v, prop_name=prop_name, entity_type=entity_type, _depth=_depth + 1)) for v in value)
    if isinstance(value, (str, int, float, bool)):
        return value
    for attr in ("value", "Value", "NominalValue"):
        if hasattr(value, attr):
            try:
                nested = getattr(value, attr)
                if nested is not value:
                    return _normalize_ifc_value(nested, prop_name=prop_name, entity_type=entity_type, _depth=_depth + 1)
            except Exception:
                continue
    try:
        APP_LOGGER.warning(
            "IFC value fallback string conversion property=%s entity_type=%s python_type=%s fallback_conversion=%s",
            prop_name,
            entity_type,
            type(value).__name__,
            True,
        )
        return str(value)
    except Exception:
        APP_LOGGER.warning(
            "IFC value conversion failed property=%s entity_type=%s python_type=%s fallback_conversion=%s",
            prop_name,
            entity_type,
            type(value).__name__,
            True,
        )
        return None


def _extract_nominal_value(prop: Any) -> Any:
    nominal = getattr(prop, "NominalValue", None)
    return _normalize_ifc_value(
        nominal,
        prop_name=getattr(prop, "Name", "") or "",
        entity_type=prop.is_a() if hasattr(prop, "is_a") else type(prop).__name__,
    )


def _line_ref(entity: Any) -> str:
    if entity is None:
        return ""
    return f"#{entity.id()}="


def _regex_check(pattern: str, value: str) -> str:
    if not pattern:
        return ""
    try:
        return "True" if re.search(pattern, value or "") else "False"
    except re.error:
        return "False"


def _regex_extract(pattern: str, value: str) -> str:
    if not pattern:
        return ""
    try:
        match = re.search(pattern, value or "")
    except re.error:
        return ""
    if not match:
        return ""
    if match.groups():
        return match.group(1)
    return match.group(0)


def _get_layers_name(entity: Any, model: Optional[ifcopenshell.file] = None) -> str:
    try:
        if model is not None:
            layers = ifcopenshell.util.element.get_layers(model, entity) or []
        else:
            try:
                layers = ifcopenshell.util.element.get_layers(entity) or []
            except TypeError:
                owner = getattr(entity, "file", None)
                if callable(owner):
                    owner = owner()
                layers = ifcopenshell.util.element.get_layers(owner, entity) if owner is not None else []
    except Exception:
        return ""
    names: List[str] = []
    seen: Set[str] = set()
    for layer in layers:
        name = (getattr(layer, "Name", "") or "").strip()
        if name and name not in seen:
            seen.add(name)
            names.append(name)
    return "; ".join(names)


def _get_object_xyz(entity: Any) -> str:
    placement = getattr(entity, "ObjectPlacement", None)
    if not placement:
        return ""
    try:
        matrix = ifcopenshell.util.placement.get_local_placement(placement)
    except Exception:
        return ""
    if matrix is None or len(matrix) < 3:
        return ""
    try:
        x = float(matrix[0][3])
        y = float(matrix[1][3])
        z = float(matrix[2][3])
    except Exception:
        return ""
    return f"{x:.3f},{y:.3f},{z:.3f}"


def _read_csv_first_column(path: Path) -> List[str]:
    values: List[str] = []
    if not path.exists():
        return values
    with open(path, newline="", encoding="utf-8") as handle:
        reader = csv.reader(handle)
        for row in reader:
            if not row:
                continue
            value = row[0].strip()
            if not value:
                continue
            if value.lower().startswith("exclude"):
                continue
            values.append(value)
    return values


def _load_pset_template(path: Path) -> Dict[str, List[str]]:
    mapping: Dict[str, List[str]] = {}
    if not path.exists():
        return mapping
    with open(path, newline="", encoding="utf-8") as handle:
        reader = csv.DictReader(handle)
        for row in reader:
            entity_type = (row.get("IFC_Entity_Occurrence_Type") or "").strip()
            psets_raw = (row.get("Pset_Dictionaries") or "").strip()
            if not entity_type:
                continue
            psets: List[str] = []
            if psets_raw:
                sep = ";" if ";" in psets_raw else ","
                psets = [p.strip() for p in psets_raw.split(sep) if p.strip()]
            mapping[entity_type] = psets
    return mapping


def _iter_object_elements(model: ifcopenshell.file) -> List[Any]:
    objects: List[Any] = []
    spatial_types = {"IfcProject", "IfcSite", "IfcBuilding", "IfcBuildingStorey", "IfcSpace"}
    for obj in model.by_type("IfcProduct"):
        if obj.is_a() in spatial_types:
            continue
        objects.append(obj)
    return objects


def _safe_get_psets(entity: Any) -> Dict[str, Dict[str, Any]]:
    try:
        return ifcopenshell.util.element.get_psets(entity, psets_only=True, include_inherited=True) or {}
    except TypeError:
        return ifcopenshell.util.element.get_psets(entity, psets_only=True) or {}


def _extract_property_single_value(prop: Any) -> Tuple[str, str]:
    if not prop:
        return "", ""
    try:
        if prop.is_a("IfcPropertySingleValue"):
            return _clean_value(_extract_nominal_value(prop)), "IfcPropertySingleValue"
        if prop.is_a("IfcPropertyEnumeratedValue"):
            values = getattr(prop, "EnumerationValues", None) or []
            clean = [
                _clean_value(
                    _normalize_ifc_value(
                        v,
                        prop_name=getattr(prop, "Name", "") or "",
                        entity_type=prop.is_a(),
                    )
                )
                for v in values
            ]
            return ", ".join([v for v in clean if v]), "IfcPropertyEnumeratedValue"
        if prop.is_a("IfcPropertyListValue"):
            values = getattr(prop, "ListValues", None) or []
            clean = [
                _clean_value(
                    _normalize_ifc_value(
                        v,
                        prop_name=getattr(prop, "Name", "") or "",
                        entity_type=prop.is_a(),
                    )
                )
                for v in values
            ]
            return ", ".join([v for v in clean if v]), "IfcPropertyListValue"
        if prop.is_a("IfcPropertyBoundedValue"):
            lower = _normalize_ifc_value(
                getattr(prop, "LowerBoundValue", None),
                prop_name=getattr(prop, "Name", "") or "",
                entity_type=prop.is_a(),
            )
            upper = _normalize_ifc_value(
                getattr(prop, "UpperBoundValue", None),
                prop_name=getattr(prop, "Name", "") or "",
                entity_type=prop.is_a(),
            )
            return f"{_clean_value(lower)}..{_clean_value(upper)}", "IfcPropertyBoundedValue"
        if prop.is_a("IfcPropertyReferenceValue"):
            value = _normalize_ifc_value(
                getattr(prop, "PropertyReference", None),
                prop_name=getattr(prop, "Name", "") or "",
                entity_type=prop.is_a(),
            )
            return _clean_value(value), "IfcPropertyReferenceValue"
        if prop.is_a("IfcComplexProperty"):
            return _clean_value(getattr(prop, "UsageName", "")), "IfcComplexProperty"
        return "", prop.is_a()
    except Exception:
        APP_LOGGER.exception(
            "Failed to extract IFC property value property=%s entity_type=%s",
            getattr(prop, "Name", "") or "",
            prop.is_a() if hasattr(prop, "is_a") else type(prop).__name__,
        )
        return "", prop.is_a() if hasattr(prop, "is_a") else type(prop).__name__


def _extract_quantity_value(quantity: Any) -> Tuple[str, str]:
    for attr in (
        "LengthValue",
        "AreaValue",
        "VolumeValue",
        "CountValue",
        "WeightValue",
        "TimeValue",
    ):
        if hasattr(quantity, attr):
            value = getattr(quantity, attr, None)
            if value is not None:
                return _clean_value(value), quantity.is_a()
    if hasattr(quantity, "NominalValue"):
        value = getattr(quantity, "NominalValue", None)
        return _clean_value(
            _normalize_ifc_value(
                value,
                prop_name=getattr(quantity, "Name", "") or "",
                entity_type=quantity.is_a(),
            )
        ), quantity.is_a()
    return "", quantity.is_a()


def _iter_occurrence_property_rows(obj: Any, allowed_psets: Optional[List[str]]) -> List[Tuple[str, str, str, str]]:
    rows: List[Tuple[str, str, str, str]] = []
    for rel in getattr(obj, "IsDefinedBy", None) or []:
        if not rel or not rel.is_a("IfcRelDefinesByProperties"):
            continue
        pdef = getattr(rel, "RelatingPropertyDefinition", None)
        if not pdef:
            continue
        if pdef.is_a("IfcPropertySet"):
            pset_name = getattr(pdef, "Name", "") or ""
            if allowed_psets is not None and allowed_psets and pset_name not in allowed_psets:
                continue
            for prop in getattr(pdef, "HasProperties", None) or []:
                value, value_type = _extract_property_single_value(prop)
                rows.append((pset_name, getattr(prop, "Name", "") or "", value, value_type))
        elif pdef.is_a("IfcElementQuantity"):
            for quantity in getattr(pdef, "Quantities", None) or []:
                value, value_type = _extract_quantity_value(quantity)
                rows.append(("BaseQuantities", getattr(quantity, "Name", "") or "", value, value_type))
    return rows


def _iter_type_property_rows(type_obj: Any, allowed_psets: Optional[List[str]]) -> List[Tuple[str, str, str, str]]:
    if not type_obj:
        return []
    rows: List[Tuple[str, str, str, str]] = []
    for pdef in getattr(type_obj, "HasPropertySets", None) or []:
        if not pdef:
            continue
        if pdef.is_a("IfcPropertySet"):
            pset_name = getattr(pdef, "Name", "") or ""
            if allowed_psets is not None and allowed_psets and pset_name not in allowed_psets:
                continue
            for prop in getattr(pdef, "HasProperties", None) or []:
                value, value_type = _extract_property_single_value(prop)
                rows.append((pset_name, getattr(prop, "Name", "") or "", value, value_type))
        elif pdef.is_a("IfcElementQuantity"):
            for quantity in getattr(pdef, "Quantities", None) or []:
                value, value_type = _extract_quantity_value(quantity)
                rows.append(("BaseQuantities", getattr(quantity, "Name", "") or "", value, value_type))
    return rows


def _iter_entity_classifications(entity: Any) -> List[Tuple[str, str, str, str]]:
    rows: List[Tuple[str, str, str, str]] = []
    for rel in getattr(entity, "HasAssociations", None) or []:
        if not rel or not rel.is_a("IfcRelAssociatesClassification"):
            continue
        classification = getattr(rel, "RelatingClassification", None)
        if not classification:
            continue
        sys_name = getattr(classification, "Name", "") or ""
        code = getattr(classification, "Identification", "") or ""
        desc = getattr(classification, "Description", "") or ""
        name = getattr(classification, "Name", "") or ""
        if classification.is_a("IfcClassificationReference"):
            source = getattr(classification, "ReferencedSource", None)
            sys_name = (getattr(source, "Name", "") if source else "") or name
            code = getattr(classification, "Identification", "") or getattr(classification, "ItemReference", "") or ""
        rows.append((sys_name, name, code, desc))
    return rows


def _qa_override_dir(session_id: str) -> Path:
    root = Path(SESSION_STORE.ensure(session_id))
    override_dir = root / "ifc_qa_overrides"
    override_dir.mkdir(parents=True, exist_ok=True)
    return override_dir


def _qa_default_paths() -> Dict[str, Path]:
    return {
        "qa_rules": QA_RESOURCE_DIR / "qa_rules.template.csv",
        "qa_property_requirements": QA_RESOURCE_DIR / "qa_property_requirements.template.csv",
        "qa_unacceptable_values": QA_RESOURCE_DIR / "qa_unacceptable_values.template.csv",
        "regex_patterns": QA_RESOURCE_DIR / "regex_patterns.template.csv",
        "exclude_filter": QA_RESOURCE_DIR / "exclude_filter.template.csv",
        "pset_template": QA_RESOURCE_DIR / "pset_template.template.csv",
    }


def _qa_config_path(session_id: Optional[str], key: str, override_path: Optional[Path] = None) -> Path:
    defaults = _qa_default_paths()
    if override_path and override_path.exists():
        return override_path
    if session_id:
        candidate = _qa_override_dir(session_id) / f"{key}.csv"
        if candidate.exists():
            return candidate
    return defaults[key]


def _load_regex_patterns(path: Path) -> Dict[str, Dict[str, str]]:
    patterns: Dict[str, Dict[str, str]] = {}
    if not path.exists():
        return patterns
    with open(path, newline="", encoding="utf-8") as handle:
        reader = csv.DictReader(handle)
        for row in reader:
            key = (row.get("key") or "").strip()
            if not key:
                continue
            patterns[key] = {
                "pattern": row.get("pattern", ""),
                "enabled": (row.get("enabled", "true") or "true").strip().lower(),
            }
    return patterns


# ----------------------------------------------------------------------------
# IFC cleaner (from original app.py)
# ----------------------------------------------------------------------------

def clean_ifc_file(
    in_path: str,
    out_path: str,
    prefix: str = "InfoDrainage",
    case_insensitive: bool = True,
    delete_psets_with_prefix: bool = True,
    delete_properties_in_other_psets: bool = True,
    drop_empty_psets: bool = True,
    also_remove_loose_props: bool = True,
) -> Dict[str, Any]:
    def starts_with(s: str) -> bool:
        if s is None:
            return False
        if case_insensitive:
            return s.lower().startswith(prefix.lower())
        return s.startswith(prefix)

    report = {
        "input": os.path.basename(in_path),
        "output": os.path.basename(out_path),
        "prefix": prefix,
        "case_insensitive": case_insensitive,
        "removed": {
            "IfcPropertySet": 0,
            "IfcRelDefinesByProperties": 0,
            "emptied_psets": 0,
            "IfcPropertySingleValue": 0,
            "IfcComplexProperty": 0,
            "IfcPropertyEnumeratedValue": 0,
            "IfcPropertyReferenceValue": 0,
            "IfcPropertyListValue": 0,
            "IfcPropertyTableValue": 0,
            "loose_properties": 0,
        },
        "notes": [],
        "status": "started",
    }

    f = ifcopenshell.open(in_path)

    psets_to_delete = set()
    if delete_psets_with_prefix:
        for pset in f.by_type("IfcPropertySet"):
            try:
                if starts_with(getattr(pset, "Name", None)):
                    psets_to_delete.add(pset)
            except Exception:
                pass

    prop_types = [
        "IfcPropertySingleValue",
        "IfcComplexProperty",
        "IfcPropertyEnumeratedValue",
        "IfcPropertyReferenceValue",
        "IfcPropertyListValue",
        "IfcPropertyTableValue",
    ]
    prop_removed_count = {t: 0 for t in prop_types}
    emptied_pset_count = 0

    if delete_properties_in_other_psets:
        for pset in f.by_type("IfcPropertySet"):
            try:
                if pset in psets_to_delete:
                    continue

                props = list(getattr(pset, "HasProperties", []) or [])
                if not props:
                    continue

                to_keep = []
                for p in props:
                    nm = getattr(p, "Name", None)
                    if nm and starts_with(nm):
                        try:
                            kind = p.is_a()
                        except Exception:
                            kind = None
                        try:
                            f.remove(p)
                            if kind in prop_removed_count:
                                prop_removed_count[kind] += 1
                        except Exception:
                            if kind in prop_removed_count:
                                prop_removed_count[kind] += 1
                    else:
                        to_keep.append(p)

                try:
                    pset.HasProperties = tuple(to_keep)
                except Exception:
                    pass

                if drop_empty_psets and len(getattr(pset, "HasProperties", []) or []) == 0:
                    psets_to_delete.add(pset)
                    emptied_pset_count += 1
            except Exception:
                pass

    rel_del_count = 0
    pset_del_count = 0

    for pset in list(psets_to_delete):
        inverses = []
        try:
            inverses = f.get_inverse(pset)
        except Exception:
            pass

        for rel in list(inverses or []):
            try:
                if rel.is_a("IfcRelDefinesByProperties") and rel.RelatingPropertyDefinition == pset:
                    f.remove(rel)
                    rel_del_count += 1
            except Exception:
                pass

        try:
            f.remove(pset)
            pset_del_count += 1
        except Exception:
            pass

    loose_removed = 0
    if also_remove_loose_props:
        for t in prop_types:
            for p in list(f.by_type(t)):
                nm = getattr(p, "Name", None)
                if nm and starts_with(nm):
                    try:
                        f.remove(p)
                        prop_removed_count[t] += 1
                        loose_removed += 1
                    except Exception:
                        pass

    try:
        f.write(out_path)
        status = "success"
    except Exception as e:
        status = "save_failed"
        report["notes"].append(f"Save error: {e!r}")

    report["removed"]["IfcPropertySet"] = pset_del_count
    report["removed"]["IfcRelDefinesByProperties"] = rel_del_count
    report["removed"]["emptied_psets"] = emptied_pset_count
    report["removed"]["loose_properties"] = loose_removed
    for t, c in prop_removed_count.items():
        report["removed"][t] = c

    report["status"] = status
    return report


# ----------------------------------------------------------------------------
# Excel extractor/updater (from app (1).py)
# ----------------------------------------------------------------------------

COBIE_MAPPING = {
    "COBie_Specification": {"scope": "T", "props": [
        ("NominalLength", "Length"),
        ("NominalWidth", "Length"),
        ("NominalHeight", "Length"),
        ("Shape", "Text"),
        ("Size", "Text"),
        ("Color", "Text"),
        ("Finish", "Text"),
        ("Grade", "Text"),
        ("Material", "Text"),
        ("Constituents", "Text"),
        ("Features", "Text"),
        ("AccessibilityPerformance", "Text"),
        ("CodePerformance", "Text"),
        ("SustainabilityPerformance", "Text"),
    ]},
    "COBie_Component": {"scope": "I", "props": [
        ("COBie", "Boolean"),
        ("InstallationDate", "Text"),
        ("WarrantyStartDate", "Text"),
        ("TagNumber", "Text"),
        ("AssetIdentifier", "Text"),
        ("Space", "Text"),
        ("CreatedBy", "Text"),
        ("CreatedOn", "Text"),
        ("Name", "Text"),
        ("Description", "Text"),
        ("Area", "Area"),
        ("Length", "Length"),
    ]},
    "COBie_Asset": {"scope": "T", "props": [
        ("AssetType", "Text"),
    ]},
    "COBie_Warranty": {"scope": "T", "props": [
        ("WarrantyDurationParts", "Real"),
        ("WarrantyGuarantorLabor", "Text"),
        ("WarrantyDurationLabor", "Real"),
        ("WarrantyDurationDescription", "Text"),
        ("WarrantyDurationUnit", "Text"),
        ("WarrantyGuarantorParts", "Text"),
    ]},
    "Pset_ManufacturerOccurence": {"scope": "I", "props": [
        ("SerialNumber", "Text"),
        ("BarCode", "Text"),
    ]},
    "COBie_ServiceLife": {"scope": "T", "props": [
        ("ServiceLifeDuration", "Real"),
        ("DurationUnit", "Text"),
    ]},
    "COBie_EconomicalImpactValues": {"scope": "T", "props": [
        ("ReplacementCost", "Real"),
    ]},
    "COBie_Type": {"scope": "T", "props": [
        ("COBie", "Boolean"),
        ("CreatedBy", "Text"),
        ("CreatedOn", "Text"),
        ("Name", "Text"),
        ("Description", "Text"),
        ("Category", "Text"),
        ("Area", "Area"),
        ("Length", "Length"),
    ]},
    "COBie_System": {"scope": "I", "props": [
        ("Name", "Text"),
        ("Description", "Text"),
        ("Category", "Text"),
    ]},
    "Classification_General": {"scope": "T", "props": [
        ("Classification.Uniclass.Pr.Number", "Text"),
        ("Classification.Uniclass.Pr.Description", "Text"),
        ("Classification.Uniclass.Ss.Number", "Text"),
        ("Classification.Uniclass.Ss.Description", "Text"),
        ("Classification.NRM1.Number", "Text"),
        ("Classification.NRM1.Description", "Text"),
    ]},
    "Pset_ManufacturerTypeInformation": {"scope": "T", "props": [
        ("Manufacturer", "Text"),
        ("ModelNumber", "Text"),
        ("ModelReference", "Text"),
    ]},
    "PPset_DoorCommon": {"scope": "T", "props": [
        ("FireRating", "Text"),
    ]},
    "Pset_BuildingCommon": {"scope": "T", "props": [
        ("NumberOfStoreys", "Text"),
    ]},
    "COBie_Space": {"scope": "T", "props": [
        ("RoomTag", "Text"),
    ]},
    "COBie_BuildingCommon_UK": {"scope": "T", "props": [
        ("UPRN", "Text"),
    ]},
    "Additional_Pset_BuildingCommon": {"scope": "T", "props": [
        ("BlockConstructionType", "Text"),
        ("MaximumBlockHeight", "Text"),
    ]},
    "Additional_Pset_SystemCommon": {"scope": "T", "props": [
        ("SystemCategory", "Text"),
        ("SystemDescription", "Text"),
        ("SystemName", "Text"),
    ]},
}


RE_SPLIT_LIST = re.compile(r"[;,|\n]+|\s{2,}")

CIVIL3D_EXTENDED_FIELDS: Tuple[str, ...] = (
    "Classification.Uniclass.Pr.Description",
    "Classification.Uniclass.Pr.Number",
    "ClassificationCode",
    "ExtObject",
    "IFC_Enumeration",
    "IFC Name",
    "IFCPresentationLayer",
    "Name",
    "Structural Material",
    "SystemCategory",
    "SystemDescription",
    "SystemName",
    "Type (User Defined)",
    "Uniclass2015_Pr",
)

EXTRACTION_PROFILES: Dict[str, Dict[str, Any]] = {
    "civil3d_extended": {
        "include_sheets": {"ProjectData", "Elements", "Properties", "COBieMapping", "Uniclass_Pr", "Uniclass_Ss", "Uniclass_EF"},
        "include_type_properties": True,
        "include_spatial_fields": True,
        "include_classifications": True,
        "civil3d_extended": True,
    }
}


def path_of(f):
    return f if isinstance(f, str) else getattr(f, "name", f)


def clean_value(v):
    if pd.isna(v):
        return None
    if isinstance(v, str):
        v = v.strip()
        if v == "":
            return None
    return v


def ensure_aggregates(parent, child, ifc):
    rel = None
    for r in parent.IsDecomposedBy or []:
        if r.is_a("IfcRelAggregates"):
            rel = r
            break
    if rel is None:
        ifc.create_entity(
            "IfcRelAggregates",
            GlobalId=new_guid(),
            RelatingObject=parent,
            RelatedObjects=[child],
        )
    else:
        if child not in rel.RelatedObjects:
            rel.RelatedObjects = list(rel.RelatedObjects) + [child]


def reassign_aggregate(parent, child, ifc):
    for rel in list(child.Decomposes or []):
        if not rel.is_a("IfcRelAggregates"):
            continue
        if rel.RelatingObject == parent:
            continue
        related = list(rel.RelatedObjects)
        if child in related:
            related.remove(child)
            if related:
                rel.RelatedObjects = related
            else:
                ifc.remove(rel)
    ensure_aggregates(parent, child, ifc)


def parse_required_pairs(raw):
    if not raw or not isinstance(raw, str):
        return []
    items = [t.strip() for t in RE_SPLIT_LIST.split(raw) if t.strip()]
    pairs = []
    for token in items:
        if "." in token:
            pset, prop = token.split(".", 1)
            pset, prop = pset.strip(), prop.strip()
            if pset and prop:
                pairs.append((pset, prop))
    return pairs


@dataclass
class ExtractionPlan:
    include_sheets: Set[str] = field(default_factory=lambda: {"ProjectData", "Elements", "Types", "RawEntities", "Properties", "COBieMapping", "Uniclass_Pr", "Uniclass_Ss", "Uniclass_EF", "ChangeLog"})
    entity_classes: Set[str] = field(default_factory=set)
    property_sets: Set[str] = field(default_factory=set)
    quantity_sets: Set[str] = field(default_factory=set)
    cobie_pairs: Set[Tuple[str, str]] = field(default_factory=set)
    include_type_properties: bool = True
    include_spatial_fields: bool = True
    include_classifications: bool = True
    civil3d_extended: bool = False


class StageTimer:
    def __init__(self) -> None:
        self._marks: Dict[str, float] = {}
        self.timings: Dict[str, float] = {}

    def start(self, name: str) -> None:
        self._marks[name] = time.perf_counter()

    def stop(self, name: str) -> None:
        start = self._marks.pop(name, None)
        if start is not None:
            self.timings[name] = round((time.perf_counter() - start) * 1000.0, 2)

    def as_payload(self) -> Dict[str, float]:
        return dict(self.timings)


def _extract_uniclass(entity: Any, target_name: str, is_ifc2x3: bool) -> Tuple[str, str]:
    for rel in getattr(entity, "HasAssociations", []) or []:
        if not rel.is_a("IfcRelAssociatesClassification"):
            continue
        classification_ref = rel.RelatingClassification
        if not classification_ref or not classification_ref.is_a("IfcClassificationReference"):
            continue
        if is_ifc2x3:
            if getattr(classification_ref, "Name", "") == target_name:
                return getattr(classification_ref, "ItemReference", ""), getattr(classification_ref, "Name", "")
            continue
        src = getattr(classification_ref, "ReferencedSource", None)
        if src and getattr(src, "Name", "") == target_name:
            return getattr(classification_ref, "ItemReference", ""), getattr(classification_ref, "Name", "")
    return "", ""


_EN_ENTITIES_NOOP_VALUES = {"", "n/a", "na", "none", "null"}


def _normalize_en_entities_key(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip().lower()


def _read_projectdata_en_entities(project_df: pd.DataFrame) -> Tuple[Optional[str], Optional[str]]:
    if project_df is None or project_df.empty:
        return None, None

    normalized_columns = {_normalize_en_entities_key(col): col for col in project_df.columns}
    ref_col = normalized_columns.get("uniclassenreference")
    name_col = normalized_columns.get("uniclassenname")
    if ref_col is None:
        return None, None

    project_row = None
    data_type_col = normalized_columns.get("datatype")
    if data_type_col is not None:
        for _, row in project_df.iterrows():
            if _normalize_en_entities_key(row.get(data_type_col)) == "project":
                project_row = row
                break
    if project_row is None:
        project_row = project_df.iloc[0]

    raw_ref = clean_value(project_row.get(ref_col))
    if raw_ref is None:
        return None, None
    ref_value = str(raw_ref).strip()
    if ref_value.lower() in _EN_ENTITIES_NOOP_VALUES:
        return None, None

    raw_name = clean_value(project_row.get(name_col)) if name_col is not None else None
    if raw_name is None:
        return ref_value, ref_value
    name_value = str(raw_name).strip()
    if name_value.lower() in _EN_ENTITIES_NOOP_VALUES:
        return ref_value, ref_value
    return ref_value, name_value


def _ensure_en_entities_classification_rel(
    ifc: Any,
    target_entity: Any,
    en_entities_value: str,
    en_entities_name: Optional[str] = None,
    *,
    source_name: str = "Uniclass En Entities",
) -> Tuple[Any, bool]:
    schema_upper = (ifc.schema or "").upper()
    is_ifc2x3 = schema_upper == "IFC2X3"
    owner_history = next(iter(ifc.by_type("IfcOwnerHistory")), None)

    cls_src = None
    for c in ifc.by_type("IfcClassification"):
        if str(getattr(c, "Name", "")).strip().lower() == source_name.strip().lower():
            cls_src = c
            break
    if cls_src is None:
        cls_src = ifc.create_entity("IfcClassification", Name=source_name)

    existing_rel = None
    existing_ref = None
    for rel in getattr(target_entity, "HasAssociations", []) or []:
        if not rel.is_a("IfcRelAssociatesClassification"):
            continue
        cref = rel.RelatingClassification
        if not cref or not cref.is_a("IfcClassificationReference"):
            continue
        src = getattr(cref, "ReferencedSource", None)
        src_name = str(getattr(src, "Name", "") or getattr(cref, "Name", "")).strip().lower()
        if src_name == source_name.strip().lower():
            existing_rel = rel
            existing_ref = cref
            break
    APP_LOGGER.info(
        "EN Entities existing classification association found=%s target_id=%s",
        existing_rel is not None,
        target_entity.id() if target_entity is not None else None,
    )

    if existing_ref is None:
        if is_ifc2x3:
            existing_ref = ifc.create_entity("IfcClassificationReference", ItemReference=en_entities_value, Name=en_entities_name or en_entities_value)
        else:
            existing_ref = ifc.create_entity("IfcClassificationReference", Identification=en_entities_value, Name=en_entities_name or en_entities_value)
        existing_ref.ReferencedSource = cls_src
    else:
        if is_ifc2x3 and hasattr(existing_ref, "ItemReference"):
            existing_ref.ItemReference = en_entities_value
        elif not is_ifc2x3 and hasattr(existing_ref, "Identification"):
            existing_ref.Identification = en_entities_value
        elif hasattr(existing_ref, "ItemReference"):
            existing_ref.ItemReference = en_entities_value
        else:
            existing_ref.Name = en_entities_value
        if en_entities_name:
            existing_ref.Name = en_entities_name
        if getattr(existing_ref, "ReferencedSource", None) is None:
            existing_ref.ReferencedSource = cls_src

    if existing_rel is not None:
        if target_entity not in (existing_rel.RelatedObjects or []):
            existing_rel.RelatedObjects = list(existing_rel.RelatedObjects or []) + [target_entity]
        return existing_rel, False

    rel_kwargs: Dict[str, Any] = {
        "GlobalId": new_guid(),
        "RelatedObjects": [target_entity],
        "RelatingClassification": existing_ref,
    }
    if is_ifc2x3 and owner_history is not None:
        rel_kwargs["OwnerHistory"] = owner_history
    created_rel = ifc.create_entity("IfcRelAssociatesClassification", **rel_kwargs)
    return created_rel, True


def _validate_en_entities_writeback(ifc_path: str, expected_value: str) -> Tuple[bool, str]:
    model = ifcopenshell.open(ifc_path)
    expected_value_lower = expected_value.strip().lower()
    for rel in model.by_type("IfcRelAssociatesClassification"):
        cref = getattr(rel, "RelatingClassification", None)
        if not cref or not cref.is_a("IfcClassificationReference"):
            continue
        src = getattr(cref, "ReferencedSource", None)
        src_name = str(getattr(src, "Name", "") or getattr(cref, "Name", "")).strip().lower()
        if src_name != "uniclass en entities":
            continue
        ident = str(getattr(cref, "Identification", "") or getattr(cref, "ItemReference", "") or "").strip().lower()
        if ident != expected_value_lower:
            continue
        related = list(getattr(rel, "RelatedObjects", []) or [])
        if not related:
            continue
        all_related_valid = all(model.by_id(obj.id()) is not None for obj in related)
        if not all_related_valid:
            return False, "EN Entities classification found but includes invalid RelatedObjects."
        return True, f"Validated relation #{rel.id()} with {len(related)} related object(s)."
    return False, "No IFCRELASSOCIATESCLASSIFICATION found for the EN Entities value."


def _parse_excel_extraction_plan(payload: Optional[Dict[str, Any]]) -> ExtractionPlan:
    if not payload:
        return ExtractionPlan()
    profile_name = (payload.get("profile") or "").strip()
    if profile_name:
        profile = EXTRACTION_PROFILES.get(profile_name)
        if not profile:
            APP_LOGGER.warning("Unknown Excel extraction profile '%s'; using defaults", profile_name)
            profile = {}
    else:
        profile = {}
    include_sheets = set(payload.get("include_sheets") or []) or ExtractionPlan().include_sheets
    include_sheets = set(profile.get("include_sheets", include_sheets))
    entity_classes = set(payload.get("entity_classes") or [])
    property_sets = set(payload.get("property_sets") or [])
    quantity_sets = set(payload.get("quantity_sets") or [])
    include_type_properties = bool(payload.get("include_type_properties", True))
    include_spatial_fields = bool(payload.get("include_spatial_fields", True))
    include_classifications = bool(payload.get("include_classifications", True))
    if "include_type_properties" in profile:
        include_type_properties = bool(profile["include_type_properties"])
    if "include_spatial_fields" in profile:
        include_spatial_fields = bool(profile["include_spatial_fields"])
    if "include_classifications" in profile:
        include_classifications = bool(profile["include_classifications"])
    raw_pairs = payload.get("cobie_pairs") or []
    cobie_pairs: Set[Tuple[str, str]] = set()
    for item in raw_pairs:
        if isinstance(item, str) and "." in item:
            pset, prop = item.split(".", 1)
            if pset and prop:
                cobie_pairs.add((pset.strip(), prop.strip()))
        elif isinstance(item, dict):
            pset = (item.get("pset") or "").strip()
            prop = (item.get("property") or "").strip()
            if pset and prop:
                cobie_pairs.add((pset, prop))
    return ExtractionPlan(
        include_sheets=include_sheets,
        entity_classes=entity_classes,
        property_sets=property_sets,
        quantity_sets=quantity_sets,
        cobie_pairs=cobie_pairs,
        include_type_properties=include_type_properties,
        include_spatial_fields=include_spatial_fields,
        include_classifications=include_classifications,
        civil3d_extended=bool(payload.get("civil3d_extended", profile.get("civil3d_extended", False))),
    )


_IFC2X3_ENTITY_MAPPING_CACHE: Optional[Dict[str, Any]] = None


def load_ifc2x3_entity_mapping() -> Dict[str, Any]:
    global _IFC2X3_ENTITY_MAPPING_CACHE
    if _IFC2X3_ENTITY_MAPPING_CACHE is not None:
        return _IFC2X3_ENTITY_MAPPING_CACHE
    payload = json.loads((BACKEND_CONFIG_DIR / "ifc2x3_entities_and_predefined_types.json").read_text(encoding="utf-8"))
    _IFC2X3_ENTITY_MAPPING_CACHE = payload
    return payload


def _excel_range(sheet_name: str, column_letter: str, start_row: int, end_row: int) -> str:
    return f"'{sheet_name}'!${column_letter}${start_row}:${column_letter}${end_row}"


def _upsert_workbook_defined_name(workbook: Any, name: str, attr_text: str) -> None:
    # Workbook-level named ranges back DataValidation dropdowns. This avoids long inline list
    # formulas, which can generate invalid XML and trigger Excel repair warnings.
    existing = workbook.defined_names.get(name)
    if existing is not None:
        workbook.defined_names.delete(name)
    workbook.defined_names.add(DefinedName(name=name, attr_text=attr_text))


def detect_ifc_schema_from_header(ifc_path: str) -> Tuple[str, str]:
    try:
        with open(ifc_path, "r", encoding="utf-8", errors="ignore") as handle:
            for _ in range(80):
                line = handle.readline()
                if not line:
                    break
                upper = line.upper()
                if "FILE_SCHEMA" not in upper:
                    continue
                if "IFC2X3" in upper:
                    return "IFC2X3", ""
                if "IFC4X3" in upper:
                    return "IFC4X3", ""
                if "IFC4" in upper:
                    return "IFC4", ""
    except Exception:
        pass
    return "IFC4", "Unable to parse FILE_SCHEMA from IFC header; using IFC4 fallback lists."


def _sanitize_excel_text(value: Any) -> Any:
    if isinstance(value, str):
        return ILLEGAL_CHARACTERS_RE.sub("", value)
    return value


def _sanitize_dataframe_for_excel(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return df
    return df.map(_sanitize_excel_text)


def validate_workbook_after_export(path: str) -> Optional[str]:
    try:
        wb = load_workbook(path, data_only=False)
        if not wb.sheetnames:
            return "Workbook has no sheets"
        seen = set()
        for ws in wb.worksheets:
            if ws.title in seen:
                return f"Duplicate worksheet title detected: {ws.title}"
            seen.add(ws.title)
            _ = ws.max_row
            _ = ws.max_column
            if ws.data_validations is not None:
                for dv in ws.data_validations.dataValidation:
                    if not dv.sqref:
                        return f"Invalid data validation range on sheet {ws.title}"
        for dn in wb.defined_names.values():
            if not dn.name:
                return "Workbook contains an unnamed defined name"
        wb.close()
        return None
    except Exception as exc:
        return f"Workbook validation failed: {exc}"


def _normalize_field_key(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", (value or "").strip().lower())


def _resolve_field_value(
    elem: Any,
    type_obj: Any,
    field_name: str,
    get_pset_value: Callable[[Any, str, str], Any],
) -> str:
    if field_name == "IFCPresentationLayer":
        return _get_layers_name(elem, None)
    if field_name == "Name":
        return getattr(elem, "Name", "") or ""
    if field_name == "IFC Name":
        candidates = ["IFC Name", "IFC_Name", "IfcName"]
        for pset_name in ("Additional_Pset_GeneralCommon", "Additional_Pset_SystemCommon", "Pset_ManufacturerTypeInformation", "Pset_ManufacturerOccurrence"):
            for candidate in candidates:
                val = get_pset_value(elem, pset_name, candidate)
                if clean_value(val) is not None:
                    return str(val)
        direct = getattr(elem, "Name", "") or ""
        if direct:
            return direct
    if field_name == "Type (User Defined)":
        return getattr(elem, "ObjectType", "") or (getattr(type_obj, "ObjectType", "") if type_obj else "") or ""
    if field_name == "ClassificationCode":
        candidates = ["ClassificationCode", "Classification Code"]
    elif field_name == "Classification.Uniclass.Pr.Number":
        candidates = ["Classification.Uniclass.Pr.Number", "Uniclass2015_Pr", "Uniclass Pr Number"]
    elif field_name == "Classification.Uniclass.Pr.Description":
        candidates = ["Classification.Uniclass.Pr.Description", "Uniclass Pr Description"]
    else:
        candidates = [field_name]

    for pset_name in ("Additional_Pset_GeneralCommon", "Additional_Pset_SystemCommon", "Pset_ManufacturerTypeInformation", "Pset_ManufacturerOccurrence"):
        for candidate in candidates:
            val = get_pset_value(elem, pset_name, candidate)
            if clean_value(val) is not None:
                return str(val)

    psets = _safe_get_psets(elem)
    if type_obj is not None:
        t_psets = _safe_get_psets(type_obj)
    else:
        t_psets = {}
    target_keys = {_normalize_field_key(c) for c in candidates}
    for source_psets in (psets, t_psets):
        for _, props in source_psets.items():
            for prop_name, prop_value in (props or {}).items():
                if _normalize_field_key(prop_name) in target_keys and clean_value(prop_value) is not None:
                    return str(prop_value)
    return ""


def _resolve_name_with_priority(elem: Any, type_obj: Any, get_pset_value: Callable[[Any, str, str], Any]) -> str:
    for source in ("IFC Name", "Name"):
        value = _resolve_field_value(elem, type_obj, source, get_pset_value)
        if clean_value(value) is not None:
            APP_LOGGER.info("Name resolution for %s used source '%s'", getattr(elem, "GlobalId", elem.id()), source)
            return str(value)
    fallback = getattr(type_obj, "Name", "") if type_obj else ""
    APP_LOGGER.info("Name resolution for %s fell back to legacy type/block name", getattr(elem, "GlobalId", elem.id()))
    return fallback or ""


def _resolve_type_name_with_priority(elem: Any, type_obj: Any, get_pset_value: Callable[[Any, str, str], Any]) -> str:
    for source in ("IFC Name", "Type (User Defined)", "Name"):
        value = _resolve_field_value(elem, type_obj, source, get_pset_value)
        if clean_value(value) is not None:
            APP_LOGGER.info("IFCElementType.Name resolution for %s used source '%s'", getattr(elem, "GlobalId", elem.id()), source)
            return str(value)
    fallback = getattr(type_obj, "Name", "") if type_obj else ""
    APP_LOGGER.info("IFCElementType.Name for %s fell back to legacy type/block name", getattr(elem, "GlobalId", elem.id()))
    return fallback or ""


def scan_model_for_excel_preview(ifc_path: str, timer: Optional[StageTimer] = None) -> Dict[str, Any]:
    timer = timer or StageTimer()
    timer.start("model_load")
    ifc = ifcopenshell.open(ifc_path)
    timer.stop("model_load")

    timer.start("entity_index")
    elements = list(ifc.by_type("IfcElement"))
    type_cache: Dict[int, Any] = {}
    pset_names: Dict[str, int] = {}
    quantity_names: Dict[str, Dict[str, int]] = {}
    property_names_by_pset: Dict[str, Dict[str, int]] = {}
    class_counts: Dict[str, int] = {}
    classification_names: Dict[str, int] = {}
    classification_targets = {
        "Uniclass Pr Products",
        "Uniclass Ss Systems",
        "Uniclass EF Elements Functions",
        "Uniclass En Entities",
    }
    cobie_dynamic_pairs: Set[Tuple[str, str]] = set()
    for elem in elements:
        class_name = elem.is_a()
        class_counts[class_name] = class_counts.get(class_name, 0) + 1

        elem_id = elem.id()
        type_obj = type_cache.get(elem_id)
        if type_obj is None:
            type_obj = ifcopenshell.util.element.get_type(elem)
            type_cache[elem_id] = type_obj

        psets_elem = _safe_get_psets(elem)
        add_pset = psets_elem.get("Additional_Pset_GeneralCommon", {})
        cobie_dynamic_pairs.update(parse_required_pairs(add_pset.get("RequiredForCOBie", "")))
        cobie_dynamic_pairs.update(parse_required_pairs(add_pset.get("RequiredForCOBieComponent", "")))

        for pset_name, values in psets_elem.items():
            pset_names[pset_name] = pset_names.get(pset_name, 0) + 1
            property_names_by_pset.setdefault(pset_name, {})
            for prop_name in values.keys():
                property_names_by_pset[pset_name][prop_name] = property_names_by_pset[pset_name].get(prop_name, 0) + 1

        if type_obj:
            type_psets = _safe_get_psets(type_obj)
            add_pset_t = type_psets.get("Additional_Pset_GeneralCommon", {})
            cobie_dynamic_pairs.update(parse_required_pairs(add_pset_t.get("RequiredForCOBie", "")))
            cobie_dynamic_pairs.update(parse_required_pairs(add_pset_t.get("RequiredForCOBieComponent", "")))

        for rel in getattr(elem, "IsDefinedBy", []) or []:
            if not rel.is_a("IfcRelDefinesByProperties"):
                continue
            qset = rel.RelatingPropertyDefinition
            if not qset or not qset.is_a("IfcElementQuantity"):
                continue
            q_name = getattr(qset, "Name", "") or ""
            if not q_name:
                continue
            quantity_names.setdefault(q_name, {})
            for qty in getattr(qset, "Quantities", []) or []:
                q_prop = getattr(qty, "Name", "") or ""
                if q_prop:
                    quantity_names[q_name][q_prop] = quantity_names[q_name].get(q_prop, 0) + 1

        for rel in getattr(elem, "HasAssociations", []) or []:
            if not rel.is_a("IfcRelAssociatesClassification"):
                continue
            class_ref = rel.RelatingClassification
            if class_ref and class_ref.is_a("IfcClassificationReference"):
                src = getattr(class_ref, "ReferencedSource", None)
                c_name = (getattr(src, "Name", "") or getattr(class_ref, "Name", "") or "").strip()
                if c_name:
                    classification_names[c_name] = classification_names.get(c_name, 0) + 1
    timer.stop("entity_index")

    timer.start("spatial_summary")
    storeys = len(ifc.by_type("IfcBuildingStorey"))
    spaces = len(ifc.by_type("IfcSpace"))
    buildings = len(ifc.by_type("IfcBuilding"))
    sites = len(ifc.by_type("IfcSite"))
    timer.stop("spatial_summary")

    mapping_pairs: List[Tuple[str, str]] = []
    if COBIE_MAPPING:
        for pset, info in COBIE_MAPPING.items():
            for pname, _ in info["props"]:
                mapping_pairs.append((pset, pname))
    all_cobie_pairs = mapping_pairs + sorted(cobie_dynamic_pairs - set(mapping_pairs))

    return {
        "schema": ifc.schema,
        "model_info": {
            "projects": len(ifc.by_type("IfcProject")),
            "sites": sites,
            "buildings": buildings,
            "storeys": storeys,
            "spaces": spaces,
            "elements": len(elements),
        },
        "available_classes": [{"name": name, "count": count} for name, count in sorted(class_counts.items(), key=lambda x: (-x[1], x[0]))],
        "available_psets": [{"name": name, "count": count} for name, count in sorted(pset_names.items(), key=lambda x: (-x[1], x[0]))],
        "properties_by_pset": {name: sorted(props.keys()) for name, props in sorted(property_names_by_pset.items())},
        "quantities_by_set": {name: sorted(props.keys()) for name, props in sorted(quantity_names.items())},
        "classification_systems": [{"name": name, "count": count, "is_uniclass_target": name in classification_targets} for name, count in sorted(classification_names.items(), key=lambda x: (-x[1], x[0]))],
        "cobie_pairs": [{"pset": pset, "property": prop} for pset, prop in all_cobie_pairs],
        "default_include_sheets": sorted(list(ExtractionPlan().include_sheets)),
        "timings_ms": timer.as_payload(),
    }


def extract_to_excel_with_plan(ifc_path: str, output_path: str, plan: Optional[ExtractionPlan] = None) -> Dict[str, Any]:
    plan = plan or ExtractionPlan()
    timer = StageTimer()
    timer.start("model_load")
    ifc = ifcopenshell.open(ifc_path)
    timer.stop("model_load")

    timer.start("entity_index")
    all_objects = [e for e in ifc.by_type("IfcObject") if getattr(e, "GlobalId", None)]
    all_types = [e for e in ifc.by_type("IfcTypeObject") if getattr(e, "GlobalId", None)]
    all_export_objects = sorted({*all_objects, *all_types}, key=lambda e: e.id())
    elements = [e for e in all_objects if e.is_a("IfcProduct")]
    if plan.entity_classes:
        elements = [e for e in elements if e.is_a() in plan.entity_classes]
    type_by_elem_id: Dict[int, Any] = {}
    psets_cache: Dict[int, Dict[str, Dict[str, Any]]] = {}
    type_psets_cache: Dict[int, Dict[str, Dict[str, Any]]] = {}
    spatial_cache: Dict[int, Tuple[str, str, str, str]] = {}

    def _element_type_obj(elem: Any) -> Any:
        elem_id = elem.id()
        if elem_id not in type_by_elem_id:
            type_by_elem_id[elem_id] = ifcopenshell.util.element.get_type(elem)
        return type_by_elem_id[elem_id]

    def _spatial_context(elem: Any) -> Tuple[str, str, str, str]:
        elem_id = elem.id()
        if elem_id in spatial_cache:
            return spatial_cache[elem_id]
        container = ifcopenshell.util.element.get_container(elem)
        space = storey = building = site = None
        current = container
        while current:
            if current.is_a("IfcSpace"):
                space = space or getattr(current, "Name", "")
            elif current.is_a("IfcBuildingStorey"):
                storey = storey or getattr(current, "Name", "")
            elif current.is_a("IfcBuilding"):
                building = building or getattr(current, "Name", "")
            elif current.is_a("IfcSite"):
                site = site or getattr(current, "Name", "")
            current = ifcopenshell.util.element.get_container(current)
        spatial_cache[elem_id] = (space or "", storey or "", building or "", site or "")
        return spatial_cache[elem_id]

    def _resolved_psets(elem: Any) -> Dict[str, Dict[str, Any]]:
        elem_id = elem.id()
        if elem_id not in psets_cache:
            psets_cache[elem_id] = _safe_get_psets(elem)
        return psets_cache[elem_id]

    def _resolved_type_psets(type_obj: Any) -> Dict[str, Dict[str, Any]]:
        if not type_obj:
            return {}
        type_id = type_obj.id()
        if type_id not in type_psets_cache:
            type_psets_cache[type_id] = _safe_get_psets(type_obj)
        return type_psets_cache[type_id]

    def _get_pset_value(elem: Any, pset_name: str, prop_name: str) -> Any:
        psets = _resolved_psets(elem)
        if pset_name in psets and prop_name in psets[pset_name]:
            return psets[pset_name][prop_name]
        if not plan.include_type_properties:
            return ""
        type_obj = _element_type_obj(elem)
        if type_obj is not None:
            type_psets = _resolved_type_psets(type_obj)
            if pset_name in type_psets and prop_name in type_psets[pset_name]:
                return type_psets[pset_name][prop_name]
        return ""
    timer.stop("entity_index")

    timer.start("project_data")
    project_data = []
    projects = ifc.by_type("IfcProject")
    project = projects[0] if projects else None
    sites = ifc.by_type("IfcSite")
    site = sites[0] if sites else None
    buildings = ifc.by_type("IfcBuilding")
    building = buildings[0] if buildings else None
    detected_schema, schema_warning = detect_ifc_schema_from_header(ifc_path)
    schema_for_lookup = (detected_schema or ifc.schema or "IFC4").upper()
    is_ifc2x3 = schema_for_lookup == "IFC2X3"
    b_en_ref, b_en_name = ("", "")
    if building is not None:
        b_en_ref, b_en_name = _extract_uniclass(building, "Uniclass En Entities", is_ifc2x3)
    project_number = ""
    if project is not None:
        project_number = getattr(project, "LongName", "") or ""
        for pset_name in ("Additional_Pset_ProjectCommon", "Pset_ProjectCommon"):
            psets = _safe_get_psets(project)
            values = psets.get(pset_name, {})
            for key in ("Project Number", "ProjectNumber", "ProjectNo"):
                if clean_value(values.get(key)) is not None:
                    project_number = str(values.get(key))
                    break
            if project_number:
                break
    project_data.append({"DataType": "Project", "Name": getattr(project, "Name", "") if project else "", "Description": getattr(project, "Description", "") if project else "", "Phase": getattr(project, "Phase", "") if project else "", "ProjectNumber": project_number, "UniclassEnReference": "", "UniclassEnName": ""})
    project_data.append({"DataType": "Site", "Name": getattr(site, "Name", "") if site else "", "Description": getattr(site, "Description", "") if site else "", "Phase": "", "ProjectNumber": "", "UniclassEnReference": "", "UniclassEnName": ""})
    project_data.append({"DataType": "Building", "Name": getattr(building, "Name", "") if building else "", "Description": getattr(building, "Description", "") if building else "", "Phase": "", "ProjectNumber": "", "UniclassEnReference": b_en_ref, "UniclassEnName": b_en_name})
    project_data.append({"DataType": "DetectedSchema", "Name": schema_for_lookup, "Description": "Parsed from FILE_SCHEMA in IFC header", "Phase": "", "ProjectNumber": "", "UniclassEnReference": "", "UniclassEnName": ""})
    if schema_warning:
        project_data.append({"DataType": "SchemaWarning", "Name": schema_warning, "Description": "Fallback applied", "Phase": "", "ProjectNumber": "", "UniclassEnReference": "", "UniclassEnName": ""})
    project_df = pd.DataFrame(project_data)
    timer.stop("project_data")

    timer.start("elements_table")
    element_rows: List[Dict[str, Any]] = []
    type_rows: List[Dict[str, Any]] = []
    raw_entity_rows: List[Dict[str, Any]] = []
    def _build_roundtrip_row(obj: Any, is_type_object: bool) -> Dict[str, Any]:
        current_entity = obj.is_a()
        current_predefined = getattr(obj, "PredefinedType", "") if hasattr(obj, "PredefinedType") else ""
        return {
            "RowKey": f"{obj.GlobalId or ''}:{obj.id()}",
            "StepId": obj.id(),
            "GlobalId": getattr(obj, "GlobalId", ""),
            "CurrentEntity": current_entity,
            "TargetEntity": current_entity,
            "CurrentPredefinedType": current_predefined or "",
            "TargetPredefinedType": current_predefined or "",
            "Name": getattr(obj, "Name", "") or "",
            "ObjectType or ElementType": (getattr(obj, "ElementType", "") if is_type_object else getattr(obj, "ObjectType", "")) or "",
            "ApplicableOccurrence": getattr(obj, "ApplicableOccurrence", "") if is_type_object else "",
            "IsTypeObject": bool(is_type_object),
            "Validation": "",
            "ApplyChange": "No",
            "SuggestedEntity": "",
            "SuggestedPredefinedType": "",
            "SuggestionConfidence": 0.0,
            "SuggestionReason": "",
        }

    source_file_name = os.path.basename(ifc_path)
    for elem in elements:
        type_obj = _element_type_obj(elem)
        current_predefined = getattr(elem, "PredefinedType", "") if hasattr(elem, "PredefinedType") else ""
        row = {
            "GlobalId": elem.GlobalId,
            "Class": elem.is_a(),
            "OccurrenceName": getattr(elem, "Name", ""),
            "OccurrenceType": getattr(elem, "ObjectType", ""),
            "TypeName": getattr(type_obj, "Name", "") if type_obj else "",
            "TypeDescription": getattr(elem, "Description", ""),
            "IFCPresentationLayer": _get_layers_name(elem, ifc),
            "ExpressLine": str(elem),
            "IfcEntity": elem.is_a(),
            "PredefinedType": current_predefined or "",
            "Name": getattr(elem, "Name", "") or "",
            "ObjectType": getattr(elem, "ObjectType", "") or "",
            "SourceFile": source_file_name,
        }
        if plan.civil3d_extended:
            for field_name in CIVIL3D_EXTENDED_FIELDS:
                row[field_name] = _resolve_field_value(elem, type_obj, field_name, _get_pset_value)
        row.update(_build_roundtrip_row(elem, is_type_object=False))
        row.update(_build_classification_suggestion(row))
        element_rows.append(row)
    for type_obj in all_types:
        type_row = _build_roundtrip_row(type_obj, is_type_object=True)
        type_row.update(_build_classification_suggestion(type_row))
        type_rows.append(type_row)
    for obj in all_export_objects:
        raw_entity_rows.append(
            {
                "StepId": obj.id(),
                "GlobalId": getattr(obj, "GlobalId", "") or "",
                "Entity": obj.is_a(),
                "Name": getattr(obj, "Name", "") or "",
                "RawStepLine": str(obj),
            }
        )
    elements_df = pd.DataFrame(element_rows)
    types_df = pd.DataFrame(type_rows)
    elements_df, types_df = _merge_existing_excel_overrides(output_path, elements_df, types_df)
    raw_entities_df = pd.DataFrame(raw_entity_rows)
    changelog_df = pd.DataFrame(columns=["RowKey", "GlobalId", "StepId", "Status", "Message", "FromEntity", "ToEntity", "FromPredefinedType", "ToPredefinedType"])
    timer.stop("elements_table")

    timer.start("properties_table")
    prop_rows: List[List[Any]] = []
    if "Properties" in plan.include_sheets:
        for elem in elements:
            if plan.include_spatial_fields:
                space_name, storey_name, building_name, site_name = _spatial_context(elem)
            else:
                space_name = storey_name = building_name = site_name = ""
            for definition in elem.IsDefinedBy or []:
                if not definition.is_a("IfcRelDefinesByProperties"):
                    continue
                pset = definition.RelatingPropertyDefinition
                if pset.is_a("IfcPropertySet"):
                    pset_name = getattr(pset, "Name", "") or ""
                    if plan.property_sets and pset_name not in plan.property_sets:
                        continue
                    for prop in pset.HasProperties:
                        val = None
                        try:
                            if prop.is_a("IfcPropertySingleValue"):
                                val = _extract_nominal_value(prop)
                            elif prop.is_a("IfcPropertyEnumeratedValue") and prop.EnumerationValues:
                                val = ", ".join(
                                    _clean_value(
                                        _normalize_ifc_value(
                                            v,
                                            prop_name=getattr(prop, "Name", "") or "",
                                            entity_type=prop.is_a(),
                                        )
                                    )
                                    for v in prop.EnumerationValues
                                )
                            elif prop.is_a("IfcPropertyListValue") and getattr(prop, "ListValues", None):
                                val = ", ".join(
                                    _clean_value(
                                        _normalize_ifc_value(
                                            v,
                                            prop_name=getattr(prop, "Name", "") or "",
                                            entity_type=prop.is_a(),
                                        )
                                    )
                                    for v in prop.ListValues
                                )
                        except Exception:
                            APP_LOGGER.exception(
                                "Best-effort property extraction failed property=%s entity_type=%s",
                                getattr(prop, "Name", "") or "",
                                prop.is_a() if hasattr(prop, "is_a") else type(prop).__name__,
                            )
                            val = None
                        prop_rows.append([elem.GlobalId, elem.is_a(), getattr(elem, "Name", ""), getattr(elem, "ObjectType", ""), space_name, storey_name, building_name, site_name, pset_name, prop.Name, val])
                elif pset.is_a("IfcElementQuantity"):
                    q_name = getattr(pset, "Name", "") or ""
                    if plan.quantity_sets and q_name not in plan.quantity_sets:
                        continue
                    for qty in getattr(pset, "Quantities", []) or []:
                        prop_rows.append([elem.GlobalId, elem.is_a(), getattr(elem, "Name", ""), getattr(elem, "ObjectType", ""), space_name, storey_name, building_name, site_name, q_name, getattr(qty, "Name", ""), getattr(qty, "NominalValue", "")])
    props_df = pd.DataFrame(prop_rows, columns=["GlobalId", "Class", "ObjectName", "ObjectType", "ContainerSpace", "ContainerStorey", "ContainerBuilding", "ContainerSite", "PropertySet", "Property", "Value"])
    timer.stop("properties_table")

    timer.start("classification_extract")
    pr_rows, ss_rows, ef_rows = [], [], []
    if plan.include_classifications and any(sheet in plan.include_sheets for sheet in {"Uniclass_Pr", "Uniclass_Ss", "Uniclass_EF"}):
        for elem in elements:
            pr_ref, pr_name = _extract_uniclass(elem, "Uniclass Pr Products", is_ifc2x3)
            ss_ref, ss_name = _extract_uniclass(elem, "Uniclass Ss Systems", is_ifc2x3)
            ef_ref, ef_name = _extract_uniclass(elem, "Uniclass EF Elements Functions", is_ifc2x3)
            pr_rows.append({"GlobalId": elem.GlobalId, "Reference": pr_ref, "Name": pr_name})
            ss_rows.append({"GlobalId": elem.GlobalId, "Reference": ss_ref, "Name": ss_name})
            ef_rows.append({"GlobalId": elem.GlobalId, "Reference": ef_ref, "Name": ef_name})
    uniclass_pr_df = pd.DataFrame(pr_rows)
    uniclass_ss_df = pd.DataFrame(ss_rows)
    uniclass_ef_df = pd.DataFrame(ef_rows)
    timer.stop("classification_extract")

    timer.start("cobie_extract")
    cobie_rows = []
    if "COBieMapping" in plan.include_sheets:
        mapping_pairs = []
        if COBIE_MAPPING:
            for pset, info in COBIE_MAPPING.items():
                for pname, _ in info["props"]:
                    mapping_pairs.append((pset, pname))
        dynamic_pairs = set()
        if plan.cobie_pairs:
            dynamic_pairs = set(plan.cobie_pairs)
        else:
            for elem in elements:
                add_pset = _resolved_psets(elem).get("Additional_Pset_GeneralCommon", {})
                dynamic_pairs.update(parse_required_pairs(add_pset.get("RequiredForCOBie", "")))
                dynamic_pairs.update(parse_required_pairs(add_pset.get("RequiredForCOBieComponent", "")))
                type_obj = _element_type_obj(elem)
                if type_obj is not None and plan.include_type_properties:
                    add_pset_t = _resolved_type_psets(type_obj).get("Additional_Pset_GeneralCommon", {})
                    dynamic_pairs.update(parse_required_pairs(add_pset_t.get("RequiredForCOBie", "")))
                    dynamic_pairs.update(parse_required_pairs(add_pset_t.get("RequiredForCOBieComponent", "")))
        all_pairs = mapping_pairs + sorted(dynamic_pairs - set(mapping_pairs))
        extra_cols = list(CIVIL3D_EXTENDED_FIELDS) if plan.civil3d_extended else []
        cobie_cols = ["GlobalId", "IFCElement.Name", "IFCElementType.Name"] + extra_cols + [f"{pset}.{pname}" for pset, pname in all_pairs]
        for elem in elements:
            type_obj = _element_type_obj(elem)
            row = {
                "GlobalId": elem.GlobalId,
                "IFCElement.Name": _resolve_name_with_priority(elem, type_obj, _get_pset_value),
                "IFCElementType.Name": _resolve_type_name_with_priority(elem, type_obj, _get_pset_value),
            }
            for field_name in extra_cols:
                row[field_name] = _resolve_field_value(elem, type_obj, field_name, _get_pset_value)
            for pset, pname in all_pairs:
                row[f"{pset}.{pname}"] = _get_pset_value(elem, pset, pname)
            cobie_rows.append(row)
        cobie_df = pd.DataFrame(cobie_rows, columns=cobie_cols)
    else:
        cobie_df = pd.DataFrame()
    timer.stop("cobie_extract")

    timer.start("excel_write")
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        if "ProjectData" in plan.include_sheets:
            _sanitize_dataframe_for_excel(project_df).to_excel(writer, sheet_name="ProjectData", index=False)
        if "Elements" in plan.include_sheets:
            _sanitize_dataframe_for_excel(elements_df).to_excel(writer, sheet_name="Elements", index=False)
        if "Types" in plan.include_sheets:
            _sanitize_dataframe_for_excel(types_df).to_excel(writer, sheet_name="Types", index=False)
        if "RawEntities" in plan.include_sheets:
            _sanitize_dataframe_for_excel(raw_entities_df).to_excel(writer, sheet_name="RawEntities", index=False)
        if "Properties" in plan.include_sheets:
            _sanitize_dataframe_for_excel(props_df).to_excel(writer, sheet_name="Properties", index=False)
        if "COBieMapping" in plan.include_sheets:
            _sanitize_dataframe_for_excel(cobie_df).to_excel(writer, sheet_name="COBieMapping", index=False)
        if "Uniclass_Pr" in plan.include_sheets:
            _sanitize_dataframe_for_excel(uniclass_pr_df).to_excel(writer, sheet_name="Uniclass_Pr", index=False)
        if "Uniclass_Ss" in plan.include_sheets:
            _sanitize_dataframe_for_excel(uniclass_ss_df).to_excel(writer, sheet_name="Uniclass_Ss", index=False)
        if "Uniclass_EF" in plan.include_sheets:
            _sanitize_dataframe_for_excel(uniclass_ef_df).to_excel(writer, sheet_name="Uniclass_EF", index=False)
        if "ChangeLog" in plan.include_sheets:
            _sanitize_dataframe_for_excel(changelog_df).to_excel(writer, sheet_name="ChangeLog", index=False)

        if schema_for_lookup == "IFC2X3":
            mapping = load_ifc2x3_entity_mapping()
            entities = sorted((mapping.get("entities") or {}).keys())
            predefined_values = sorted(
                {
                    (predefined or "").strip()
                    for entity in entities
                    for predefined in ((mapping.get("entities", {}).get(entity, {}) or {}).get("predefined_types", []) or [])
                    if (predefined or "").strip()
                }
            )
        else:
            entities = sorted(_entity_names(_schema_definition(schema_for_lookup)))
            predefined_values = sorted(
                {
                    lit
                    for entity in entities
                    for lit in (_predefined_type_info(schema_for_lookup, entity).get("enum_items", []) or [])
                    if lit
                }
            )
        entity_lookup_df = pd.DataFrame({"IfcEntity": entities})
        predefined_lookup_df = pd.DataFrame({"PredefinedType": predefined_values})
        entity_predefined_map_rows: List[Dict[str, str]] = []
        for entity in entities:
            if schema_for_lookup == "IFC2X3":
                enum_items = (mapping.get("entities", {}).get(entity, {}) or {}).get("predefined_types", []) or []
            else:
                enum_items = _predefined_type_info(schema_for_lookup, entity).get("enum_items", []) or []
            for val in enum_items:
                cleaned = (val or "").strip()
                if cleaned:
                    entity_predefined_map_rows.append({"IfcEntity": entity, "PredefinedType": cleaned})
        entity_predefined_map_df = pd.DataFrame(entity_predefined_map_rows, columns=["IfcEntity", "PredefinedType"])
        lookup_suffix = "IFC2X3" if schema_for_lookup == "IFC2X3" else "IFC4"
        entities_sheet = f"_Lookups_{lookup_suffix}_Entities"
        predefs_sheet = f"_Lookups_{lookup_suffix}_Predefs"
        mapping_sheet = f"_Lookups_{lookup_suffix}_Map"
        entity_lookup_df.to_excel(writer, sheet_name=entities_sheet, index=False)
        predefined_lookup_df.to_excel(writer, sheet_name=predefs_sheet, index=False)
        entity_predefined_map_df.to_excel(writer, sheet_name=mapping_sheet, index=False)

        workbook = writer.book
        if len(entities) > 0:
            _upsert_workbook_defined_name(
                workbook,
                "IfcEntityList",
                _excel_range(entities_sheet, "A", 2, len(entities) + 1),
            )
        if len(predefined_values) > 0:
            _upsert_workbook_defined_name(
                workbook,
                "PredefinedTypeList",
                _excel_range(predefs_sheet, "A", 2, len(predefined_values) + 1),
            )

        if "Elements" in workbook.sheetnames:
            ws = workbook["Elements"]
            header = [c.value for c in ws[1]]
            if "IfcEntity" in header and len(entities) > 0 and ws.max_row >= 2:
                entity_col = get_column_letter(header.index("IfcEntity") + 1)
                entity_dv = DataValidation(type="list", formula1="=IfcEntityList", allow_blank=True)
                ws.add_data_validation(entity_dv)
                entity_dv.add(f"{entity_col}2:{entity_col}{ws.max_row}")
            if "PredefinedType" in header and len(predefined_values) > 0 and ws.max_row >= 2:
                predef_col = get_column_letter(header.index("PredefinedType") + 1)
                predef_dv = DataValidation(type="list", formula1="=PredefinedTypeList", allow_blank=True)
                ws.add_data_validation(predef_dv)
                predef_dv.add(f"{predef_col}2:{predef_col}{ws.max_row}")

        for lookup_sheet in (entities_sheet, predefs_sheet, mapping_sheet):
            workbook[lookup_sheet].sheet_state = "hidden"
    validation_error = validate_workbook_after_export(output_path)
    if validation_error:
        raise HTTPException(status_code=500, detail={"message": "Excel export validation failed", "error": validation_error})
    timer.stop("excel_write")
    return {
        "path": output_path,
        "timings_ms": timer.as_payload(),
        "counts": {"elements": len(elements), "types": len(all_types), "properties": len(prop_rows), "cobie_rows": len(cobie_rows)},
        "schema_detected": schema_for_lookup,
        "schema_warning": schema_warning,
    }


def extract_to_excel(ifc_path: str, output_path: str, plan_payload: Optional[Dict[str, Any]] = None) -> Dict[str, Any]:
    plan = _parse_excel_extraction_plan(plan_payload)
    return extract_to_excel_with_plan(ifc_path, output_path, plan=plan)


def _set_element_presentation_layer(ifc, elem, target_layer_name: str):
    target = clean_value(target_layer_name)

    def _expand_item(item, visited):
        if not item:
            return []
        key = item.id() if hasattr(item, "id") else id(item)
        if key in visited:
            return []
        visited.add(key)
        out = [item]
        if item.is_a("IfcMappedItem"):
            source = getattr(item, "MappingSource", None)
            mapped = getattr(source, "MappedRepresentation", None) if source else None
            for child in getattr(mapped, "Items", []) or []:
                out.extend(_expand_item(child, visited))
        elif item.is_a("IfcBooleanResult") or item.is_a("IfcBooleanClippingResult"):
            out.extend(_expand_item(getattr(item, "FirstOperand", None), visited))
            out.extend(_expand_item(getattr(item, "SecondOperand", None), visited))
        elif item.is_a("IfcCsgSolid"):
            out.extend(_expand_item(getattr(item, "TreeExpression", None), visited))
        return out

    def _collect_items_from_representation(rep_holder):
        collected = []
        for rep in getattr(rep_holder, "Representations", []) or []:
            for item in getattr(rep, "Items", []) or []:
                collected.extend(_expand_item(item, set()))
        return collected

    items = []
    representation = getattr(elem, "Representation", None)
    items.extend(_collect_items_from_representation(representation))

    type_obj = ifcopenshell.util.element.get_type(elem)
    for rep_map in getattr(type_obj, "RepresentationMaps", []) or []:
        mapped = getattr(rep_map, "MappedRepresentation", None)
        items.extend(_collect_items_from_representation(mapped))

    deduped = []
    seen = set()
    for item in items:
        key = item.id() if hasattr(item, "id") else id(item)
        if key in seen:
            continue
        seen.add(key)
        deduped.append(item)
    items = deduped

    if not items:
        return

    empty_layers = set()
    for item in items:
        for inv in ifc.get_inverse(item) or []:
            if not inv or not inv.is_a("IfcPresentationLayerAssignment"):
                continue
            assigned = list(getattr(inv, "AssignedItems", []) or [])
            if item in assigned:
                assigned = [a for a in assigned if a != item]
                inv.AssignedItems = assigned
                if len(assigned) == 0:
                    empty_layers.add(inv)

    for layer in empty_layers:
        try:
            ifc.remove(layer)
        except Exception:
            pass

    if not target:
        return

    layer = None
    for candidate in ifc.by_type("IfcPresentationLayerAssignment"):
        if (getattr(candidate, "Name", "") or "") == target:
            layer = candidate
            break
    if layer is None:
        layer = ifc.create_entity("IfcPresentationLayerAssignment", Name=target, AssignedItems=[])

    assigned = list(getattr(layer, "AssignedItems", []) or [])
    for item in items:
        if item not in assigned:
            assigned.append(item)
    layer.AssignedItems = assigned


def validate_excel_import_data(
    ifc: ifcopenshell.file,
    elements_df: pd.DataFrame,
    cobie_df: pd.DataFrame,
    project_df: pd.DataFrame,
) -> List[str]:
    issues: List[str] = []
    if "DataType" not in project_df.columns:
        issues.append("ProjectData sheet is missing DataType column.")
    else:
        project_rows = project_df[project_df["DataType"] == "Project"]
        if project_rows.empty:
            issues.append("ProjectData sheet does not contain a Project row.")
        elif "ProjectNumber" in project_df.columns:
            project_number = clean_value(project_rows.iloc[0].get("ProjectNumber"))
            if project_number is None:
                issues.append("ProjectData.ProjectNumber is blank for Project row.")
        else:
            issues.append("ProjectData sheet is missing ProjectNumber column.")

    if "GlobalId" not in elements_df.columns:
        issues.append("Elements sheet is missing GlobalId column.")
    else:
        missing_guid_count = int(elements_df["GlobalId"].isna().sum())
        if missing_guid_count:
            issues.append(f"Elements sheet has {missing_guid_count} rows with blank GlobalId.")

    for col in ("IFCElement.Name", "IFCElementType.Name"):
        if col in cobie_df.columns and cobie_df[col].isna().any():
            issues.append(f"COBieMapping.{col} has blank values.")
    return issues


def _resolve_class_mapping_candidate(elem: Any, row: pd.Series) -> Tuple[str, str]:
    ext_object = clean_value(row.get("ExtObject")) or clean_value(row.get("Ext Object"))
    if ext_object:
        return str(ext_object), "ExtObject"
    ifc_enum = clean_value(row.get("IFC_Enumeration"))
    if ifc_enum:
        return str(ifc_enum), "IFC_Enumeration"
    object_type = clean_value(row.get("OccurrenceType")) or clean_value(getattr(elem, "ObjectType", ""))
    if object_type:
        return str(object_type), "existing_proxy_mapping"
    return "IfcBuildingElementProxy", "fallback_proxy"


def _normalize_override_value(value: Any) -> str:
    cleaned = clean_value(value)
    if cleaned is None:
        return ""
    return str(cleaned).strip()


def _keyword_contains(text: str, keywords: List[str]) -> bool:
    lowered = (text or "").lower()
    return any(keyword in lowered for keyword in keywords)


def _build_classification_suggestion(row: Dict[str, Any]) -> Dict[str, Any]:
    text_fields = [
        row.get("Name", ""),
        row.get("ObjectType", ""),
        row.get("TypeDescription", ""),
        row.get("IFCName", ""),
        row.get("Type_(UserDefined)", ""),
        row.get("IFCDescription", ""),
        row.get("ExtObject", ""),
        row.get("Layer", ""),
        row.get("BlockName", ""),
    ]
    blob = " | ".join(str(v) for v in text_fields if clean_value(v) is not None)
    ext_object = _normalize_override_value(row.get("ExtObject"))
    if ext_object.upper().startswith("IFC"):
        return {
            "SuggestedEntity": ext_object,
            "SuggestedPredefinedType": "",
            "SuggestionConfidence": 0.95,
            "SuggestionReason": "ExtObject signal from COBie_Occurrence_(Component)",
        }
    if _keyword_contains(blob, ["pipe", "drain", "sewer"]):
        return {
            "SuggestedEntity": "IfcPipeSegment",
            "SuggestedPredefinedType": "",
            "SuggestionConfidence": 0.78,
            "SuggestionReason": "Keyword match: pipe/drain/sewer",
        }
    if _keyword_contains(blob, ["chamber", "manhole", "gully", "catchpit"]):
        return {
            "SuggestedEntity": "IfcFlowStorageDevice",
            "SuggestedPredefinedType": "",
            "SuggestionConfidence": 0.74,
            "SuggestionReason": "Keyword match: chamber/manhole/gully/catchpit",
        }
    if _keyword_contains(blob, ["headwall"]):
        return {
            "SuggestedEntity": "IfcBuildingElementProxy",
            "SuggestedPredefinedType": "",
            "SuggestionConfidence": 0.35,
            "SuggestionReason": "Keyword match: headwall (low-confidence fallback)",
        }
    return {
        "SuggestedEntity": "",
        "SuggestedPredefinedType": "",
        "SuggestionConfidence": 0.0,
        "SuggestionReason": "",
    }


def _merge_existing_excel_overrides(output_path: str, elements_df: pd.DataFrame, types_df: pd.DataFrame) -> Tuple[pd.DataFrame, pd.DataFrame]:
    if not os.path.exists(output_path):
        return elements_df, types_df
    try:
        prior = pd.ExcelFile(output_path)
    except Exception:
        return elements_df, types_df

    def _merge(sheet_name: str, fresh_df: pd.DataFrame) -> pd.DataFrame:
        if fresh_df is None or fresh_df.empty:
            return fresh_df
        try:
            prior_df = pd.read_excel(prior, sheet_name=sheet_name, usecols=lambda c: c is not None)
        except Exception:
            return fresh_df
        if prior_df.empty:
            return fresh_df

        key_cols = [col for col in ("GlobalId", "StepId") if col in prior_df.columns and col in fresh_df.columns]
        if not key_cols:
            return fresh_df
        override_cols = [col for col in ("TargetEntity", "TargetPredefinedType", "ApplyChange") if col in prior_df.columns and col in fresh_df.columns]
        if not override_cols:
            return fresh_df

        prior_subset = prior_df[key_cols + override_cols].copy()
        renamed = {col: f"__prev_{col}" for col in override_cols}
        prior_subset = prior_subset.rename(columns=renamed)
        merged = fresh_df.merge(prior_subset, on=key_cols, how="left")
        for col in override_cols:
            prev_col = f"__prev_{col}"
            if prev_col not in merged.columns:
                continue
            has_prev = merged[prev_col].apply(lambda value: clean_value(value) is not None)
            merged.loc[has_prev, col] = merged.loc[has_prev, prev_col]
            merged = merged.drop(columns=[prev_col])
        return merged

    return _merge("Elements", elements_df), _merge("Types", types_df)


def _check_heavy_timeout(started_at: float, endpoint: str) -> None:
    elapsed = time.monotonic() - started_at
    if elapsed > HEAVY_JOB_TIMEOUT_SECONDS:
        raise TimeoutError(f"{endpoint} exceeded HEAVY_JOB_TIMEOUT_SECONDS ({HEAVY_JOB_TIMEOUT_SECONDS}s)")


def update_ifc_from_excel(
    ifc_file,
    excel_file,
    output_path: str,
    update_mode="update",
    add_new="no",
    *,
    session_id: Optional[str] = None,
    endpoint: str = "excel/update",
):
    started_at = time.monotonic()
    ifc_path = path_of(ifc_file)
    xls_path = path_of(excel_file)
    file_size = os.path.getsize(ifc_path) if os.path.exists(ifc_path) else None
    log_memory_stage(stage="file upload received", session_id=session_id, file_name=os.path.basename(ifc_path), file_size=file_size, endpoint=endpoint, started_at=started_at)
    enforce_upload_limits(ifc_path, endpoint=endpoint)
    enforce_upload_limits(xls_path, endpoint=endpoint)

    ifc = ifcopenshell.open(ifc_path)
    log_memory_stage(stage="IFC file open", session_id=session_id, file_name=os.path.basename(ifc_path), file_size=file_size, endpoint=endpoint, started_at=started_at)
    _check_heavy_timeout(started_at, endpoint)

    xls = pd.ExcelFile(xls_path)
    elements_df = pd.read_excel(xls, "Elements", usecols=lambda c: c is not None)
    try:
        types_df = pd.read_excel(xls, "Types", usecols=lambda c: c is not None)
    except Exception:
        types_df = pd.DataFrame()
    props_df = pd.read_excel(xls, "Properties")
    cobie_df = pd.read_excel(xls, "COBieMapping")
    project_df = pd.read_excel(xls, "ProjectData")
    log_memory_stage(stage="workbook load", session_id=session_id, file_name=os.path.basename(xls_path), file_size=os.path.getsize(xls_path), endpoint=endpoint, started_at=started_at)
    _check_heavy_timeout(started_at, endpoint)

    change_log_rows: List[Dict[str, Any]] = []

    try:
        uniclass_pr_df = pd.read_excel(xls, "Uniclass_Pr")
    except Exception:
        uniclass_pr_df = None
    try:
        uniclass_ss_df = pd.read_excel(xls, "Uniclass_Ss")
    except Exception:
        uniclass_ss_df = None
    try:
        uniclass_ef_df = pd.read_excel(xls, "Uniclass_EF")
    except Exception:
        try:
            uniclass_ef_df = pd.read_excel(xls, "Uniclass_En")
        except Exception:
            uniclass_ef_df = None

    validation_issues = validate_excel_import_data(ifc, elements_df, cobie_df, project_df)
    if validation_issues:
        raise ValueError("Excel validation failed: " + "; ".join(validation_issues))

    project = ifc.by_type("IfcProject")[0]
    site = ifc.by_type("IfcSite")[0] if ifc.by_type("IfcSite") else None
    building = ifc.by_type("IfcBuilding")[0] if ifc.by_type("IfcBuilding") else None
    detected_schema = (ifc.schema or "").upper()
    APP_LOGGER.info("EN Entities write-back detected schema=%s", detected_schema)
    en_entities_value, en_entities_name = _read_projectdata_en_entities(project_df)
    APP_LOGGER.info("ProjectData EN Entities value read=%r name=%r", en_entities_value, en_entities_name)
    en_entities_rel: Optional[Any] = None

    def set_entity_uniclass(entity, source_name, ref_value, name_value):
        if entity is None:
            return
        ref = clean_value(ref_value)
        nm = clean_value(name_value)
        if ref is None and nm is None:
            return
        cls_src = None
        for c in ifc.by_type("IfcClassification"):
            if getattr(c, "Name", "") == source_name:
                cls_src = c
                break
        if cls_src is None and add_new == "yes":
            cls_src = ifc.create_entity(
                "IfcClassification",
                Name=source_name,
                Source="https://www.thenbs.com/our-tools/uniclass-2015",
                Edition="2015",
            )

        existing_ref = None
        for rel in getattr(entity, "HasAssociations", []) or []:
            if rel.is_a("IfcRelAssociatesClassification"):
                cref = rel.RelatingClassification
                if cref and cref.is_a("IfcClassificationReference"):
                    src = getattr(cref, "ReferencedSource", None)
                    if src and getattr(src, "Name", "") == source_name:
                        existing_ref = cref
                        break
        if existing_ref:
            if ref is not None:
                existing_ref.ItemReference = str(ref)
            if nm is not None:
                existing_ref.Name = str(nm)
        elif add_new == "yes" and (ref is not None or nm is not None) and cls_src is not None:
            cref = ifc.create_entity(
                "IfcClassificationReference",
                ItemReference=str(ref) if ref is not None else None,
                Name=str(nm) if nm is not None else None,
            )
            cref.ReferencedSource = cls_src
            ifc.create_entity(
                "IfcRelAssociatesClassification",
                GlobalId=new_guid(),
                RelatedObjects=[entity],
                RelatingClassification=cref,
            )

    for _, row in project_df.iterrows():
        dt = row["DataType"]
        if dt == "Project":
            if pd.notna(row.get("Name")):
                project.Name = clean_value(row["Name"]) or project.Name
            if pd.notna(row.get("Description")):
                project.Description = clean_value(row["Description"]) or project.Description
            if pd.notna(row.get("Phase")):
                project.Phase = clean_value(row["Phase"]) or project.Phase
            if pd.notna(row.get("ProjectNumber")):
                project_number = clean_value(row.get("ProjectNumber"))
                if project_number is not None:
                    project.LongName = str(project_number)
                    APP_LOGGER.info("Updated project metadata ProjectNumber=%s", project_number)
        elif dt == "Site":
            name = clean_value(row.get("Name"))
            desc = clean_value(row.get("Description"))
            if site is None and (add_new == "yes" or name or desc):
                site = ifc.create_entity("IfcSite", GlobalId=new_guid(), Name=name or "Site")
            if site is not None:
                if name is not None:
                    site.Name = name
                if desc is not None:
                    site.Description = desc
                ensure_aggregates(project, site, ifc)
                if building is not None:
                    reassign_aggregate(site, building, ifc)
        elif dt == "Building":
            if building is None and add_new == "yes":
                building = ifc.create_entity(
                    "IfcBuilding",
                    GlobalId=new_guid(),
                    Name=clean_value(row.get("Name")) or "Building",
                )
            if building is not None:
                if pd.notna(row.get("Name")):
                    building.Name = clean_value(row["Name"]) or building.Name
                if pd.notna(row.get("Description")):
                    building.Description = clean_value(row["Description"]) or building.Description
                if site is not None:
                    reassign_aggregate(site, building, ifc)
                else:
                    ensure_aggregates(project, building, ifc)
                if en_entities_value:
                    en_entities_rel, was_created = _ensure_en_entities_classification_rel(
                        ifc,
                        building,
                        en_entities_value,
                        en_entities_name,
                        source_name="Uniclass En Entities",
                    )
                    APP_LOGGER.info(
                        "EN Entities classification association %s relation_id=%s related_count=%s related_ids=%s",
                        "created" if was_created else "updated",
                        en_entities_rel.id() if en_entities_rel else None,
                        len(getattr(en_entities_rel, "RelatedObjects", []) or []) if en_entities_rel else 0,
                        [obj.id() for obj in (getattr(en_entities_rel, "RelatedObjects", []) or [])] if en_entities_rel else [],
                    )
                else:
                    APP_LOGGER.info("EN Entities value is blank/no-op; no classification write-back performed.")

    schema_name = (ifc.schema or "IFC4").upper()
    schema_def = _schema_definition(schema_name)
    entities_cfg = (load_ifc2x3_entity_mapping().get("entities") or {}) if schema_name == "IFC2X3" else {}

    def _schema_has_entity(entity_name: str) -> bool:
        if not entity_name:
            return False
        if schema_def is None:
            return True
        try:
            return schema_def.declaration_by_name(entity_name) is not None
        except Exception:
            return False

    def _validate_instance_shape(entity_instance: Any) -> bool:
        if entity_instance is None or schema_def is None:
            return True
        try:
            decl = schema_def.declaration_by_name(entity_instance.is_a())
            expected = len(decl.all_attributes())
            return len(entity_instance) == expected
        except Exception:
            return True

    def _enum_items_for_entity(entity_name: str) -> List[str]:
        enum_items = _predefined_type_info(schema_name, entity_name).get("enum_items", []) or []
        if enum_items:
            return enum_items
        if schema_name == "IFC2X3":
            return (entities_cfg.get(entity_name, {}) or {}).get("predefined_types", []) or []
        return []

    def _lookup_by_row(row: pd.Series) -> Any:
        guid = clean_value(row.get("GlobalId"))
        if guid:
            try:
                found = ifc.by_guid(str(guid))
                if found:
                    return found
            except Exception:
                pass
        step_id = row.get("StepId")
        if pd.notna(step_id):
            try:
                return ifc.by_id(int(step_id))
            except Exception:
                return None
        return None

    def _apply_entity_predefined_updates(rows_df: pd.DataFrame, expected_kind: str) -> None:
        if rows_df is None or rows_df.empty:
            return
        if "Validation" in rows_df.columns:
            rows_df["Validation"] = rows_df["Validation"].astype(object)
        for row_idx, row in rows_df.iterrows():
            target = _lookup_by_row(row)
            if target is None:
                continue
            apply_change = str(clean_value(row.get("ApplyChange")) or "").strip().lower() in {"yes", "y", "true", "1"}
            current_entity = str(clean_value(row.get("CurrentEntity")) or target.is_a())
            requested_entity = str(clean_value(row.get("TargetEntity")) or current_entity)
            current_predef = str(clean_value(row.get("CurrentPredefinedType")) or getattr(target, "PredefinedType", "") or "")
            requested_predef_raw = clean_value(row.get("TargetPredefinedType"))
            requested_predef = str(requested_predef_raw if requested_predef_raw is not None else current_predef)
            if not apply_change:
                continue

            result = {
                "RowKey": row.get("RowKey", ""),
                "GlobalId": getattr(target, "GlobalId", ""),
                "StepId": target.id(),
                "FromEntity": current_entity,
                "ToEntity": requested_entity,
                "FromPredefinedType": current_predef,
                "ToPredefinedType": requested_predef,
            }
            if not _schema_has_entity(requested_entity):
                result.update({"Status": "Rejected", "Message": "Invalid mapping: target entity not in schema"})
                rows_df.at[row_idx, "Validation"] = "Invalid mapping"
                change_log_rows.append(result)
                continue
            expected_type = requested_entity.upper().endswith("TYPE")
            if (expected_kind == "type" and not expected_type) or (expected_kind == "occurrence" and expected_type):
                result.update({"Status": "Rejected", "Message": "attempt to convert type to occurrence or occurrence to type"})
                rows_df.at[row_idx, "Validation"] = "Invalid mapping"
                change_log_rows.append(result)
                continue

            migrated = target
            if requested_entity != current_entity:
                try:
                    migrated = ifcopenshell.api.run("root.reassign_class", ifc, product=target, ifc_class=requested_entity)
                except Exception as exc:
                    result.update({"Status": "Rejected", "Message": f"failed entity migration: {exc}"})
                    rows_df.at[row_idx, "Validation"] = "Invalid mapping"
                    change_log_rows.append(result)
                    continue

            row_validation = ""
            fallback_userdefined = False
            suggestion_used = clean_value(row.get("SuggestedEntity")) is not None
            if not _validate_instance_shape(migrated):
                result.update({"Status": "Rejected", "Message": "Invalid mapping: entity argument structure mismatch"})
                rows_df.at[row_idx, "Validation"] = "Invalid mapping"
                change_log_rows.append(result)
                continue

            if requested_predef and hasattr(migrated, "PredefinedType"):
                enum_items = _enum_items_for_entity(migrated.is_a())
                normalized_lookup = {normalize_token(item): item for item in enum_items}
                normalized_requested = normalize_token(requested_predef)
                if normalized_requested and normalized_requested in normalized_lookup:
                    migrated.PredefinedType = normalized_lookup[normalized_requested]
                else:
                    if "USERDEFINED" in normalized_lookup or hasattr(migrated, "PredefinedType"):
                        userdefined_literal = normalized_lookup.get("USERDEFINED", "USERDEFINED")
                        migrated.PredefinedType = userdefined_literal
                        if hasattr(migrated, "ObjectType"):
                            migrated.ObjectType = requested_predef
                        elif hasattr(migrated, "ElementType"):
                            migrated.ElementType = requested_predef
                        fallback_userdefined = True
                    else:
                        result.update({"Status": "Rejected", "Message": "Invalid mapping: predefined type incompatible with entity"})
                        rows_df.at[row_idx, "Validation"] = "Invalid mapping"
                        change_log_rows.append(result)
                        continue
            elif requested_predef and not hasattr(migrated, "PredefinedType"):
                APP_LOGGER.warning(
                    "Ignoring TargetPredefinedType for entity without PredefinedType GlobalId=%s entity=%s target_predefined=%s schema=%s",
                    getattr(migrated, "GlobalId", ""),
                    migrated.is_a(),
                    requested_predef,
                    schema_name,
                )
                row_validation = clean_value(rows_df.at[row_idx, "Validation"]) or ""
            result.update({"Status": "Applied", "Message": "ok"})
            change_log_rows.append(result)
            rows_df.at[row_idx, "Validation"] = row_validation
            APP_LOGGER.info(
                "excel_row_apply global_id=%s entity=%s->%s predefined=%s->%s suggestion_used=%s userdefined_fallback=%s schema=%s",
                getattr(migrated, "GlobalId", ""),
                current_entity,
                requested_entity,
                current_predef,
                requested_predef,
                bool(suggestion_used),
                fallback_userdefined,
                schema_name,
            )

    _apply_entity_predefined_updates(elements_df, "occurrence")
    _apply_entity_predefined_updates(types_df, "type")

    for _, row in elements_df.iterrows():
        guid = row.get("GlobalId")
        if pd.isna(guid):
            continue
        try:
            elem = ifc.by_guid(guid)
        except Exception as exc:
            APP_LOGGER.warning("IFC by_guid failed guid=%s error=%s", guid, exc)
            continue
        if not elem:
            continue
        if pd.notna(row.get("OccurrenceName")):
            elem.Name = clean_value(row["OccurrenceName"]) or elem.Name
        if pd.notna(row.get("OccurrenceType")):
            elem.ObjectType = clean_value(row["OccurrenceType"]) or elem.ObjectType
        if pd.notna(row.get("TypeDescription")):
            elem.Description = clean_value(row["TypeDescription"]) or elem.Description
        if "IFCPresentationLayer" in elements_df.columns and pd.notna(row.get("IFCPresentationLayer")):
            _set_element_presentation_layer(ifc, elem, row.get("IFCPresentationLayer"))
        if pd.notna(row.get("TypeName")):
            type_name = str(clean_value(row["TypeName"]))
            type_obj = None
            for rel in ifc.get_inverse(elem):
                if rel.is_a("IfcRelDefinesByType"):
                    type_obj = rel.RelatingType
            if not type_obj and add_new == "yes":
                type_class = (
                    elem.is_a() + "Type"
                    if elem.is_a().endswith("Element") or elem.is_a().endswith("Door")
                    else "IfcTypeObject"
                )
                try:
                    type_obj = ifcopenshell.api.run("type.create_type", ifc, ifc_class=type_class, name=type_name)
                except Exception:
                    type_obj = ifc.create_entity(
                        "IfcBuildingElementType",
                        GlobalId=new_guid(),
                        Name=type_name,
                    )
                ifcopenshell.api.run("type.assign_type", ifc, related_objects=[elem], relating_type=type_obj)
            elif type_obj:
                type_obj.Name = type_name
        class_candidate, source = _resolve_class_mapping_candidate(elem, row)
        APP_LOGGER.info(
            "Class mapping candidate for %s resolved to %s via %s",
            getattr(elem, "GlobalId", ""),
            class_candidate,
            source,
        )

    log_memory_stage(stage="row iteration/update", session_id=session_id, file_name=os.path.basename(ifc_path), file_size=file_size, endpoint=endpoint, started_at=started_at)
    _check_heavy_timeout(started_at, endpoint)

    if cobie_df is not None:
        mapping_keys = set()
        if COBIE_MAPPING is not None:
            for pset, info in COBIE_MAPPING.items():
                for pname, _ in info["props"]:
                    mapping_keys.add(f"{pset}.{pname}")

        candidate_cols = [
            c
            for c in cobie_df.columns
            if c not in ("GlobalId", "IFCElement.Name", "IFCElementType.Name") and "." in c
        ]

        for _, row in cobie_df.iterrows():
            guid = row.get("GlobalId")
            if pd.isna(guid):
                continue
            try:
                elem = ifc.by_guid(guid)
            except Exception as exc:
                APP_LOGGER.warning("IFC by_guid failed during COBie mapping guid=%s error=%s", guid, exc)
                continue
            if not elem:
                continue

            for col in candidate_cols:
                if pd.isna(row.get(col)):
                    continue
                val = row[col]
                pset, pname = col.split(".", 1)
                pset, pname = pset.strip(), pname.strip()

                psets = ifcopenshell.util.element.get_psets(elem)
                if pset not in psets and add_new == "no":
                    continue

                pset_entity = None
                for rel in elem.IsDefinedBy or []:
                    if (
                        rel.is_a("IfcRelDefinesByProperties")
                        and rel.RelatingPropertyDefinition
                        and getattr(rel.RelatingPropertyDefinition, "Name", "") == pset
                    ):
                        pset_entity = rel.RelatingPropertyDefinition
                        break
                if pset_entity is None and add_new == "yes":
                    pset_entity = ifcopenshell.api.run("pset.add_pset", ifc, product=elem, name=pset)

                if pset_entity:
                    try:
                        ifcopenshell.api.run("pset.edit_pset", ifc, pset=pset_entity, properties={pname: val})
                    except Exception:
                        pass
            for field_name in CIVIL3D_EXTENDED_FIELDS:
                if field_name not in row or pd.isna(row.get(field_name)):
                    continue
                val = row.get(field_name)
                try:
                    psets = ifcopenshell.util.element.get_psets(elem)
                    if "Additional_Pset_GeneralCommon" not in psets and add_new == "yes":
                        pset_entity = ifcopenshell.api.run("pset.add_pset", ifc, product=elem, name="Additional_Pset_GeneralCommon")
                    else:
                        pset_entity = None
                        for rel in elem.IsDefinedBy or []:
                            if rel.is_a("IfcRelDefinesByProperties") and getattr(rel.RelatingPropertyDefinition, "Name", "") == "Additional_Pset_GeneralCommon":
                                pset_entity = rel.RelatingPropertyDefinition
                                break
                    if pset_entity:
                        ifcopenshell.api.run("pset.edit_pset", ifc, pset=pset_entity, properties={field_name: val})
                except Exception as exc:
                    APP_LOGGER.warning("Failed to write Civil3D field %s on %s: %s", field_name, guid, exc)

    def set_uniclass(df, source_name):
        if df is None:
            return
        cls_src = None
        for c in ifc.by_type("IfcClassification"):
            if getattr(c, "Name", "") == source_name:
                cls_src = c
                break
        if cls_src is None and add_new == "yes":
            cls_src = ifc.create_entity(
                "IfcClassification",
                Name=source_name,
                Source="https://www.thenbs.com/our-tools/uniclass-2015",
                Edition="2015",
            )
        for _, r in df.iterrows():
            guid = r.get("GlobalId")
            if pd.isna(guid):
                continue
            try:
                elem = ifc.by_guid(guid)
            except Exception as exc:
                APP_LOGGER.warning("IFC by_guid failed during Uniclass mapping guid=%s error=%s", guid, exc)
                continue
            if not elem:
                continue
            ref = clean_value(r.get("Reference"))
            nm = clean_value(r.get("Name"))
            if ref is None and nm is None:
                continue
            existing_ref = None
            for rel in getattr(elem, "HasAssociations", []) or []:
                if rel.is_a("IfcRelAssociatesClassification"):
                    cref = rel.RelatingClassification
                    if cref and cref.is_a("IfcClassificationReference"):
                        src = getattr(cref, "ReferencedSource", None)
                        if src and getattr(src, "Name", "") == source_name:
                            existing_ref = cref
                            break
            if existing_ref:
                if ref is not None:
                    existing_ref.ItemReference = str(ref)
                if nm is not None:
                    existing_ref.Name = str(nm)
            elif add_new == "yes" and (ref is not None or nm is not None) and cls_src is not None:
                cref = ifc.create_entity(
                    "IfcClassificationReference",
                    ItemReference=str(ref) if ref is not None else None,
                    Name=str(nm) if nm is not None else None,
                )
                cref.ReferencedSource = cls_src
                ifc.create_entity(
                    "IfcRelAssociatesClassification",
                    GlobalId=new_guid(),
                    RelatedObjects=[elem],
                    RelatingClassification=cref,
                )

    set_uniclass(uniclass_pr_df, "Uniclass Pr Products")
    set_uniclass(uniclass_ss_df, "Uniclass Ss Systems")
    set_uniclass(uniclass_ef_df, "Uniclass EF Elements Functions")

    ifc.write(output_path)
    log_memory_stage(stage="IFC write/export", session_id=session_id, file_name=os.path.basename(output_path), file_size=os.path.getsize(output_path), endpoint=endpoint, started_at=started_at)
    if en_entities_value:
        valid_en_entities, validation_message = _validate_en_entities_writeback(output_path, en_entities_value)
        APP_LOGGER.info("EN Entities write-back validation result=%s detail=%s", valid_en_entities, validation_message)
        if not valid_en_entities:
            raise ValueError(validation_message)

    xls.close()
    if change_log_rows:
        with pd.ExcelWriter(xls_path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
            pd.DataFrame(change_log_rows).to_excel(writer, sheet_name="ChangeLog", index=False)
    if gc is not None:
        gc.collect()
    return output_path


# ----------------------------------------------------------------------------
# Storey / Global Z + BaseQuantities (from app (2).py)
# ----------------------------------------------------------------------------

_SI_PREFIX_TO_M = {None: 1.0, "MILLI": 1e-3, "CENTI": 1e-2, "DECI": 1e-1, "KILO": 1e3}


def model_length_unit_in_m(model) -> float:
    try:
        projs = model.by_type("IfcProject")
        if not projs:
            return 1.0
        units = projs[0].UnitsInContext
        if units and getattr(units, "Units", None):
            for u in units.Units:
                if u.is_a("IfcSIUnit") and getattr(u, "UnitType", None) == "LENGTHUNIT":
                    prefix = getattr(u, "Prefix", None)
                    factor = _SI_PREFIX_TO_M.get(prefix, 1.0)
                    return factor
    except Exception:
        pass
    return 1.0


def to_model_units_length(value, input_unit_code, model) -> float:
    if value in (None, ""):
        return None
    try:
        val = float(value)
    except Exception:
        return None
    if input_unit_code == "m":
        return val / model_length_unit_in_m(model)
    if input_unit_code == "mm":
        return (val * 0.001) / model_length_unit_in_m(model)
    return val


def ui_to_meters(value, units_code) -> float:
    if value in (None, ""):
        return 0.0
    val = float(value)
    return val if units_code == "m" else val / 1000.0


def meters_to_model_units(val_m, model) -> float:
    mu = model_length_unit_in_m(model)
    return val_m / mu if mu else val_m


def get_first_owner_history(model):
    oh = model.by_type("IfcOwnerHistory")
    return oh[0] if oh else None


def find_storeys(model):
    storeys = []
    for st in model.by_type("IfcBuildingStorey"):
        label = f"{st.Name or '(unnamed)'} — Elev: {getattr(st, 'Elevation', None)}"
        storeys.append((st.id(), label, st, getattr(st, "Elevation", None)))
    storeys.sort(key=lambda s: (s[3] is None, s[3]))
    return storeys


def get_existing_elq(model, storey):
    if not storey or not storey.IsDefinedBy:
        return None
    for rel in storey.IsDefinedBy:
        if rel.is_a("IfcRelDefinesByProperties"):
            pdef = rel.RelatingPropertyDefinition
            if pdef and pdef.is_a("IfcElementQuantity") and getattr(pdef, "Name", "") == "BaseQuantities":
                return pdef
    return None


def find_qtylength(elq, name):
    if not elq or not getattr(elq, "Quantities", None):
        return None
    for q in elq.Quantities:
        if q.is_a("IfcQuantityLength") and getattr(q, "Name", "") == name:
            return q
    return None


def ensure_qtylength(model, elq, name, value_model_units, description=None):
    ql = find_qtylength(elq, name)
    if ql:
        ql.LengthValue = float(value_model_units)
        if description is not None:
            ql.Description = description
        return ql
    ql = model.create_entity(
        "IfcQuantityLength",
        Name=name,
        LengthValue=float(value_model_units),
        Description=description,
        Unit=None,
    )
    if getattr(elq, "Quantities", None):
        elq.Quantities = list(elq.Quantities) + [ql]
    else:
        elq.Quantities = [ql]
    return ql


def create_or_update_storey_basequantities(
    model,
    storey,
    gross_val_ui=None,
    net_val_ui=None,
    input_unit_code="m",
    method_of_measurement=None,
    mirror_to_qto=False,
):
    owner_history = get_first_owner_history(model)
    elq = get_existing_elq(model, storey)
    if not elq:
        elq = model.create_entity(
            "IfcElementQuantity",
            GlobalId=new_guid(),
            OwnerHistory=owner_history,
            Name="BaseQuantities",
            MethodOfMeasurement=method_of_measurement,
        )
        rel = model.create_entity(
            "IfcRelDefinesByProperties",
            GlobalId=new_guid(),
            OwnerHistory=owner_history,
            Name="BaseQuantities",
            Description=None,
            RelatingPropertyDefinition=elq,
            RelatedObjects=[storey],
        )
        if getattr(storey, "IsDefinedBy", None):
            storey.IsDefinedBy = list(storey.IsDefinedBy) + [rel]
        else:
            storey.IsDefinedBy = [rel]
    else:
        if method_of_measurement is not None:
            elq.MethodOfMeasurement = method_of_measurement

    mu_factor = model_length_unit_in_m(model)
    if gross_val_ui is not None:
        gross_in_model = to_model_units_length(gross_val_ui, input_unit_code, model)
        ensure_qtylength(model, elq, "GrossHeight", gross_in_model)
    if net_val_ui is not None:
        net_in_model = to_model_units_length(net_val_ui, input_unit_code, model)
        ensure_qtylength(model, elq, "NetHeight", net_in_model)

    if mirror_to_qto:
        elq.Name = "Qto_BuildingStoreyBaseQuantities"


def ascend_to_root_local_placement(lp):
    cur = lp
    while cur and getattr(cur, "PlacementRelTo", None) is not None:
        cur = cur.PlacementRelTo
    return cur


def get_location_cartesian_point(lp):
    if lp is None:
        return None
    loc = getattr(lp, "Location", None)
    if loc and loc.is_a("IfcCartesianPoint"):
        return loc
    return None


def get_all_map_conversions(model):
    mcs = []
    for site in model.by_type("IfcSite"):
        if getattr(site, "RefLatitude", None) is not None:
            for rel in site.HasCoordinateOperation or []:
                if rel.is_a("IfcMapConversion"):
                    mcs.append(rel)
    return mcs


def countershift_product_local_points(model, delta_model):
    c = 0
    for prod in model.by_type("IfcProduct"):
        lp = getattr(prod, "ObjectPlacement", None)
        if lp and lp.is_a("IfcLocalPlacement"):
            loc = get_location_cartesian_point(lp.RelativePlacement)
            if loc is None:
                continue
            coords = list(loc.Coordinates)
            if len(coords) < 3:
                coords += [0.0] * (3 - len(coords))
            try:
                coords[2] = float(coords[2] or 0.0) - delta_model
                new_pt = model.create_entity(
                    "IfcCartesianPoint",
                    Coordinates=(
                        float(coords[0]) if coords[0] is not None else 0.0,
                        float(coords[1]) if coords[1] is not None else 0.0,
                        coords[2],
                    ),
                )
                lp.RelativePlacement.Location = new_pt
                c += 1
            except Exception:
                pass
    return c


def adjust_local_placement_z(lp, delta_model):
    if lp is None or not lp.is_a("IfcLocalPlacement"):
        return
    rel = getattr(lp, "RelativePlacement", None)
    if rel and rel.is_a("IfcAxis2Placement3D"):
        loc = getattr(rel, "Location", None)
        if loc and loc.is_a("IfcCartesianPoint"):
            coords = list(loc.Coordinates)
            if len(coords) < 3:
                coords += [0.0] * (3 - len(coords))
            coords[2] = float(coords[2] or 0.0) + delta_model
            loc.Coordinates = tuple(coords)


def parse_ifc_storeys(ifc_path: str) -> Dict[str, Any]:
    size = os.path.getsize(ifc_path)
    model = ifcopenshell.open(ifc_path)
    storeys = find_storeys(model)
    unit_m = model_length_unit_in_m(model)
    unit_label = "m" if abs(unit_m - 1.0) < 1e-12 else ("mm" if abs(unit_m - 1e-3) < 1e-12 else f"{unit_m} m/unit")
    mc_list = get_all_map_conversions(model)
    choices = [
        {
            "id": sid,
            "label": lbl,
        }
        for (sid, lbl, _ent, _elev) in storeys
    ]
    return {
        "storeys": choices,
        "summary": f"{len(choices)} storey level(s); model unit: {unit_label}; size {human_size(size)}",
        "map_conversions": len(mc_list),
        "unit_label": unit_label,
    }


def apply_storey_changes(
    ifc_path: str,
    storey_id: Optional[int],
    units_code: str,
    gross_val: Optional[float],
    net_val: Optional[float],
    mom_txt: Optional[str],
    mirror: bool,
    target_z: Optional[float],
    countershift_geometry: bool,
    use_crs_mode: bool,
    update_all_mcs: bool,
    show_diag: bool,
    crs_set_storey_elev: bool,
    output_path: str,
) -> Tuple[str, str]:
    model = ifcopenshell.open(ifc_path)
    storey = model.by_id(int(storey_id)) if storey_id else None
    if not storey:
        raise ValueError("Selected storey not found")

    gross_maybe = gross_val if gross_val not in (None, "") else None
    net_maybe = net_val if net_val not in (None, "") else None
    if gross_maybe is not None or net_maybe is not None:
        create_or_update_storey_basequantities(
            model,
            storey,
            gross_val_ui=gross_maybe,
            net_val_ui=net_maybe,
            input_unit_code=units_code,
            method_of_measurement=(mom_txt or None),
            mirror_to_qto=bool(mirror),
        )

    delta_model = 0.0
    used_path = "root-local"
    diag_lines: List[str] = []

    mc_list = []
    if use_crs_mode and target_z not in (None, ""):
        all_mcs = get_all_map_conversions(model)
        if all_mcs:
            mc_list = all_mcs if update_all_mcs else [all_mcs[0]]

    if mc_list:
        new_m = ui_to_meters(target_z, units_code)
        delta_m = new_m
        for mc in mc_list:
            old_height = float(getattr(mc, "OrthogonalHeight", 0.0) or 0.0)
            mc.OrthogonalHeight = new_m
            if show_diag:
                diag_lines.append(f"MapConversion {mc.id()} OrthogonalHeight {old_height} → {new_m} m")
        if crs_set_storey_elev:
            delta_model = meters_to_model_units(delta_m, model)
            old_storey_elev = float(getattr(storey, "Elevation", 0.0) or 0.0)
            storey.Elevation = old_storey_elev + delta_model
            if show_diag:
                diag_lines.append(f"Storey.Elevation {old_storey_elev} mu → {storey.Elevation} mu")
        else:
            delta_model = meters_to_model_units(delta_m, model)
        used_path = "crs-mapconversion(all)" if (update_all_mcs and len(mc_list) > 1) else "crs-mapconversion"
    else:
        if target_z not in (None, ""):
            root_lp = ascend_to_root_local_placement(storey.ObjectPlacement)
            root_pt = get_location_cartesian_point(root_lp)
            if root_pt is None:
                raise ValueError("Could not find root CartesianPoint for the storey's placement.")
            coords = list(root_pt.Coordinates)
            if len(coords) < 3:
                raise ValueError("Root CartesianPoint has no Z coordinate.")
            old_z = float(coords[2]) if coords[2] is not None else 0.0
            new_z = to_model_units_length(target_z, units_code, model)
            delta_model = new_z - old_z
            coords[2] = new_z
            root_pt.Coordinates = tuple(coords)
            old_storey_elev = float(getattr(storey, "Elevation", 0.0) or 0.0)
            storey.Elevation = float(new_z)
            if show_diag:
                diag_lines.append(f"RootLP {old_z} mu → {new_z} mu (Δ={delta_model} mu)")
                diag_lines.append(f"Storey.Elevation (Root mode) {old_storey_elev} mu → {storey.Elevation} mu")
            used_path = "root-local"

    shifted = 0
    if countershift_geometry and abs(delta_model) > 0:
        shifted = countershift_product_local_points(model, delta_model)

    model.write(output_path)

    mu_m = model_length_unit_in_m(model)
    mu_label = "m" if abs(mu_m - 1.0) < 1e-12 else ("mm" if abs(mu_m - 1e-3) < 1e-12 else f"{mu_m} m/unit")
    parts = [
        "Done ✅",
        f"Schema: {model.schema}",
        f"Model length unit: {mu_label}",
        f"Mode: {'CRS (IfcMapConversion)' if used_path.startswith('crs-mapconversion') else 'Root LocalPlacement'}",
        f"Target Z = {target_z if target_z not in (None,'') else ''} {units_code}",
        f"Δ applied (model units) = {delta_model}",
        (f"Counter-shifted {shifted} product placements by −Δ (kept world positions)." if shifted else None),
    ]
    try:
        site = (model.by_type("IfcSite") or [None])[0]
        site_ref = float(getattr(site, "RefElevation", 0.0) or 0.0) if site else 0.0
        parts.append(f"Site.RefElevation = {site_ref} mu")
        parts.append(f"Storey.Elevation = {float(getattr(storey,'Elevation',0.0) or 0.0)} mu")
        mcs = get_all_map_conversions(model)
        if mcs:
            parts.append(f"MapConversion.OrthogonalHeight = {float(getattr(mcs[0],'OrthogonalHeight',0.0) or 0.0)} m")
    except Exception:
        pass
    if show_diag and diag_lines:
        parts.append("Diagnostics:")
        parts.extend([" • " + d for d in diag_lines])
    if gross_maybe is not None:
        parts.append(f" • GrossHeight (UI {units_code}) = {gross_maybe}")
    if net_maybe is not None:
        parts.append(f" • NetHeight  (UI {units_code}) = {net_maybe}")
    if mom_txt:
        parts.append(f" • MethodOfMeasurement = '{mom_txt}'")
    parts.append(f"Output: {os.path.basename(output_path)}")
    return "\n".join([p for p in parts if p]), output_path


# ----------------------------------------------------------------------------
# Level manager: list, update, delete (with reassignment), add
# ----------------------------------------------------------------------------


def storey_elevation(storey) -> float:
    try:
        return float(getattr(storey, "Elevation", 0.0) or 0.0)
    except Exception:
        return 0.0


def storey_comp_height(storey) -> Optional[float]:
    # IFC4 adds ElevationOfRefHeight; fall back to None
    if hasattr(storey, "ElevationOfRefHeight"):
        val = getattr(storey, "ElevationOfRefHeight", None)
        return float(val) if val not in (None, "") else None
    return None


def list_storey_objects(storey) -> List[Any]:
    objs = []
    for rel in storey.ContainsElements or []:
        if rel.is_a("IfcRelContainedInSpatialStructure"):
            for el in rel.RelatedElements or []:
                objs.append(el)
    # Include spatial children (e.g., IfcSpace) aggregated beneath the storey
    for rel in storey.IsDecomposedBy or []:
        if rel.is_a("IfcRelAggregates"):
            for child in rel.RelatedObjects or []:
                if child.is_a("IfcSpace"):
                    objs.append(child)
    # Deduplicate by express ID to avoid double-counting if an element appears in multiple relations
    seen = {}
    for obj in objs:
        seen[obj.id()] = obj
    return list(seen.values())


def ensure_storey_associations(storey) -> List[Any]:
    try:
        assoc = getattr(storey, "HasAssociations", None)
    except Exception:
        assoc = None
    if assoc in (None, (), []):
        assoc = []
    else:
        assoc = list(assoc)
    try:
        storey.HasAssociations = assoc
    except Exception:
        try:
            setattr(storey, "HasAssociations", assoc)
        except Exception:
            pass
    return assoc


def containment_rels(model, obj) -> List[Any]:
    rels = []
    if hasattr(obj, "ContainedInStructure"):
        try:
            rels = list(obj.ContainedInStructure or [])
        except Exception:
            rels = []
    if not rels:
        try:
            rels = [
                r
                for r in model.get_inverse(obj) or []
                if r.is_a("IfcRelContainedInSpatialStructure") and obj in (r.RelatedElements or [])
            ]
        except Exception:
            rels = []
    return rels


def ensure_storey_associations(storey) -> List[Any]:
    try:
        assoc = getattr(storey, "HasAssociations", None)
    except Exception:
        assoc = None
    if assoc in (None, (), []):
        assoc = []
    else:
        assoc = list(assoc)
    try:
        storey.HasAssociations = assoc
    except Exception:
        try:
            setattr(storey, "HasAssociations", assoc)
        except Exception:
            pass
    return assoc


def move_objects_to_storey(model, objects, source_storey, target_storey):
    if target_storey is None:
        return
    delta = storey_elevation(source_storey) - storey_elevation(target_storey)
    # Ensure target containment relation
    target_rel = None
    for rel in target_storey.ContainsElements or []:
        if rel.is_a("IfcRelContainedInSpatialStructure"):
            target_rel = rel
            break
    if target_rel is None:
        target_rel = model.create_entity(
            "IfcRelContainedInSpatialStructure",
            GlobalId=new_guid(),
            RelatedElements=[],
            RelatingStructure=target_storey,
        )
        if getattr(target_storey, "ContainsElements", None):
            target_storey.ContainsElements = list(target_storey.ContainsElements) + [target_rel]
        else:
            target_storey.ContainsElements = [target_rel]

    for obj in objects:
        # Remove from current containment
        for rel in containment_rels(model, obj):
            if rel.RelatingStructure == source_storey and rel.is_a("IfcRelContainedInSpatialStructure"):
                rel.RelatedElements = [e for e in rel.RelatedElements if e != obj]
        # Adjust placement to keep world position
        adjust_local_placement_z(getattr(obj, "ObjectPlacement", None), delta)
        # Add to target relation
        if obj not in target_rel.RelatedElements:
            target_rel.RelatedElements = list(target_rel.RelatedElements) + [obj]


def cleanup_empty_containment(model, storey):
    for rel in list(storey.ContainsElements or []):
        if rel.is_a("IfcRelContainedInSpatialStructure") and not rel.RelatedElements:
            try:
                storey.ContainsElements = [r for r in storey.ContainsElements if r != rel]
                model.remove(rel)
            except Exception:
                pass


def remove_storey_from_aggregates(model, storey):
    for rel in list(storey.Decomposes or []):
        if rel.is_a("IfcRelAggregates"):
            rel.RelatedObjects = [o for o in rel.RelatedObjects if o != storey]
            if not rel.RelatedObjects:
                model.remove(rel)


COBIE_FLOOR_CLASS_NAME = "COBie Floors"


def _get_cobie_floor_classification(storey) -> Tuple[Optional[str], Optional[str]]:
    """Return (item_reference, name) for the COBie Floors classification reference if present."""
    for rel in ensure_storey_associations(storey):
        if not rel.is_a("IfcRelAssociatesClassification"):
            continue
        cref = getattr(rel, "RelatingClassification", None)
        if not cref or not cref.is_a("IfcClassificationReference"):
            continue
        source = getattr(cref, "ReferencedSource", None)
        if getattr(source, "Name", "") == COBIE_FLOOR_CLASS_NAME or getattr(cref, "Name", "") == COBIE_FLOOR_CLASS_NAME:
            return getattr(cref, "ItemReference", None), getattr(cref, "Name", None)
    return None, None


def list_levels(ifc_path: str) -> Dict[str, Any]:
    model = ifcopenshell.open(ifc_path)
    result = []
    for st in model.by_type("IfcBuildingStorey"):
        objs = list_storey_objects(st)
        cobie_ref, cobie_name = _get_cobie_floor_classification(st)
        result.append(
            {
                "id": st.id(),
                "name": getattr(st, "Name", ""),
                "description": getattr(st, "Description", ""),
                "elevation": storey_elevation(st),
                "comp_height": storey_comp_height(st),
                "global_id": getattr(st, "GlobalId", None),
                "cobie_floor": cobie_ref or cobie_name,
                "object_count": len(objs),
                "objects": [
                    {"id": o.id(), "name": getattr(o, "Name", ""), "type": o.is_a()}
                    for o in objs
                ],
            }
        )
    return {"levels": result}


def update_level(ifc_path: str, storey_id: int, payload: Dict[str, Any], output_path: str) -> str:
    model = ifcopenshell.open(ifc_path)
    storey = model.by_id(int(storey_id))
    if not storey:
        raise ValueError("Storey not found")
    if "name" in payload:
        storey.Name = payload.get("name")
    if "description" in payload:
        storey.Description = payload.get("description")
    if "global_id" in payload and payload.get("global_id"):
        storey.GlobalId = payload.get("global_id")
    if "elevation" in payload and payload.get("elevation") not in (None, ""):
        storey.Elevation = float(payload.get("elevation"))
    if "comp_height" in payload and hasattr(storey, "ElevationOfRefHeight"):
        comp = payload.get("comp_height")
        if comp not in (None, ""):
            storey.ElevationOfRefHeight = float(comp)

    if "cobie_floor" in payload:
        desired_ref = payload.get("cobie_floor")
        existing_rel = None
        existing_cref = None
        associations = ensure_storey_associations(storey)
        for rel in list(associations):
            if rel.is_a("IfcRelAssociatesClassification"):
                cref = getattr(rel, "RelatingClassification", None)
                source = getattr(cref, "ReferencedSource", None) if cref else None
                if cref and cref.is_a("IfcClassificationReference") and (
                    getattr(source, "Name", "") == COBIE_FLOOR_CLASS_NAME
                    or getattr(cref, "Name", "") == COBIE_FLOOR_CLASS_NAME
                ):
                    existing_rel = rel
                    existing_cref = cref
                    break

        if desired_ref in (None, ""):
            if existing_rel:
                try:
                    storey.HasAssociations = [r for r in associations if r != existing_rel]
                    model.remove(existing_rel)
                except Exception:
                    pass
        else:
            classification = None
            for cls in model.by_type("IfcClassification"):
                if getattr(cls, "Name", "") == COBIE_FLOOR_CLASS_NAME:
                    classification = cls
                    break
            if classification is None:
                classification = model.create_entity(
                    "IfcClassification",
                    Name=COBIE_FLOOR_CLASS_NAME,
                    Source="COBie",
                )

            if existing_cref is None:
                existing_cref = model.create_entity(
                    "IfcClassificationReference",
                    ItemReference=str(desired_ref),
                    Name=COBIE_FLOOR_CLASS_NAME,
                    ReferencedSource=classification,
                )
                existing_rel = model.create_entity(
                    "IfcRelAssociatesClassification",
                    GlobalId=new_guid(),
                    RelatedObjects=[storey],
                    RelatingClassification=existing_cref,
                )
                associations = ensure_storey_associations(storey)
                storey.HasAssociations = list(associations) + [existing_rel]
            else:
                existing_cref.ItemReference = str(desired_ref)
                if getattr(existing_cref, "Name", "") in (None, "", COBIE_FLOOR_CLASS_NAME):
                    existing_cref.Name = COBIE_FLOOR_CLASS_NAME

    model.write(output_path)
    return output_path


def delete_level(ifc_path: str, storey_id: int, target_storey_id: int, object_ids: Optional[List[int]], output_path: str) -> str:
    model = ifcopenshell.open(ifc_path)
    storey = model.by_id(int(storey_id))
    target = model.by_id(int(target_storey_id)) if target_storey_id else None
    if not storey or not target:
        raise ValueError("Storey or target storey not found")

    objs = list_storey_objects(storey)
    if object_ids:
        objs = [o for o in objs if o.id() in object_ids]
    move_objects_to_storey(model, objs, storey, target)
    cleanup_empty_containment(model, storey)
    remove_storey_from_aggregates(model, storey)
    try:
        model.remove(storey)
    except Exception:
        pass
    model.write(output_path)
    return output_path


def add_level(
    ifc_path: str,
    name: str,
    description: Optional[str],
    elevation: Optional[float],
    comp_height: Optional[float],
    object_ids: Optional[List[int]],
    output_path: str,
) -> str:
    model = ifcopenshell.open(ifc_path)
    building = (model.by_type("IfcBuilding") or [None])[0]
    site = (model.by_type("IfcSite") or [None])[0]
    parent = building or site
    if parent is None:
        raise ValueError("No Building or Site found to host the new level")

    storey = model.create_entity(
        "IfcBuildingStorey",
        GlobalId=new_guid(),
        Name=name or "New Storey",
        Description=description,
        Elevation=float(elevation) if elevation not in (None, "") else None,
    )
    if comp_height not in (None, "") and hasattr(storey, "ElevationOfRefHeight"):
        storey.ElevationOfRefHeight = float(comp_height)

    ensure_aggregates(parent, storey, model)

    if object_ids:
        objs = [model.by_id(int(oid)) for oid in object_ids if model.by_id(int(oid))]
        for obj in objs:
            # find original storey for delta
            origin_storey = None
            for rel in containment_rels(model, obj):
                if rel.is_a("IfcRelContainedInSpatialStructure"):
                    origin_storey = rel.RelatingStructure
                    break
            delta = 0.0
            if origin_storey:
                delta = storey_elevation(origin_storey) - storey_elevation(storey)
            adjust_local_placement_z(getattr(obj, "ObjectPlacement", None), delta)
            # remove from old relations
            for rel in containment_rels(model, obj):
                if rel.is_a("IfcRelContainedInSpatialStructure"):
                    rel.RelatedElements = [e for e in rel.RelatedElements if e != obj]
            # add to new relation
            target_rel = None
            for rel in storey.ContainsElements or []:
                if rel.is_a("IfcRelContainedInSpatialStructure"):
                    target_rel = rel
                    break
            if target_rel is None:
                target_rel = model.create_entity(
                    "IfcRelContainedInSpatialStructure",
                    GlobalId=new_guid(),
                    RelatedElements=[],
                    RelatingStructure=storey,
                )
                storey.ContainsElements = list(storey.ContainsElements or []) + [target_rel]
            target_rel.RelatedElements = list(target_rel.RelatedElements) + [obj]

    model.write(output_path)
    return output_path


def reassign_objects(
    ifc_path: str,
    source_storey_id: int,
    target_storey_id: int,
    object_ids: Optional[List[int]],
    output_path: str,
) -> str:
    model = ifcopenshell.open(ifc_path)
    source = model.by_id(int(source_storey_id))
    target = model.by_id(int(target_storey_id))
    if not source or not target:
        raise ValueError("Source or target storey not found")
    objs = list_storey_objects(source)
    if object_ids:
        objs = [o for o in objs if o.id() in object_ids]
    move_objects_to_storey(model, objs, source, target)
    cleanup_empty_containment(model, source)
    model.write(output_path)
    return output_path


def apply_level_actions(ifc_path: str, actions: List[Dict[str, Any]], output_path: str) -> str:
    if not actions:
        raise ValueError("No actions supplied")
    work_path = output_path
    current_path = ifc_path
    for idx, action in enumerate(actions):
        kind = action.get("type")
        if kind == "update":
            update_level(
                current_path,
                int(action["storey_id"]),
                action.get("payload", {}),
                work_path,
            )
        elif kind == "delete":
            delete_level(
                current_path,
                int(action["storey_id"]),
                int(action["target_storey_id"]),
                action.get("object_ids"),
                work_path,
            )
        elif kind == "add":
            add_level(
                current_path,
                name=action.get("name"),
                description=action.get("description"),
                elevation=action.get("elevation"),
                comp_height=action.get("comp_height"),
                object_ids=action.get("object_ids"),
                output_path=work_path,
            )
        elif kind == "reassign":
            reassign_objects(
                current_path,
                int(action["source_storey_id"]),
                int(action["target_storey_id"]),
                action.get("object_ids"),
                work_path,
            )
        else:
            raise ValueError(f"Unsupported action type: {kind}")
        current_path = work_path
    return work_path


# ----------------------------------------------------------------------------
# Proxy / Type mapper (from app (3).py)
# ----------------------------------------------------------------------------

FALLBACK_ENUM_LIBRARY = {
    "IfcWasteTerminalTypeEnum": {
        "FLOORTRAP",
        "FLOORWASTE",
        "GULLYSUMP",
        "GULLYTRAP",
        "GREASEINTERCEPTOR",
        "OILINTERCEPTOR",
        "PETROLINTERCEPTOR",
        "ROOFDRAIN",
        "WASTEDISPOSALUNIT",
        "WASTETRAP",
        "USERDEFINED",
        "NOTDEFINED",
    },
    "IfcPipeSegmentTypeEnum": {
        "CULVERT",
        "FLEXIBLESEGMENT",
        "GUTTER",
        "RIGIDSEGMENT",
        "SPOOL",
        "USERDEFINED",
        "NOTDEFINED",
    },
    "IfcDistributionChamberElementTypeEnum": {
        "FORMEDDUCT",
        "INSPECTIONCHAMBER",
        "INSPECTIONPIT",
        "MANHOLE",
        "METERCHAMBER",
        "SUMP",
        "TRENCH",
        "VALVECHAMBER",
        "USERDEFINED",
        "NOTDEFINED",
    },
    "IfcTankTypeEnum": {
        "PREFORMED",
        "SECTIONAL",
        "EXPANSION",
        "PRESSUREVESSEL",
        "FEEDANDEXPANSION",
        "USERDEFINED",
        "NOTDEFINED",
    },
}

LEGACY_OCCURRENCE_FALLBACK = {
    "IFCWASTETERMINALTYPE": "IFCFLOWTERMINAL",
    "IFCPIPESEGMENTTYPE": "IFCFLOWSEGMENT",
    "IFCTANKTYPE": "IFCFLOWSTORAGEDEVICE",
}


def normalize_token(value: str) -> str:
    return re.sub(r"[^a-z0-9]", "", (value or "").lower())


def split_meaningful_tokens(type_name: str) -> List[str]:
    return [part.strip() for part in (type_name or "").split("_") if part.strip()]


def parse_name_parts(type_name: str) -> Dict[str, Any]:
    raw_tokens = split_meaningful_tokens(type_name)
    ordinal_index = None
    for idx in range(len(raw_tokens) - 1, -1, -1):
        if re.fullmatch(r"type\s*\d*", raw_tokens[idx], re.IGNORECASE):
            ordinal_index = idx
            break

    tokens_without_ordinal = raw_tokens[:ordinal_index] if ordinal_index is not None else raw_tokens
    classish_raw = tokens_without_ordinal[0] if tokens_without_ordinal else ""
    predef_candidate_raw = ""
    if len(tokens_without_ordinal) > 1:
        predef_candidate_raw = "_".join(tokens_without_ordinal[1:])

    return {
        "raw_tokens": raw_tokens,
        "tokens_without_ordinal": tokens_without_ordinal,
        "classish_raw": classish_raw,
        "predef_candidate_raw": predef_candidate_raw,
        "ordinal_raw": raw_tokens[ordinal_index] if ordinal_index is not None else "",
    }


def _schema_definition(schema_name: str):
    try:
        return ifcopenshell.ifcopenshell_wrapper.schema_by_name(schema_name)
    except Exception:
        return None


def _entity_names(schema_def) -> set:
    if schema_def is None:
        return set()
    out = set()
    try:
        for decl in schema_def.declarations():
            try:
                if decl.as_entity() is not None:
                    out.add(decl.name())
            except Exception:
                continue
    except Exception:
        pass
    return out


def build_type_class_lookup(schema_name: str) -> Dict[str, str]:
    schema_def = _schema_definition(schema_name)
    lookup = {}
    if schema_def is None:
        return lookup
    for entity_name in _entity_names(schema_def):
        if not entity_name.startswith("Ifc") or not entity_name.endswith("Type"):
            continue
        key = normalize_token(entity_name[3:-4])
        if key:
            lookup[key] = entity_name
    return lookup


def resolve_type_class_from_name(type_name: str, type_lookup: Dict[str, str]) -> Dict[str, Any]:
    parsed = parse_name_parts(type_name)
    raw_tokens = parsed["raw_tokens"]
    tokens_without_ordinal = parsed["tokens_without_ordinal"]

    classish_tokens_used = 0
    resolved_type_class = None
    parsed_classish = parsed["classish_raw"]
    parsed_predef = parsed["predef_candidate_raw"]

    if parsed_classish:
        direct_key = normalize_token(parsed_classish)
        resolved_type_class = type_lookup.get(direct_key)
        if resolved_type_class:
            classish_tokens_used = 1

    if not resolved_type_class:
        for i in range(1, len(tokens_without_ordinal) + 1):
            candidate = "".join(tokens_without_ordinal[:i])
            resolved = type_lookup.get(normalize_token(candidate))
            if resolved:
                resolved_type_class = resolved
                classish_tokens_used = i
                parsed_classish = "_".join(tokens_without_ordinal[:i])
                parsed_predef = "_".join(tokens_without_ordinal[i:])
                break

    return {
        **parsed,
        "parsed_classish": parsed_classish,
        "parsed_predef": parsed_predef,
        "resolved_type_class": resolved_type_class,
        "classish_tokens_used": classish_tokens_used,
        "raw_tokens": raw_tokens,
    }


def _predefined_type_info(schema_name: str, entity_class: str) -> Dict[str, Any]:
    schema_def = _schema_definition(schema_name)
    if schema_def is None:
        return {"has_predefined": False, "enum_name": None, "enum_items": []}
    try:
        decl = schema_def.declaration_by_name(entity_class)
    except Exception:
        return {"has_predefined": False, "enum_name": None, "enum_items": []}

    attr = None
    try:
        for a in decl.all_attributes():
            if a.name().lower() == "predefinedtype":
                attr = a
                break
    except Exception:
        attr = None
    if attr is None:
        return {"has_predefined": False, "enum_name": None, "enum_items": []}

    enum_name = None
    enum_items: List[str] = []
    try:
        attr_type = attr.type_of_attribute()
        declared = attr_type.declared_type() if hasattr(attr_type, "declared_type") else None
        if declared is not None and hasattr(declared, "enumeration_items"):
            enum_name = declared.name()
            enum_items = [str(item) for item in declared.enumeration_items()]
    except Exception:
        pass
    return {"has_predefined": True, "enum_name": enum_name, "enum_items": enum_items}


def resolve_predefined_literal(predef_candidate_raw: str, enum_items: List[str]) -> Dict[str, str]:
    enum_lookup = {normalize_token(item): item for item in enum_items}
    normalized_candidate = normalize_token(predef_candidate_raw)
    if normalized_candidate and normalized_candidate in enum_lookup:
        return {"value": enum_lookup[normalized_candidate], "reason": "enum matched"}
    if "USERDEFINED" in enum_items:
        return {"value": "USERDEFINED", "reason": "no enum match → USERDEFINED"}
    return {"value": "", "reason": "no enum match"}


def resolve_occurrence_from_type_class(schema_name: str, type_class: Optional[str]) -> Optional[str]:
    if not type_class:
        return None
    if not type_class.upper().endswith("TYPE"):
        return None
    base = type_class[:-4]
    entity_names = _entity_names(_schema_definition(schema_name))
    if base in entity_names:
        return base
    return LEGACY_OCCURRENCE_FALLBACK.get(type_class.upper())

_ENUM_LIBRARY_CACHE: Dict[str, Dict[Tuple[str, str, str], str]] = {}
_PSET_APPLICABILITY_CACHE: Optional[Dict[Tuple[str, str], Dict[str, str]]] = None


def build_entity_predefined_enum_library(schema_name: str) -> Dict[Tuple[str, str, str], str]:
    if schema_name in _ENUM_LIBRARY_CACHE:
        return _ENUM_LIBRARY_CACHE[schema_name]
    lookup: Dict[Tuple[str, str, str], str] = {}
    schema_def = _schema_definition(schema_name)
    if schema_def is None:
        _ENUM_LIBRARY_CACHE[schema_name] = lookup
        return lookup

    for entity_name in _entity_names(schema_def):
        info = _predefined_type_info(schema_name, entity_name)
        if not info.get("has_predefined"):
            continue
        for lit in info.get("enum_items", []):
            lookup[(entity_name, "PredefinedType", normalize_token(lit))] = lit

    _ENUM_LIBRARY_CACHE[schema_name] = lookup
    return lookup


def _extract_applicable_type_value(pset_name: str, ifc_class: str) -> str:
    stem = pset_name[len("Pset_") :] if pset_name.startswith("Pset_") else pset_name
    class_stem = ifc_class[3:] if ifc_class.startswith("Ifc") else ifc_class
    if stem.startswith(class_stem):
        return stem[len(class_stem) :]
    return ""


def load_ifc2x3_pset_applicability_library() -> Dict[Tuple[str, str], Dict[str, str]]:
    global _PSET_APPLICABILITY_CACHE
    if _PSET_APPLICABILITY_CACHE is not None:
        return _PSET_APPLICABILITY_CACHE

    lookup: Dict[Tuple[str, str], Dict[str, str]] = {}
    p = DATA_DIR / "ifc2x3_pset_applicability.json"
    if p.exists():
        try:
            records = json.loads(p.read_text(encoding="utf-8"))
            for row in records:
                ifc_class = row.get("ifc_class", "")
                value = row.get("applicable_type_value", "")
                if not ifc_class or not value:
                    continue
                lookup[(ifc_class, normalize_token(value))] = {
                    "pset_name": row.get("pset_name", ""),
                    "applicable_type_value": value,
                }
        except Exception:
            lookup = {}

    if not lookup:
        template = ifcopenshell.util.pset.get_template("IFC2X3")
        for template_file in template.templates:
            for pset in template_file.by_type("IfcPropertySetTemplate"):
                pset_name = getattr(pset, "Name", "") or ""
                ifc_class = getattr(pset, "ApplicableEntity", "") or ""
                if not pset_name.startswith("Pset_"):
                    continue
                if not ifc_class.startswith("Ifc") or not ifc_class.endswith("Type"):
                    continue
                value = _extract_applicable_type_value(pset_name, ifc_class)
                if not value:
                    continue
                lookup[(ifc_class, normalize_token(value))] = {
                    "pset_name": pset_name,
                    "applicable_type_value": value,
                }

    _PSET_APPLICABILITY_CACHE = lookup
    return lookup


def resolve_pset_applicability(resolved_type_class: str, parsed_predef_token: str) -> Optional[Dict[str, str]]:
    if not resolved_type_class or not parsed_predef_token:
        return None
    lib = load_ifc2x3_pset_applicability_library()
    return lib.get((resolved_type_class, normalize_token(parsed_predef_token)))


def resolve_type_and_predefined_for_name(type_name: str, schema_name: str) -> Dict[str, Any]:
    type_lookup = build_type_class_lookup(schema_name)
    resolved = resolve_type_class_from_name(type_name, type_lookup)
    resolved_type_class = resolved.get("resolved_type_class")
    predef_info = _predefined_type_info(schema_name, resolved_type_class) if resolved_type_class else {
        "has_predefined": False,
        "enum_name": None,
        "enum_items": [],
    }
    predef_resolution = resolve_predefined_literal(resolved.get("parsed_predef", ""), predef_info.get("enum_items", []))
    pset_match = resolve_pset_applicability(resolved_type_class, resolved.get("parsed_predef", "")) if schema_name.upper() == "IFC2X3" else None
    return {
        **resolved,
        "resolved_predefined_type": predef_resolution.get("value", ""),
        "resolved_predefined_reason": predef_resolution.get("reason", ""),
        "predef_info": predef_info,
        "pset_applicability_match": pset_match,
    }


def rewrite_proxy_types(in_path: str, out_path: str) -> Tuple[str, str]:
    with open(in_path, "r", encoding="utf-8", errors="ignore") as f:
        lines = f.readlines()

    model = None
    schema_name = "IFC4"
    try:
        model = ifcopenshell.open(in_path)
        schema_name = model.schema
    except Exception:
        schema_name = "IFC4"

    stats = {
        "proxy_types_total": 0,
        "building_types_total": 0,
        "proxy_types_converted": 0,
        "building_types_converted": 0,
        "left_as_proxy_type": 0,
        "left_as_building_type": 0,
        "occurrences_converted": 0,
    }

    typeid_to_occ_entity = {}

    proxy_type_re = re.compile(
        r"^(?P<ws>\s*)(?P<id>#\d+)=IFCBUILDINGELEMENTPROXYTYPE"
        r"\('(?P<guid>[^']*)',"
        r"(?P<own>[^,]*),"
        r"'(?P<name>[^']*)',"
        r"(?P<mid>.*),"
        r"\.(?P<enum>\w+)\.\);",
        re.IGNORECASE,
    )

    building_type_re = re.compile(
        r"^(?P<ws>\s*)(?P<id>#\d+)=IFCBUILDINGELEMENTTYPE"
        r"\('(?P<guid>[^']*)',"
        r"(?P<own>[^,]*),"
        r"'(?P<name>[^']*)',"
        r"(?P<mid>.*)\);",
        re.IGNORECASE,
    )

    updated_lines = []

    for line in lines:
        m_proxy = proxy_type_re.match(line)
        if m_proxy:
            stats["proxy_types_total"] += 1
            g = m_proxy.groupdict()
            ws = g["ws"]
            type_id = g["id"]
            guid = g["guid"]
            owner = g["own"]
            type_name = g["name"]
            mid = g["mid"]

            resolved = resolve_type_and_predefined_for_name(type_name, schema_name)
            target_type = resolved.get("resolved_type_class")
            if not target_type:
                stats["left_as_proxy_type"] += 1
                updated_lines.append(line)
                continue
            enum_val = resolved.get("resolved_predefined_type") or "USERDEFINED"

            new_line = (
                f"{ws}{type_id}={target_type}('{guid}',{owner},"
                f"'{type_name}',{mid},.{enum_val}.);"
            )
            updated_lines.append(new_line)
            stats["proxy_types_converted"] += 1

            occ_entity = resolve_occurrence_from_type_class(schema_name, target_type) or "IFCBUILDINGELEMENTPROXY"
            typeid_to_occ_entity[type_id] = occ_entity.upper()
            continue

        m_build = building_type_re.match(line)
        if m_build:
            stats["building_types_total"] += 1
            g = m_build.groupdict()
            ws = g["ws"]
            type_id = g["id"]
            guid = g["guid"]
            owner = g["own"]
            type_name = g["name"]
            mid = g["mid"]

            resolved = resolve_type_and_predefined_for_name(type_name, schema_name)
            target_type = resolved.get("resolved_type_class")
            if not target_type:
                stats["left_as_building_type"] += 1
                updated_lines.append(line)
                continue
            enum_val = resolved.get("resolved_predefined_type") or "USERDEFINED"

            new_line = (
                f"{ws}{type_id}={target_type}('{guid}',{owner},"
                f"'{type_name}',{mid},.{enum_val}.);"
            )
            updated_lines.append(new_line)
            stats["building_types_converted"] += 1

            occ_entity = resolve_occurrence_from_type_class(schema_name, target_type) or "IFCBUILDINGELEMENTPROXY"
            typeid_to_occ_entity[type_id] = occ_entity.upper()
            continue

        updated_lines.append(line)

    rel_def_type_re = re.compile(
        r"^(?P<ws>\s*)#(?P<relid>\d+)=IFCRELDEFINESBYTYPE\("
        r"'(?P<guid>[^']*)',"
        r"(?P<owner>[^,]*),"
        r"(?P<name>[^,]*),"
        r"(?P<desc>[^,]*),"
        r"\((?P<objs>[^)]*)\),"
        r"(?P<typeid>#\d+)\);",
        re.IGNORECASE,
    )

    occid_to_entity = {}
    for line in updated_lines:
        m = rel_def_type_re.match(line)
        if not m:
            continue
        d = m.groupdict()
        type_id = d["typeid"]
        if type_id not in typeid_to_occ_entity:
            continue
        occ_entity = typeid_to_occ_entity[type_id]
        objs_raw = d["objs"]
        obj_ids = [o.strip() for o in objs_raw.split(",") if o.strip()]
        for oid in obj_ids:
            occid_to_entity[oid] = occ_entity

    final_lines = []

    occ_re = re.compile(
        r"^(?P<ws>\s*)(?P<id>#\d+)=IFCBUILDINGELEMENTPROXY\(",
        re.IGNORECASE,
    )

    for line in updated_lines:
        m = occ_re.match(line)
        if not m:
            final_lines.append(line)
            continue

        ws = m.group("ws")
        occ_id = m.group("id")
        if occ_id not in occid_to_entity:
            final_lines.append(line)
            continue

        new_entity = occid_to_entity[occ_id]
        rest = line[m.end():]
        new_line = f"{ws}{occ_id}={new_entity}({rest}"
        final_lines.append(new_line)
        stats["occurrences_converted"] += 1

    with open(out_path, "w", encoding="utf-8") as f:
        f.writelines(final_lines)

    base = os.path.basename(in_path)
    summary = (
        f"Input file:  {base}\n"
        f"Output file: {os.path.basename(out_path)}\n\n"
        f"Proxy types (IFCBUILDINGELEMENTPROXYTYPE) found: {stats['proxy_types_total']}\n"
        f"  → converted to specific IFC types: {stats['proxy_types_converted']}\n"
        f"  → left as IFCBUILDINGELEMENTPROXYTYPE: {stats['left_as_proxy_type']}\n\n"
        f"Building types (IFCBUILDINGELEMENTTYPE) found: {stats['building_types_total']}\n"
        f"  → converted to specific IFC types: {stats['building_types_converted']}\n"
        f"  → left as IFCBUILDINGELEMENTTYPE: {stats['left_as_building_type']}\n\n"
        f"Occurrences converted from IfcBuildingElementProxy to typed entities: "
        f"{stats['occurrences_converted']}\n\n"
        "Occurrences are retyped only when an IfcRelDefinesByType exists and the "
        "referenced type could be mapped. Mapping is IFC2x3-compliant: waste terminals "
        "→ IfcFlowTerminal/IfcWasteTerminalType, pipe segments → "
        "IfcFlowSegment/IfcPipeSegmentType, tanks → "
        "IfcFlowStorageDevice/IfcTankType, distribution chambers → "
        "IfcDistributionChamberElement/IfcDistributionChamberElementType.\n"
    )

    return out_path, summary


# ----------------------------------------------------------------------------
# Presentation layer purge
# ----------------------------------------------------------------------------

DEFAULT_ALLOWED_LAYERS_CSV = Path(__file__).resolve().parent / "config" / "presentation_layers" / "default_allowed_layers.csv"


def _normalize_layer(value: str) -> str:
    return re.sub(r"\s+", " ", (value or "").strip()).lower()


def _score_layer_similarity(existing: str, candidate: str) -> float:
    simplified_existing = re.sub(r"[^a-z0-9]", "", _normalize_layer(existing))
    simplified_candidate = re.sub(r"[^a-z0-9]", "", _normalize_layer(candidate))
    return difflib.SequenceMatcher(a=simplified_existing, b=simplified_candidate).ratio()


def _combine_layer(code: str, description: str) -> str:
    return f"{(code or '').strip()}--{(description or '').strip()}"


def parse_allowed_layers_csv_text(csv_text: str) -> Dict[str, Any]:
    rows: List[Dict[str, str]] = []
    errors: List[Dict[str, Any]] = []
    if not csv_text or not csv_text.strip():
        return {"rows": rows, "errors": errors}
    reader = csv.DictReader(csv_text.splitlines())
    required = {"layer_code", "layer_description"}
    field_names = {f.strip() for f in (reader.fieldnames or []) if f}
    if not required.issubset(field_names):
        return {
            "rows": [],
            "errors": [{"row": 0, "message": "CSV requires columns: layer_code, layer_description"}],
        }
    dedupe = set()
    for idx, row in enumerate(reader, start=2):
        code = (row.get("layer_code") or "").strip()
        description = (row.get("layer_description") or "").strip()
        if not code and not description:
            continue
        if not code or not description:
            errors.append({"row": idx, "message": "layer_code and layer_description are required"})
            continue
        combined = _combine_layer(code, description)
        dedupe_key = _normalize_layer(combined)
        if dedupe_key in dedupe:
            continue
        dedupe.add(dedupe_key)
        rows.append({"layer_code": code, "layer_description": description, "full_layer": combined})
    return {"rows": rows, "errors": errors}


def get_default_allowed_layers() -> List[Dict[str, str]]:
    parsed = parse_allowed_layers_csv_text(DEFAULT_ALLOWED_LAYERS_CSV.read_text(encoding="utf-8"))
    return parsed["rows"]


def build_allowed_layers(uploaded_csv_text: str = "", use_uploaded_only: bool = False) -> Dict[str, Any]:
    default_rows = [] if use_uploaded_only else get_default_allowed_layers()
    uploaded_parse = parse_allowed_layers_csv_text(uploaded_csv_text)
    merged: Dict[str, Dict[str, str]] = {}
    for row in default_rows + uploaded_parse["rows"]:
        merged[_normalize_layer(row["full_layer"])] = row
    merged_rows = sorted(merged.values(), key=lambda row: row["full_layer"])
    return {
        "rows": merged_rows,
        "full_values": [row["full_layer"] for row in merged_rows],
        "errors": uploaded_parse["errors"],
    }


def parse_allowed_layers(text_or_file: Optional[str]) -> set:
    if not text_or_file:
        return set()
    parsed = parse_allowed_layers_csv_text(text_or_file)
    return {row["full_layer"] for row in parsed["rows"]}


def compute_shallow_layer(layer_value: str) -> Optional[str]:
    if not layer_value:
        return None
    match = re.match(r"^([A-Z]-Ss)(\d{4})\d+", layer_value)
    if not match:
        return None
    prefix, digits = match.groups()
    return f"{prefix}{digits}0--"


def propose_layer_mapping(
    current_value: str,
    allowed_set: set,
    explicit_map: Dict[str, str],
    auto_shallow: bool,
) -> Optional[Dict[str, str]]:
    if not current_value:
        return None
    if current_value in explicit_map:
        target = explicit_map[current_value]
        return {"target": target, "reason": "Explicit", "allowed": "yes" if target in allowed_set else "no"}
    if current_value in allowed_set:
        return {"target": current_value, "reason": "Exact", "allowed": "yes", "confidence": 1.0}
    if not allowed_set:
        return {"target": "", "reason": "Manual", "allowed": "unknown", "confidence": 0.0}
    best = max(allowed_set, key=lambda a: _score_layer_similarity(current_value, a))
    confidence = _score_layer_similarity(current_value, best)
    if confidence < 0.55:
        return {"target": "", "reason": "Unmatched", "allowed": "no", "confidence": confidence}
    return {"target": best, "reason": "Suggested", "allowed": "yes", "confidence": confidence}


def extract_presentation_layers(ifc_path: str) -> List[Dict[str, Any]]:
    model = ifcopenshell.open(ifc_path)
    counts: Dict[str, int] = {}
    layer_ids: Dict[str, List[int]] = {}
    for layer in model.by_type("IfcPresentationLayerAssignment"):
        name = (getattr(layer, "Name", "") or "").strip()
        if not name:
            continue
        counts[name] = counts.get(name, 0) + len(getattr(layer, "AssignedItems", []) or [])
        layer_ids.setdefault(name, []).append(layer.id())
    return [
        {"existing_layer": name, "count": counts[name], "layer_ids": layer_ids.get(name, [])}
        for name in sorted(counts.keys())
    ]


def extract_uniclass_ss_classifications(ifc_path: str) -> List[Dict[str, Any]]:
    model = ifcopenshell.open(ifc_path)
    counts: Dict[str, int] = {}

    def classification_parts(classification_ref) -> Tuple[str, str, str]:
        code = (getattr(classification_ref, "Identification", "") or getattr(classification_ref, "ItemReference", "") or "").strip()
        name = (getattr(classification_ref, "Name", "") or "").strip()
        source = getattr(classification_ref, "ReferencedSource", None)
        source_name = (getattr(source, "Name", "") or "").strip()
        return code, name, source_name

    for rel in model.by_type("IfcRelAssociatesClassification"):
        classification_ref = getattr(rel, "RelatingClassification", None)
        if not classification_ref:
            continue
        code, name, source_name = classification_parts(classification_ref)
        code_lower = code.lower()
        source_lower = source_name.lower()
        looks_like_ss = bool(re.search(r"(^|[^a-z])ss[-_ ]?\d", code_lower)) or code_lower.startswith("ss")
        if not looks_like_ss and "uniclass" not in source_lower:
            continue
        existing = f"{code}--{name}" if code and name else (code or name)
        if not existing:
            continue
        related = getattr(rel, "RelatedObjects", None) or []
        counts[existing] = counts.get(existing, 0) + len(related)

    return [{"existing_layer": key, "count": counts[key], "layer_ids": []} for key in sorted(counts.keys())]


def extract_unassigned_elements_by_ifcclass(ifc_path: str) -> List[Dict[str, Any]]:
    model = ifcopenshell.open(ifc_path)
    grouped: Dict[str, Dict[str, Any]] = {}

    for element in model.by_type("IfcProduct"):
        global_id = getattr(element, "GlobalId", None)
        if not global_id:
            continue

        presentation_layers = [
            layer
            for layer in ifcopenshell.util.element.get_layers(model, element)
            if (getattr(layer, "Name", "") or "").strip()
        ]
        if presentation_layers:
            continue

        has_layer_property = any((prop.get("value") or "").strip() for prop in find_layer_properties(element))
        if has_layer_property:
            continue

        has_classification = any(rel.is_a("IfcRelAssociatesClassification") for rel in (getattr(element, "HasAssociations", None) or []))
        if has_classification:
            continue

        ifc_class = element.is_a()
        bucket = grouped.setdefault(
            ifc_class,
            {
                "existing_layer": f"{ifc_class} (unassigned)",
                "count": 0,
                "layer_ids": [],
                "ifc_class": ifc_class,
                "globalids": [],
            },
        )
        bucket["count"] += 1
        bucket["globalids"].append(global_id)

    return [grouped[key] for key in sorted(grouped.keys())]


def build_layer_review(ifc_path: str, allowed_values: List[str], confidence_threshold: float = 0.7) -> Dict[str, Any]:
    allowed_set = set(allowed_values)
    rows = []

    extracted = extract_presentation_layers(ifc_path)
    source_mode = "presentation_layers"
    if not extracted:
        extracted = extract_uniclass_ss_classifications(ifc_path)
        source_mode = "uniclass_ss_fallback" if extracted else "none"

    for item in extracted:
        mapping = propose_layer_mapping(item["existing_layer"], allowed_set, {}, True) or {}
        suggestion = mapping.get("target", "")
        confidence = float(mapping.get("confidence", 1.0 if mapping.get("reason") == "Exact" else 0.0))
        exact = item["existing_layer"] in allowed_set
        if source_mode == "uniclass_ss_fallback":
            status = "classification_candidate"
            final_value = suggestion if confidence >= confidence_threshold else ""
        else:
            status = "exact" if exact else ("suggested" if suggestion else "unmatched")
            final_value = suggestion if (exact or confidence >= confidence_threshold) else ""
        rows.append(
            {
                "existing_layer": item["existing_layer"],
                "normalized_existing_layer": re.sub(r"[^a-z0-9]", "", _normalize_layer(item["existing_layer"])),
                "count": item["count"],
                "layer_ids": item["layer_ids"],
                "exact_match": exact,
                "suggested_layer": suggestion,
                "suggested_confidence": round(confidence, 4),
                "final_layer": final_value,
                "status": status,
                "source_type": source_mode,
                "apply_change": bool(final_value and final_value != item["existing_layer"]),
            }
        )

    unassigned = extract_unassigned_elements_by_ifcclass(ifc_path)
    for item in unassigned:
        rows.append(
            {
                "existing_layer": item["existing_layer"],
                "normalized_existing_layer": re.sub(r"[^a-z0-9]", "", _normalize_layer(item["existing_layer"])),
                "count": item["count"],
                "layer_ids": [],
                "exact_match": False,
                "suggested_layer": "",
                "suggested_confidence": 0.0,
                "final_layer": "",
                "status": "unassigned_ifcclass",
                "source_type": "unassigned_ifcclass",
                "ifc_class": item["ifc_class"],
                "globalids": item["globalids"],
                "apply_change": False,
            }
        )

    if source_mode == "none" and unassigned:
        source_mode = "ifcclass_unassigned_fallback"

    summary = {
        "layers_found": len(rows),
        "exact_matches": sum(1 for r in rows if r["status"] == "exact"),
        "suggested": sum(1 for r in rows if r["status"] == "suggested"),
        "unmatched": sum(1 for r in rows if r["status"] == "unmatched"),
        "classification_candidates": sum(1 for r in rows if r["status"] == "classification_candidate"),
        "unassigned_ifcclass": sum(1 for r in rows if r["status"] == "unassigned_ifcclass"),
        "source_mode": source_mode,
    }
    return {"rows": rows, "summary": summary}


def _extract_property_value(prop: ifcopenshell.entity_instance) -> Optional[str]:
    if not prop or not prop.is_a("IfcPropertySingleValue"):
        return None
    return _extract_nominal_value(prop)


def find_layer_properties(element: ifcopenshell.entity_instance) -> List[Dict[str, Any]]:
    layer_props = []
    ifc_file = element.file

    def collect_from_definition(definition, source: str):
        if not definition or not definition.is_a("IfcPropertySet"):
            return
        for prop in definition.HasProperties or []:
            if not prop.is_a("IfcPropertySingleValue"):
                continue
            name = getattr(prop, "Name", "") or ""
            if name.lower() != "layer":
                continue
            value = _extract_property_value(prop)
            layer_props.append({"id": prop.id(), "value": value, "source": source})

    for rel in getattr(element, "IsDefinedBy", None) or []:
        if rel.is_a("IfcRelDefinesByProperties"):
            collect_from_definition(rel.RelatingPropertyDefinition, "occurrence")

    element_type = ifcopenshell.util.element.get_type(element)
    if element_type:
        for definition in getattr(element_type, "HasPropertySets", None) or []:
            collect_from_definition(definition, "type")

    return layer_props


def scan_layers(
    ifc_path: str,
    allowed_set: set,
    explicit_map: Dict[str, str],
    options: Dict[str, Any],
) -> Tuple[Dict[str, Any], List[Dict[str, Any]]]:
    model = ifcopenshell.open(ifc_path)
    auto_shallow = bool(options.get("auto_shallow", True))
    rows = []
    elements = [e for e in model.by_type("IfcProduct") if getattr(e, "GlobalId", None)]
    for element in elements:
        presentation_layers = []
        for layer in ifcopenshell.util.element.get_layers(model, element):
            name = getattr(layer, "Name", None)
            if name:
                presentation_layers.append({"id": layer.id(), "name": name})
        property_layers = find_layer_properties(element)
        property_values = [p["value"] for p in property_layers if p.get("value")]
        property_display = "; ".join([v for v in property_values if v])
        if presentation_layers:
            for layer_info in presentation_layers:
                mapping = propose_layer_mapping(layer_info["name"], allowed_set, explicit_map, auto_shallow)
                if not mapping:
                    continue
                row = {
                    "row_id": uuid.uuid4().hex,
                    "globalid": element.GlobalId,
                    "ifc_class": element.is_a(),
                    "source": "presentation",
                    "presentation_layer": layer_info["name"],
                    "property_layer": property_display,
                    "presentation_layer_id": layer_info["id"],
                    "property_layer_id": None,
                    "presentation_layers": presentation_layers,
                    "property_layers": property_layers,
                    "target_layer": mapping["target"],
                    "mapping_reason": mapping["reason"],
                    "allowed_status": mapping["allowed"],
                    "apply_default": mapping["allowed"] == "yes",
                }
                rows.append(row)
        else:
            for prop_info in property_layers:
                mapping = propose_layer_mapping(prop_info.get("value"), allowed_set, explicit_map, auto_shallow)
                if not mapping:
                    continue
                row = {
                    "row_id": uuid.uuid4().hex,
                    "globalid": element.GlobalId,
                    "ifc_class": element.is_a(),
                    "source": "property",
                    "presentation_layer": "",
                    "property_layer": prop_info.get("value") or "",
                    "presentation_layer_id": None,
                    "property_layer_id": prop_info.get("id"),
                    "presentation_layers": presentation_layers,
                    "property_layers": property_layers,
                    "target_layer": mapping["target"],
                    "mapping_reason": mapping["reason"],
                    "allowed_status": mapping["allowed"],
                    "apply_default": mapping["allowed"] == "yes",
                }
                rows.append(row)

    stats = {
        "schema": model.schema,
        "elements": len(elements),
        "presentation_layers": sum(len(ifcopenshell.util.element.get_layers(model, e)) for e in elements),
        "property_layers": sum(len(find_layer_properties(e)) for e in elements),
        "rows": len(rows),
    }
    return stats, rows


def _update_property_value(model: ifcopenshell.file, prop: ifcopenshell.entity_instance, new_value: str) -> None:
    nominal = getattr(prop, "NominalValue", None)
    value_type = "IfcLabel"
    if nominal is not None and hasattr(nominal, "is_a"):
        value_type = nominal.is_a()
    prop.NominalValue = model.create_entity(value_type, new_value)


def apply_layer_changes(
    ifc_path: str,
    rows_to_apply: List[Dict[str, Any]],
    options: Dict[str, Any],
) -> Tuple[str, str, str, Dict[str, Any]]:
    model = ifcopenshell.open(ifc_path)
    update_both = bool(options.get("update_both", False))
    remove_orphans = bool(options.get("remove_orphans", True))
    updated_layers = set()
    change_log = []
    now = utc_now().isoformat() + "Z"
    mapping = {}
    for row in rows_to_apply:
        existing = (row.get("existing_layer") or "").strip()
        target = (row.get("final_layer") or "").strip()
        applied = bool(row.get("apply_change", False))
        if not existing:
            continue
        mapping[existing] = {"target": target, "applied": applied, "suggested_layer": (row.get("suggested_layer") or "").strip()}

    def _assign_items_to_layer(target_name: str, items: List[Any]) -> int:
        if not target_name or not items:
            return 0
        target_layer = None
        for layer in model.by_type("IfcPresentationLayerAssignment"):
            if (getattr(layer, "Name", "") or "").strip() == target_name:
                target_layer = layer
                break
        if target_layer is None:
            target_layer = model.create_entity("IfcPresentationLayerAssignment", Name=target_name, AssignedItems=[])
        existing_items = list(getattr(target_layer, "AssignedItems", []) or [])
        existing_ids = {item.id() for item in existing_items if hasattr(item, "id")}
        added = 0
        for item in items:
            if not hasattr(item, "id"):
                continue
            item_id = item.id()
            if item_id in existing_ids:
                continue
            existing_items.append(item)
            existing_ids.add(item_id)
            added += 1
        if added:
            target_layer.AssignedItems = existing_items
        return added

    for row in rows_to_apply:
        if row.get("source_type") != "unassigned_ifcclass":
            continue
        target = (row.get("final_layer") or "").strip()
        if not target or not bool(row.get("apply_change", False)):
            continue
        globalids = [gid for gid in (row.get("globalids") or []) if gid]
        representation_items: List[Any] = []
        for globalid in globalids:
            try:
                element = model.by_guid(globalid)
            except Exception:
                element = None
            if element is None:
                continue
            representation = getattr(element, "Representation", None)
            for shape in getattr(representation, "Representations", None) or []:
                representation_items.extend(getattr(shape, "Items", None) or [])
        added = _assign_items_to_layer(target, representation_items)
        if added:
            change_log.append({
                "existing_layer": row.get("existing_layer", ""),
                "suggested_layer": row.get("suggested_layer", ""),
                "final_layer": target,
                "applied": True,
                "status": "assigned_from_ifcclass",
                "timestamp": now,
            })

    for layer in model.by_type("IfcPresentationLayerAssignment"):
        old_value = (getattr(layer, "Name", "") or "").strip()
        map_row = mapping.get(old_value)
        if not map_row or not map_row["applied"] or not map_row["target"]:
            continue
        target = map_row["target"]
        if old_value == target:
            continue
        layer.Name = target
        updated_layers.add(layer.id())
        change_log.append({
            "existing_layer": old_value,
            "suggested_layer": map_row.get("suggested_layer", ""),
            "final_layer": target,
            "applied": True,
            "status": "changed",
            "timestamp": now,
        })

    if update_both:
        for prop in model.by_type("IfcPropertySingleValue"):
            if (getattr(prop, "Name", "") or "").lower() != "layer":
                continue
            old_value = (_extract_property_value(prop) or "").strip()
            map_row = mapping.get(old_value)
            if not map_row or not map_row["applied"] or not map_row["target"] or old_value == map_row["target"]:
                continue
            _update_property_value(model, prop, map_row["target"])

    if remove_orphans:
        for layer in model.by_type("IfcPresentationLayerAssignment"):
            if not (getattr(layer, "AssignedItems", []) or []):
                model.remove(layer)

    base_dir = os.path.dirname(ifc_path)
    ts = utc_now().strftime("%Y%m%d_%H%M%S")
    base = os.path.splitext(os.path.basename(ifc_path))[0]
    out_path = os.path.join(base_dir, f"{base}_layer_purged_{ts}.ifc")
    model.write(out_path)

    json_path = os.path.join(base_dir, f"{base}_layer_purge_log_{ts}.json")
    csv_path = os.path.join(base_dir, f"{base}_layer_purge_log_{ts}.csv")
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(change_log, f, indent=2)
    with open(csv_path, "w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=["existing_layer", "suggested_layer", "final_layer", "applied", "status", "timestamp"],
        )
        writer.writeheader()
        writer.writerows(change_log)

    summary = {
        "layers_found": len(mapping),
        "changed": len(change_log),
        "unchanged": max(0, len(mapping) - len(change_log)),
        "unmatched": sum(1 for row in rows_to_apply if not (row.get("final_layer") or "").strip()),
    }

    return out_path, json_path, csv_path, summary


def match_type_name_for_proxy(type_name: str) -> bool:
    if not type_name:
        return False
    return bool(resolve_type_and_predefined_for_name(type_name, "IFC4").get("resolved_type_class"))


def list_instance_classes(ifc_path: str) -> List[str]:
    model = ifcopenshell.open(ifc_path)
    classes = sorted({e.is_a() for e in model.by_type("IfcObject") if getattr(e, "GlobalId", None)})
    return classes


def scan_predefined_types(
    ifc_path: str,
    class_filter: Optional[List[str]] = None,
) -> Tuple[Dict[str, Any], List[Dict[str, Any]]]:
    model = ifcopenshell.open(ifc_path)
    schema_name = model.schema
    class_set = {c for c in (class_filter or []) if c}
    elements = [e for e in model.by_type("IfcObject") if getattr(e, "GlobalId", None)]
    if class_set:
        elements = [e for e in elements if e.is_a() in class_set]
    rows = []
    for element in elements:
        element_type = ifcopenshell.util.element.get_type(element)
        type_name = getattr(element_type, "Name", "") if element_type else ""
        resolved = resolve_type_and_predefined_for_name(type_name, schema_name)
        resolved_type_class = resolved.get("resolved_type_class")
        parsed_predef_token = resolved.get("parsed_predef", "")

        predef_target = None
        predef_target_source = "none"
        predef_target_info = {"has_predefined": False, "enum_name": None, "enum_items": []}
        alt_target = None
        alt_target_info = {"has_predefined": False, "enum_name": None, "enum_items": []}

        if element_type is not None:
            type_info = _predefined_type_info(schema_name, element_type.is_a())
            if type_info["has_predefined"]:
                predef_target = element_type
                predef_target_source = "type"
                predef_target_info = type_info
            occ_info = _predefined_type_info(schema_name, element.is_a())
            if occ_info["has_predefined"]:
                alt_target = element
                alt_target_info = occ_info

        if predef_target is None:
            occ_info = _predefined_type_info(schema_name, element.is_a())
            if occ_info["has_predefined"]:
                predef_target = element
                predef_target_source = "occurrence"
                predef_target_info = occ_info
            if element_type is not None:
                type_info = _predefined_type_info(schema_name, element_type.is_a())
                if type_info["has_predefined"]:
                    alt_target = element_type
                    alt_target_info = type_info

        match_source = "none"
        matched_pset_name = ""
        predef_reason = ""
        proposed = "USERDEFINED"

        if not resolved_type_class:
            predef_reason = "class resolution failed"
        else:
            enum_lookup = build_entity_predefined_enum_library(schema_name)
            normalized_token = normalize_token(parsed_predef_token)

            if predef_target is not None and normalized_token:
                key = (predef_target.is_a(), "PredefinedType", normalized_token)
                enum_val = enum_lookup.get(key)
                if enum_val:
                    proposed = enum_val
                    match_source = "enum"
                    predef_reason = "enum matched on primary target"

            if match_source == "none" and alt_target is not None and normalized_token:
                key = (alt_target.is_a(), "PredefinedType", normalized_token)
                enum_val = enum_lookup.get(key)
                if enum_val:
                    proposed = enum_val
                    match_source = "enum_alt_target"
                    predef_reason = "enum matched on alternate target"

            if match_source == "none" and schema_name.upper() == "IFC2X3":
                pset_match = resolve_pset_applicability(resolved_type_class, parsed_predef_token)
                if pset_match:
                    proposed = pset_match["applicable_type_value"]
                    matched_pset_name = pset_match["pset_name"]
                    match_source = "pset_applicability"
                    predef_reason = "matched IFC2X3 Pset applicability"

            if match_source == "none" and predef_target is not None and predef_target_info.get("enum_items"):
                if "USERDEFINED" in predef_target_info.get("enum_items", []):
                    proposed = "USERDEFINED"
                    match_source = "enum_fallback_userdefined"
                    predef_reason = "no enum match → USERDEFINED"
                else:
                    predef_reason = "no enum match"
            elif match_source == "none":
                predef_reason = "no enum or IFC2X3 pset applicability match"

        match_found = match_source != "none"

        rows.append(
            {
                "row_id": uuid.uuid4().hex,
                "globalid": element.GlobalId,
                "ifc_class": element.is_a(),
                "type_name": type_name or "",
                "match_found": match_found,
                "proposed_predefined_type": proposed,
                "apply_default": predef_target is not None and proposed not in ("", "N/A"),
                "predef_target_source": predef_target_source,
                "predef_target_globalid": getattr(predef_target, "GlobalId", None) if predef_target else None,
                "predef_target_id": int(predef_target.id()) if predef_target else None,
                "predef_target_class": predef_target.is_a() if predef_target else None,
                "parsed_class": resolved.get("parsed_classish", ""),
                "resolved_type_class": resolved_type_class,
                "parsed_predef_token": parsed_predef_token,
                "resolved_predefined_type": proposed,
                "target_source": predef_target_source,
                "target_globalid": getattr(predef_target, "GlobalId", None) if predef_target else None,
                "target_ifc_id": int(predef_target.id()) if predef_target else None,
                "target_class": predef_target.is_a() if predef_target else None,
                "predef_supported": bool(predef_target_info.get("has_predefined")),
                "predef_reason": predef_reason,
                "match_source": match_source,
                "matched_pset_name": matched_pset_name,
                "schema": schema_name,
            }
        )
    stats = {"schema": model.schema, "elements": len(elements), "rows": len(rows)}
    return stats, rows


def apply_predefined_type_changes(
    ifc_path: str,
    rows_to_apply: List[Dict[str, Any]],
) -> Tuple[str, str, str]:
    model = ifcopenshell.open(ifc_path)
    change_log = []
    now = utc_now().isoformat() + "Z"
    updated = 0

    def _ensure_pset_on_entity(entity, pset_name: str) -> bool:
        if not entity or not pset_name:
            return False
        try:
            psets = ifcopenshell.util.element.get_psets(entity, psets_only=True, include_inherited=False) or {}
        except TypeError:
            psets = ifcopenshell.util.element.get_psets(entity, psets_only=True) or {}
        if pset_name in psets:
            return False
        owner_history = next(iter(model.by_type("IfcOwnerHistory")), None)
        if owner_history is None:
            return False
        pset = model.create_entity(
            "IfcPropertySet",
            GlobalId=new_guid(),
            OwnerHistory=owner_history,
            Name=pset_name,
            Description=None,
            HasProperties=[],
        )
        if hasattr(entity, "HasPropertySets"):
            existing = list(getattr(entity, "HasPropertySets", []) or [])
            existing.append(pset)
            entity.HasPropertySets = existing
        else:
            model.create_entity(
                "IfcRelDefinesByProperties",
                GlobalId=new_guid(),
                OwnerHistory=owner_history,
                Name=None,
                Description=None,
                RelatedObjects=[entity],
                RelatingPropertyDefinition=pset,
            )
        return True

    for row in rows_to_apply:
        target = row.get("proposed_predefined_type")
        if target in (None, ""):
            continue
        if row.get("match_source") == "pset_applicability":
            type_entity = None
            type_gid = row.get("target_globalid")
            if type_gid:
                type_entity = model.by_guid(type_gid)
            if type_entity is None:
                element = model.by_guid(row.get("globalid")) if row.get("globalid") else None
                if element is not None:
                    type_entity = ifcopenshell.util.element.get_type(element)
            if type_entity is not None:
                pset_name = row.get("matched_pset_name") or ""
                added = _ensure_pset_on_entity(type_entity, pset_name)
                change_log.append(
                    {
                        "globalid": row.get("globalid"),
                        "ifc_class": row.get("ifc_class"),
                        "target": "pset_applicability",
                        "target_source": "type",
                        "target_globalid": getattr(type_entity, "GlobalId", None),
                        "target_ifc_id": int(type_entity.id()),
                        "old_value": "",
                        "new_value": target,
                        "mapping_reason": row.get("predef_reason") or "matched IFC2X3 Pset applicability",
                        "target_class": type_entity.is_a(),
                        "applied_pset": pset_name,
                        "applied_mode": "pset_only",
                        "allowed_status": "n/a",
                        "timestamp": now,
                    }
                )
                if added:
                    updated += 1
            continue

        target_entity = None
        target_gid = row.get("target_globalid") or row.get("predef_target_globalid")
        if target_gid:
            target_entity = model.by_guid(target_gid)

        if target_entity is None:
            target_id = row.get("target_ifc_id") or row.get("predef_target_id")
            if target_id is not None:
                try:
                    candidate = model.by_id(int(target_id))
                except Exception:
                    candidate = None
                if candidate is not None and hasattr(candidate, "PredefinedType"):
                    target_entity = candidate

        if not target_entity or not hasattr(target_entity, "PredefinedType"):
            continue

        old_value = getattr(target_entity, "PredefinedType", None)
        if old_value == target:
            continue

        target_entity.PredefinedType = target
        updated += 1
        change_log.append(
            {
                "globalid": row.get("globalid"),
                "ifc_class": row.get("ifc_class"),
                "target": f"predefined_type:{row.get('target_source') or row.get('predef_target_source', 'none')}",
                "target_source": row.get("target_source") or row.get("predef_target_source", "none"),
                "target_globalid": row.get("target_globalid") or row.get("predef_target_globalid"),
                "target_ifc_id": row.get("target_ifc_id") or row.get("predef_target_id"),
                "old_value": old_value,
                "new_value": target,
                "mapping_reason": row.get("predef_reason") or ("Type name match" if row.get("match_found") else "No match"),
                "target_class": row.get("target_class") or row.get("predef_target_class") or (target_entity.is_a() if target_entity else None),
                "applied_pset": "",
                "applied_mode": "predefined_type",
                "allowed_status": "n/a",
                "timestamp": now,
            }
        )

    base_dir = os.path.dirname(ifc_path)
    ts = utc_now().strftime("%Y%m%d_%H%M%S")
    base = os.path.splitext(os.path.basename(ifc_path))[0]
    out_path = os.path.join(base_dir, f"{base}_predefined_{ts}.ifc")
    model.write(out_path)

    json_path = os.path.join(base_dir, f"{base}_predefined_log_{ts}.json")
    csv_path = os.path.join(base_dir, f"{base}_predefined_log_{ts}.csv")
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(change_log, f, indent=2)
    with open(csv_path, "w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=[
                "globalid",
                "ifc_class",
                "target",
                "target_source",
                "target_globalid",
                "target_ifc_id",
                "old_value",
                "new_value",
                "mapping_reason",
                "target_class",
                "applied_pset",
                "applied_mode",
                "allowed_status",
                "timestamp",
            ],
        )
        writer.writeheader()
        writer.writerows(change_log)

    return out_path, json_path, csv_path


# ----------------------------------------------------------------------------
# Model checking helpers
# ----------------------------------------------------------------------------

CHECK_CACHE: Dict[str, Any] = {"definitions": None, "summary": {}}


def _expression_lookup() -> Dict[str, str]:
    expr_cfg = load_expression_config()
    merged: Dict[str, str] = {}
    merged.update(expr_cfg.get("by_classification_system", {}))
    merged.update(expr_cfg.get("by_check_id", {}))
    return merged


def load_definitions() -> List[Any]:
    mapping_cfg = load_mapping_config()
    definitions = load_check_definitions(mapping_cfg, _expression_lookup())
    CHECK_CACHE["definitions"] = definitions
    CHECK_CACHE["summary"] = summarize_sections(definitions)
    return definitions


def _ensure_definitions() -> List[Any]:
    return CHECK_CACHE["definitions"] or load_definitions()


def serialize_field_descriptor(fd: FieldDescriptor) -> Dict[str, Any]:
    return {
        "kind": fd.kind.value,
        "attribute": fd.attribute_name,
        "property": fd.property_name,
        "pset": fd.pset_name,
        "quantity": fd.quantity_name,
        "qto": fd.qto_name,
        "classification_system": fd.classification_system,
        "expression": fd.expression,
    }


def serialize_definition(defn) -> Dict[str, Any]:
    return {
        "check_id": defn.check_id,
        "description": defn.description,
        "entity_scope": defn.entity_scope,
        "info": defn.info_to_check,
        "applicable_models": defn.applicable_models,
        "milestones": defn.milestones,
        "section": defn.section,
        "mapping_status": defn.mapping_status,
        "field": serialize_field_descriptor(defn.field) if defn.field else None,
    }


def _to_serializable(val: Any):
    if isinstance(val, (int, float, str)) or val is None:
        return val
    try:
        return float(val)
    except Exception:
        try:
            return str(val)
        except Exception:
            return None


def _filter_defs(defs: List[Any], section: Optional[str], riba_stage: Optional[str]) -> List[Any]:
    filtered = []
    for d in defs:
        if section and d.section != section:
            continue
        if riba_stage and d.milestones and riba_stage not in d.milestones:
            continue
        filtered.append(d)
    return filtered


def _collect_targets(model, defs: List[Any], entity_filter: Optional[str], entity_filters: Optional[List[str]]) -> List[Any]:
    targets = []
    entity_names = set()
    filters = entity_filters or []
    if entity_filter:
        filters.append(entity_filter)
    if filters:
        entity_names.update(filters)
    else:
        for d in defs:
            for ent in d.entity_scope:
                entity_names.add(ent)
    for ent in sorted(entity_names):
        try:
            for obj in model.by_type(ent):
                targets.append(obj)
        except Exception:
            continue
    return targets


def _row_id(obj) -> str:
    gid = getattr(obj, "GlobalId", None) or getattr(obj, "GlobalID", None)
    return f"{gid or obj.id()}"


def build_table_data(model, section: str, riba_stage: Optional[str], entity_filter: Optional[str], entity_filters: Optional[List[str]]) -> Dict[str, Any]:
    defs = _filter_defs(_ensure_definitions(), section, riba_stage)
    targets = _collect_targets(model, defs, entity_filter, entity_filters)
    expr_engine = ExpressionEngine(model)
    columns = [serialize_definition(d) for d in defs]
    rows = []
    for obj in targets:
        row_values: Dict[str, Dict[str, Any]] = {}
        issues_count = 0
        for d in defs:
            if d.field is None:
                continue
            val = get_value(obj, d.field)
            generated = expr_engine.evaluate(d.field.expression, obj) if d.field.expression else None
            validation = validate_value(model, obj, d.field, val, check_id=d.check_id)
            issues_count += len(validation)
            row_values[d.check_id] = {
                "value": _to_serializable(val),
                "generated": generated,
                "issues": [vars(v) for v in validation],
                "descriptor": serialize_field_descriptor(d.field),
            }
        rows.append(
            {
                "id": obj.id(),
                "global_id": getattr(obj, "GlobalId", None),
                "name": getattr(obj, "Name", None) or getattr(obj, "LongName", None),
                "type": obj.is_a(),
                "issues": issues_count,
                "values": row_values,
            }
        )
    return {
        "columns": columns,
        "rows": rows,
        "summary": {
            "rows": len(rows),
            "columns": len(columns),
            "issues": sum(r["issues"] for r in rows),
        },
    }


def _change_log_path(session_id: str) -> Path:
    return Path(SESSION_STORE.session_path(session_id)) / "change_log.json"


def read_change_log(session_id: str) -> List[Dict[str, Any]]:
    path = _change_log_path(session_id)
    if path.exists():
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)
    return []


def append_change_log(session_id: str, entries: List[Dict[str, Any]]):
    path = _change_log_path(session_id)
    path.parent.mkdir(parents=True, exist_ok=True)
    existing = read_change_log(session_id)
    existing.extend(entries)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(existing, f, indent=2)


def update_step2ifc_job(job_id: str, **updates: Any) -> None:
    job = STEP2IFC_JOBS.get(job_id)
    if not job:
        return
    job.update(updates)


def run_step2ifc_auto_job(job_id: str, session_id: str, input_path: Path, output_path: Path) -> None:
    update_step2ifc_job(job_id, status="running", progress=5, message="Starting auto conversion")

    def progress_cb(percent: int, message: str) -> None:
        update_step2ifc_job(job_id, progress=percent, message=message)

    if not STEP2IFC_AVAILABLE:
        update_step2ifc_job(
            job_id,
            status="error",
            progress=100,
            message=STEP2IFC_IMPORT_ERROR or "step2ifc dependencies unavailable",
            error=True,
            done=True,
        )
        APP_LOGGER.error("STEP2IFC auto conversion unavailable: %s", STEP2IFC_IMPORT_ERROR)
        return

    if STEP2IFC_RUN_CONVERT is None:
        update_step2ifc_job(
            job_id,
            status="error",
            progress=100,
            message="step2ifc converter unavailable",
            error=True,
            done=True,
        )
        APP_LOGGER.error("STEP2IFC converter unavailable")
        return

    mapping_path = STEP2IFC_JOBS.get(job_id, {}).get("mapping_path")
    if mapping_path:
        update_step2ifc_job(job_id, status="running", progress=15, message="Running mapped conversion")
        try:
            STEP2IFC_RUN_CONVERT(
                argparse.Namespace(
                    input_path=str(input_path),
                    output_path=str(output_path),
                    schema="IFC4",
                    units="mm",
                    project="Project",
                    site="Site",
                    building="Building",
                    storey="Storey",
                    geom="brep",
                    mesh_deflection=0.5,
                    mesh_angle=0.5,
                    merge_by_name=False,
                    split_by_assembly=False,
                    default_type="IfcBuildingElementProxy",
                    class_map=mapping_path,
                    log_path=None,
                )
            )
        except Exception as exc:  # pragma: no cover - background task guard
            APP_LOGGER.exception("STEP2IFC mapped conversion failed")
            update_step2ifc_job(
                job_id,
                status="error",
                progress=100,
                message=str(exc),
                error=True,
                done=True,
            )
            return
        update_step2ifc_job(job_id, progress=85, message="Mapped conversion complete; gathering outputs")
    else:
        try:
            auto_convert(input_path, output_path, progress_cb=progress_cb)
        except Exception as exc:  # pragma: no cover - background task guard
            APP_LOGGER.exception("STEP2IFC auto conversion failed")
            update_step2ifc_job(
                job_id,
                status="error",
                progress=100,
                message=str(exc),
                error=True,
                done=True,
            )
            return

    outputs = []
    for path in [
        output_path,
        output_path.with_suffix(".qc.json"),
        output_path.with_suffix(".qc.txt"),
        output_path.with_suffix(".log.jsonl"),
        output_path.parent / "classmap.autogen.yaml",
    ]:
        if path.exists():
            outputs.append({"name": path.name, "url": f"/api/session/{session_id}/download?name={path.name}"})

    update_step2ifc_job(
        job_id,
        status="done",
        progress=100,
        message="Auto conversion complete",
        done=True,
        error=False,
        outputs=outputs,
    )


def update_data_extract_job(job_id: str, **updates: Any) -> None:
    job = DATA_EXTRACT_JOBS.get(job_id)
    if not job:
        return
    job.update(updates)


def _write_csv_rows(path: Path, header: List[str], rows: List[List[Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", newline="", encoding="utf-8") as handle:
        writer = csv.writer(handle)
        writer.writerow(header)
        for row in rows:
            writer.writerow([_clean_value(v) for v in row])


def _write_model_table(model: ifcopenshell.file, path: Path, source_file: str) -> None:
    header = ["Source_File", "IFC_Schema", "IFC_Description", "Entity_Count"]
    rows = [
        [
            source_file,
            getattr(model, "schema", ""),
            _clean_value(getattr(model, "description", "")),
            len(model),
        ]
    ]
    _write_csv_rows(path, header, rows)


def _format_child_entities(entity: Any) -> str:
    try:
        children = ifcopenshell.util.element.get_decomposition(entity) or []
    except Exception:
        children = []
    parts = []
    for child in children:
        name = getattr(child, "Name", "") or getattr(child, "LongName", "") or ""
        parts.append(f"{_line_ref(child)}{child.is_a()}({name})")
    return "; ".join(parts)


def _write_project_table(model: ifcopenshell.file, path: Path, source_file: str) -> None:
    header = [
        "IFC_Line_Ref",
        "Source_File",
        "IFC_Name",
        "IFC_GlobalId",
        "IFC_Entity",
        "IFC_Description",
        "IFC_LongName",
        "Child_Entities",
    ]
    rows: List[List[Any]] = []
    for ifc_type in ["IfcProject", "IfcSite", "IfcBuilding", "IfcBuildingStorey"]:
        for entity in model.by_type(ifc_type):
            rows.append(
                [
                    _line_ref(entity),
                    source_file,
                    getattr(entity, "Name", "") or "",
                    getattr(entity, "GlobalId", "") or "",
                    entity.is_a(),
                    getattr(entity, "Description", "") or "",
                    getattr(entity, "LongName", "") or "",
                    _format_child_entities(entity),
                ]
            )
    _write_csv_rows(path, header, rows)


def _write_object_table(
    objects: List[Any],
    path: Path,
    source_file: str,
    regexes: Dict[str, str],
) -> Dict[str, Any]:
    header = [
        "IFC_Line_Ref",
        "Source_File",
        "IFC_Name",
        "IFC_GlobalId",
        "IFC_Entity",
        "IFC_Description",
        "IFC_ObjectType",
        "IFC_Type",
        "IFC_Type_Name",
        "IFC_Predefined_Type",
        "IFC_Layer",
        "IFC_Tag",
        "IFC_Name_Syntax_Check",
        "IFC_Name_Short_Code",
        "IFC_Type_Syntax_Check",
        "IFC_Name_Type_Code",
        "IFC_Layer_Syntax_Check",
        "IFC_Name_Duplicate",
        "IFC_LongName",
        "IFC_Type_Line_Ref",
        "IFC_Type_Description",
        "IFC_TypeId",
        "Coordinates_xyz",
    ]
    name_counts: Dict[str, int] = {}
    for obj in objects:
        name = getattr(obj, "Name", "") or ""
        if name:
            name_counts[name] = name_counts.get(name, 0) + 1
    rows: List[List[Any]] = []
    for obj in objects:
        name = getattr(obj, "Name", "") or ""
        type_obj = ifcopenshell.util.element.get_type(obj)
        type_name = getattr(type_obj, "Name", "") if type_obj else ""
        layer_name = _get_layers_name(obj)
        rows.append(
            [
                _line_ref(obj),
                source_file,
                name,
                getattr(obj, "GlobalId", "") or "",
                obj.is_a(),
                getattr(obj, "Description", "") or "",
                getattr(obj, "ObjectType", "") or "",
                type_obj.is_a() if type_obj else "",
                type_name or "",
                getattr(obj, "PredefinedType", "") or "",
                layer_name,
                getattr(obj, "Tag", "") or "",
                _regex_check(regexes.get("regex_ifc_name", ""), name),
                _regex_extract(regexes.get("regex_ifc_name_code", ""), name),
                _regex_check(regexes.get("regex_ifc_type", ""), type_name),
                _regex_extract(regexes.get("regex_ifc_type_code", ""), type_name),
                _regex_check(regexes.get("regex_ifc_layer", ""), layer_name),
                "True" if name and name_counts.get(name, 0) > 1 else "False",
                getattr(obj, "LongName", "") or "",
                _line_ref(type_obj),
                getattr(type_obj, "Description", "") if type_obj else "",
                getattr(type_obj, "GlobalId", "") if type_obj else "",
                _get_object_xyz(obj),
            ]
        )
    _write_csv_rows(path, header, rows)
    return {"objects": objects, "name_counts": name_counts}


def _write_classification_table(
    model: ifcopenshell.file,
    path: Path,
    source_file: str,
    include_ids: Optional[set] = None,
    objects: Optional[List[Any]] = None,
) -> None:
    header = [
        "Source_File",
        "IFC_GlobalId",
        "IFC_Entity",
        "Classification_System",
        "Classification_Name",
        "Classification_Code",
        "Classification_Description",
    ]
    rows: List[List[Any]] = []

    target_objects = objects or []
    if not target_objects:
        target_objects = [obj for obj in model.by_type("IfcProduct") if include_ids is None or obj.id() in include_ids]

    for obj in target_objects:
        if include_ids is not None and obj.id() not in include_ids:
            continue
        seen = set()
        for sys_name, name, code, desc in _iter_entity_classifications(obj):
            key = (getattr(obj, "GlobalId", "") or "", obj.is_a(), sys_name, name, code, desc)
            if key in seen:
                continue
            seen.add(key)
            rows.append([source_file, key[0], key[1], key[2], key[3], key[4], key[5]])

        type_obj = ifcopenshell.util.element.get_type(obj)
        if type_obj:
            for sys_name, name, code, desc in _iter_entity_classifications(type_obj):
                key = (getattr(type_obj, "GlobalId", "") or "", type_obj.is_a(), sys_name, name, code, desc)
                if key in seen:
                    continue
                seen.add(key)
                rows.append([source_file, key[0], key[1], key[2], key[3], key[4], key[5]])

    _write_csv_rows(path, header, rows)


def _write_system_table(
    model: ifcopenshell.file,
    path: Path,
    source_file: str,
    include_ids: Optional[set] = None,
) -> None:
    header = [
        "Source_File",
        "IFC_GlobalId",
        "IFC_Entity",
        "System_Name",
        "System_GlobalId",
        "System_Description",
    ]
    rows: List[List[Any]] = []
    for rel in model.by_type("IfcRelAssignsToGroup"):
        group = getattr(rel, "RelatingGroup", None)
        if not group or not group.is_a("IfcSystem"):
            continue
        for obj in getattr(rel, "RelatedObjects", []) or []:
            if include_ids is not None and obj.id() not in include_ids:
                continue
            rows.append(
                [
                    source_file,
                    getattr(obj, "GlobalId", "") or "",
                    obj.is_a(),
                    getattr(group, "Name", "") or "",
                    getattr(group, "GlobalId", "") or "",
                    getattr(group, "Description", "") or "",
                ]
            )
    _write_csv_rows(path, header, rows)


def _write_spatial_table(
    model: ifcopenshell.file,
    path: Path,
    source_file: str,
    objects: List[Any],
) -> None:
    header = [
        "Source_File",
        "IFC_GlobalId",
        "IFC_Entity",
        "Container_Space",
        "Container_Storey",
        "Container_Building",
        "Container_Site",
        "Container_Project",
    ]
    rows: List[List[Any]] = []
    for obj in objects:
        container = ifcopenshell.util.element.get_container(obj)
        space = storey = building = site = project = None
        current = container
        while current:
            if current.is_a("IfcSpace"):
                space = current
            elif current.is_a("IfcBuildingStorey"):
                storey = current
            elif current.is_a("IfcBuilding"):
                building = current
            elif current.is_a("IfcSite"):
                site = current
            elif current.is_a("IfcProject"):
                project = current
            current = ifcopenshell.util.element.get_container(current)
        rows.append(
            [
                source_file,
                getattr(obj, "GlobalId", "") or "",
                obj.is_a(),
                _line_ref(space),
                _line_ref(storey),
                _line_ref(building),
                _line_ref(site),
                _line_ref(project),
            ]
        )
    _write_csv_rows(path, header, rows)


def _write_pset_template_table(
    path: Path,
    source_file: str,
    template_map: Dict[str, List[str]],
    object_type_counts: Dict[str, int],
) -> None:
    header = ["Source_File", "IFC_Entity_Occurrence_Type", "Pset_Name", "Applied_Count"]
    rows: List[List[Any]] = []
    for entity_type, psets in template_map.items():
        if not psets:
            rows.append([source_file, entity_type, "", object_type_counts.get(entity_type, 0)])
        for pset in psets:
            rows.append([source_file, entity_type, pset, object_type_counts.get(entity_type, 0)])
    _write_csv_rows(path, header, rows)


def _write_property_table(
    model: ifcopenshell.file,
    path: Path,
    source_file: str,
    objects: List[Any],
    template_map: Dict[str, List[str]],
) -> None:
    header = [
        "Source_File",
        "IFC_GlobalId",
        "IFC_Entity",
        "Pset_Name",
        "Property_Name",
        "Property_Value",
        "Property_Value_Type",
        "Unit",
        "IFC_TypeId",
    ]
    rows: List[List[Any]] = []
    type_cache: Dict[int, Any] = {}
    occ_cache: Dict[int, List[Tuple[str, str, str, str]]] = {}
    type_prop_cache: Dict[int, List[Tuple[str, str, str, str]]] = {}

    for obj in objects:
        obj_id = obj.id()
        type_obj = type_cache.get(obj_id)
        if obj_id not in type_cache:
            type_obj = ifcopenshell.util.element.get_type(obj)
            type_cache[obj_id] = type_obj
        obj_type = obj.is_a()
        type_name = type_obj.is_a() if type_obj else ""

        allowed = None
        if obj_type in template_map:
            allowed = template_map[obj_type]
        elif type_name in template_map:
            allowed = template_map[type_name]
        if allowed is None:
            continue

        occ_rows = occ_cache.get(obj_id)
        if occ_rows is None:
            occ_rows = _iter_occurrence_property_rows(obj, allowed)
            occ_cache[obj_id] = occ_rows

        for pset_name, prop_name, prop_value, prop_type in occ_rows:
            rows.append([
                source_file,
                getattr(obj, "GlobalId", "") or "",
                obj.is_a(),
                pset_name,
                prop_name,
                prop_value,
                prop_type,
                "",
                getattr(type_obj, "GlobalId", "") if type_obj else "",
            ])

        if type_obj:
            type_id = type_obj.id()
            t_rows = type_prop_cache.get(type_id)
            if t_rows is None:
                t_rows = _iter_type_property_rows(type_obj, allowed)
                type_prop_cache[type_id] = t_rows
            for pset_name, prop_name, prop_value, prop_type in t_rows:
                rows.append([
                    source_file,
                    getattr(type_obj, "GlobalId", "") or "",
                    type_obj.is_a(),
                    pset_name,
                    prop_name,
                    prop_value,
                    prop_type,
                    "",
                    getattr(type_obj, "GlobalId", "") or "",
                ])

    _write_csv_rows(path, header, rows)


def run_data_extractor_job(
    job_id: str,
    session_id: str,
    ifc_files: List[str],
    exclude_filter: Optional[str],
    pset_template: Optional[str],
    tables: List[str],
    regexes: Dict[str, str],
    progress_callback: Optional[Callable[[Dict[str, Any]], None]] = None,
) -> None:
    APP_LOGGER.info(
        "data_extractor_job_start session_id=%s job_id=%s file_count=%s files=%s",
        session_id,
        job_id,
        len(ifc_files or []),
        ifc_files,
    )
    session_root = Path(SESSION_STORE.ensure(session_id))
    work_dir = Path(tempfile.mkdtemp(prefix="data_extractor_", dir=session_root))
    timeout_seconds = int(os.getenv("IFC_JOB_TIMEOUT_SECONDS", "1200"))
    started_at = time.time()
    log_lines: List[str] = []

    def emit(**payload: Any) -> None:
        update_data_extract_job(job_id, **payload)
        if progress_callback:
            progress_callback(payload)

    def log(message: str) -> None:
        log_lines.append(message)
        if len(log_lines) > 250:
            del log_lines[:-250]
        emit(logs=log_lines)

    total_tables = max(len(tables), 1)
    emit(status="running", progress=2, message="Starting extraction", logs=log_lines)

    exclude_path = Path(exclude_filter) if exclude_filter else RESOURCE_DIR / "Exclude_Filter_Template.csv"
    pset_path = Path(pset_template) if pset_template else RESOURCE_DIR / "GPA_Pset_Template.csv"
    exclude_terms = _read_csv_first_column(exclude_path)
    template_map = _load_pset_template(pset_path)

    outputs: List[Dict[str, Any]] = []
    preview_payload: Optional[Dict[str, Any]] = None

    for file_index, file_name in enumerate(ifc_files, start=1):
        if time.time() - started_at > timeout_seconds:
            raise TimeoutError(f"IFC extraction exceeded {timeout_seconds}s timeout")
        safe_name = sanitize_filename(file_name)
        input_path = session_root / safe_name
        if not input_path.exists():
            log(f"[{safe_name}] Missing IFC file.")
            continue
        base_name = os.path.splitext(safe_name)[0]
        file_dir = work_dir / base_name
        file_dir.mkdir(parents=True, exist_ok=True)
        log(f"[{safe_name}] Opening IFC...")
        try:
            model = ifcopenshell.open(str(input_path))
        except Exception as exc:
            log(f"[{safe_name}] ERROR opening IFC: {exc}")
            continue

        progress_base = int((file_index - 1) / max(len(ifc_files), 1) * 100)
        per_file_step = int(100 / max(len(ifc_files), 1))

        all_objects = _iter_object_elements(model)
        objects = [o for o in all_objects if not any(t in (getattr(o, "Name", "") or "") for t in exclude_terms)]
        include_ids = {obj.id() for obj in objects}
        object_type_counts: Dict[str, int] = {}
        for obj in objects:
            object_type_counts[obj.is_a()] = object_type_counts.get(obj.is_a(), 0) + 1

        for table_index, table_name in enumerate(tables, start=1):
            emit(
                progress=min(progress_base + int((table_index / total_tables) * per_file_step), 99),
                message=f"{safe_name}: {table_name}",
            )
            try:
                if table_name == "Model Data Table":
                    out_path = file_dir / f"IFC MODEL - {base_name}.csv"
                    _write_model_table(model, out_path, safe_name)
                elif table_name == "Project Data Table":
                    out_path = file_dir / f"IFC PROJECT - {base_name}.csv"
                    _write_project_table(model, out_path, safe_name)
                elif table_name == "Object Data Table":
                    out_path = file_dir / f"IFC OBJECT TYPE - {base_name}.csv"
                    _write_object_table(objects, out_path, safe_name, regexes)
                elif table_name == "Property Data Table":
                    out_path = file_dir / f"IFC PROPERTY - {base_name}.csv"
                    _write_property_table(model, out_path, safe_name, objects, template_map)
                elif table_name == "Classification Data Table":
                    out_path = file_dir / f"IFC CLASSIFICATION - {base_name}.csv"
                    _write_classification_table(model, out_path, safe_name, include_ids if include_ids else None, objects)
                elif table_name == "Spatial Structure Data Table":
                    out_path = file_dir / f"IFC SPATIAL STRUCTURE - {base_name}.csv"
                    _write_spatial_table(model, out_path, safe_name, objects)
                elif table_name == "System Data Table":
                    out_path = file_dir / f"IFC SYSTEM - {base_name}.csv"
                    _write_system_table(model, out_path, safe_name, include_ids if include_ids else None)
                elif table_name == "Pset Template Data Table":
                    out_path = file_dir / f"IFC PSET TEMPLATE - {base_name}.csv"
                    _write_pset_template_table(out_path, safe_name, template_map, object_type_counts)
            except Exception as exc:
                log(f"[{safe_name}] ERROR writing {table_name}: {exc}")

        log(f"[{safe_name}] Extraction complete.")

        if preview_payload is None:
            preview_candidates = [
                file_dir / f"IFC OBJECT TYPE - {base_name}.csv",
                file_dir / f"IFC PROJECT - {base_name}.csv",
                file_dir / f"IFC MODEL - {base_name}.csv",
            ]
            for candidate in preview_candidates:
                if candidate.exists():
                    df = pd.read_csv(candidate, nrows=200)
                    preview_payload = {"columns": list(df.columns), "rows": df.fillna("").values.tolist()}
                    break

    zip_name = f"ifc_data_extract_{uuid.uuid4().hex}.zip"
    zip_path = session_root / zip_name
    with zipfile.ZipFile(zip_path, "w", compression=zipfile.ZIP_DEFLATED) as zipf:
        for file_path in work_dir.rglob("*.csv"):
            arcname = file_path.relative_to(work_dir)
            zipf.write(file_path, arcname.as_posix())
    outputs.append({"name": zip_name, "url": f"/api/session/{session_id}/download?name={zip_name}"})

    emit(
        status="done",
        progress=100,
        message="Extraction complete",
        done=True,
        error=False,
        outputs=outputs,
        preview=preview_payload,
    )
    APP_LOGGER.info("data_extractor_job_end session_id=%s job_id=%s status=done outputs=%s", session_id, job_id, len(outputs))

def update_ifc_qa_job(job_id: str, **updates: Any) -> None:
    job = IFC_QA_JOBS.get(job_id)
    if not job:
        return
    job.update(updates)


def _read_csv_rows(path: Path) -> List[Dict[str, str]]:
    if not path.exists():
        return []
    with open(path, newline="", encoding="utf-8") as handle:
        reader = csv.DictReader(handle)
        return [dict(row) for row in reader]


def _evaluate_qa_rules(
    tables: Dict[str, List[Dict[str, str]]],
    rules: List[Dict[str, str]],
    property_requirements: List[Dict[str, str]],
    unacceptable_values: List[Dict[str, str]],
) -> Tuple[List[Dict[str, Any]], Dict[str, Dict[str, int]], int]:
    failures: List[Dict[str, Any]] = []
    totals_by_page: Dict[str, Dict[str, int]] = {}
    total_checks = 0

    def bump(page: str, checks: int = 0, fails: int = 0) -> None:
        totals_by_page.setdefault(page, {"checks": 0, "fails": 0})
        totals_by_page[page]["checks"] += checks
        totals_by_page[page]["fails"] += fails

    for rule in rules:
        page = (rule.get("page") or "qa_summary").strip()
        table_name = (rule.get("table") or "").strip()
        field = (rule.get("field") or "").strip()
        check_type = (rule.get("check_type") or "required").strip().lower()
        pattern = rule.get("pattern") or ""
        severity = rule.get("severity") or "medium"
        message = rule.get("message") or "Rule failed"
        rows = tables.get(table_name, [])
        for row in rows:
            total_checks += 1
            value = row.get(field, "")
            passed = True
            if check_type == "required":
                passed = bool(value)
            elif check_type == "regex":
                try:
                    passed = bool(re.search(pattern, value or ""))
                except re.error:
                    passed = False
            elif check_type == "equals":
                passed = value == pattern
            elif check_type == "not_equals":
                passed = value != pattern
            elif check_type == "contains":
                passed = pattern in (value or "")
            if not passed:
                failures.append(
                    {
                        "page": page,
                        "rule_id": rule.get("rule_id") or "",
                        "severity": severity,
                        "source_file": row.get("Source_File", ""),
                        "ifc_globalid": row.get("IFC_GlobalId", ""),
                        "table_name": table_name,
                        "field": field,
                        "actual_value": value,
                        "message": message,
                    }
                )
                bump(page, checks=1, fails=1)
            else:
                bump(page, checks=1, fails=0)

    prop_rows = tables.get("IFC PROPERTY", [])
    prop_index = {}
    for row in prop_rows:
        key = (
            row.get("IFC_GlobalId", ""),
            row.get("IFC_Entity", ""),
            row.get("Pset_Name", ""),
            row.get("Property_Name", ""),
        )
        prop_index.setdefault(key, []).append(row)
    for req in property_requirements:
        if not (req.get("required") or "").strip().lower() == "true":
            continue
        total_checks += 1
        target_entity = req.get("ifc_entity", "")
        pset_name = req.get("pset_name", "")
        prop_name = req.get("property_name", "")
        severity = req.get("severity", "medium")
        message = req.get("message", "Required property missing")
        page = "property_values"
        matched = any(
            key[1] == target_entity and key[2] == pset_name and key[3] == prop_name for key in prop_index.keys()
        )
        if not matched:
            failures.append(
                {
                    "page": page,
                    "rule_id": req.get("rule_id", ""),
                    "severity": severity,
                    "source_file": "",
                    "ifc_globalid": "",
                    "table_name": "IFC PROPERTY",
                    "field": prop_name,
                    "actual_value": "",
                    "message": message,
                }
            )
            bump(page, checks=1, fails=1)
        else:
            bump(page, checks=1, fails=0)

    for bad in unacceptable_values:
        field = (bad.get("field") or "").strip()
        bad_value = bad.get("unacceptable_value", "")
        severity = bad.get("severity", "medium")
        message = bad.get("message", "Unacceptable value")
        for table_name, rows in tables.items():
            for row in rows:
                if field not in row:
                    continue
                total_checks += 1
                value = row.get(field, "")
                if value == bad_value:
                    page = "property_values" if table_name == "IFC PROPERTY" else "occurrence_naming"
                    failures.append(
                        {
                            "page": page,
                            "rule_id": bad.get("rule_id", ""),
                            "severity": severity,
                            "source_file": row.get("Source_File", ""),
                            "ifc_globalid": row.get("IFC_GlobalId", ""),
                            "table_name": table_name,
                            "field": field,
                            "actual_value": value,
                            "message": message,
                        }
                    )
                    bump(page, checks=1, fails=1)
                else:
                    bump("qa_summary", checks=1, fails=0)

    return failures, totals_by_page, total_checks


def run_ifc_qa_job(
    job_id: str,
    ifc_paths: List[Path],
    override_paths: Dict[str, Optional[Path]],
    session_id: Optional[str],
) -> None:
    job_dir = Path(tempfile.mkdtemp(prefix="ifc_qa_"))
    output_dir = job_dir / "outputs"
    output_dir.mkdir(parents=True, exist_ok=True)

    logs: List[str] = []

    def log(msg: str) -> None:
        logs.append(msg)
        update_ifc_qa_job(job_id, logs=logs)

    try:
        update_ifc_qa_job(job_id, status="running", percent=2, currentStep="Initializing", logs=logs)

        regex_path = _qa_config_path(session_id, "regex_patterns", override_paths.get("regex_patterns"))
        exclude_path = _qa_config_path(session_id, "exclude_filter", override_paths.get("exclude_filter"))
        pset_path = _qa_config_path(session_id, "pset_template", override_paths.get("pset_template"))
        rules_path = _qa_config_path(session_id, "qa_rules", override_paths.get("qa_rules"))
        props_path = _qa_config_path(
            session_id, "qa_property_requirements", override_paths.get("qa_property_requirements")
        )
        bads_path = _qa_config_path(session_id, "qa_unacceptable_values", override_paths.get("qa_unacceptable_values"))

        regex_patterns = _load_regex_patterns(regex_path)
        regexes = {
            "regex_ifc_name": regex_patterns.get("regex_ifc_name", {}).get("pattern", ""),
            "regex_ifc_type": regex_patterns.get("regex_ifc_type", {}).get("pattern", ""),
            "regex_ifc_system": regex_patterns.get("regex_ifc_system", {}).get("pattern", ""),
            "regex_ifc_layer": regex_patterns.get("regex_ifc_layer", {}).get("pattern", ""),
            "regex_ifc_name_code": regex_patterns.get("regex_ifc_name_code", {}).get("pattern", ""),
            "regex_ifc_type_code": regex_patterns.get("regex_ifc_type_code", {}).get("pattern", ""),
            "regex_ifc_system_code": regex_patterns.get("regex_ifc_system_code", {}).get("pattern", ""),
        }
        for key, meta in regex_patterns.items():
            if meta.get("enabled") == "false":
                regexes[key] = ""

        exclude_terms = _read_csv_first_column(exclude_path)
        pset_template = _load_pset_template(pset_path)

        qa_rules = _read_csv_rows(rules_path)
        qa_property_requirements = _read_csv_rows(props_path)
        qa_unacceptable = _read_csv_rows(bads_path)

        for index, ifc_path in enumerate(ifc_paths, start=1):
            safe_name = sanitize_filename(ifc_path.name)
            base_name = ifc_path.stem
            file_out_dir = output_dir / base_name
            file_out_dir.mkdir(parents=True, exist_ok=True)
            update_ifc_qa_job(
                job_id,
                currentFile=safe_name,
                currentStep="Opening IFC",
                percent=min(int(index / max(len(ifc_paths), 1) * 100), 95),
            )
            log(f"[{safe_name}] Opening IFC...")
            try:
                model = ifcopenshell.open(str(ifc_path))
            except Exception as exc:
                log(f"[{safe_name}] ERROR opening IFC: {exc}")
                continue

            all_objects = _iter_object_elements(model)
            objects = [
                o for o in all_objects if not any(t in (getattr(o, "Name", "") or "") for t in exclude_terms)
            ]
            include_ids = {obj.id() for obj in objects}
            object_type_counts: Dict[str, int] = {}
            for obj in objects:
                object_type_counts[obj.is_a()] = object_type_counts.get(obj.is_a(), 0) + 1

            update_ifc_qa_job(job_id, currentStep="Extracting IFC PROJECT")
            _write_project_table(model, file_out_dir / f"IFC PROJECT - {base_name}.csv", safe_name)

            update_ifc_qa_job(job_id, currentStep="Extracting IFC OBJECT TYPE")
            _write_object_table(objects, file_out_dir / f"IFC OBJECT TYPE - {base_name}.csv", safe_name, regexes)

            update_ifc_qa_job(job_id, currentStep="Extracting IFC CLASSIFICATION")
            _write_classification_table(
                model,
                file_out_dir / f"IFC CLASSIFICATION - {base_name}.csv",
                safe_name,
                include_ids if include_ids else None,
            )

            update_ifc_qa_job(job_id, currentStep="Extracting IFC SPATIAL STRUCTURE")
            _write_spatial_table(
                model,
                file_out_dir / f"IFC SPATIAL STRUCTURE - {base_name}.csv",
                safe_name,
                objects,
            )

            update_ifc_qa_job(job_id, currentStep="Extracting IFC SYSTEM")
            _write_system_table(
                model,
                file_out_dir / f"IFC SYSTEM - {base_name}.csv",
                safe_name,
                include_ids if include_ids else None,
            )

            update_ifc_qa_job(job_id, currentStep="Extracting IFC PSET TEMPLATE")
            _write_pset_template_table(
                file_out_dir / f"IFC PSET TEMPLATE - {base_name}.csv",
                safe_name,
                pset_template,
                object_type_counts,
            )

            update_ifc_qa_job(job_id, currentStep="Extracting IFC PROPERTY")
            _write_property_table(
                model,
                file_out_dir / f"IFC PROPERTY - {base_name}.csv",
                safe_name,
                objects,
                pset_template,
            )
            log(f"[{safe_name}] Extraction complete.")

        tables: Dict[str, List[Dict[str, str]]] = {}
        for csv_path in output_dir.rglob("*.csv"):
            stem = csv_path.name.split(" - ")[0].strip()
            tables.setdefault(stem, [])
            tables[stem].extend(_read_csv_rows(csv_path))

        failures, totals_by_page, total_checks = _evaluate_qa_rules(
            tables,
            qa_rules,
            qa_property_requirements,
            qa_unacceptable,
        )
        objects_checked = len(tables.get("IFC OBJECT TYPE", []))
        total_failures = len(failures)
        pass_percent = 100.0 if total_checks == 0 else round(100.0 * (total_checks - total_failures) / total_checks, 2)

        pages = [
            "project_naming",
            "occurrence_naming",
            "type_naming",
            "classification_template",
            "classification_values",
            "pset_template",
            "property_values",
            "system",
            "zone",
        ]
        per_page = {}
        for page in pages:
            stats = totals_by_page.get(page, {"checks": 0, "fails": 0})
            checks = stats["checks"]
            fails = stats["fails"]
            per_page[page] = {
                "checks": checks,
                "fails": fails,
                "pass_percent": 100.0 if checks == 0 else round(100.0 * (checks - fails) / checks, 2),
            }

        qa_summary = {
            "overall": {
                "files_checked": len(ifc_paths),
                "objects_checked": objects_checked,
                "total_checks": total_checks,
                "total_failures": total_failures,
                "pass_percent": pass_percent,
            },
            "per_page": per_page,
            "failures": failures,
        }
        summary_path = job_dir / "qa_summary.json"
        with open(summary_path, "w", encoding="utf-8") as handle:
            json.dump(qa_summary, handle, indent=2)

        zip_path = job_dir / f"ifc_qa_{job_id}.zip"
        with zipfile.ZipFile(zip_path, "w", compression=zipfile.ZIP_DEFLATED) as zipf:
            for file_path in output_dir.rglob("*.csv"):
                arcname = file_path.relative_to(output_dir)
                zipf.write(file_path, arcname.as_posix())
            zipf.write(summary_path, summary_path.name)

        update_ifc_qa_job(
            job_id,
            status="complete",
            percent=100,
            currentStep="Complete",
            result_path=str(zip_path),
            summary=qa_summary,
        )
    except Exception as exc:  # pragma: no cover - job-level guard
        log(f"ERROR: {exc}")
        update_ifc_qa_job(job_id, status="failed", percent=100, currentStep="Failed")


def apply_edits(session_id: str, in_path: str, edits: List[Dict[str, Any]]) -> Tuple[str, List[Dict[str, Any]]]:
    if not edits:
        raise HTTPException(status_code=400, detail="No edits supplied")
    model = ifcopenshell.open(in_path)
    defs = _ensure_definitions()
    def_by_id = {d.check_id: d for d in defs}
    expr_engine = ExpressionEngine(model)
    audits = []
    for e in edits:
        check_id = e.get("check_id")
        obj_id = e.get("object_id")
        if check_id not in def_by_id:
            continue
        desc = def_by_id[check_id].field
        if desc is None:
            continue
        target = model.by_id(int(obj_id))
        if target is None:
            continue
        mode = e.get("mode", "manual")
        value = e.get("value")
        if mode == "generated" and desc.expression:
            value = expr_engine.evaluate(desc.expression, target)
        old_val, new_val = set_value(model, target, desc, value)
        audits.append(
            {
                "object_id": target.id(),
                "global_id": getattr(target, "GlobalId", None),
                "check_id": check_id,
                "field": desc.path_label(),
                "old": _to_serializable(old_val),
                "new": _to_serializable(new_val),
                "timestamp": utc_now().isoformat() + "Z",
            }
        )
    base, ext = os.path.splitext(os.path.basename(in_path))
    ts = utc_now().strftime("%Y%m%d_%H%M%S")
    out_name = f"{base}_checked_{ts}{ext or '.ifc'}"
    out_path = os.path.join(os.path.dirname(in_path), out_name)
    model.write(out_path)
    append_change_log(session_id, audits)
    return out_name, audits



def _tail_text(value: str, max_chars: int = 4000) -> str:
    return (value or "")[-max_chars:]


def _run_cobieqc_job(job_id: str) -> None:
    started_at = time.monotonic()
    try:
        with single_flight_heavy_job("/api/tools/cobieqc/run"):
            job = COBIEQC_JOB_STORE.update_job(
                job_id,
                status=STATUS_RUNNING,
                progress=0.1,
                started_at=utc_now().isoformat() + "Z",
                message="Running COBieQC reporter",
            )
            job_dir = COBIEQC_JOB_STORE.get_job_dir(job_id)
            input_path = job_dir / job.get("input_filename", "input.xlsx")
            stage = str(job.get("stage", "D")).upper()
            COBIEQC_JOB_STORE.append_log(job_id, f"Running stage {stage} for {input_path.name}")
            log_memory_stage(stage="COBieQC launch", session_id=job_id, file_name=input_path.name, file_size=input_path.stat().st_size if input_path.exists() else None, endpoint="/api/tools/cobieqc/run", started_at=started_at)
            result = run_cobieqc(str(input_path), stage, str(job_dir))
        COBIEQC_JOB_STORE.append_log(job_id, "--- STDOUT ---")
        COBIEQC_JOB_STORE.append_log(job_id, result.get("stdout", ""))
        COBIEQC_JOB_STORE.append_log(job_id, "--- STDERR ---")
        COBIEQC_JOB_STORE.append_log(job_id, result.get("stderr", ""))

        if result.get("ok"):
            COBIEQC_JOB_STORE.update_job(
                job_id,
                status=STATUS_DONE,
                progress=1.0,
                message="Report generated",
                finished_at=utc_now().isoformat() + "Z",
                output_filename=result.get("output_filename", "report.html"),
                cobie_xml_filename=Path(str(result.get("cobie_xml", ""))).name if result.get("cobie_xml") else None,
                svrl_xml_filename=Path(str(result.get("svrl_xml", ""))).name if result.get("svrl_xml") else None,
                validation_summary=result.get("summary", {}),
            )
            log_memory_stage(stage="response complete", session_id=job_id, file_name=result.get("output_filename", "report.html"), file_size=None, endpoint="/api/tools/cobieqc/run", started_at=started_at)
        else:
            COBIEQC_JOB_STORE.update_job(
                job_id,
                status=STATUS_ERROR,
                progress=1.0,
                message=result.get("error") or "COBieQC failed",
                finished_at=utc_now().isoformat() + "Z",
            )
    except Exception as exc:
        COBIEQC_JOB_STORE.append_log(job_id, f"Unhandled error: {exc}")
        COBIEQC_JOB_STORE.update_job(
            job_id,
            status=STATUS_ERROR,
            progress=1.0,
            message=str(exc),
            finished_at=utc_now().isoformat() + "Z",
        )


# ----------------------------------------------------------------------------
# FastAPI app + routes
# ----------------------------------------------------------------------------

def startup_cleanup() -> None:
    SESSION_STORE.cleanup_stale()
    COBIEQC_JOB_STORE.cleanup_old_jobs()
    host, port = resolve_server_host_port()
    APP_LOGGER.info("Startup network binding host=%s port=%s", host, port)
    bootstrap_cobieqc_assets()
    runtime_diag = get_cobieqc_runtime_diagnostics()
    engine = get_cobieqc_engine()
    java_xmx_mb = os.getenv("COBIEQC_JAVA_XMX_MB", "512")
    APP_LOGGER.info(
        "COBieQC startup health engine=%s enabled=%s jar_exists=%s resource_dir_exists=%s jar_path=%s resource_dir=%s java_xmx_mb=%s xml_count=%s xsl_count=%s",
        engine,
        runtime_diag["enabled"],
        runtime_diag["jar_exists"],
        runtime_diag["resource_dir_exists"],
        runtime_diag["jar_path"],
        runtime_diag["resource_dir"],
        java_xmx_mb,
        runtime_diag["xml_count"],
        runtime_diag["xsl_count"],
    )
    if runtime_diag["jar_error"]:
        APP_LOGGER.warning("COBieQC startup JAR issue: %s", runtime_diag["jar_error"])
    if runtime_diag["resource_error"]:
        APP_LOGGER.warning("COBieQC startup resource issue: %s", runtime_diag["resource_error"])
    APP_LOGGER.info("COBieQC Java diagnostics PATH=%s", os.getenv("PATH", ""))
    app_file = str(Path(__file__).resolve())
    runner_file = str(Path(cobieqc_runner_module.__file__).resolve())
    git_sha = _resolve_git_commit_sha()
    cobieqc_marker = getattr(cobieqc_runner_module, "COBIEQC_RUNNER_BUILD_MARKER", "unknown")
    cobieqc_flags = getattr(cobieqc_runner_module, "COBIEQC_RUNNER_FLAG_MARKER", "unknown")
    APP_LOGGER.info(
        "Runtime build diagnostics git_sha=%s app_file=%s cobieqc_runner_file=%s cobieqc_marker=%s cobieqc_flags=%s",
        git_sha,
        app_file,
        runner_file,
        cobieqc_marker,
        cobieqc_flags,
    )
    if engine != "java":
        APP_LOGGER.info("COBieQC engine is '%s'; skipping Java runtime checks", engine)
        return

    which_java = subprocess.run(
        ["which", "java"],
        capture_output=True,
        text=True,
        check=False,
        timeout=10,
    )
    which_java_output = (which_java.stdout or which_java.stderr or "").strip()
    APP_LOGGER.info(
        "COBieQC Java diagnostics which java (rc=%s): %s",
        which_java.returncode,
        which_java_output or "<not found>",
    )
    try:
        java_bin = resolve_java_executable()
        proc = subprocess.run([java_bin, "-version"], capture_output=True, text=True, check=False, timeout=10)
        java_version_output = (proc.stderr or proc.stdout or "").strip()
        APP_LOGGER.info("COBieQC Java diagnostics java -version (rc=%s): %s", proc.returncode, java_version_output)
        if proc.returncode == 0:
            APP_LOGGER.info("COBieQC Java runtime detected at %s", java_bin)
        else:
            APP_LOGGER.warning("COBieQC Java runtime check failed: %s", java_version_output)
            APP_LOGGER.warning("COBieQC Java runtime unavailable")
    except Exception as exc:
        APP_LOGGER.warning("COBieQC Java runtime unavailable: %s", exc)


def shutdown_cleanup() -> None:
    for sid in list(SESSION_STORE.sessions.keys()):
        SESSION_STORE.drop(sid)


def _collect_session_route_lines() -> List[str]:
    lines: List[str] = []
    for route in app.routes:
        path = getattr(route, "path", "")
        if not path.startswith("/api/session"):
            continue
        methods = sorted((getattr(route, "methods", set()) or set()) - {"HEAD", "OPTIONS"})
        for method in methods:
            lines.append(f"- {method} {path}")
    return sorted(set(lines))


def _log_session_route_registration() -> None:
    lines = _collect_session_route_lines()
    if lines:
        APP_LOGGER.info("SESSION_UPLOAD_ROUTES_REGISTERED:\n%s", "\n".join(lines))
    else:
        APP_LOGGER.warning("SESSION_UPLOAD_ROUTES_REGISTERED:\n- <none>")


def _log_area_spaces_route_registration() -> None:
    lines: List[str] = []
    for route in app.routes:
        path = getattr(route, "path", "")
        if "area-spaces" not in path:
            continue
        methods = sorted((getattr(route, "methods", set()) or set()) - {"HEAD", "OPTIONS"})
        for method in methods:
            lines.append(f"- {method} {path}")
    APP_LOGGER.info("AREA_SPACES_ROUTES_REGISTERED:\n%s", "\n".join(sorted(lines)) if lines else "- <none>")


@asynccontextmanager
async def lifespan(_: FastAPI):
    startup_cleanup()
    _log_session_route_registration()
    _log_area_spaces_route_registration()
    try:
        yield
    finally:
        shutdown_cleanup()


app = FastAPI(title="IFC Toolkit Hub", lifespan=lifespan)
app.mount("/static", CacheControlledStaticFiles(directory="static"), name="static")
templates = Jinja2Templates(directory="templates")
templates.env.globals["asset_url"] = resolve_asset_url
templates.env.globals["frontend_build_id"] = FRONTEND_BUILD_ID


@app.middleware("http")
async def upload_request_size_guard(request: Request, call_next):
    upload_paths = {
        "/api/ifc-qa/run",
        "/api/ifc-qa/add-to-zip",
        "/api/tools/cobieqc/run",
    }
    path = request.url.path
    is_upload_path = path in upload_paths or path.endswith("/upload")
    if is_upload_path:
        content_length_header = request.headers.get("content-length")
        if content_length_header:
            try:
                content_length = int(content_length_header)
            except ValueError:
                content_length = None
            if content_length and content_length > MAX_REQUEST_BODY_BYTES:
                log_upload_rejection(
                    endpoint=path,
                    filename="request-body",
                    content_length=content_length,
                    actual_size=None,
                    rejection_reason="request_content_length_exceeds_limit",
                )
                return JSONResponse(status_code=413, content=upload_too_large_payload())
    return await call_next(request)


@app.middleware("http")
async def html_cache_control_guard(request: Request, call_next):
    response = await call_next(request)
    content_type = (response.headers.get("content-type") or "").lower()
    if content_type.startswith("text/html"):
        response.headers["Cache-Control"] = "no-cache, must-revalidate"
        response.headers["Pragma"] = "no-cache"
        response.headers["Expires"] = "0"
    return response


@app.exception_handler(HTTPException)
async def http_exception_handler(request: Request, exc: HTTPException):
    if exc.status_code == 413 and isinstance(exc.detail, dict) and exc.detail.get("code") == "UPLOAD_TOO_LARGE":
        return JSONResponse(status_code=413, content=exc.detail)
    return JSONResponse(status_code=exc.status_code, content={"detail": exc.detail})


@app.get("/health")
def health():
    runtime_diag = get_cobieqc_runtime_diagnostics()
    bootstrap_status = get_cobieqc_bootstrap_status()
    engine = runtime_diag.get("engine", "python")
    jar_ready = bool(bootstrap_status.get("jar_ready", runtime_diag["jar_exists"])) if engine == "java" else True
    payload = {
        "status": "ok",
        "service": "ifc-tools",
        "cobieqc": {
            "enabled": runtime_diag["enabled"],
            "engine": engine,
            "jar_exists": runtime_diag["jar_exists"],
            "jar_ready": jar_ready,
            "resource_dir_exists": runtime_diag["resource_dir_exists"],
            "resources_ready": bool(bootstrap_status.get("resources_ready", runtime_diag["resource_dir_exists"])),
            "jar_source": bootstrap_status.get("jar_source", ""),
            "jar_validation_error": bootstrap_status.get("jar_validation_error", ""),
            "jar_path": runtime_diag.get("jar_path"),
            "resource_dir": runtime_diag.get("resource_dir"),
            "missing_files": bootstrap_status.get("missing_files", []),
            "warnings": bootstrap_status.get("warnings", []),
            "errors": bootstrap_status.get("errors", []),
            "source_mode": bootstrap_status.get("source_mode", "unknown"),
        },
    }
    last_error = bootstrap_status.get("last_error") or runtime_diag.get("jar_error") or runtime_diag.get("resource_error")
    if last_error:
        payload["cobieqc"]["last_error"] = last_error
    return payload


@app.get("/api/upload/limits")
def upload_limits():
    return {
        "max_upload_bytes": MAX_UPLOAD_BYTES,
        "max_upload_gb": MAX_UPLOAD_GB,
        "max_upload_display": MAX_UPLOAD_DISPLAY,
        "max_request_body_bytes": MAX_REQUEST_BODY_BYTES,
    }


@app.get("/", response_class=HTMLResponse)
def upload_page(request: Request):
    return templates.TemplateResponse(request=request, name="upload.html", context={"request": request, "active": "upload"})


@app.get("/cleaner", response_class=HTMLResponse)
def cleaner_page(request: Request):
    return templates.TemplateResponse(request=request, name="cleaner.html", context={"request": request, "active": "cleaner"})


@app.get("/excel", response_class=HTMLResponse)
def excel_page(request: Request):
    return templates.TemplateResponse(request=request, name="excel.html", context={"request": request, "active": "excel"})


@app.get("/ifc-qa/extractor", response_class=HTMLResponse)
def ifc_qa_extractor_page(request: Request):
    return templates.TemplateResponse(
        request=request,
        name="ifc_qa.html",
        context={"request": request, "active": "ifc-qa", "qa_page": "extractor"},
    )


@app.get("/ifc-qa/dashboard", response_class=HTMLResponse)
def ifc_qa_dashboard_page(request: Request):
    return templates.TemplateResponse(
        request=request,
        name="ifc_qa.html",
        context={"request": request, "active": "ifc-qa", "qa_page": "dashboard"},
    )


@app.get("/ifc-qa/config", response_class=HTMLResponse)
def ifc_qa_config_page(request: Request):
    return templates.TemplateResponse(
        request=request,
        name="ifc_qa.html",
        context={"request": request, "active": "ifc-qa", "qa_page": "config"},
    )


@app.get("/data-extractor", response_class=HTMLResponse)
def data_extractor_page(request: Request):
    return templates.TemplateResponse(
        request=request,
        name="data_extractor.html",
        context={"request": request, "active": "data-extractor", "regex_defaults": load_default_config()},
    )


@app.get("/storeys", response_class=HTMLResponse)
def storeys_page(request: Request):
    return templates.TemplateResponse(request=request, name="storeys.html", context={"request": request, "active": "storeys"})


@app.get("/proxy", response_class=HTMLResponse)
def proxy_page(request: Request):
    return templates.TemplateResponse(request=request, name="proxy.html", context={"request": request, "active": "proxy"})


@app.get("/presentation-layer", response_class=HTMLResponse)
def presentation_layer_page(request: Request):
    return templates.TemplateResponse(
        request=request,
        name="presentation_layer.html",
        context={"request": request, "active": "presentation-layer"},
    )


@app.get("/step2ifc", response_class=HTMLResponse)
def step2ifc_page(request: Request):
    return templates.TemplateResponse(request=request, name="step2ifc.html", context={"request": request, "active": "step2ifc"})


@app.get("/tools/cobieqc", response_class=HTMLResponse)
def cobieqc_page(request: Request):
    return templates.TemplateResponse(request=request, name="cobieqc.html", context={"request": request, "active": "cobieqc"})


@app.get("/files", response_class=HTMLResponse)
def files_page(request: Request):
    return templates.TemplateResponse(request=request, name="files.html", context={"request": request, "active": "files"})


@app.get("/levels", response_class=HTMLResponse)
def levels_page(request: Request):
    return templates.TemplateResponse(request=request, name="levels.html", context={"request": request, "active": "levels"})


@app.get("/viewer", response_class=HTMLResponse)
def viewer_page(request: Request):
    return HTMLResponse(
        "<html><body><h2>IFC Viewer temporarily disabled</h2><p>The viewer is not included in this build.</p></body></html>",
        status_code=503,
    )


@app.get("/model-checking", response_class=HTMLResponse)
def model_checking_page(request: Request):
    return templates.TemplateResponse(request=request, name="model_checking.html", context={"request": request, "active": "model-checking"})


@app.get("/admin/mappings", response_class=HTMLResponse)
def admin_mappings_page(request: Request):
    return templates.TemplateResponse(request=request, name="mappings.html", context={"request": request, "active": "mappings"})


@app.get("/wip/ifc-move-rotate", response_class=HTMLResponse)
def ifc_move_rotate_page(request: Request):
    return templates.TemplateResponse(request=request, name="ifc_move_rotate.html", context={"request": request, "active": "ifc-move-rotate"})


@app.get("/tools/reduce-file-size", response_class=HTMLResponse)
def reduce_file_size_page(request: Request):
    return templates.TemplateResponse(request=request, name="ifc_file_size_reducer.html", context={"request": request, "active": "reduce-file-size"})


@app.get("/tools/purge-area-spaces", response_class=HTMLResponse)
def purge_area_spaces_page(request: Request):
    return templates.TemplateResponse(request=request, name="purge_area_spaces.html", context={"request": request, "active": "purge-area-spaces"})


@app.post("/api/session")
def create_session(payload: Dict[str, Any] = Body(default=None)):
    SESSION_STORE.cleanup_stale()
    incoming = payload.get("session_id") if payload else None
    if incoming and SESSION_STORE.exists(incoming):
        SESSION_STORE.touch(incoming)
        session_id = incoming
        APP_LOGGER.info("session_reused session_id=%s", session_id)
    else:
        session_id = SESSION_STORE.create()
        APP_LOGGER.info("session_created session_id=%s root=%s", session_id, SESSION_STORE.session_path(session_id))
    expiry = utc_now() + datetime.timedelta(hours=SESSION_STORE.ttl_hours)
    return {"session_id": session_id, "expires_at": expiry.isoformat() + "Z"}


@app.get("/api/session/{session_id}")
def get_session(session_id: str):
    normalized = _require_valid_session_id(session_id)
    root = _ensure_session_dir_for_upload(normalized)
    APP_LOGGER.info("session_lookup session_id=%s", session_id)
    files = [name for name in os.listdir(root) if os.path.isfile(os.path.join(root, name))]
    expiry = utc_now() + datetime.timedelta(hours=SESSION_STORE.ttl_hours)
    return {
        "session_id": session_id,
        "expires_at": expiry.isoformat() + "Z",
        "upload_root": root,
        "file_count": len(files),
    }


@app.delete("/api/session/{session_id}")
def end_session(session_id: str):
    normalized = _require_valid_session_id(session_id)
    SESSION_STORE.drop(normalized)
    return {"status": "deleted"}


@app.get("/api/session/{session_id}/files")
def list_files(session_id: str, request: Request = None):
    normalized = _require_valid_session_id(session_id)
    root = _ensure_session_dir_for_upload(normalized)
    list_start = time.perf_counter()
    APP_LOGGER.info("file_index_refresh_start session_id=%s root=%s", session_id, root)
    files = []
    for fname in sorted(os.listdir(root)):
        fpath = os.path.join(root, fname)
        if os.path.isfile(fpath):
            files.append(_build_session_file_metadata(root, fname))
    list_duration_ms = int((time.perf_counter() - list_start) * 1000)
    APP_LOGGER.info(
        "file_index_refresh_complete session_id=%s duration_ms=%s files_returned=%s",
        session_id,
        list_duration_ms,
        len(files),
    )
    ifc_count = sum(1 for item in files if _is_ifc_compatible(item.get("name", "")))
    page_name = "unknown"
    if request is not None:
        page_name = (request.query_params.get("page") or request.headers.get("x-ifc-tool-page") or "unknown").strip() or "unknown"
    APP_LOGGER.info(
        "session_file_list page=%s session_id=%s root=%s files_returned=%s ifc_files=%s",
        page_name,
        session_id,
        root,
        len(files),
        ifc_count,
    )
    return {"files": files}


@app.get("/api/sessions/{session_id}/files")
def list_files_alias_plural(session_id: str):
    return list_files(session_id)


@app.delete("/api/session/{session_id}/files/{file_id}")
def delete_session_file(session_id: str, file_id: str):
    normalized = _require_valid_session_id(session_id)
    root = Path(_ensure_session_dir_for_upload(normalized))
    safe_name = sanitize_filename(os.path.basename(file_id))
    target = root / safe_name
    if not target.exists() or not target.is_file():
        raise HTTPException(status_code=404, detail="File not found in session")
    target.unlink(missing_ok=True)
    APP_LOGGER.info("session_file_deleted session_id=%s file=%s", session_id, safe_name)
    return {"deleted": True, "id": safe_name}


@app.get("/api/session/{session_id}/debug")
def session_debug(session_id: str):
    normalized = _require_valid_session_id(session_id)
    root = Path(_ensure_session_dir_for_upload(normalized))
    file_count = sum(1 for item in root.iterdir() if item.is_file())
    return {
        "active_session_id": session_id,
        "upload_root": str(root),
        "max_upload_bytes": MAX_UPLOAD_BYTES,
        "session_directory_exists": root.exists() and root.is_dir(),
        "session_file_count": file_count,
    }


@app.get("/api/session/debug/routes")
def session_debug_routes():
    session_routes = []
    for route in app.routes:
        path = getattr(route, "path", "")
        if "session" not in path or not path.startswith("/api/"):
            continue
        methods = sorted((getattr(route, "methods", set()) or set()) - {"HEAD", "OPTIONS"})
        for method in methods:
            session_routes.append(f"{method} {path}")
    upload_root = Path(SESSION_STORE.base_dir)
    return {
        "session_routes": sorted(set(session_routes)),
        "upload_root": str(upload_root),
        "upload_root_exists": upload_root.exists() and upload_root.is_dir(),
        "max_upload_bytes": MAX_UPLOAD_BYTES,
    }


def _is_ifc_compatible(name: str) -> bool:
    lower = str(name or "").lower()
    return lower.endswith(".ifc") or lower.endswith(".ifczip") or lower.endswith(".ifcxml")


def _session_file_kind(name: str) -> str:
    extension = Path(name or "").suffix.lower()
    if extension in {".ifc", ".ifczip", ".ifcxml"}:
        return "ifc"
    if extension in {".xlsx", ".xls"}:
        return "excel"
    if extension == ".zip":
        return "zip"
    if extension in {".json", ".xml", ".csv"}:
        return "data"
    return "file"


def _build_session_file_metadata(root: str, file_name: str) -> Dict[str, Any]:
    fpath = Path(root) / file_name
    stat = fpath.stat()
    created_at = datetime.datetime.fromtimestamp(stat.st_ctime, tz=UTC).isoformat()
    modified = datetime.datetime.fromtimestamp(stat.st_mtime, tz=UTC).isoformat()
    extension = fpath.suffix.lower()
    mime_type = mimetypes.guess_type(file_name)[0] or "application/octet-stream"
    return {
        "id": file_name,
        "filename": file_name,
        "name": file_name,
        "size": stat.st_size,
        "kind": _session_file_kind(file_name),
        "modified": modified,
        "mime_type": mime_type,
        "extension": extension,
        "created_at": created_at,
        "path": file_name,
    }


def _resolve_session_ifc_records(session_id: str, file_names: List[str]) -> List[Tuple[str, str]]:
    root = Path(SESSION_STORE.ensure(session_id)).resolve()
    records: List[Tuple[str, str]] = []
    for incoming in file_names:
        safe = sanitize_filename(os.path.basename(str(incoming or "")))
        if not safe:
            raise HTTPException(status_code=400, detail=f"Invalid file reference: {incoming!r}")
        if not _is_ifc_compatible(safe):
            raise HTTPException(status_code=400, detail=f"Unsupported IFC file type: {safe}")
        source_path = (root / safe).resolve()
        if root not in source_path.parents and source_path != root:
            raise HTTPException(status_code=400, detail=f"Session file path escapes session directory: {safe}")
        if not source_path.exists() or not source_path.is_file():
            raise HTTPException(status_code=404, detail=f"Session file not found: {safe}")
        records.append((safe, str(source_path)))
    if not records:
        raise HTTPException(status_code=400, detail="At least one IFC session file must be selected")
    return records


@app.post("/api/session/{session_id}/ifc-qa/extract")
def ifc_qa_extract_from_session(
    session_id: str,
    payload: Dict[str, Any] = Body(...),
):
    if has_active_ifc_qa_job():
        raise HTTPException(status_code=429, detail="An IFC QA job is already running on this replica. Please retry shortly.")

    raw_ids = payload.get("file_ids")
    raw_names = payload.get("file_names")
    requested = raw_ids if isinstance(raw_ids, list) and raw_ids else raw_names
    if not isinstance(requested, list):
        raise HTTPException(status_code=400, detail="Provide file_ids or file_names as an array")

    options = payload.get("options") if isinstance(payload.get("options"), dict) else {}
    options["max_workers"] = 1
    options["session_id"] = session_id

    cfg = load_ifc_qa_default_config()
    override = payload.get("config_override")
    if isinstance(override, dict):
        cfg = merge_config_override(cfg, override)
        cfg["_indexes"] = build_config_indexes(cfg)

    file_records = _resolve_session_ifc_records(session_id, [str(item) for item in requested])
    APP_LOGGER.info(
        "ifc_qa_session_extract_start session_id=%s file_count=%s selected=%s",
        session_id,
        len(file_records),
        ",".join(name for name, _ in file_records),
    )
    session_root = Path(SESSION_STORE.ensure(session_id)) / "ifc_qa_session"
    session_root.mkdir(parents=True, exist_ok=True)
    job_id = IFC_QA_JOB_STARTER(session_root, session_id, file_records, options, cfg, mode="replace")
    return {
        "success": True,
        "job_id": job_id,
        "session_id": session_id,
        "selected_files": [{"file_id": name, "file_name": name} for name, _ in file_records],
        "status_url": f"/api/ifc-qa/status/{job_id}",
    }


@app.post("/api/ifc-data-qa/extract")
def ifc_data_qa_extract(payload: Dict[str, Any] = Body(...)):
    session_id = str(payload.get("session_id") or "").strip()
    if not session_id:
        raise HTTPException(status_code=400, detail="session_id is required")
    return ifc_qa_extract_from_session(session_id=session_id, payload=payload)


@app.post("/api/ifc-tools/reduce-file-size/analyse")
def api_reduce_file_size_analyse(payload: Dict[str, Any] = Body(...)):
    session_id = payload.get("session_id")
    source_file = payload.get("source_file")
    if not session_id or not source_file:
        raise HTTPException(status_code=400, detail="session_id and source_file are required")
    root = Path(SESSION_STORE.ensure(str(session_id)))
    source_path = root / sanitize_filename(os.path.basename(str(source_file)))
    if not source_path.exists():
        raise HTTPException(status_code=404, detail="Selected source file not found")
    if not source_path.name.lower().endswith((".ifc", ".ifczip")):
        raise HTTPException(status_code=400, detail="Input must be .ifc or .ifczip")
    try:
        return {"status": "ok", "analysis": analyze_ifc_file(source_path)}
    except IfcFileSizeReducerError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Analysis failed: {exc}") from exc


@app.post("/api/ifc-tools/reduce-file-size/run")
def api_reduce_file_size_run(payload: Dict[str, Any] = Body(...)):
    session_id = payload.get("session_id")
    source_file = payload.get("source_file")
    if not session_id or not source_file:
        raise HTTPException(status_code=400, detail="session_id and source_file are required")
    root = Path(SESSION_STORE.ensure(str(session_id)))
    try:
        result = run_ifc_size_reduction(root, sanitize_filename(os.path.basename(str(source_file))), payload)
        return {"status": "ok", "result": result}
    except IfcFileSizeReducerError as exc:
        detail = str(exc)
        try:
            parsed = json.loads(detail)
            raise HTTPException(status_code=400, detail=parsed) from exc
        except json.JSONDecodeError:
            raise HTTPException(status_code=400, detail=detail) from exc
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Reduction failed: {exc}") from exc


def _resolve_session_ifc_file_paths(session_id: str, file_names: List[str]) -> List[Tuple[str, Path]]:
    records = _resolve_session_ifc_records(session_id, file_names)
    output: List[Tuple[str, Path]] = []
    for name, resolved in records:
        if not name.lower().endswith(".ifc"):
            raise HTTPException(status_code=400, detail=f"Only .ifc files are supported: {name}")
        output.append((name, Path(resolved)))
    return output


def area_spaces_session_files(session_id: str):
    normalized = _require_valid_session_id(session_id)
    root = Path(_ensure_session_dir_for_upload(normalized))
    files = []
    for file_path in sorted(root.glob("*.ifc")):
        stat = file_path.stat()
        files.append({"name": file_path.name, "size": stat.st_size, "modified": datetime.datetime.fromtimestamp(stat.st_mtime, tz=UTC).isoformat()})
    return {"session_id": normalized, "files": files, "count": len(files)}


async def area_spaces_scan(payload: Dict[str, Any] = Body(...)):
    session_id = str(payload.get("session_id") or "").strip()
    requested = payload.get("file_names") or payload.get("file_ids")
    debug_mode = bool(payload.get("debug_mode") or payload.get("debug"))
    try:
        if not session_id:
            return JSONResponse(status_code=400, content={"ok": False, "error": "AREA_SPACE_SCAN_FAILED", "message": "session_id is required", "stage": "ifc_open"})
        if not isinstance(requested, list) or not requested:
            return JSONResponse(status_code=400, content={"ok": False, "error": "AREA_SPACE_SCAN_FAILED", "message": "file_names (or file_ids) array is required", "stage": "ifc_open"})
        if len(requested) != 1:
            return JSONResponse(status_code=400, content={"ok": False, "error": "AREA_SPACE_SCAN_FAILED", "message": "Process one IFC at a time for memory safety.", "stage": "scan_spaces"})

        semaphore_locked = AREA_SPACE_JOB_SEMAPHORE.locked()
        if semaphore_locked:
            APP_LOGGER.info("area_spaces_scan_waiting_for_slot")

        async with AREA_SPACE_JOB_SEMAPHORE:
            ifc_records = _resolve_session_ifc_file_paths(session_id, [str(item) for item in requested])
            source_name, path = ifc_records[0]
            file_size_mb = round(path.stat().st_size / (1024 * 1024), 2)
            if file_size_mb > float(os.getenv("AREA_SPACE_MAX_INLINE_MB", "300")):
                return JSONResponse(
                    status_code=400,
                    content={
                        "status": "error",
                        "message": f"File too large for inline processing ({file_size_mb:.1f}MB). Use chunked processing.",
                        "file": source_name,
                    },
                )

            result = scan_ifc_for_area_spaces(path, debug_mode=debug_mode)
            APP_LOGGER.info("area_spaces_scan %s", area_space_log_payload(result))
            scan_payload = {
                "source_file": result.source_file,
                "total_spaces": result.total_spaces,
                "candidates": [candidate.__dict__ for candidate in result.candidates],
            }
            return {
                "ok": True,
                "session_id": session_id,
                "progress": "candidates found",
                "status_messages": (["Another IFC job is running. Waiting for available processing slot..."] if semaphore_locked else []) + ["ifc_open complete", "scan_spaces complete"],
                "files_scanned": 1,
                "total_spaces": result.total_spaces,
                "total_candidates": len(result.candidates),
                "results": [scan_payload],
            }
    except AreaSpaceError as exc:
        APP_LOGGER.exception("area_spaces_scan_failed")
        try:
            parsed = json.loads(str(exc))
            return JSONResponse(status_code=int(parsed.get("http_status", 400)), content=parsed)
        except json.JSONDecodeError:
            return JSONResponse(status_code=400, content={"ok": False, "error": "AREA_SPACE_SCAN_FAILED", "message": str(exc), "stage": "scan_spaces"})
    except Exception as exc:
        APP_LOGGER.exception("area_spaces_scan_failed")
        return JSONResponse(status_code=500, content={"ok": False, "error": "AREA_SPACE_SCAN_FAILED", "message": str(exc), "stage": "scan_spaces"})


async def area_spaces_purge(payload: Dict[str, Any] = Body(...)):
    session_id = str(payload.get("session_id") or "").strip()
    selected = payload.get("selected_candidates")
    if not session_id:
        return JSONResponse(status_code=400, content={"ok": False, "error": "AREA_SPACE_PURGE_FAILED", "message": "session_id is required", "stage": "purge"})
    if not isinstance(selected, list) or not selected:
        return JSONResponse(status_code=400, content={"ok": False, "error": "AREA_SPACE_PURGE_FAILED", "message": "selected_candidates must contain at least one GlobalId/STEP id", "stage": "purge"})

    requested_files = payload.get("file_names") or sorted({str(item.get("source_file")) for item in selected if isinstance(item, dict) and item.get("source_file")})
    if not isinstance(requested_files, list) or not requested_files:
        return JSONResponse(status_code=400, content={"ok": False, "error": "AREA_SPACE_PURGE_FAILED", "message": "file_names array is required", "stage": "purge"})
    if len(requested_files) != 1:
        return JSONResponse(status_code=400, content={"ok": False, "error": "AREA_SPACE_PURGE_FAILED", "message": "Process one IFC at a time for memory safety.", "stage": "purge"})

    grouped_selections: Dict[str, List[str]] = {}
    for item in selected:
        if isinstance(item, dict):
            source = str(item.get("source_file") or "")
            key = str(item.get("global_id") or item.get("step_id") or "").strip()
            if source and key:
                grouped_selections.setdefault(source, []).append(key)
        else:
            grouped_selections.setdefault(str(requested_files[0]), []).append(str(item))

    session_root = Path(SESSION_STORE.ensure(session_id))
    ifc_records = _resolve_session_ifc_file_paths(session_id, [str(item) for item in requested_files])
    purge_results = []
    produced_names: List[str] = []
    for source_name, source_path in ifc_records:
        APP_LOGGER.info("purge_request_received filename=%s selected=%s", source_name, len(selected))
        selected_keys = grouped_selections.get(source_name, [])
        if not selected_keys:
            continue
        size_mb = round(source_path.stat().st_size / (1024 * 1024), 2)
        allow_oversize = str(os.getenv("AREA_SPACE_ALLOW_OVERSIZE", "false")).lower() == "true"
        if size_mb > AREA_SPACE_MAX_PURGE_FILE_MB and not allow_oversize:
            return JSONResponse(
                status_code=413,
                content={
                    "ok": False,
                    "error": "IFC_TOO_LARGE_FOR_SAFE_PROCESSING",
                    "message": "This IFC is above the configured safe processing size. Split the IFC or increase AREA_SPACE_MAX_FILE_MB.",
                    "file_size_mb": size_mb,
                },
            )
        if is_memory_high():
            return JSONResponse(
                status_code=503,
                content={
                    "ok": False,
                    "error": "INSUFFICIENT_MEMORY_FOR_SAFE_PURGE",
                    "message": "Server memory is too high to safely write the cleaned IFC. Try again when other jobs finish.",
                    "memory": get_memory_status(),
                },
            )
        output_name = f"{source_path.stem}.area-spaces-purged.ifc"
        output_path = session_root / output_name
        audit_path = output_path.with_name(output_path.stem.replace(".area-spaces-purged", "") + ".area-spaces-purge-report.csv")
        selected_json_path = session_root / f"{source_path.stem}.area-spaces-selection.json"
        selected_json_path.write_text(json.dumps(selected_keys), encoding="utf-8")
        wait_start = time.perf_counter()
        semaphore_locked = AREA_SPACE_JOB_SEMAPHORE.locked()
        if semaphore_locked:
            APP_LOGGER.info("area_space_job_wait_start stage=purge filename=%s", source_name)
        try:
            async with AREA_SPACE_JOB_SEMAPHORE:
                wait_ms = int((time.perf_counter() - wait_start) * 1000)
                APP_LOGGER.info("area_space_job_acquired stage=purge wait_ms=%s", wait_ms)
                timeout_seconds = int(os.getenv("AREA_SPACE_PURGE_TIMEOUT_SECONDS", "900"))
                APP_LOGGER.info("purge_child_start filename=%s timeout_seconds=%s", source_name, timeout_seconds)
                worker_cmd = [
                    sys.executable,
                    "-m",
                    "ifc_app.workers.area_space_purge_worker",
                    "--source-ifc",
                    str(source_path),
                    "--output-ifc",
                    str(output_path),
                    "--selected-json",
                    str(selected_json_path),
                    "--audit-csv",
                    str(audit_path),
                ]
                try:
                    proc = subprocess.run(worker_cmd, capture_output=True, text=True, check=False, timeout=timeout_seconds)
                except subprocess.TimeoutExpired:
                    APP_LOGGER.error("purge_child_timeout filename=%s", source_name)
                    return JSONResponse(
                        status_code=504,
                        content={
                            "ok": False,
                            "error": "AREA_SPACE_PURGE_TIMEOUT",
                            "message": "Purge exceeded timeout. Split the IFC or increase timeout/memory.",
                        },
                    )
                APP_LOGGER.info("purge_child_returncode filename=%s returncode=%s", source_name, proc.returncode)
                if proc.returncode != 0:
                    APP_LOGGER.error("purge_child_failed filename=%s", source_name)
                    return JSONResponse(
                        status_code=500,
                        content={
                            "ok": False,
                            "error": "AREA_SPACE_PURGE_SUBPROCESS_FAILED",
                            "returncode": proc.returncode,
                            "message": "Purge failed in isolated worker. Main app remained online.",
                            "stderr_tail": (proc.stderr or "")[-4000:],
                        },
                    )
                result = json.loads((proc.stdout or "").strip().splitlines()[-1])
                APP_LOGGER.info("purge_child_success filename=%s", source_name)
                APP_LOGGER.info("area_space_job_released stage=purge")
        except Exception as exc:
            APP_LOGGER.exception("area_spaces_purge_failed")
            return JSONResponse(status_code=500, content={"ok": False, "error": "AREA_SPACE_PURGE_FAILED", "message": f"Purge failed for {source_name}: {exc}", "stage": "purge"})
        APP_LOGGER.info("area_spaces_purge %s", json.dumps(result, sort_keys=True))
        purge_results.append(result)
        produced_names.extend([result["output_ifc"], result["report_csv"]])

    if not purge_results:
        return JSONResponse(status_code=400, content={"ok": False, "error": "AREA_SPACE_PURGE_FAILED", "message": "No selected candidates matched the requested files", "stage": "purge"})

    output_files = [{"name": name, "download_url": f"/api/session/{session_id}/download?name={name}"} for name in produced_names]
    if len(purge_results) > 1:
        zip_name = "area-spaces-purged.outputs.zip"
        archive = package_area_space_outputs(session_root, produced_names, zip_name)
        output_files.append({"name": archive, "download_url": f"/api/session/{session_id}/download?name={archive}"})

    return {
        "ok": True,
        "session_id": session_id,
        "progress": "complete",
        "status_messages": (["Another IFC job is running. Waiting for available processing slot..."] if 'semaphore_locked' in locals() and semaphore_locked else []) + ["purge complete", "write complete"],
        "results": purge_results,
        "output_files": output_files,
    }


@app.get("/api/ifc/area-spaces/health")
def area_spaces_health():
    return {
        "ok": True,
        "routes_mounted": True,
        "max_concurrent_jobs": int(os.getenv("AREA_SPACE_MAX_CONCURRENT_JOBS", "1")),
        "child_memory_mb": int(os.getenv("AREA_SPACE_PURGE_CHILD_MEMORY_MB", "8192")),
        "timeout_seconds": int(os.getenv("AREA_SPACE_PURGE_TIMEOUT_SECONDS", "900")),
        "max_file_mb": AREA_SPACE_MAX_FILE_MB,
        "max_purge_file_mb": AREA_SPACE_MAX_PURGE_FILE_MB,
        "memory_abort_percent": AREA_SPACE_MEMORY_ABORT_PERCENT,
    }


area_spaces_router = build_area_spaces_router(
    scan_handler=area_spaces_scan,
    purge_handler=area_spaces_purge,
    files_handler=area_spaces_session_files,
)
app.include_router(area_spaces_router)


@app.get("/api/tools/cobieqc/health")
def cobieqc_health():
    runtime_diag = get_cobieqc_runtime_diagnostics()
    engine = runtime_diag.get("engine", "python")
    if engine != "java":
        return {
            "ok": runtime_diag["resource_dir_exists"],
            "engine": engine,
            "java_available": False,
            "java_path": None,
            "jar_available": runtime_diag["jar_exists"],
            "jar_path": runtime_diag["jar_path"],
            "resource_dir_available": runtime_diag["resource_dir_exists"],
            "resource_dir": runtime_diag["resource_dir"],
            "xml_count": runtime_diag["xml_count"],
            "xsl_count": runtime_diag["xsl_count"],
            "attempted_jar_paths": runtime_diag["jar_candidates"],
            "attempted_resource_dirs": runtime_diag["resource_candidates"],
            "jar_error": runtime_diag["jar_error"],
            "resource_error": runtime_diag["resource_error"],
            "detail": "Python-native COBieQC engine selected",
        }
    java_bin = resolve_java_executable()
    try:
        proc = subprocess.run([java_bin, "-version"], capture_output=True, text=True, check=False, timeout=10)
    except Exception as exc:
        return {
            "ok": False,
            "java_available": False,
            "java_path": java_bin,
            "jar_available": runtime_diag["jar_exists"],
            "jar_path": runtime_diag["jar_path"],
            "resource_dir_available": runtime_diag["resource_dir_exists"],
            "resource_dir": runtime_diag["resource_dir"],
            "xml_count": runtime_diag["xml_count"],
            "xsl_count": runtime_diag["xsl_count"],
            "attempted_jar_paths": runtime_diag["jar_candidates"],
            "attempted_resource_dirs": runtime_diag["resource_candidates"],
            "jar_error": runtime_diag["jar_error"],
            "resource_error": runtime_diag["resource_error"],
            "detail": str(exc),
        }
    return {
        "ok": proc.returncode == 0 and runtime_diag["jar_exists"] and runtime_diag["resource_dir_exists"],
        "java_available": proc.returncode == 0,
        "java_path": java_bin,
        "jar_available": runtime_diag["jar_exists"],
        "jar_path": runtime_diag["jar_path"],
        "resource_dir_available": runtime_diag["resource_dir_exists"],
        "resource_dir": runtime_diag["resource_dir"],
        "xml_count": runtime_diag["xml_count"],
        "xsl_count": runtime_diag["xsl_count"],
        "attempted_jar_paths": runtime_diag["jar_candidates"],
        "attempted_resource_dirs": runtime_diag["resource_candidates"],
        "jar_error": runtime_diag["jar_error"],
        "resource_error": runtime_diag["resource_error"],
        "detail": (proc.stderr or proc.stdout or "").strip(),
    }


@app.get("/api/runtime/build-info")
def runtime_build_info():
    return {
        "git_sha": _resolve_git_commit_sha(),
        "app_file": str(Path(__file__).resolve()),
        "cobieqc_runner_file": str(Path(cobieqc_runner_module.__file__).resolve()),
        "cobieqc_build_marker": getattr(cobieqc_runner_module, "COBIEQC_RUNNER_BUILD_MARKER", "unknown"),
        "cobieqc_flag_marker": getattr(cobieqc_runner_module, "COBIEQC_RUNNER_FLAG_MARKER", "unknown"),
        "timestamp": utc_now().isoformat() + "Z",
    }


@app.post("/api/tools/cobieqc/run")
async def cobieqc_run(
    background_tasks: BackgroundTasks,
    request: Request,
    file: UploadFile = File(...),
    stage: str = Form("D"),
):
    assert_heavy_capacity("/api/tools/cobieqc/run")
    runtime_diag = get_cobieqc_runtime_diagnostics()
    if not runtime_diag["enabled"]:
        raise HTTPException(
            status_code=503,
            detail="COBieQC runtime package is not installed or could not be restored from configured asset sources.",
        )

    stage = (stage or "D").upper()
    if stage not in {"D", "C"}:
        raise HTTPException(status_code=400, detail="Stage must be D or C")

    safe_name = sanitize_upload_filename(file.filename or "input.xlsx")
    job = COBIEQC_JOB_STORE.create_job(stage=stage, original_filename=safe_name)
    job_dir = COBIEQC_JOB_STORE.get_job_dir(job["job_id"])
    input_path = job_dir / "input.xlsx"
    written = 0
    with input_path.open("wb") as dst:
        while True:
            chunk = await file.read(1024 * 1024)
            if not chunk:
                break
            written += len(chunk)
            if written > MAX_UPLOAD_BYTES:
                input_path.unlink(missing_ok=True)
                content_length = request.headers.get("content-length")
                raise_upload_too_large(
                    endpoint="/api/tools/cobieqc/run",
                    filename=safe_name,
                    actual_size=written,
                    content_length=int(content_length) if content_length and content_length.isdigit() else None,
                    rejection_reason="streamed_upload_exceeded_limit",
                )
            dst.write(chunk)
    try:
        validate_upload(safe_name, written)
    except ValueError as exc:
        input_path.unlink(missing_ok=True)
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    enforce_upload_limits(str(input_path), endpoint="/api/tools/cobieqc/run")
    COBIEQC_JOB_STORE.append_log(job["job_id"], f"Saved input file {safe_name} ({written} bytes)")

    background_tasks.add_task(_run_cobieqc_job, job["job_id"])
    return {"job_id": job["job_id"], "status": "queued"}


@app.get("/api/tools/cobieqc/jobs/{job_id}")
def cobieqc_job_status(job_id: str):
    if not COBIEQC_JOB_STORE.job_exists(job_id):
        raise HTTPException(status_code=404, detail="Job not found")
    job = COBIEQC_JOB_STORE.read_job(job_id)
    return {
        "job_id": job["job_id"],
        "status": job.get("status"),
        "stage": job.get("stage"),
        "started_at": job.get("started_at"),
        "finished_at": job.get("finished_at"),
        "message": job.get("message"),
        "progress": job.get("progress", 0.0),
        "logs_tail": COBIEQC_JOB_STORE.logs_tail(job_id),
    }


@app.get("/api/tools/cobieqc/jobs/{job_id}/result")
def cobieqc_job_result(job_id: str):
    if not COBIEQC_JOB_STORE.job_exists(job_id):
        raise HTTPException(status_code=404, detail="Job not found")
    job = COBIEQC_JOB_STORE.read_job(job_id)
    if job.get("status") not in {STATUS_DONE, STATUS_ERROR}:
        raise HTTPException(status_code=409, detail="Job not finished")

    job_dir = COBIEQC_JOB_STORE.get_job_dir(job_id)
    output_path = job_dir / (job.get("output_filename") or "report.html")
    preview_html = ""
    if output_path.exists() and output_path.stat().st_size > 0:
        preview_html = output_path.read_text(encoding="utf-8", errors="replace")[:200000]

    logs = COBIEQC_JOB_STORE.logs_tail(job_id, max_chars=24000)
    return {
        "ok": job.get("status") == STATUS_DONE,
        "output_filename": output_path.name,
        "cobie_xml_filename": job.get("cobie_xml_filename"),
        "svrl_xml_filename": job.get("svrl_xml_filename"),
        "validation_summary": job.get("validation_summary", {}),
        "preview_html": preview_html,
        "stdout_tail": _tail_text(logs),
        "stderr_tail": _tail_text(logs),
        "message": job.get("message"),
    }


@app.get("/api/tools/cobieqc/jobs/{job_id}/download")
def cobieqc_job_download(job_id: str):
    if not COBIEQC_JOB_STORE.job_exists(job_id):
        raise HTTPException(status_code=404, detail="Job not found")
    job = COBIEQC_JOB_STORE.read_job(job_id)
    output_path = COBIEQC_JOB_STORE.get_job_dir(job_id) / (job.get("output_filename") or "report.html")
    if not output_path.exists():
        raise HTTPException(status_code=404, detail="Report not found")
    return FileResponse(
        path=str(output_path),
        media_type="text/html",
        filename=output_path.name,
    )


@app.post("/api/session/{session_id}/step2ifc/auto")
def run_step2ifc_auto(
    session_id: str,
    payload: Dict[str, Any] = Body(...),
    background_tasks: BackgroundTasks = None,
):
    root = SESSION_STORE.ensure(session_id)
    if not STEP2IFC_AVAILABLE:
        raise HTTPException(status_code=503, detail=STEP2IFC_IMPORT_ERROR or "step2ifc dependencies unavailable")
    input_name = payload.get("input_file")
    if not input_name:
        raise HTTPException(status_code=400, detail="input_file is required")
    input_path = Path(root) / sanitize_filename(input_name)
    if not input_path.exists():
        raise HTTPException(status_code=404, detail="Input STEP file not found")

    output_name = payload.get("output_name") or f"{input_path.stem}.ifc"
    output_name = sanitize_filename(output_name)
    if not output_name.lower().endswith(".ifc"):
        output_name = f"{output_name}.ifc"
    output_path = Path(root) / output_name

    mapping_name = payload.get("mapping_file")
    mapping_path = None
    if mapping_name:
        safe_mapping = sanitize_filename(mapping_name)
        mapping_path = Path(root) / safe_mapping
        if not mapping_path.exists():
            raise HTTPException(status_code=404, detail="Mapping file not found")

    job_id = uuid.uuid4().hex
    STEP2IFC_JOBS[job_id] = {
        "job_id": job_id,
        "status": "queued",
        "progress": 0,
        "message": "Queued",
        "done": False,
        "error": False,
        "outputs": [],
        "mapping_path": str(mapping_path) if mapping_path else None,
    }
    APP_LOGGER.info("STEP2IFC job queued", extra={"job_id": job_id, "input": input_path.name, "output": output_name})

    if background_tasks is None:
        background_tasks = BackgroundTasks()
    background_tasks.add_task(run_step2ifc_auto_job, job_id, session_id, input_path, output_path)
    return {"job_id": job_id, "status_url": f"/api/session/{session_id}/step2ifc/auto/{job_id}"}


@app.get("/api/session/{session_id}/step2ifc/auto/{job_id}")
def get_step2ifc_status(session_id: str, job_id: str):
    SESSION_STORE.ensure(session_id)
    job = STEP2IFC_JOBS.get(job_id)
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")
    return job


@app.post("/api/session/{session_id}/upload")
async def upload_files(session_id: str, files: List[UploadFile] = File(...)):
    endpoint_start = time.perf_counter()
    APP_LOGGER.info("upload_endpoint_start session_id=%s file_count=%s", session_id, len(files or []))
    normalized = _require_valid_session_id(session_id)
    root = _ensure_session_dir_for_upload(normalized)
    APP_LOGGER.info("session_upload_start session_id=%s file_count=%s root=%s", session_id, len(files or []), root)
    saved = []
    APP_LOGGER.info("stream_write_start session_id=%s root=%s", session_id, root)
    write_start = time.perf_counter()
    bytes_written_total = 0
    for f in files:
        safe = sanitize_filename(os.path.basename(f.filename))
        dest = os.path.join(root, safe)
        written = 0
        with open(dest, "wb") as dst:
            while True:
                chunk = await f.read(1024 * 1024)
                if not chunk:
                    break
                written += len(chunk)
                if written > MAX_UPLOAD_BYTES:
                    dst.close()
                    os.remove(dest)
                    APP_LOGGER.warning(
                        "session_upload_rejected session_id=%s file=%s bytes=%s reason=max_upload_exceeded",
                        session_id,
                        safe,
                        written,
                    )
                    raise_upload_too_large(
                        endpoint=f"/api/session/{session_id}/upload",
                        filename=safe,
                        actual_size=written,
                        rejection_reason="streamed_upload_exceeded_limit",
                    )
                dst.write(chunk)
        bytes_written_total += written
        enforce_upload_limits(dest, endpoint="/api/session/{session_id}/upload")
        saved.append(_build_session_file_metadata(root, safe))
    write_duration_ms = int((time.perf_counter() - write_start) * 1000)
    APP_LOGGER.info(
        "stream_write_complete session_id=%s duration_ms=%s bytes_written=%s files_saved=%s",
        session_id,
        write_duration_ms,
        bytes_written_total,
        len(saved),
    )
    APP_LOGGER.info("file_index_refresh_start session_id=%s root=%s", session_id, root)
    index_start = time.perf_counter()
    saved = [_build_session_file_metadata(root, item["name"]) for item in saved]
    index_duration_ms = int((time.perf_counter() - index_start) * 1000)
    APP_LOGGER.info("file_index_refresh_complete session_id=%s duration_ms=%s", session_id, index_duration_ms)
    APP_LOGGER.info("session_upload_complete session_id=%s saved=%s", session_id, len(saved))
    total_duration_ms = int((time.perf_counter() - endpoint_start) * 1000)
    APP_LOGGER.info("upload_response_sent session_id=%s duration_ms=%s", session_id, total_duration_ms)
    return {"ok": True, "session_id": normalized, "files": saved}


@app.post("/api/upload/session/{session_id}/upload")
async def upload_files_alias(session_id: str, files: List[UploadFile] = File(...)):
    return await upload_files(session_id, files)


def _public_ifc_qa_config(config: Dict[str, Any]) -> Dict[str, Any]:
    return {key: config.get(key) for key in REQUIRED_TOP_LEVEL_KEYS}


@app.post("/api/ifc-qa/run")
async def ifc_qa_run(
    files: List[UploadFile] = File(...),
    session_id: Optional[str] = Form(None),
    options_json: Optional[str] = Form(None),
    config_override_json: Optional[str] = Form(None),
):
    try:
        if has_active_ifc_qa_job():
            APP_LOGGER.warning("ifc_qa_run_rejected reason=active_job")
            raise HTTPException(
                status_code=429,
                detail="An IFC QA job is already running on this replica. Please retry shortly.",
            )
        if not files:
            APP_LOGGER.warning("ifc_qa_run_rejected reason=no_files")
            raise HTTPException(status_code=400, detail="At least one IFC file is required")

        upload_dir = Path(tempfile.mkdtemp(prefix="ifc_qa_uploads_v2_"))
        file_records: List[Tuple[str, str]] = []
        total_request_bytes = 0
        for upload in files:
            original_name = upload.filename or "upload.ifc"
            dest = upload_dir / sanitize_filename(original_name)
            with open(dest, "wb") as handle:
                written = 0
                while True:
                    chunk = await upload.read(1024 * 1024)
                    if not chunk:
                        break
                    written += len(chunk)
                    if written > MAX_UPLOAD_BYTES:
                        handle.close()
                        dest.unlink(missing_ok=True)
                        raise_upload_too_large(
                            endpoint="/api/ifc-qa/run",
                            filename=original_name,
                            actual_size=written,
                            rejection_reason="streamed_upload_exceeded_limit",
                        )
                    handle.write(chunk)
            enforce_upload_limits(str(dest), endpoint="/api/ifc-qa/run")
            file_size = dest.stat().st_size
            total_request_bytes += file_size
            file_records.append((original_name, str(dest)))

        options: Dict[str, Any] = {}
        if options_json:
            try:
                options = json.loads(options_json)
            except json.JSONDecodeError as exc:
                APP_LOGGER.warning("ifc_qa_run_rejected reason=invalid_options_json detail=%s", exc)
                raise HTTPException(status_code=400, detail="Invalid options_json payload") from exc
        options["max_workers"] = 1
        cfg = load_ifc_qa_default_config()
        if config_override_json:
            try:
                override = json.loads(config_override_json)
            except json.JSONDecodeError as exc:
                APP_LOGGER.warning("ifc_qa_run_rejected reason=invalid_config_override_json detail=%s", exc)
                raise HTTPException(status_code=400, detail="Invalid config_override_json payload") from exc
            if isinstance(override, dict):
                cfg = merge_config_override(cfg, override)
                cfg["_indexes"] = build_config_indexes(cfg)

        selected_outputs = options.get("selected_sheets", {}) if isinstance(options, dict) else {}
        selected_output_keys = [key for key, enabled in selected_outputs.items() if enabled]
        session_id = session_id or (options.get("session_id") if isinstance(options, dict) else None)
        if not session_id:
            detail = "Missing session_id for IFC QA run"
            APP_LOGGER.warning("ifc_qa_run_rejected reason=missing_session_id detail=%s", detail)
            raise HTTPException(status_code=400, detail=detail)
        options["session_id"] = session_id
        session_root = Path(SESSION_STORE.ensure(session_id)) / "ifc_qa_session"
        session_root.mkdir(parents=True, exist_ok=True)
        job_id = IFC_QA_JOB_STARTER(session_root, session_id, file_records, options, cfg, mode="replace")
        APP_LOGGER.info(
            "ifc_qa_run_started session_id=%s job_id=%s file_count=%s outputs=%s request_bytes=%s",
            session_id or "n/a",
            job_id,
            len(file_records),
            ",".join(selected_output_keys) if selected_output_keys else "none",
            total_request_bytes,
        )
        return {
            "success": True,
            "job_id": job_id,
            "session_id": session_id or "",
            "message": "Job started",
        }
    except HTTPException as exc:
        if exc.status_code == 413 and isinstance(exc.detail, dict) and exc.detail.get("code") == "UPLOAD_TOO_LARGE":
            return JSONResponse(status_code=413, content=exc.detail)
        return JSONResponse(
            status_code=exc.status_code,
            content={
                "success": False,
                "error": exc.detail if isinstance(exc.detail, str) else "IFC QA run request failed",
                "detail": exc.detail,
            },
        )
    except Exception as exc:  # pragma: no cover - defensive path
        APP_LOGGER.exception("ifc_qa_run_rejected reason=unexpected_error detail=%s", exc)
        return JSONResponse(
            status_code=500,
            content={
                "success": False,
                "error": "Failed to start IFC QA job",
                "detail": str(exc),
            },
        )


@app.post("/api/ifc-qa/add-to-zip")
async def ifc_qa_add_to_zip(
    files: List[UploadFile] = File(...),
    session_id: str = Form(...),
    options_json: Optional[str] = Form(None),
    config_override_json: Optional[str] = Form(None),
):
    if has_active_ifc_qa_job():
        raise HTTPException(status_code=429, detail="An IFC QA job is already running on this replica. Please retry shortly.")
    if not files:
        raise HTTPException(status_code=400, detail="At least one IFC file is required")
    upload_dir = Path(tempfile.mkdtemp(prefix="ifc_qa_uploads_v2_"))
    file_records: List[Tuple[str, str]] = []
    for upload in files:
        original_name = upload.filename or "upload.ifc"
        dest = upload_dir / sanitize_filename(original_name)
        with open(dest, "wb") as handle:
            written = 0
            while True:
                chunk = await upload.read(1024 * 1024)
                if not chunk:
                    break
                written += len(chunk)
                if written > MAX_UPLOAD_BYTES:
                    handle.close()
                    dest.unlink(missing_ok=True)
                    raise_upload_too_large(
                        endpoint="/api/ifc-qa/add-to-zip",
                        filename=original_name,
                        actual_size=written,
                        rejection_reason="streamed_upload_exceeded_limit",
                    )
                handle.write(chunk)
        enforce_upload_limits(str(dest), endpoint="/api/ifc-qa/add-to-zip")
        file_records.append((original_name, str(dest)))

    options: Dict[str, Any] = json.loads(options_json) if options_json else {}
    options["max_workers"] = 1
    cfg = load_ifc_qa_default_config()
    if config_override_json:
        override = json.loads(config_override_json)
        if isinstance(override, dict):
            cfg = merge_config_override(cfg, override)
            cfg["_indexes"] = build_config_indexes(cfg)
    session_root = Path(SESSION_STORE.ensure(session_id)) / "ifc_qa_session"
    session_root.mkdir(parents=True, exist_ok=True)
    job_id = start_ifc_qa_session_job(session_root, session_id, file_records, options, cfg, mode="append")
    return {"success": True, "job_id": job_id, "session_id": session_id, "files": [{"name": name} for name, _ in file_records]}


@app.get("/api/ifc-qa/status/{job_id}")
def ifc_qa_status(job_id: str):
    job = IFC_QA_V2_REGISTRY.get(job_id)
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")
    return {
        "status": job.get("status"),
        "percent": job.get("percent", 0),
        "overall_percent": job.get("overall_percent", job.get("percent", 0)),
        "currentStep": job.get("currentStep", ""),
        "currentFile": job.get("currentFile", ""),
        "logs": job.get("logs", []),
        "files": job.get("files", []),
        "per_file_percent": job.get("per_file_percent", {}),
        "per_file_stage": job.get("per_file_stage", {}),
        "processed_counts": job.get("processed_counts", {}),
        "manifest_summary": job.get("manifest_summary", {}),
        "session_id": job.get("session_id"),
    }


@app.get("/api/ifc-qa/session/{session_id}/summary")
def ifc_qa_session_summary(session_id: str):
    session_root = Path(SESSION_STORE.ensure(session_id)) / "ifc_qa_session"
    return read_ifc_qa_session_summary(session_root, session_id)


@app.get("/api/ifc-qa/result/{ref_id}")
def ifc_qa_result(ref_id: str):
    job = IFC_QA_V2_REGISTRY.get(ref_id)
    if job and job.get("result_path"):
        path = Path(job.get("result_path"))
        if path.exists():
            return FileResponse(path, media_type="application/zip", filename="IFC Output.zip")
    session_root = Path(SESSION_STORE.ensure(ref_id)) / "ifc_qa_session"
    zip_path = session_root / "IFC Output.zip"
    if not zip_path.exists():
        raise HTTPException(status_code=404, detail="Result not ready")
    return FileResponse(zip_path, media_type="application/zip", filename="IFC Output.zip")


@app.get("/api/ifc-qa/config")
def ifc_qa_config():
    return _public_ifc_qa_config(load_ifc_qa_default_config())


@app.get("/api/ifc-qa/build-info")
def ifc_qa_build_info():
    return {
        "git_sha": GIT_SHA,
        "build_timestamp_utc": BUILD_TIMESTAMP_UTC,
        "frontend_bundle_version": FRONTEND_BUNDLE_VERSION,
        "frontend_build_id": FRONTEND_BUILD_ID,
    }


@app.post("/api/ifc-qa/config/validate")
def ifc_qa_config_validate(config: Dict[str, Any] = Body(...)):
    errors = validate_config_structure(config)
    return {"valid": len(errors) == 0, "errors": errors}


@app.post("/api/ifc-qa/config/merge")
def ifc_qa_config_merge(override_config: Dict[str, Any] = Body(...)):
    default_config = load_ifc_qa_default_config()
    merged = merge_config_override(default_config, override_config)
    return _public_ifc_qa_config(merged)


@app.get("/api/ifc-qa/config/default")
def ifc_qa_default_config():
    return _public_ifc_qa_config(load_ifc_qa_default_config())


@app.get("/api/ifc-qa/config/{session_id}")
def ifc_qa_config_session(session_id: str):
    return {"session_id": session_id, "config": _public_ifc_qa_config(load_ifc_qa_default_config()), "overrides": {}}


@app.get("/api/ifc-qa/config/{session_id}/regex")
def ifc_qa_config_regex_compat(session_id: str):
    return {
        "regex": [],
        "message": "Regex extractor deprecated; using baked-in JSON config.",
    }


@app.post("/api/ifc-qa/config/import")
async def ifc_qa_config_import(config_json: UploadFile = File(...)):
    payload = json.loads((await config_json.read()).decode("utf-8"))
    return {"config": payload}
def _build_ifc_job_payload(session_id: str, payload: Dict[str, Any]) -> Dict[str, Any]:
    SESSION_STORE.cleanup_stale()
    normalized = _require_valid_session_id(session_id)
    if not SESSION_STORE.exists(normalized):
        raise HTTPException(
            status_code=404,
            detail={
                "error": "SESSION_NOT_FOUND",
                "message": "Session does not exist or has expired.",
                "session_id": normalized,
            },
        )
    root = Path(SESSION_STORE.ensure(normalized))
    ifc_files = payload.get("ifc_files") or payload.get("file_ids") or []
    tables = payload.get("tables") or []
    project_slug = (payload.get("project_slug") or payload.get("slug") or "").strip() or None
    if not project_slug and isinstance(payload.get("project"), dict):
        project_slug = str(payload["project"].get("slug") or "").strip() or None

    if not tables and project_slug:
        resolved_tables = get_tables_for_project_slug(project_slug)
        if resolved_tables:
            tables = resolved_tables

    session_records = []
    for fname in sorted(os.listdir(root)):
        fpath = root / fname
        if fpath.is_file():
            session_records.append(_build_session_file_metadata(str(root), fname))
    APP_LOGGER.info(
        "data_extractor_session_files session_id=%s files_found=%s",
        normalized,
        len(session_records),
    )
    if not session_records:
        raise HTTPException(
            status_code=400,
            detail={
                "error": "SESSION_EMPTY",
                "message": "Session contains no files. Upload IFC files first.",
                "session_id": normalized,
            },
        )

    available_ifc = {record["name"]: record for record in session_records if _is_ifc_compatible(record.get("name", ""))}
    if not available_ifc:
        raise HTTPException(
            status_code=400,
            detail={
                "error": "NO_IFC_FILES",
                "message": "No IFC-compatible files found in session (.ifc, .ifczip, .ifcxml).",
                "session_id": normalized,
            },
        )

    if not ifc_files:
        ifc_files = list(available_ifc.keys())
    APP_LOGGER.info("data_extractor_selected_files session_id=%s selected=%s", normalized, ifc_files)
    if not tables:
        hint = (
            f" No table mapping found for project slug '{project_slug}' in PROJECT_DATABASES_JSON."
            if project_slug
            else ""
        )
        raise HTTPException(status_code=400, detail=f"Select at least one table.{hint}")

    max_files = int(os.getenv("IFC_MAX_FILES_PER_JOB", "5"))
    if len(ifc_files) > max_files:
        raise HTTPException(status_code=400, detail=f"Max {max_files} IFC files per job")

    input_files = []
    total_bytes = 0
    skipped_files = []
    for name in ifc_files:
        safe = sanitize_filename(name)
        if not safe:
            skipped_files.append({"file": name, "reason": "empty_or_invalid_name"})
            continue
        if not _is_ifc_compatible(safe):
            skipped_files.append({"file": safe, "reason": "unsupported_extension"})
            continue
        record = available_ifc.get(safe)
        if record is None:
            raise HTTPException(
                status_code=400,
                detail={
                    "error": "SELECTED_FILE_NOT_IN_SESSION",
                    "message": f"Selected file '{safe}' is not present in the active session.",
                    "session_id": normalized,
                    "file": safe,
                },
            )
        candidate = (root / safe).resolve()
        if root.resolve() not in candidate.parents:
            raise HTTPException(status_code=400, detail=f"Invalid session file path: {safe}")
        size = int(record.get("size", 0))
        total_bytes += size
        input_files.append({"name": safe, "size": size})
        APP_LOGGER.info("data_extractor_resolved_path session_id=%s file=%s path=%s", normalized, safe, candidate)

    if skipped_files:
        APP_LOGGER.info("data_extractor_skipped_files session_id=%s skipped=%s", normalized, skipped_files)
    if not input_files:
        raise HTTPException(
            status_code=400,
            detail={
                "error": "NO_VALID_IFC_SELECTION",
                "message": "No valid IFC files were selected for extraction.",
                "session_id": normalized,
                "skipped": skipped_files,
            },
        )

    max_bytes = int(os.getenv("IFC_MAX_TOTAL_BYTES", "500000000"))
    if total_bytes > max_bytes:
        raise HTTPException(status_code=400, detail=f"Total upload size exceeds limit ({max_bytes} bytes)")

    exclude_filter_name = payload.get("exclude_filter")
    pset_template_name = payload.get("pset_template")
    pset_template_default = payload.get("pset_template_default") or "GPA_Pset_Template.csv"
    regex_overrides = payload.get("regex_overrides") or {}

    defaults = load_default_config()
    defaults.update({k: v for k, v in regex_overrides.items() if v is not None})

    exclude_path = None
    if exclude_filter_name:
        safe = sanitize_filename(exclude_filter_name)
        exclude_candidate = root / safe
        if exclude_candidate.exists():
            exclude_path = str(exclude_candidate)

    pset_path = None
    if pset_template_name:
        safe = sanitize_filename(pset_template_name)
        pset_candidate = root / safe
        if pset_candidate.exists():
            pset_path = str(pset_candidate)
    if not pset_path and pset_template_default:
        default_candidate = RESOURCE_DIR / sanitize_filename(pset_template_default)
        if default_candidate.exists():
            pset_path = str(default_candidate)

    return {
        "session_id": normalized,
        "ifc_files": [f["name"] for f in input_files],
        "input_files": input_files,
        "tables": tables,
        "project_slug": project_slug,
        "exclude_path": exclude_path,
        "pset_path": pset_path,
        "regexes": defaults,
    }


@app.post("/api/ifc/jobs")
def create_ifc_extraction_job(payload: Dict[str, Any] = Body(...), request: Request = None):
    session_id = payload.get("session_id")
    if not session_id:
        raise HTTPException(status_code=400, detail="session_id is required")
    job_payload = _build_ifc_job_payload(session_id, payload)
    user = None
    if request is not None:
        user = request.headers.get("x-user") or request.headers.get("x-user-email")
    job = create_ifc_job(requested_by=user, input_files=job_payload["input_files"], options={
        "session_id": session_id,
        "tables": job_payload["tables"],
        "project_slug": job_payload.get("project_slug"),
        "exclude_path": job_payload["exclude_path"],
        "pset_path": job_payload["pset_path"],
        "regexes": job_payload["regexes"],
    })
    return {"jobId": str(job["id"]), "status": job["status"], "progress": job["progress"], "message": job["message"]}


@app.get("/api/ifc/jobs/{job_id}")
def get_ifc_extraction_job(job_id: str):
    job = get_ifc_job(job_id)
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")
    return {
        "jobId": str(job["id"]),
        "status": job["status"],
        "progress": job.get("progress", 0),
        "message": job.get("message") or "",
        "error": job.get("error"),
    }


@app.get("/api/ifc/jobs/{job_id}/result")
def get_ifc_extraction_result(job_id: str):
    job = get_ifc_job(job_id)
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")
    if job.get("status") != "done":
        raise HTTPException(status_code=409, detail="Result not ready")
    return job.get("result") or {}


@app.post("/api/ifc/jobs/{job_id}/cancel")
def cancel_ifc_extraction_job(job_id: str):
    job = get_ifc_job(job_id)
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")
    if job.get("status") in {"done", "failed", "canceled"}:
        return {"jobId": str(job["id"]), "status": job.get("status")}
    update_ifc_job(job_id, status="canceled", message="Canceled by user")
    return {"jobId": str(job["id"]), "status": "canceled"}


@app.post("/api/session/{session_id}/data-extractor/start")
def start_data_extractor(session_id: str, payload: Dict[str, Any] = Body(...)):
    payload = dict(payload)
    payload["session_id"] = session_id
    job = create_ifc_extraction_job(payload)
    job_id = job["jobId"]
    return {
        "job_id": job_id,
        "jobId": job_id,
        "status": job["status"],
        "progress": job["progress"],
        "message": job["message"],
        "status_url": f"/api/ifc/jobs/{job_id}",
        "result_url": f"/api/ifc/jobs/{job_id}/result",
        "deprecated": True,
        "deprecation_note": "Use POST /api/ifc/jobs and GET /api/ifc/jobs/{id}",
    }


@app.get("/api/session/{session_id}/data-extractor/{job_id}")
def get_data_extractor_status(session_id: str, job_id: str):
    SESSION_STORE.ensure(session_id)
    job = get_ifc_job(job_id)
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")
    result = job.get("result") or {}
    return {
        "job_id": str(job["id"]),
        "status": job["status"],
        "progress": job.get("progress", 0),
        "message": job.get("message") or "",
        "done": job.get("status") in {"done", "failed", "canceled"},
        "error": bool(job.get("error")) or job.get("status") == "failed",
        "outputs": result.get("outputs", []),
        "preview": result.get("preview"),
        "logs": [],
    }


@app.post("/api/extract")
def api_extract_compat(payload: Dict[str, Any] = Body(...)):
    job = create_ifc_extraction_job(payload)
    job["deprecated"] = True
    job["deprecation_note"] = "Use /api/ifc/jobs"
    return job


@app.get("/api/session/{session_id}/download")
def download_file(session_id: str, name: str):
    root = SESSION_STORE.ensure(session_id)
    safe = sanitize_filename(os.path.basename(name))
    path = os.path.join(root, safe)
    if not os.path.isfile(path):
        raise HTTPException(status_code=404, detail="File not found")
    return FileResponse(path, filename=safe)


@app.post("/api/session/{session_id}/ifc-move-rotate")
async def api_ifc_move_rotate(
    session_id: str,
    source_file: Optional[str] = Form(None),
    upload_file: Optional[UploadFile] = File(None),
    current_x: float = Form(...),
    current_y: float = Form(...),
    current_z: float = Form(...),
    target_x: float = Form(...),
    target_y: float = Form(...),
    target_z: float = Form(...),
    rotation_deg: float = Form(...),
    tolerance: float = Form(0.001),
    rotate_about_global_z: bool = Form(True),
    preserve_metadata: bool = Form(True),
    output_suffix: str = Form("_moved_rotated"),
):
    root = Path(SESSION_STORE.ensure(session_id))
    if tolerance <= 0:
        raise HTTPException(status_code=400, detail="Tolerance must be > 0")

    input_path: Optional[Path] = None
    source_name: Optional[str] = None
    if upload_file and upload_file.filename:
        source_name = sanitize_filename(os.path.basename(upload_file.filename))
        input_path = root / source_name
        input_path.write_bytes(await upload_file.read())
    elif source_file:
        source_name = sanitize_filename(os.path.basename(source_file))
        candidate = root / source_name
        if not candidate.exists():
            raise HTTPException(status_code=404, detail="Selected source file not found in session")
        input_path = candidate
    else:
        raise HTTPException(status_code=400, detail="Provide either a session source file or upload_file")

    if not source_name.lower().endswith((".ifc", ".ifczip")):
        raise HTTPException(status_code=400, detail="Input must be an IFC file (.ifc or .ifczip)")

    out_stem, out_ext = os.path.splitext(source_name)
    cleaned_suffix = re.sub(r"[^A-Za-z0-9_-]", "_", output_suffix or "_moved_rotated")
    output_name = f"{out_stem}{cleaned_suffix}{out_ext}"
    if (root / output_name).exists():
        output_name = f"{out_stem}{cleaned_suffix}_{int(time.time())}{out_ext}"

    req = TransformRequest(
        current_xyz=(current_x, current_y, current_z),
        target_xyz=(target_x, target_y, target_z),
        rotation_deg=rotation_deg,
        tolerance=tolerance,
        output_suffix=cleaned_suffix,
        rotate_about_global_z=rotate_about_global_z,
        preserve_metadata=preserve_metadata,
    )

    try:
        summary = transform_ifc_file(str(input_path), str(root / output_name), req, APP_LOGGER)
    except Exception as exc:
        APP_LOGGER.exception("IFC Move/Rotate failed")
        raise HTTPException(status_code=500, detail=f"IFC Move/Rotate failed: {exc}") from exc

    return {
        "status": "ok",
        "ifc": {"name": output_name, "url": f"/api/session/{session_id}/download?name={output_name}"},
        "summary": summary,
    }


@app.get("/api/checks/definitions")
def api_check_definitions():
    defs = load_definitions()
    return {"definitions": [serialize_definition(d) for d in defs], "sections": CHECK_CACHE.get("summary", {})}


@app.get("/api/checks/mappings")
def api_check_mappings():
    return {
        "mapping_path": str(MAPPINGS_PATH),
        "expression_path": str(EXPRESSIONS_PATH),
        "mapping": load_mapping_config(),
        "expressions": load_expression_config(),
    }


@app.post("/api/checks/mappings")
def api_save_mapping(payload: Dict[str, Any] = Body(...)):
    check_id = payload.get("check_id")
    mapping = payload.get("mapping")
    if not check_id or not mapping:
        raise HTTPException(status_code=400, detail="check_id and mapping are required")
    data = save_mapping_for_check(check_id, mapping)
    load_definitions()
    return {"status": "ok", "mapping": data}


@app.post("/api/checks/expressions")
def api_save_expression(payload: Dict[str, Any] = Body(...)):
    check_id = payload.get("check_id")
    expression = payload.get("expression", "")
    if not check_id:
        raise HTTPException(status_code=400, detail="check_id is required")
    data = save_expression_for_check(check_id, expression)
    load_definitions()
    return {"status": "ok", "expressions": data}


@app.post("/api/session/{session_id}/checks/data")
def api_checks_data(session_id: str, payload: Dict[str, Any] = Body(...)):
    root = SESSION_STORE.ensure(session_id)
    ifc_file = payload.get("ifc_file")
    if not ifc_file:
        raise HTTPException(status_code=400, detail="ifc_file is required")
    path = os.path.join(root, sanitize_filename(ifc_file))
    if not os.path.isfile(path):
        raise HTTPException(status_code=404, detail="IFC file not found")
    model = ifcopenshell.open(path)
    section = payload.get("section", "Spaces")
    riba = payload.get("riba_stage")
    ent_filter = payload.get("entity_filter")
    ent_filters = payload.get("entity_filters") or None
    table = build_table_data(model, section, riba, ent_filter, ent_filters)
    table["change_log"] = read_change_log(session_id)
    return table


@app.post("/api/session/{session_id}/checks/apply")
def api_checks_apply(session_id: str, payload: Dict[str, Any] = Body(...)):
    root = SESSION_STORE.ensure(session_id)
    ifc_file = payload.get("ifc_file")
    edits = payload.get("edits", [])
    if not ifc_file:
        raise HTTPException(status_code=400, detail="ifc_file is required")
    path = os.path.join(root, sanitize_filename(ifc_file))
    if not os.path.isfile(path):
        raise HTTPException(status_code=404, detail="IFC file not found")
    out_name, audits = apply_edits(session_id, path, edits)
    return {
        "ifc": {"name": out_name, "url": f"/api/session/{session_id}/download?name={out_name}"},
        "audit": audits,
    }


@app.get("/api/session/{session_id}/checks/log")
def api_checks_log(session_id: str):
    SESSION_STORE.ensure(session_id)
    return {"entries": read_change_log(session_id)}


@app.post("/api/session/{session_id}/clean")
def run_cleaner(session_id: str, payload: Dict[str, Any] = Body(...)):
    root = SESSION_STORE.ensure(session_id)
    files = payload.get("files", [])
    if not files:
        raise HTTPException(status_code=400, detail="No files provided")
    prefix = payload.get("prefix", "InfoDrainage")
    case_insensitive = bool(payload.get("case_insensitive", True))
    delete_psets_with_prefix = bool(payload.get("delete_psets_with_prefix", True))
    delete_properties_in_other_psets = bool(payload.get("delete_properties_in_other_psets", True))
    drop_empty_psets = bool(payload.get("drop_empty_psets", True))
    also_remove_loose_props = bool(payload.get("also_remove_loose_props", True))

    reports = []
    outputs = []
    ts = utc_now().strftime("%Y%m%d_%H%M%S")
    for fname in files:
        in_path = os.path.join(root, sanitize_filename(fname))
        if not os.path.isfile(in_path):
            raise HTTPException(status_code=404, detail=f"File not found: {fname}")
        base, ext = os.path.splitext(os.path.basename(in_path))
        out_name = f"{base}_cleaned_{ts}{ext or '.ifc'}"
        out_path = os.path.join(root, out_name)
        report = clean_ifc_file(
            in_path=in_path,
            out_path=out_path,
            prefix=prefix,
            case_insensitive=case_insensitive,
            delete_psets_with_prefix=delete_psets_with_prefix,
            delete_properties_in_other_psets=delete_properties_in_other_psets,
            drop_empty_psets=drop_empty_psets,
            also_remove_loose_props=also_remove_loose_props,
        )
        reports.append(report)
        outputs.append({"name": out_name, "url": f"/api/session/{session_id}/download?name={out_name}"})
    return {"reports": reports, "outputs": outputs}


@app.post("/api/session/{session_id}/excel/extract")
def excel_extract(session_id: str, payload: Dict[str, Any] = Body(...)):
    root = SESSION_STORE.ensure(session_id)
    source = payload.get("ifc_file")
    if not source:
        raise HTTPException(status_code=400, detail="No IFC file provided")
    in_path = os.path.join(root, sanitize_filename(source))
    if not os.path.isfile(in_path):
        raise HTTPException(status_code=404, detail="IFC file not found")
    base = os.path.splitext(os.path.basename(in_path))[0]
    out_name = f"{base}_extracted.xlsx"
    out_path = os.path.join(root, out_name)
    plan_payload = payload.get("plan") if isinstance(payload.get("plan"), dict) else None
    cache_key = f"{session_id}:{sanitize_filename(source)}"
    cached_preview = EXCEL_SCAN_CACHE.get(cache_key)
    if cached_preview is not None:
        plan_payload = dict(plan_payload or {})
        if "cobie_pairs" not in plan_payload and cached_preview.get("cobie_pairs"):
            plan_payload["cobie_pairs"] = [f"{item.get('pset')}.{item.get('property')}" for item in cached_preview.get("cobie_pairs", [])]
    with single_flight_heavy_job("/api/session/{session_id}/excel/extract"):
        result = extract_to_excel(in_path, out_path, plan_payload=plan_payload)
    APP_LOGGER.info("excel_extract timings_ms=%s counts=%s source=%s", result.get("timings_ms", {}), result.get("counts", {}), source)
    return {
        "excel": {"name": out_name, "url": f"/api/session/{session_id}/download?name={out_name}"},
        "timings_ms": result.get("timings_ms", {}),
        "counts": result.get("counts", {}),
        "schema_detected": result.get("schema_detected"),
        "schema_warning": result.get("schema_warning", ""),
    }


@app.post("/api/session/{session_id}/excel/scan")
def excel_scan(session_id: str, payload: Dict[str, Any] = Body(...)):
    root = SESSION_STORE.ensure(session_id)
    source = payload.get("ifc_file")
    if not source:
        raise HTTPException(status_code=400, detail="No IFC file provided")
    in_path = os.path.join(root, sanitize_filename(source))
    if not os.path.isfile(in_path):
        raise HTTPException(status_code=404, detail="IFC file not found")
    timer = StageTimer()
    with single_flight_heavy_job("/api/session/{session_id}/excel/scan"):
        preview = scan_model_for_excel_preview(in_path, timer=timer)
    EXCEL_SCAN_CACHE[f"{session_id}:{sanitize_filename(source)}"] = preview
    APP_LOGGER.info("excel_scan timings_ms=%s source=%s", preview.get("timings_ms", {}), source)
    return {"preview": preview}


@app.post("/api/session/{session_id}/excel/update")
def excel_apply(session_id: str, payload: Dict[str, Any] = Body(...)):
    started_at = time.monotonic()
    root = SESSION_STORE.ensure(session_id)
    ifc_name = payload.get("ifc_file")
    excel_name = payload.get("excel_file")
    if not ifc_name or not excel_name:
        raise HTTPException(status_code=400, detail="IFC and Excel files are required")
    in_path = os.path.join(root, sanitize_filename(ifc_name))
    xls_path = os.path.join(root, sanitize_filename(excel_name))
    if not os.path.isfile(in_path) or not os.path.isfile(xls_path):
        raise HTTPException(status_code=404, detail="Input file(s) not found")
    enforce_upload_limits(in_path, endpoint="/api/session/{session_id}/excel/update")
    enforce_upload_limits(xls_path, endpoint="/api/session/{session_id}/excel/update")
    base = os.path.splitext(os.path.basename(in_path))[0]
    out_name = f"{base}_updated.ifc"
    out_path = os.path.join(root, out_name)
    with single_flight_heavy_job("/api/session/{session_id}/excel/update"):
        try:
            update_ifc_from_excel(
                in_path,
                xls_path,
                out_path,
                update_mode=payload.get("update_mode", "update"),
                add_new=payload.get("add_new", "no"),
                session_id=session_id,
                endpoint="/api/session/{session_id}/excel/update",
            )
        except ValueError as exc:
            raise HTTPException(status_code=400, detail=str(exc)) from exc
        except TimeoutError as exc:
            raise HTTPException(status_code=408, detail=str(exc)) from exc
    log_memory_stage(stage="response complete", session_id=session_id, file_name=out_name, file_size=os.path.getsize(out_path), endpoint="/api/session/{session_id}/excel/update", started_at=started_at)
    return {"ifc": {"name": out_name, "url": f"/api/session/{session_id}/download?name={out_name}"}}


@app.post("/api/session/{session_id}/storeys/parse")
def parse_storeys(session_id: str, payload: Dict[str, Any] = Body(...)):
    root = SESSION_STORE.ensure(session_id)
    src = payload.get("ifc_file")
    if not src:
        raise HTTPException(status_code=400, detail="No IFC file provided")
    in_path = os.path.join(root, sanitize_filename(src))
    if not os.path.isfile(in_path):
        raise HTTPException(status_code=404, detail="IFC file not found")
    return parse_ifc_storeys(in_path)


@app.post("/api/session/{session_id}/storeys/apply")
def apply_storeys(session_id: str, payload: Dict[str, Any] = Body(...)):
    root = SESSION_STORE.ensure(session_id)
    src = payload.get("ifc_file")
    if not src:
        raise HTTPException(status_code=400, detail="No IFC file provided")
    in_path = os.path.join(root, sanitize_filename(src))
    if not os.path.isfile(in_path):
        raise HTTPException(status_code=404, detail="IFC file not found")
    storey_id = payload.get("storey_id")
    if storey_id is None:
        raise HTTPException(status_code=400, detail="storey_id is required")
    base = os.path.splitext(os.path.basename(in_path))[0]
    out_name = f"{base}_gsb_adjusted.ifc"
    out_path = os.path.join(root, out_name)
    try:
        summary, path = apply_storey_changes(
            ifc_path=in_path,
            storey_id=storey_id,
            units_code=payload.get("units", "m"),
            gross_val=payload.get("gross"),
            net_val=payload.get("net"),
            mom_txt=payload.get("mom"),
            mirror=bool(payload.get("mirror", False)),
            target_z=payload.get("target_z"),
            countershift_geometry=bool(payload.get("countershift_geometry", True)),
            use_crs_mode=bool(payload.get("use_crs_mode", True)),
            update_all_mcs=bool(payload.get("update_all_mcs", True)),
            show_diag=bool(payload.get("show_diag", True)),
            crs_set_storey_elev=bool(payload.get("crs_set_storey_elev", True)),
            output_path=out_path,
        )
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))
    return {
        "summary": summary,
        "ifc": {"name": out_name, "url": f"/api/session/{session_id}/download?name={out_name}"},
    }


@app.post("/api/session/{session_id}/proxy")
def proxy_mapper(session_id: str, payload: Dict[str, Any] = Body(...)):
    root = SESSION_STORE.ensure(session_id)
    src = payload.get("ifc_file")
    if not src:
        raise HTTPException(status_code=400, detail="No IFC file provided")
    in_path = os.path.join(root, sanitize_filename(src))
    if not os.path.isfile(in_path):
        raise HTTPException(status_code=404, detail="IFC file not found")
    base, ext = os.path.splitext(os.path.basename(in_path))
    out_name = f"{base}_typed{ext or '.ifc'}"
    out_path = os.path.join(root, out_name)
    _, summary = rewrite_proxy_types(in_path, out_path)
    return {
        "summary": summary,
        "ifc": {"name": out_name, "url": f"/api/session/{session_id}/download?name={out_name}"},
    }


@app.post("/api/session/{session_id}/proxy/predefined/classes")
def proxy_predefined_classes(session_id: str, payload: Dict[str, Any] = Body(...)):
    root = SESSION_STORE.ensure(session_id)
    src = payload.get("ifc_file")
    if not src:
        raise HTTPException(status_code=400, detail="No IFC file provided")
    in_path = os.path.join(root, sanitize_filename(src))
    if not os.path.isfile(in_path):
        raise HTTPException(status_code=404, detail="IFC file not found")
    classes = list_instance_classes(in_path)
    return {"classes": classes}


@app.post("/api/session/{session_id}/proxy/predefined/scan")
def proxy_predefined_scan(session_id: str, payload: Dict[str, Any] = Body(...)):
    root = SESSION_STORE.ensure(session_id)
    src = payload.get("ifc_file")
    if not src:
        raise HTTPException(status_code=400, detail="No IFC file provided")
    in_path = os.path.join(root, sanitize_filename(src))
    if not os.path.isfile(in_path):
        raise HTTPException(status_code=404, detail="IFC file not found")
    class_filter = payload.get("classes") or []
    stats, rows = scan_predefined_types(in_path, class_filter=class_filter)
    return {"stats": stats, "rows": rows}


@app.post("/api/session/{session_id}/proxy/predefined/apply")
def proxy_predefined_apply(session_id: str, payload: Dict[str, Any] = Body(...)):
    root = SESSION_STORE.ensure(session_id)
    src = payload.get("ifc_file")
    rows = payload.get("rows", [])
    if not src:
        raise HTTPException(status_code=400, detail="No IFC file provided")
    in_path = os.path.join(root, sanitize_filename(src))
    if not os.path.isfile(in_path):
        raise HTTPException(status_code=404, detail="IFC file not found")
    out_path, json_path, csv_path = apply_predefined_type_changes(in_path, rows)
    return {
        "ifc": {"name": os.path.basename(out_path), "url": f"/api/session/{session_id}/download?name={os.path.basename(out_path)}"},
        "log_json": {"name": os.path.basename(json_path), "url": f"/api/session/{session_id}/download?name={os.path.basename(json_path)}"},
        "log_csv": {"name": os.path.basename(csv_path), "url": f"/api/session/{session_id}/download?name={os.path.basename(csv_path)}"},
    }


@app.get("/api/presentation-layers/template")
def presentation_layers_template():
    if not DEFAULT_ALLOWED_LAYERS_CSV.exists():
        raise HTTPException(status_code=404, detail="Template CSV not found")
    return FileResponse(str(DEFAULT_ALLOWED_LAYERS_CSV), media_type="text/csv", filename="default_allowed_layers.csv")


@app.post("/api/presentation-layers/allowed-layers/parse")
def parse_allowed_layers_api(payload: Dict[str, Any] = Body(...)):
    csv_text = payload.get("csv_text") or ""
    use_uploaded_only = bool(payload.get("use_uploaded_only", False))
    parsed = build_allowed_layers(csv_text, use_uploaded_only=use_uploaded_only)
    return {
        "allowed_layers": parsed["rows"],
        "allowed_full_values": parsed["full_values"],
        "errors": parsed["errors"],
        "count": len(parsed["rows"]),
    }


@app.post("/api/session/{session_id}/presentation-layer/extract")
def presentation_layer_extract(session_id: str, payload: Dict[str, Any] = Body(...)):
    root = SESSION_STORE.ensure(session_id)
    APP_LOGGER.info("Presentation layer extraction requested for session %s", session_id)
    src = payload.get("ifc_file")
    if not src:
        raise HTTPException(status_code=400, detail="No IFC file provided")
    in_path = os.path.join(root, sanitize_filename(src))
    if not os.path.isfile(in_path):
        raise HTTPException(status_code=404, detail="IFC file not found")
    csv_text = payload.get("allowed_csv_text") or ""
    use_uploaded_only = bool(payload.get("use_uploaded_only", False))
    threshold = float(payload.get("confidence_threshold", 0.7))
    allowed = build_allowed_layers(csv_text, use_uploaded_only=use_uploaded_only)
    review = build_layer_review(in_path, allowed["full_values"], confidence_threshold=threshold)
    APP_LOGGER.info(
        "Presentation layer extraction complete for %s: mode=%s rows=%s",
        src,
        review.get("summary", {}).get("source_mode"),
        len(review.get("rows", [])),
    )
    return {
        "rows": review["rows"],
        "summary": review["summary"],
        "allowed_layers": allowed["rows"],
        "allowed_full_values": allowed["full_values"],
        "allowed_errors": allowed["errors"],
    }


@app.post("/api/session/{session_id}/presentation-layer/purge/apply")
def presentation_layer_apply(session_id: str, payload: Dict[str, Any] = Body(...)):
    root = SESSION_STORE.ensure(session_id)
    src = payload.get("ifc_file")
    rows = payload.get("rows", [])
    options = payload.get("options") or {}
    if not src:
        raise HTTPException(status_code=400, detail="No IFC file provided")
    in_path = os.path.join(root, sanitize_filename(src))
    if not os.path.isfile(in_path):
        raise HTTPException(status_code=404, detail="IFC file not found")
    out_path, json_path, csv_path, summary = apply_layer_changes(in_path, rows, options)
    return {
        "summary": summary,
        "ifc": {"name": os.path.basename(out_path), "url": f"/api/session/{session_id}/download?name={os.path.basename(out_path)}"},
        "log_json": {"name": os.path.basename(json_path), "url": f"/api/session/{session_id}/download?name={os.path.basename(json_path)}"},
        "log_csv": {"name": os.path.basename(csv_path), "url": f"/api/session/{session_id}/download?name={os.path.basename(csv_path)}"},
    }


@app.post("/api/session/{session_id}/levels/list")
def api_levels_list(session_id: str, payload: Dict[str, Any] = Body(...)):
    root = SESSION_STORE.ensure(session_id)
    src = payload.get("ifc_file")
    if not src:
        raise HTTPException(status_code=400, detail="No IFC file provided")
    in_path = os.path.join(root, sanitize_filename(src))
    if not os.path.isfile(in_path):
        raise HTTPException(status_code=404, detail="IFC file not found")
    return list_levels(in_path)


@app.post("/api/session/{session_id}/levels/update")
def api_levels_update(session_id: str, payload: Dict[str, Any] = Body(...)):
    root = SESSION_STORE.ensure(session_id)
    src = payload.get("ifc_file")
    storey_id = payload.get("storey_id")
    if not src or storey_id is None:
        raise HTTPException(status_code=400, detail="ifc_file and storey_id are required")
    in_path = os.path.join(root, sanitize_filename(src))
    if not os.path.isfile(in_path):
        raise HTTPException(status_code=404, detail="IFC file not found")
    base = os.path.splitext(os.path.basename(in_path))[0]
    out_name = f"{base}_levels.ifc"
    out_path = os.path.join(root, out_name)
    try:
        update_level(in_path, int(storey_id), payload, out_path)
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))
    return {"ifc": {"name": out_name, "url": f"/api/session/{session_id}/download?name={out_name}"}}


@app.post("/api/session/{session_id}/levels/delete")
def api_levels_delete(session_id: str, payload: Dict[str, Any] = Body(...)):
    root = SESSION_STORE.ensure(session_id)
    src = payload.get("ifc_file")
    storey_id = payload.get("storey_id")
    target_id = payload.get("target_storey_id")
    object_ids = payload.get("object_ids")
    if not src or storey_id is None or target_id is None:
        raise HTTPException(status_code=400, detail="ifc_file, storey_id, and target_storey_id are required")
    in_path = os.path.join(root, sanitize_filename(src))
    if not os.path.isfile(in_path):
        raise HTTPException(status_code=404, detail="IFC file not found")
    base = os.path.splitext(os.path.basename(in_path))[0]
    out_name = f"{base}_levels.ifc"
    out_path = os.path.join(root, out_name)
    try:
        delete_level(in_path, int(storey_id), int(target_id), object_ids, out_path)
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))
    return {"ifc": {"name": out_name, "url": f"/api/session/{session_id}/download?name={out_name}"}}


@app.post("/api/session/{session_id}/levels/add")
def api_levels_add(session_id: str, payload: Dict[str, Any] = Body(...)):
    root = SESSION_STORE.ensure(session_id)
    src = payload.get("ifc_file")
    name = payload.get("name")
    if not src or not name:
        raise HTTPException(status_code=400, detail="ifc_file and name are required")
    in_path = os.path.join(root, sanitize_filename(src))
    if not os.path.isfile(in_path):
        raise HTTPException(status_code=404, detail="IFC file not found")
    base = os.path.splitext(os.path.basename(in_path))[0]
    out_name = f"{base}_levels.ifc"
    out_path = os.path.join(root, out_name)
    try:
        add_level(
            ifc_path=in_path,
            name=name,
            description=payload.get("description"),
            elevation=payload.get("elevation"),
            comp_height=payload.get("comp_height"),
            object_ids=payload.get("object_ids"),
            output_path=out_path,
        )
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))
    return {"ifc": {"name": out_name, "url": f"/api/session/{session_id}/download?name={out_name}"}}


@app.post("/api/session/{session_id}/levels/reassign")
def api_levels_reassign(session_id: str, payload: Dict[str, Any] = Body(...)):
    root = SESSION_STORE.ensure(session_id)
    src = payload.get("ifc_file")
    source_id = payload.get("source_storey_id")
    target_id = payload.get("target_storey_id")
    object_ids = payload.get("object_ids")
    if not src or source_id is None or target_id is None:
        raise HTTPException(status_code=400, detail="ifc_file, source_storey_id, and target_storey_id are required")
    in_path = os.path.join(root, sanitize_filename(src))
    if not os.path.isfile(in_path):
        raise HTTPException(status_code=404, detail="IFC file not found")
    base = os.path.splitext(os.path.basename(in_path))[0]
    out_name = f"{base}_levels.ifc"
    out_path = os.path.join(root, out_name)
    try:
        reassign_objects(in_path, int(source_id), int(target_id), object_ids, out_path)
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))
    return {"ifc": {"name": out_name, "url": f"/api/session/{session_id}/download?name={out_name}"}}


@app.post("/api/session/{session_id}/levels/batch")
def api_levels_batch(session_id: str, payload: Dict[str, Any] = Body(...)):
    root = SESSION_STORE.ensure(session_id)
    src = payload.get("ifc_file")
    actions = payload.get("actions") or []
    if not src or not actions:
        raise HTTPException(status_code=400, detail="ifc_file and actions are required")
    in_path = os.path.join(root, sanitize_filename(src))
    if not os.path.isfile(in_path):
        raise HTTPException(status_code=404, detail="IFC file not found")
    base = os.path.splitext(os.path.basename(in_path))[0]
    work_path = os.path.join(root, f"{base}_levels_work.ifc")
    final_name = f"{base}_levels_batch.ifc"
    final_path = os.path.join(root, final_name)
    try:
        apply_level_actions(in_path, actions, work_path)
        shutil.copyfile(work_path, final_path)
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))
    return {"ifc": {"name": final_name, "url": f"/api/session/{session_id}/download?name={final_name}"}, "actions": actions}


if __name__ == "__main__":
    import uvicorn

    host, port = resolve_server_host_port()
    uvicorn.run(app, host=host, port=port)
